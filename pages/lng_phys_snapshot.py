"""LNG Physical Snapshot page."""

from __future__ import annotations

import datetime as dt
from io import BytesIO

import pandas as pd
from dash import Input, Output, State, callback, dcc, html
from dash.exceptions import PreventUpdate

from utils.ag_grid_tables import create_ag_grid_from_datatable
from utils.lng_phys_snapshot_data import (
    DISPLAY_YEARS,
    build_demand_matrix,
    build_provider_metadata,
    format_storage_records,
    get_demand_snapshot,
    get_storage_snapshot,
    next_storage_endpoints,
)


VISIBLE_DEMAND_COLUMNS = [
    "Country",
    "Provider",
    *(f"{year}E" for year in DISPLAY_YEARS),
]


def build_demand_grid(row_data=None):
    columns = [
        {
            "name": ["Market", "Country"],
            "id": "Country",
            "type": "text",
            "cellClass": "phys-snapshot-country-cell",
            "cellRenderer": {
                "function": "physSnapshotCountryGroupLabel(params)"
            },
            "cellClassRules": {
                "phys-snapshot-country-label-cell": {
                    "function": (
                        "physSnapshotCountryIsContinuation(params) === false"
                    )
                },
                "phys-snapshot-country-continuation-cell": {
                    "function": "physSnapshotCountryIsContinuation(params)"
                },
            },
        },
        {
            "name": ["Market", "Provider"],
            "id": "Provider",
            "type": "text",
            "cellClass": "phys-snapshot-provider-cell",
        },
    ]
    for year in DISPLAY_YEARS:
        column_id = f"{year}E"
        columns.append(
            {
                "name": ["Annual LNG Imports (MMT)", column_id],
                "id": column_id,
                "type": "numeric",
                "precision": 1,
                "tooltipValueGetter": {
                    "function": (
                        f"params.data ? "
                        f"params.data['__{column_id}_tooltip'] : ''"
                    )
                },
                "cellClassRules": {
                    "phys-snapshot-lto-cell": {
                        "function": (
                            f"params.data && "
                            f"params.data['__{column_id}_is_lto'] === true"
                        )
                    },
                    "phys-snapshot-missing-cell": {
                        "function": (
                            "params.value === null || "
                            "params.value === undefined || "
                            "params.value === '' || "
                            "params.value === '—'"
                        )
                    },
                },
            }
        )

    grid = create_ag_grid_from_datatable(
        id="phys-snapshot-demand-grid",
        columns=columns,
        data=row_data or [],
        sort_action="native",
        filter_action="none",
        page_action="none",
        fixed_columns={"headers": True, "data": 2},
        fill_width=False,
        className="phys-snapshot-demand-grid",
        height="744px",
        style_cell_conditional=[
            {
                "if": {"column_id": "Country"},
                "width": "135px",
                "minWidth": "135px",
                "maxWidth": "135px",
                "textAlign": "left",
            },
            {
                "if": {"column_id": "Provider"},
                "width": "150px",
                "minWidth": "150px",
                "maxWidth": "150px",
                "textAlign": "left",
            },
            *[
                {
                    "if": {"column_id": f"{year}E"},
                    "width": "112px",
                    "minWidth": "112px",
                    "maxWidth": "112px",
                    "textAlign": "right",
                }
                for year in DISPLAY_YEARS
            ],
        ],
        dashGridOptions={
            "tooltipShowDelay": 120,
            "tooltipHideDelay": 12000,
            "suppressPaginationPanel": True,
            "alwaysShowHorizontalScroll": False,
        },
        rowClassRules={
            "phys-snapshot-country-start-row": (
                "params.data && "
                "params.data.__country_group_start === true"
            ),
            "phys-snapshot-provider-ea-row": (
                "params.data && params.data.Provider === 'Energy Aspects'"
            ),
            "phys-snapshot-provider-woodmac-row": (
                "params.data && params.data.Provider === 'WoodMac'"
            ),
            "phys-snapshot-provider-platts-row": (
                "params.data && params.data.Provider === 'Platts'"
            ),
        },
    )
    numeric_fields = {f"{year}E" for year in DISPLAY_YEARS}
    for column_group in grid.columnDefs:
        for leaf_def in column_group.get("children", [column_group]):
            if leaf_def.get("field") not in numeric_fields:
                continue
            leaf_def["valueFormatter"] = {
                "function": "physSnapshotOneDecimal(params)"
            }
    grid.eventListeners = {
        "sortChanged": [
            (
                "params.api.refreshCells({"
                "columns: ['Country'], force: true"
                "})"
            )
        ]
    }
    return grid


def render_provider_metadata(metadata):
    return [
        html.Div(
            [
                html.Span(item["provider"], className="phys-snapshot-source-name"),
                html.Span(
                    item["vintage"],
                    className="phys-snapshot-source-vintage",
                    title=item["vintage"],
                ),
                html.Span(
                    f"Uploaded {item['upload']}",
                    className="phys-snapshot-source-upload",
                    title=f"Uploaded {item['upload']}",
                ),
            ],
            className="phys-snapshot-source-chip",
        )
        for item in metadata
    ]


def create_storage_card(record):
    storage_pct = record.get("storage_pct")
    storage_bcm = record.get("storage_bcm")
    pct_text = "—" if storage_pct is None else f"{storage_pct:,.1f}%"
    bcm_text = "—" if storage_bcm is None else f"{storage_bcm:,.1f} BCM"
    card_class = "phys-snapshot-storage-card"
    if record.get("stockout"):
        card_class += " phys-snapshot-storage-card-stockout"

    return html.Div(
        [
            html.Div(
                [
                    html.H4(
                        record["label"],
                        className="phys-snapshot-storage-card-title",
                    ),
                    html.Span(
                        record["date"].strftime("%d %b %Y"),
                        className="phys-snapshot-storage-card-date",
                    ),
                ],
                className="phys-snapshot-storage-card-header",
            ),
            html.Div(
                [
                    html.Div(
                        pct_text,
                        className="phys-snapshot-storage-card-value",
                    ),
                    html.Div(
                        bcm_text,
                        className="phys-snapshot-storage-card-bcm",
                    ),
                    (
                        html.Div(
                            "Modelled stockout",
                            className="phys-snapshot-stockout-label",
                        )
                        if record.get("stockout")
                        else None
                    ),
                ],
                className="phys-snapshot-storage-card-body",
            ),
        ],
        className=card_class,
    )


def build_demand_export_bytes(row_data) -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    export_frame = pd.DataFrame(row_data or [])
    if export_frame.empty:
        return b""

    for column_name in VISIBLE_DEMAND_COLUMNS:
        if column_name not in export_frame.columns:
            export_frame[column_name] = None
    export_frame = export_frame[VISIBLE_DEMAND_COLUMNS].copy()
    for column_name in VISIBLE_DEMAND_COLUMNS[2:]:
        export_frame[column_name] = pd.to_numeric(
            export_frame[column_name], errors="coerce"
        )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_frame.to_excel(
            writer,
            sheet_name="LNG Physical Snapshot",
            index=False,
        )
        worksheet = writer.sheets["LNG Physical Snapshot"]
        worksheet.freeze_panes = "C2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(
            min_row=2,
            min_col=3,
            max_col=len(VISIBLE_DEMAND_COLUMNS),
        ):
            for cell in row:
                cell.number_format = "#,##0.0"
                cell.alignment = Alignment(horizontal="right")

        widths = {
            "Country": 18,
            "Provider": 20,
            **{f"{year}E": 13 for year in DISPLAY_YEARS},
        }
        for index, column_name in enumerate(VISIBLE_DEMAND_COLUMNS, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = widths[
                column_name
            ]

    return output.getvalue()


layout = html.Div(
    [
        dcc.Download(id="phys-snapshot-demand-download"),
        html.Div(
            [
                html.Div(
                    "LNG Physical Snapshot",
                    className="phys-snapshot-filter-title",
                ),
                html.Div(
                    [
                        html.Div(
                            "EU Storage Scenario",
                            className="filter-group-header",
                        ),
                        dcc.RadioItems(
                            id="phys-snapshot-storage-scenario",
                            options=[
                                {
                                    "label": "Base case",
                                    "value": "base_case",
                                },
                                {
                                    "label": "Best view",
                                    "value": "best_view",
                                },
                            ],
                            value="best_view",
                            inline=True,
                            className=(
                                "supply-dest-view-selector "
                                "exporters-sticky-selector "
                                "phys-snapshot-scenario-selector"
                            ),
                            inputStyle={"display": "none"},
                            labelStyle={"marginRight": "0"},
                        ),
                    ],
                    className=(
                        "filter-group exporters-sticky-filter-group "
                        "phys-snapshot-scenario-filter"
                    ),
                ),
                html.Div(
                    id="phys-snapshot-last-refresh",
                    className="phys-snapshot-refresh-status",
                ),
            ],
            className=(
                "professional-section-header exporters-sticky-filter-bar "
                "phys-snapshot-filter-bar"
            ),
        ),
        html.Div(
            [
                html.Div(
                    [
                        html.Div(
                            [
                                html.H3(
                                    "EU Storage Outlook",
                                    className="section-title-inline",
                                ),
                            ],
                            className="supply-rolling-title-row",
                        ),
                        html.Div(
                            id="phys-snapshot-storage-source-metadata",
                            className="phys-snapshot-storage-metadata",
                        ),
                    ],
                    className=(
                        "inline-section-header supply-rolling-section-header "
                        "phys-snapshot-section-header "
                        "phys-snapshot-storage-section-header"
                    ),
                ),
                html.Div(
                    id="phys-snapshot-storage-warning",
                    className="phys-snapshot-warning",
                    role="alert",
                ),
                dcc.Loading(
                    id="phys-snapshot-storage-loading",
                    type="default",
                    children=[
                        html.Div(
                            id="phys-snapshot-storage-cards",
                            className="phys-snapshot-storage-grid",
                        )
                    ],
                ),
            ],
            className=(
                "main-section-container supply-rolling-section "
                "phys-snapshot-storage-section"
            ),
        ),
        html.Div(
            [
                html.Div(
                    [
                        html.Div(
                            [
                                html.H3(
                                    "Annual LNG Import Forecasts (MMT)",
                                    className="section-title-inline",
                                ),
                                html.Span(
                                    "2026–2030",
                                    className="phys-snapshot-period-chip",
                                ),
                                html.Div(
                                    id="phys-snapshot-demand-source-metadata",
                                    className="phys-snapshot-source-grid",
                                ),
                            ],
                            className=(
                                "supply-rolling-title-row "
                                "phys-snapshot-demand-header-content"
                            ),
                        ),
                        html.Button(
                            "Export to Excel",
                            id="phys-snapshot-export-button",
                            n_clicks=0,
                            className="supply-rolling-export-button",
                        ),
                    ],
                    className=(
                        "inline-section-header supply-rolling-section-header "
                        "phys-snapshot-section-header"
                    ),
                ),
                html.Div(
                    id="phys-snapshot-demand-warning",
                    className="phys-snapshot-warning",
                    role="alert",
                ),
                dcc.Loading(
                    id="phys-snapshot-demand-loading",
                    type="default",
                    children=[build_demand_grid()],
                ),
                html.Div(
                    [
                        html.Span(
                            "Reading guide",
                            className="phys-snapshot-note-label",
                        ),
                        html.Span(
                            "Subtle italic cells use WoodMac Long Term Outlook data. "
                            "A dash means a complete 12-month forecast is not available.",
                        ),
                    ],
                    className="phys-snapshot-table-note",
                ),
            ],
            className=(
                "main-section-container supply-rolling-section "
                "phys-snapshot-demand-section"
            ),
        ),
    ],
    className="lng-phys-snapshot-page",
)


@callback(
    Output("phys-snapshot-demand-grid", "rowData"),
    Output("phys-snapshot-demand-source-metadata", "children"),
    Output("phys-snapshot-demand-warning", "children"),
    Output("phys-snapshot-last-refresh", "children"),
    Input("global-refresh-button", "n_clicks"),
)
def update_demand_snapshot(global_refresh_clicks):
    annual_frame, warnings = get_demand_snapshot(global_refresh_clicks)
    refreshed_at = dt.datetime.now(dt.timezone.utc).strftime(
        "%d %b %Y %H:%M UTC"
    )
    return (
        build_demand_matrix(annual_frame),
        render_provider_metadata(build_provider_metadata(annual_frame)),
        " ".join(warnings),
        f"Updated {refreshed_at}",
    )


@callback(
    Output("phys-snapshot-storage-cards", "children"),
    Output("phys-snapshot-storage-source-metadata", "children"),
    Output("phys-snapshot-storage-warning", "children"),
    Input("phys-snapshot-storage-scenario", "value"),
    Input("global-refresh-button", "n_clicks"),
)
def update_storage_snapshot(scenario, global_refresh_clicks):
    selected_scenario = scenario or "best_view"
    endpoints = next_storage_endpoints()
    try:
        storage_frame = get_storage_snapshot(
            selected_scenario,
            endpoints,
            global_refresh_clicks,
        )
        warning = ""
    except Exception:
        storage_frame = pd.DataFrame()
        warning = "EU storage is unavailable."

    records = format_storage_records(
        storage_frame,
        endpoints,
        selected_scenario,
    )
    uploads = (
        pd.to_datetime(
            storage_frame.get("upload_timestamp_utc"),
            errors="coerce",
            utc=True,
        ).dropna()
        if not storage_frame.empty
        else pd.Series(dtype="datetime64[ns, UTC]")
    )
    upload_text = (
        uploads.max().strftime("%d %b %Y %H:%M UTC")
        if not uploads.empty
        else "Unknown upload"
    )
    scenario_label = selected_scenario.replace("_", " ").title()
    return (
        [create_storage_card(record) for record in records],
        f"EU · {scenario_label} · Uploaded {upload_text}",
        warning,
    )


@callback(
    Output("phys-snapshot-demand-download", "data"),
    Input("phys-snapshot-export-button", "n_clicks"),
    State("phys-snapshot-demand-grid", "virtualRowData"),
    State("phys-snapshot-demand-grid", "rowData"),
    prevent_initial_call=True,
)
def export_demand_snapshot(
    n_clicks,
    virtual_row_data,
    row_data,
):
    if not n_clicks:
        raise PreventUpdate
    export_rows = virtual_row_data or row_data or []
    workbook = build_demand_export_bytes(export_rows)
    if not workbook:
        raise PreventUpdate
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return dcc.send_bytes(
        workbook,
        f"LNG_Physical_Snapshot_{timestamp}.xlsx",
    )

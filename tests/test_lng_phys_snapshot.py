from io import BytesIO
from pathlib import Path
from threading import Barrier

import pandas as pd
from openpyxl import load_workbook

from pages import lng_phys_snapshot as page
from utils import lng_phys_snapshot_data as data
from utils.ag_grid_tables import iter_leaf_column_defs


REPO = Path(__file__).resolve().parents[1]


def _monthly_frame(
    provider,
    country_name="China",
    year=2026,
    months=12,
    monthly_mmt=1.0,
    release_type="Forecast",
):
    return pd.DataFrame(
        {
            "country_name": [country_name] * months,
            "provider": [provider] * months,
            "month": pd.date_range(
                f"{year}-01-01", periods=months, freq="MS"
            ),
            "monthly_mmt": [monthly_mmt] * months,
            "source_vintage": ["test vintage"] * months,
            "source_upload": [pd.Timestamp("2026-07-23")] * months,
            "release_type": [release_type] * months,
        }
    )


def _walk(component):
    yield component
    children = getattr(component, "children", None)
    if children is None:
        return
    if not isinstance(children, (list, tuple)):
        children = [children]
    for child in children:
        if hasattr(child, "to_plotly_json"):
            yield from _walk(child)


def test_annualization_requires_exactly_twelve_numeric_months():
    complete = _monthly_frame("Energy Aspects", months=12)
    incomplete = _monthly_frame(
        "Platts",
        country_name="Japan",
        months=11,
    )

    result = data.annualize_monthly_demand(
        pd.concat([complete, incomplete], ignore_index=True)
    )

    ea = result[result["provider"].eq("Energy Aspects")].iloc[0]
    platts = result[result["provider"].eq("Platts")].iloc[0]
    assert bool(ea["complete_year"])
    assert ea["annual_mmt"] == 12.0
    assert not bool(platts["complete_year"])
    assert pd.isna(platts["annual_mmt"])


def test_provider_failures_are_isolated(monkeypatch):
    monkeypatch.setattr(
        data,
        "fetch_ea_monthly",
        lambda _engine: _monthly_frame("Energy Aspects"),
    )
    monkeypatch.setattr(
        data,
        "fetch_woodmac_monthly",
        lambda _engine: (_ for _ in ()).throw(RuntimeError("unavailable")),
    )
    monkeypatch.setattr(
        data,
        "fetch_platts_monthly",
        lambda _engine: _monthly_frame("Platts"),
    )

    result, warnings = data.fetch_demand_snapshot(object())

    assert set(result["provider"]) == {"Energy Aspects", "Platts"}
    assert warnings == ["WoodMac demand is unavailable."]


def test_provider_loaders_overlap_with_exact_deterministic_output(monkeypatch):
    loader_barrier = Barrier(3, timeout=2)
    provider_frames = {
        "Energy Aspects": _monthly_frame(
            "Energy Aspects",
            country_name="China",
        ),
        "WoodMac": _monthly_frame(
            "WoodMac",
            country_name="Japan",
        ),
        "Platts": _monthly_frame(
            "Platts",
            country_name="South Korea",
        ),
    }

    def load(provider):
        loader_barrier.wait()
        return provider_frames[provider]

    monkeypatch.setattr(
        data,
        "fetch_ea_monthly",
        lambda _engine: load("Energy Aspects"),
    )
    monkeypatch.setattr(
        data,
        "fetch_woodmac_monthly",
        lambda _engine: load("WoodMac"),
    )
    monkeypatch.setattr(
        data,
        "fetch_platts_monthly",
        lambda _engine: load("Platts"),
    )

    actual, warnings = data.fetch_demand_snapshot(object())
    expected = data.annualize_monthly_demand(
        pd.concat(
            [
                provider_frames[provider]
                for provider in data.PROVIDER_ORDER
            ],
            ignore_index=True,
        )
    )

    pd.testing.assert_frame_equal(actual, expected, check_exact=True)
    assert warnings == []


def test_provider_warning_order_follows_provider_order(monkeypatch):
    monkeypatch.setattr(
        data,
        "fetch_ea_monthly",
        lambda _engine: (_ for _ in ()).throw(RuntimeError("ea failed")),
    )
    monkeypatch.setattr(
        data,
        "fetch_woodmac_monthly",
        lambda _engine: data.empty_monthly_frame(),
    )
    monkeypatch.setattr(
        data,
        "fetch_platts_monthly",
        lambda _engine: None,
    )

    result, warnings = data.fetch_demand_snapshot(object())

    assert result.empty
    assert result.columns.tolist() == list(data.DEMAND_COLUMNS)
    assert warnings == [
        "Energy Aspects demand is unavailable.",
        "WoodMac returned no mapped demand data.",
        "Platts returned no mapped demand data.",
    ]


def test_ea_query_is_pinned_to_latest_accepted_full_run(monkeypatch):
    captured = {}
    monkeypatch.setattr(
        data,
        "fetch_current_ea_run",
        lambda _engine, schema: {
            "run_id": 321,
            "snapshot_at": "2026-07-23T05:32:00Z",
        },
    )

    def fake_read_sql(query, _engine, params):
        captured["query"] = str(query)
        captured["params"] = params
        return data.empty_monthly_frame()

    monkeypatch.setattr(data, "_read_sql", fake_read_sql)

    data.fetch_ea_monthly(object())

    assert captured["params"]["ea_as_of_run_id"] == 321
    assert captured["params"]["ea_source_vintage"] == "2026-07-23T05:32:00Z"
    assert "ea_values_at_run" in captured["query"]
    assert "country_mapping.country = datasets.country" in captured["query"]


def test_provider_queries_use_exact_mapping_and_source_precedence():
    source = (REPO / "utils" / "lng_phys_snapshot_data.py").read_text()
    woodmac_query = str(data.WOODMAC_DEMAND_QUERY)
    platts_query = str(data.PLATTS_DEMAND_QUERY)

    assert source.count("JOIN {DB_SCHEMA}.mappings_country") == 3
    assert "country_mapping.country = datasets.country" in str(
        data.EA_DEMAND_QUERY
    )
    assert "country_mapping.country = source.country_name" in woodmac_query
    assert "country_mapping.country = source.country_or_market" in platts_query
    assert "ROW_NUMBER() OVER" in woodmac_query
    assert "month > short_term_horizon.final_short_term_month" in woodmac_query
    assert "SUM(source.metric_value)::double precision / 12.0" in woodmac_query
    assert "latest_header" in platts_query
    assert "lnga_short_term_demand" in platts_query
    assert "Turkiye" not in source
    assert "China (Mainland)" not in source


def test_platts_query_uses_one_dataset_only_latest_header():
    platts_query = str(data.PLATTS_DEMAND_QUERY)
    header_sql, data_sql = platts_query.split(
        "SELECT\n        country_mapping.country_name",
        maxsplit=1,
    )
    normalized_header = " ".join(header_sql.split())

    assert "WITH latest_header AS MATERIALIZED" in header_sql
    assert (
        "WHERE dataset_key = 'lnga_short_term_demand' "
        "ORDER BY upload_timestamp_utc DESC NULLS LAST, "
        "vintage_date DESC NULLS LAST LIMIT 1"
    ) in normalized_header
    assert "metric =" not in header_sql
    assert "flow_type =" not in header_sql
    assert "unit =" not in header_sql
    assert "latest_upload" not in platts_query
    assert "latest_vintage" not in platts_query
    assert "CROSS JOIN latest_header" in platts_query
    assert "source.metric = 'lng_demand_forecast'" in data_sql
    assert "source.flow_type = 'demand'" in data_sql
    assert "source.unit = 'MMt'" in data_sql
    assert "GROUP BY country_mapping.country_name, source.period_start::date" in data_sql
    assert "ORDER BY country_mapping.country_name, month" in data_sql


def test_platts_loader_preserves_exact_monthly_output(monkeypatch):
    expected = _monthly_frame(
        "Platts",
        country_name="Japan",
        year=2027,
        months=2,
        monthly_mmt=1.25,
        release_type="Short Term Outlook",
    )
    captured = {}

    def fake_read_sql(query, db_engine, params):
        captured["query"] = query
        captured["engine"] = db_engine
        captured["params"] = params
        return expected

    db_engine = object()
    monkeypatch.setattr(data, "_read_sql", fake_read_sql)

    actual = data.fetch_platts_monthly(db_engine)

    pd.testing.assert_frame_equal(actual, expected, check_exact=True)
    assert captured == {
        "query": data.PLATTS_DEMAND_QUERY,
        "engine": db_engine,
        "params": {"countries": list(data.DISPLAY_COUNTRIES)},
    }


def test_demand_matrix_has_21_rows_blanks_lto_and_tooltips():
    annual = data.annualize_monthly_demand(
        pd.concat(
            [
                _monthly_frame("Energy Aspects"),
                _monthly_frame(
                    "WoodMac",
                    year=2029,
                    release_type="Long Term Outlook",
                ),
                _monthly_frame("Platts", year=2030),
            ],
            ignore_index=True,
        )
    )

    rows = data.build_demand_matrix(annual)

    assert len(rows) == 21
    assert rows[0]["Country"] == "China"
    assert rows[0]["Provider"] == "Energy Aspects"
    assert rows[0]["2026E"] == 12.0
    assert rows[0]["2030E"] is None
    assert rows[1]["__2029E_is_lto"] is True
    assert "Long Term Outlook" in rows[1]["__2029E_tooltip"]
    assert sum(bool(row["__country_group_start"]) for row in rows) == 7


def test_demand_cache_is_reused_and_global_refresh_invalidates(monkeypatch):
    calls = []
    data.clear_snapshot_caches()
    monkeypatch.setattr(data, "_cache_bucket", lambda: 10)
    monkeypatch.setattr(
        data,
        "fetch_demand_snapshot",
        lambda _engine: (
            calls.append("load")
            or pd.DataFrame(columns=data.DEMAND_COLUMNS),
            [],
        ),
    )

    data.get_demand_snapshot(0)
    data.get_demand_snapshot(0)
    data.get_demand_snapshot(1)

    assert calls == ["load", "load"]
    data.clear_snapshot_caches()


def test_next_storage_endpoints_and_negative_stockout():
    endpoints = data.next_storage_endpoints("2026-07-23")
    assert [item["date"].isoformat() for item in endpoints] == [
        "2026-10-31",
        "2027-03-31",
        "2027-10-31",
        "2028-03-31",
    ]

    frame = pd.DataFrame(
        {
            "date": [endpoints[-1]["date"]],
            "storage_pct": [-1.3],
            "storage_bcm": [-1.1],
            "upload_timestamp_utc": [pd.Timestamp("2026-07-23")],
        }
    )
    records = data.format_storage_records(frame, endpoints, "base_case")

    assert records[-1]["storage_pct"] == -1.3
    assert records[-1]["storage_bcm"] == -1.1
    assert records[-1]["stockout"]
    card = page.create_storage_card(records[-1])
    assert "phys-snapshot-storage-card-stockout" in card.className
    assert "Modelled stockout" in str(card)
    standard_card = page.create_storage_card(records[0])
    assert "Modelled endpoint" not in str(standard_card)


def test_storage_cache_separates_scenarios_and_refresh(monkeypatch):
    calls = []
    endpoints = data.next_storage_endpoints("2026-07-23")
    data.clear_snapshot_caches()
    monkeypatch.setattr(data, "_cache_bucket", lambda: 10)

    def fake_fetch(scenario, target_dates, _engine):
        calls.append((scenario, target_dates))
        return pd.DataFrame()

    monkeypatch.setattr(data, "fetch_storage_snapshot", fake_fetch)

    data.get_storage_snapshot("base_case", endpoints, 0)
    data.get_storage_snapshot("base_case", endpoints, 0)
    data.get_storage_snapshot("best_view", endpoints, 0)
    data.get_storage_snapshot("base_case", endpoints, 1)

    assert [call[0] for call in calls] == [
        "base_case",
        "best_view",
        "base_case",
    ]
    data.clear_snapshot_caches()


def test_ag_grid_contract_and_export_workbook():
    rows = data.build_demand_matrix(pd.DataFrame(columns=data.DEMAND_COLUMNS))
    rows[0]["2026E"] = 12.34
    grid = page.build_demand_grid(rows)
    leaf_columns = iter_leaf_column_defs(grid.columnDefs)

    assert len(grid.rowData) == 21
    assert grid.dashGridOptions["pagination"] is False
    assert grid.style["height"] == "744px"
    assert [column["field"] for column in leaf_columns] == (
        page.VISIBLE_DEMAND_COLUMNS
    )
    assert leaf_columns[0]["pinned"] == "left"
    assert leaf_columns[1]["pinned"] == "left"
    assert leaf_columns[2].get("pinned") is None
    assert leaf_columns[0]["cellRenderer"]["function"] == (
        "physSnapshotCountryGroupLabel(params)"
    )
    assert set(leaf_columns[0]["cellClassRules"]) == {
        "phys-snapshot-country-label-cell",
        "phys-snapshot-country-continuation-cell",
    }
    assert leaf_columns[2]["valueFormatter"]["function"] == (
        "physSnapshotOneDecimal(params)"
    )
    assert set(grid.rowClassRules) >= {
        "phys-snapshot-provider-ea-row",
        "phys-snapshot-provider-woodmac-row",
        "phys-snapshot-provider-platts-row",
    }
    assert "sortChanged" in grid.eventListeners
    assert "refreshCells" in grid.eventListeners["sortChanged"][0]
    assert grid.columnDefs[0]["headerName"] == "Market"
    assert grid.columnDefs[1]["headerName"] == "Annual LNG Imports (MMT)"

    workbook_bytes = page.build_demand_export_bytes(rows)
    workbook = load_workbook(BytesIO(workbook_bytes))
    worksheet = workbook["LNG Physical Snapshot"]

    assert worksheet.freeze_panes == "C2"
    assert [cell.value for cell in worksheet[1]] == page.VISIBLE_DEMAND_COLUMNS
    assert worksheet["C2"].value == 12.34
    assert worksheet["C2"].number_format == "#,##0.0"
    assert worksheet.max_column == 7


def test_page_uses_global_refresh_and_shipping_route_contract():
    components = {
        getattr(component, "id", None): component
        for component in _walk(page.layout)
        if getattr(component, "id", None)
    }
    assert "phys-snapshot-export-button" in components
    assert "phys-snapshot-storage-scenario" in components
    assert "phys-snapshot-demand-grid" in components
    assert components["phys-snapshot-storage-scenario"].value == "best_view"
    section_classes = [
        str(getattr(component, "className", ""))
        for component in page.layout.children
    ]
    assert next(
        index
        for index, class_name in enumerate(section_classes)
        if "phys-snapshot-storage-section" in class_name
    ) < next(
        index
        for index, class_name in enumerate(section_classes)
        if "phys-snapshot-demand-section" in class_name
    )
    demand_headers = [
        component
        for component in _walk(page.layout)
        if "phys-snapshot-section-header"
        in str(getattr(component, "className", ""))
        and "phys-snapshot-storage-section-header"
        not in str(getattr(component, "className", ""))
    ]
    assert len(demand_headers) == 1
    assert any(
        getattr(component, "id", None)
        == "phys-snapshot-demand-source-metadata"
        for component in _walk(demand_headers[0])
    )

    page_source = (REPO / "pages" / "lng_phys_snapshot.py").read_text()
    route_source = (REPO / "index_shipping_snd.py").read_text()
    styles = (REPO / "assets" / "styles.css").read_text()
    grid_functions = (
        REPO / "assets" / "lng_phys_snapshot_grid.js"
    ).read_text()

    assert page_source.count('Input("global-refresh-button"') == 2
    assert "Next four seasonal endpoints" not in page_source
    assert "phys-snapshot-refresh-button" not in page_source
    assert "import pages.lng_phys_snapshot" in route_source
    assert 'path="/lng-phys-snapshot"' in route_source
    assert "LNG Shipping - LNG Physical Snapshot" in route_source
    assert "nav-lng-phys-snapshot" in route_source
    assert "grid-template-columns: repeat(4" in styles
    assert "grid-template-columns: repeat(2" in styles
    assert "dagfuncs.physSnapshotOneDecimal" in grid_functions
    assert "dagfuncs.physSnapshotCountryGroupLabel" in grid_functions


def test_storage_failure_does_not_affect_demand_callback(monkeypatch):
    monkeypatch.setattr(
        page,
        "get_storage_snapshot",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("storage unavailable")
        ),
    )

    cards, metadata, warning = page.update_storage_snapshot(None, 0)

    assert len(cards) == 4
    assert "Best View" in metadata
    assert warning == "EU storage is unavailable."

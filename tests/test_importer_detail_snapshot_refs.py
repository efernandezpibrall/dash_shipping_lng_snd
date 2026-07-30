import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import threading
import time

from dash._utils import to_json
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import pytest

from pages import importer_detail
from utils import dashboard_snapshot_cache as snapshots


def _catalog(country="China"):
    return [{
        "destination_country_name": country,
        "country": country,
        "country_display": country,
        "continent": "Asia",
        "subcontinent": "East Asia",
        "basin": "Pacific",
        "country_classification_level1": "Asia",
        "country_classification": "Importer",
        "shipping_region": "North Asia",
    }]


def _scoped_frame():
    return pd.DataFrame(
        {
            "end_date": pd.to_datetime(
                [
                    "2025-01-02",
                    "2025-02-03",
                    "2026-01-04",
                    "2026-02-05",
                ]
            ),
            "cargo_mcm": np.array([62.25, np.nan, 74.5, 83.75]),
            "origin_country": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "origin_continent_chart": [
                "Asia",
                "Oceania",
                "North America",
                "Asia",
            ],
            "origin_continent": [
                "Asia",
                "Oceania",
                "North America",
                "Asia",
            ],
            "origin_shipping_region": [
                "Arabian Gulf",
                "Australia",
                "US Gulf",
                "Southeast Asia",
            ],
            "origin_basin": [
                "Pacific",
                "Pacific",
                "Atlantic",
                "Pacific",
            ],
            "origin_subcontinent": [
                "Middle East",
                "Oceania",
                "North America",
                None,
            ],
            "origin_classification_level1": [
                "Middle East",
                "Pacific",
                "North America",
                "Southeast Asia",
            ],
            "origin_classification": [
                "Producer",
                "Producer",
                "Producer",
                "Producer",
            ],
        }
    )


def _scoped_frame_with_status():
    frame = _scoped_frame()
    frame["status"] = [
        "Delivered",
        "Scheduled",
        "Delivered",
        "In Transit",
    ]
    return frame


def _forecast_frame():
    dates = pd.to_datetime(["2026-08-01", "2026-08-02"])
    return pd.DataFrame(
        {
            "date": dates,
            "year": dates.year,
            "day_of_year": dates.dayofyear,
            "month_day": dates.strftime("%b %d"),
            "mcmd": [334.85634408602147, 333.48172043010754],
            "is_forecast": [True, True],
            "source": ["Short Term", "Short Term"],
        }
    )


def _allocation_source():
    run_metadata = {
        "run_id": "allocation-run-importer",
        "analysis_date": pd.Timestamp("2026-07-20 12:00:00"),
        "forecast_start": pd.Timestamp("2026-07-01"),
        "forecast_end": pd.Timestamp("2028-12-01"),
        "supply_scenario": "base_view",
        "split_by_contract": True,
        "woodmac_short_term_outlook": "July 2026",
        "woodmac_long_term_outlook": "June 2026",
    }
    mapping_df = pd.DataFrame(
        {
            "country": ["Qatar", "Australia"],
            "country_name": ["Qatar", "Australia"],
            "continent": ["Asia", "Oceania"],
            "shipping_region": ["Arabian Gulf", "Australia"],
            "basin": ["Pacific", "Pacific"],
            "subcontinent": ["Middle East", "Oceania"],
            "country_classification_level1": [
                "Middle East",
                "Pacific",
            ],
            "country_classification": ["Producer", "Producer"],
        }
    )
    allocation_df = pd.DataFrame(
        {
            "date": pd.to_datetime(
                [
                    "2026-07-01",
                    "2026-08-01",
                    "2027-01-01",
                    "2028-01-01",
                ]
            ),
            "origin_country": [
                "Qatar",
                "Australia",
                "Qatar",
                "Australia",
            ],
            "allocated_volume_bcm": [1.25, 1.5, 3.0, 3.5],
            "alias": ["Qatar", "Australia", "Qatar", "Australia"],
            "country_display": [
                "Qatar",
                "Australia",
                "Qatar",
                "Australia",
            ],
            "continent": ["Asia", "Oceania", "Asia", "Oceania"],
            "country": ["Qatar", "Australia", "Qatar", "Australia"],
        }
    )
    demand_totals_df = pd.DataFrame(
        {
            "date": pd.to_datetime(
                [
                    "2026-07-01",
                    "2026-08-01",
                    "2027-01-01",
                    "2028-01-01",
                ]
            ),
            "forecast_demand_bcm": [2.8, 3.0, 6.5, 7.0],
        }
    )
    return {
        "destination_countries": ["China"],
        "selected_destination_aggregation": "country",
        "run_metadata": run_metadata,
        "mapping_df": mapping_df,
        "allocation_df": allocation_df,
        "internal_allocation_df": pd.DataFrame(
            columns=["date", "internal_allocation_bcm"]
        ),
        "demand_totals_df": demand_totals_df,
    }


def _base_payload(country="China"):
    context = {
        "display_label": country,
        "destination_countries": (country,),
    }
    payload = importer_detail._import_analysis_store_payload(
        "country",
        country,
        context,
        _scoped_frame(),
        forecast_df=pd.DataFrame(
            columns=[
                "date",
                "year",
                "day_of_year",
                "month_day",
                "mcmd",
                "is_forecast",
                "source",
            ]
        ),
    )
    payload["loaded_at"] = "2026-07-24T12:00:00"
    return payload


@pytest.mark.parametrize(
    ("aggregation", "mapping_column"),
    list(
        importer_detail.IMPORTER_SELECTION_TO_MAPPING_COLUMN.items()
    ),
)
def test_importer_net_scope_uses_only_whitelisted_mapping_columns(
    aggregation,
    mapping_column,
):
    assert (
        importer_detail._importer_net_scope_mapping_column(
            aggregation
        )
        == mapping_column
    )
    assert (
        importer_detail._importer_net_scope_mapping_column(
            "untrusted_column"
        )
        == "country_name"
    )


def test_scoped_trade_sql_maps_both_sides_nets_and_aggregates(
    monkeypatch,
):
    captured = {}
    sql_result = pd.DataFrame([{
        "end_date": pd.Timestamp("2026-07-01"),
        "cargo_mcm": 12.5,
        "origin_country": "Qatar",
        "origin_continent_chart": "Asia",
        "origin_continent": "Asia",
        "origin_shipping_region": "Arabian Gulf",
        "origin_basin": "Atlantic Basin",
        "origin_subcontinent": "Middle East",
        "origin_classification_level1": "Middle East",
        "origin_classification": "Producer",
        "destination_country_name": "Russian Federation",
        "status": "Delivered",
    }])

    def read_sql(query, _engine, params=None):
        captured["statement"] = str(query)
        captured["params"] = params
        return sql_result.copy()

    monkeypatch.setattr(importer_detail.pd, "read_sql", read_sql)
    result = importer_detail._fetch_importer_scoped_trades(
        importer_detail.engine,
        ["Russian Federation"],
        delivered_only=True,
        include_destination_context=True,
        include_status=True,
        selected_destination_aggregation=(
            "country_classification_level1"
        ),
    )

    statement = captured["statement"]
    assert (
        "INNER JOIN at_lng.mappings_country destination_map"
        in statement
    )
    assert (
        "LEFT JOIN at_lng.mappings_country origin_map"
        in statement
    )
    assert "origin_map.country_name" in statement
    assert (
        "origin_map.country_classification_level1"
        in statement
    )
    assert (
        "destination_map.country_classification_level1"
        in statement
    )
    assert "SUM(COALESCE(kt.cargo_destination_cubic_meters" in statement
    assert "GROUP BY" in statement
    assert captured["params"]["destination_countries"] == (
        "Russian Federation",
    )
    assert result["origin_country"].tolist() == ["Qatar"]
    assert result["cargo_mcm"].tolist() == [12.5]


def test_allocation_sql_excludes_internal_aliases_and_nets_footer(
    monkeypatch,
):
    statements = []
    mapping_df = pd.DataFrame([
        {
            "country": "Russian Federation",
            "country_name": "Russia",
            "continent": "Europe",
            "shipping_region": "Russia",
            "basin": "Atlantic Basin",
            "subcontinent": "Eastern Europe",
            "country_classification_level1": "Russia",
            "country_classification": "Russia",
        },
        {
            "country": "Qatar",
            "country_name": "Qatar",
            "continent": "Asia",
            "shipping_region": "Arabian Gulf",
            "basin": "Atlantic Basin",
            "subcontinent": "Middle East",
            "country_classification_level1": "Middle East",
            "country_classification": "Producer",
        },
    ])
    allocation_rows = pd.DataFrame([
        {
            "date": pd.Timestamp("2026-07-01"),
            "alias": "Russian Federation",
            "origin_country": "Russia",
            "country_display": "Russia",
            "country": "Russia",
            "continent": "Europe",
            "shipping_region": "Russia",
            "basin": "Atlantic Basin",
            "subcontinent": "Eastern Europe",
            "country_classification_level1": "Russia",
            "country_classification": "Russia",
            "is_internal_flow": True,
            "allocated_volume_bcm": 0.5,
        },
        {
            "date": pd.Timestamp("2026-07-01"),
            "alias": "Qatar",
            "origin_country": "Qatar",
            "country_display": "Qatar",
            "country": "Qatar",
            "continent": "Asia",
            "shipping_region": "Arabian Gulf",
            "basin": "Atlantic Basin",
            "subcontinent": "Middle East",
            "country_classification_level1": "Middle East",
            "country_classification": "Producer",
            "is_internal_flow": False,
            "allocated_volume_bcm": 1.5,
        },
    ])
    demand_df = pd.DataFrame({
        "date": [pd.Timestamp("2026-07-01")],
        "forecast_demand_bcm": [2.0],
    })

    def read_sql(query, _engine, **_kwargs):
        statement = str(query)
        statements.append(statement)
        if "mappings_country" in statement and (
            "fundamentals_supply_allocation" not in statement
        ):
            return mapping_df.copy()
        if "fundamentals_supply_allocation_demand_detail" in statement:
            return allocation_rows.copy()
        if "fundamentals_supply_allocation_demand_summary" in statement:
            return demand_df.copy()
        raise AssertionError(statement)

    monkeypatch.setattr(importer_detail.pd, "read_sql", read_sql)
    source = importer_detail._fetch_origin_forecast_source_data(
        importer_detail.engine,
        ["Russian Federation"],
        current_date="2026-07-28",
        run_metadata={"run_id": "test-run"},
        selected_destination_aggregation="country",
    )

    assert source["allocation_df"]["country"].tolist() == ["Qatar"]
    assert source["internal_allocation_df"].to_dict("records") == [{
        "date": pd.Timestamp("2026-07-01"),
        "internal_allocation_bcm": 0.5,
    }]
    allocation_sql = next(
        statement for statement in statements
        if "fundamentals_supply_allocation_demand_detail"
        in statement
    )
    assert "destination_map.country" in allocation_sql
    assert "origin_map.country" in allocation_sql
    assert "origin_map.country_name" in allocation_sql

    _summary, footer_rows, _metadata = (
        importer_detail._build_origin_forecast_summary_from_source(
            source,
            current_date="2026-07-28",
            origin_level="origin_country_name",
        )
    )
    assert footer_rows[0]["Continent"] == (
        "NET WOODMAC DEMAND TOTAL"
    )
    assert footer_rows[1]["Continent"] == (
        "NET ALLOCATED SUPPLY TOTAL"
    )
    assert footer_rows[0]["Jul'26"] == footer_rows[1]["Jul'26"]
    assert footer_rows[2]["Jul'26"] == 0.0


@pytest.fixture
def persistent_importer_detail_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "importer-detail-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(snapshots.LOCAL_CACHE_DIR_ENV, str(cache_directory))
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    monkeypatch.setattr(
        importer_detail,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: "2026-07-24T00:00:00Z",
    )
    monkeypatch.setattr(
        importer_detail,
        "fetch_woodmac_country_import_forecast_data",
        lambda _countries: pd.DataFrame(),
    )
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_reference(monkeypatch, country="China", frame=None):
    scoped_frame = frame.copy() if frame is not None else _scoped_frame()
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: scoped_frame.copy(),
    )
    reference = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        country,
        _catalog(country),
    )
    return _base_payload(country), reference


def _delete_or_corrupt_reference(reference, mode):
    stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if mode == "missing":
        stores.cache.delete(record_key, retry=True)
    else:
        stores.cache.set(record_key, b"corrupt", retry=True)
    snapshots.clear_local_snapshots()


def _workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: [
            [
                (cell.value, cell.data_type, cell.number_format)
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]
        for worksheet in workbook.worksheets
    }


@pytest.mark.parametrize("country", ["China", "Kuwait"])
def test_base_loader_emits_small_resolvable_reference_and_preserves_payload(
    country,
    monkeypatch,
    persistent_importer_detail_cache,
):
    payload, reference = _load_reference(monkeypatch, country)

    assert snapshots.is_snapshot_reference(
        reference,
        importer_detail.IMPORTER_DETAIL_BASE_NAMESPACE,
    )
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert len(to_json(reference).encode("utf-8")) < 50_000

    legacy = importer_detail._resolve_import_analysis_base_data(payload)
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    restarted = importer_detail._resolve_import_analysis_base_data(reference)

    pd.testing.assert_frame_equal(
        restarted[0],
        legacy[0],
        check_dtype=True,
        check_exact=True,
    )
    assert list(restarted[0].columns) == list(legacy[0].columns)
    assert restarted[0].isna().equals(legacy[0].isna())
    assert restarted[1:] == legacy[1:]


def test_allocation_reference_is_small_exact_and_control_paths_are_sql_free(
    monkeypatch,
    persistent_importer_detail_cache,
):
    source_data = _allocation_source()
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_data["run_metadata"]),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        lambda *_args, **_kwargs: copy.deepcopy(source_data),
    )

    reference = importer_detail.refresh_origin_forecast_source(
        0,
        "country",
        "China",
        _catalog(),
    )
    assert snapshots.is_snapshot_reference(
        reference,
        importer_detail.IMPORTER_ALLOCATION_SOURCE_NAMESPACE,
    )
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    resolved = importer_detail._resolve_origin_forecast_source_data(
        reference
    )
    for frame_key in (
        "mapping_df",
        "allocation_df",
        "internal_allocation_df",
        "demand_totals_df",
    ):
        pd.testing.assert_frame_equal(
            resolved[frame_key],
            source_data[frame_key],
            check_exact=True,
        )

    monkeypatch.setattr(
        importer_detail,
        "fetch_origin_forecast_summary_data",
        lambda *_args, **_kwargs: pytest.fail(
            "allocation controls queried PostgreSQL"
        ),
    )
    levels = [
        "origin_country_name",
        "origin_basin",
        "origin_subcontinent",
        "origin_classification_level1",
        "origin_classification",
    ]
    for level in levels:
        direct = importer_detail.update_origin_forecast_summary_table(
            "country",
            "China",
            ["Asia"],
            _catalog(),
            level,
            "mcm_d",
            source_data,
        )
        cached = importer_detail.update_origin_forecast_summary_table(
            "country",
            "China",
            ["Asia"],
            _catalog(),
            level,
            "mcm_d",
            reference,
        )
        assert to_json(cached) == to_json(direct)


def test_allocation_demand_only_source_matches_legacy_for_every_origin_level(
    monkeypatch,
):
    source_fixture = _allocation_source()
    raw_empty_allocation = pd.DataFrame(
        columns=[
            "date",
            "origin_country",
            "destination",
            "allocated_volume_bcm",
        ]
    )

    def read_sql(query, _engine, **_kwargs):
        statement = str(query)
        if "mappings_country" in statement:
            return source_fixture["mapping_df"].copy()
        if "fundamentals_supply_allocation_demand_summary" in statement:
            return source_fixture["demand_totals_df"].copy()
        if "fundamentals_supply_allocation_demand_detail" in statement:
            return raw_empty_allocation.copy()
        raise AssertionError(statement)

    monkeypatch.setattr(importer_detail.pd, "read_sql", read_sql)
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_fixture["run_metadata"]),
    )
    source_data = importer_detail._fetch_origin_forecast_source_data(
        importer_detail.engine,
        ["China"],
        run_metadata=copy.deepcopy(source_fixture["run_metadata"]),
    )

    levels = [
        "origin_country_name",
        "origin_shipping_region",
        "continent_origin_name",
        "origin_basin",
        "origin_subcontinent",
        "origin_classification_level1",
        "origin_classification",
    ]
    for level in levels:
        legacy = importer_detail.fetch_origin_forecast_summary_data(
            importer_detail.engine,
            ["China"],
            origin_level=level,
        )
        cached = importer_detail._build_origin_forecast_summary_from_source(
            source_data,
            origin_level=level,
        )
        pd.testing.assert_frame_equal(
            cached[0],
            legacy[0],
            check_exact=True,
        )
        assert cached[1] == legacy[1]
        assert cached[2] == legacy[2]


def test_allocation_no_run_uses_one_metadata_read_and_consistent_identity(
    monkeypatch,
    persistent_importer_detail_cache,
):
    metadata_reads = []

    def no_run(_engine):
        metadata_reads.append(True)
        if len(metadata_reads) > 1:
            raise AssertionError("allocation metadata was queried twice")
        return None

    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        no_run,
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        lambda *_args, **_kwargs: pytest.fail(
            "no-run refresh invoked the source loader"
        ),
    )

    reference = importer_detail.refresh_origin_forecast_source(
        0,
        "country",
        "China",
        _catalog(),
    )
    current_month = pd.Timestamp.today().replace(day=1).date().isoformat()
    expected_key = snapshots.build_source_key(
        importer_detail.IMPORTER_ALLOCATION_SOURCE_NAMESPACE,
        None,
        None,
        "country",
        ["China"],
        current_month,
    )
    assert metadata_reads == [True]
    assert reference["source_key"] == expected_key
    resolved = importer_detail._resolve_origin_forecast_source_data(
        reference
    )
    assert resolved["run_metadata"] is None
    assert resolved["allocation_df"].empty
    assert resolved["demand_totals_df"].empty
    manifest = snapshots._LOCAL_MANIFESTS[
        (
            reference["namespace"],
            reference["source_key"],
            reference["revision"],
        )
    ]
    assert manifest["run_id"] is None
    output = importer_detail.update_origin_forecast_summary_table(
        "country",
        "China",
        [],
        _catalog(),
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        reference,
    )
    assert "No compatible WoodMac supply-allocation SQL run" in to_json(
        output
    )


def test_route_reference_is_small_exact_and_controls_and_export_are_sql_free(
    monkeypatch,
    persistent_importer_detail_cache,
):
    route_source_version = {
        "kpler_watermark": "kpler-v1",
        "distance_watermark": "distance-v1",
    }
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_route_source_version",
        lambda: copy.deepcopy(route_source_version),
    )
    processed_df = pd.DataFrame(
        {
            "year": [2025, 2025, 2026, 2026],
            "month": [1, 4, 7, 10],
            "season": ["W", "S", "S", "W"],
            "quarter": ["Q1", "Q2", "Q3", "Q4"],
            "voyage_id": ["v1", "v2", "v3", "v4"],
            "origin_country_name": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "distanceDirect": [100.0, 100.0, 100.0, 100.0],
            "distanceViaSuez": [120.0, np.nan, 120.0, np.nan],
            "distanceViaPanama": [np.nan, 130.0, 130.0, np.nan],
            "selected_route": [
                "Direct",
                "ViaPanama",
                "ViaSuez",
                "Direct",
            ],
        }
    )
    mapping_df = pd.DataFrame(
        {
            "country": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "country_name": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "continent": [
                "Asia",
                "Oceania",
                "North America",
                "Asia",
            ],
            "basin": ["Pacific", "Pacific", "Atlantic", "Pacific"],
            "shipping_region": [
                "Arabian Gulf",
                "Australia",
                "US Gulf",
                "Southeast Asia",
            ],
            "subcontinent": [
                "Middle East",
                "Oceania",
                "North America",
                "Southeast Asia",
            ],
            "country_classification_level1": [
                "Middle East",
                "Pacific",
                "North America",
                "Southeast Asia",
            ],
            "country_classification": [
                "Producer",
                "Producer",
                "Producer",
                "Producer",
            ],
        }
    )
    source_data = {
        "destination_countries": ["China"],
        "processed_df": processed_df,
        "mapping_df": mapping_df,
    }
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_route_source_payload",
        lambda _countries: copy.deepcopy(source_data),
    )
    reference = importer_detail.refresh_importer_route_analysis_source(
        "country",
        "China",
        _catalog(),
        0,
    )
    assert snapshots.is_snapshot_reference(
        reference,
        importer_detail.IMPORTER_ROUTE_SOURCE_NAMESPACE,
    )
    assert len(to_json(reference).encode("utf-8")) < 10_000

    monkeypatch.setattr(
        importer_detail,
        "process_trade_and_distance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "route controls or export queried PostgreSQL"
        ),
    )
    monkeypatch.setattr(
        importer_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pytest.fail(
            "route controls queried mappings"
        ),
    )
    direct = importer_detail.update_route_analysis_charts_and_tables(
        "Year",
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "country",
        "China",
        _catalog(),
        source_data,
    )
    cached = importer_detail.update_route_analysis_charts_and_tables(
        "Year",
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "country",
        "China",
        _catalog(),
        reference,
    )
    assert to_json(cached) == to_json(direct)

    direct_export = (
        importer_detail.export_importer_route_analysis_to_excel(
            1,
            "Year",
            importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
            "country",
            "China",
            _catalog(),
            source_data,
        )
    )
    cached_export = (
        importer_detail.export_importer_route_analysis_to_excel(
            1,
            "Year",
            importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
            "country",
            "China",
            _catalog(),
            reference,
        )
    )
    assert _workbook_cells(cached_export) == _workbook_cells(
        direct_export
    )


def test_diversion_reference_is_small_exact_and_sql_free(
    monkeypatch,
    persistent_importer_detail_cache,
):
    source_data = {
        "destination_label": "China",
        "destination_countries": ["China"],
        "main_data": [{
            "Diversion date": "2026-06-01",
            "Vessel": "Test Vessel",
            "State": "Loaded",
            "Charterer": "Test Charterer",
            "Cubic Meters": 174500.0,
            "Origin location": "Sabine Pass",
            "Origin country": "United States",
            "Origin date": "2026-05-20",
            "Diverted from location": "Gate",
            "Diverted from country": "Netherlands",
            "Diverted from date": "2026-06-10",
            "New destination location": "Tianjin",
            "New destination country": "China",
            "New destination date": "2026-06-20",
            "Added shipping days": 10,
        }],
        "charts_data": [{
            "Diversion_month": "2026-06-01 00:00:00",
            "basin_combo": "Atlantic -> Pacific",
            "region_combo": "Northwest Europe -> North Asia",
            "country_combo": "Netherlands -> China",
            "Added shipping days": 10,
            "Cubic Meters": 174500.0,
        }],
    }
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_diversion_source_version",
        lambda: "diversion-v1",
    )
    captured_versions = []

    def build_diversion(*_args, **kwargs):
        captured_versions.append(kwargs.get("source_version"))
        return copy.deepcopy(source_data)

    monkeypatch.setattr(
        importer_detail,
        "_build_importer_diversion_payload",
        build_diversion,
    )
    reference = importer_detail.refresh_importer_diversion_source(
        0,
        "country",
        "China",
        _catalog(),
    )
    assert snapshots.is_snapshot_reference(
        reference,
        importer_detail.IMPORTER_DIVERSION_SOURCE_NAMESPACE,
    )
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert captured_versions == ["diversion-v1"]
    assert to_json(importer_detail.update_diversion_ui(None, "basin_combo")) == to_json(
        importer_detail.update_diversion_ui(
            {"main_data": [], "charts_data": []},
            "basin_combo",
        )
    )

    monkeypatch.setattr(
        importer_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pytest.fail(
            "diversion controls and export must not query SQL"
        ),
    )
    direct_ui = importer_detail.update_diversion_ui(
        source_data,
        "country_combo",
    )
    cached_ui = importer_detail.update_diversion_ui(
        reference,
        "country_combo",
    )
    assert to_json(cached_ui) == to_json(direct_ui)
    direct_export = (
        importer_detail.export_importer_diversion_summary_to_excel(
            1,
            source_data,
            direct_ui[1],
        )
    )
    cached_export = (
        importer_detail.export_importer_diversion_summary_to_excel(
            1,
            reference,
            cached_ui[1],
        )
    )
    assert _workbook_cells(cached_export) == _workbook_cells(
        direct_export
    )

    corrupt_reference = dict(reference, revision="missing")
    corrupt_ui = importer_detail.update_diversion_ui(
        corrupt_reference,
        "basin_combo",
    )
    assert importer_detail.IMPORTER_DIVERSION_RECOVERY_MESSAGE in to_json(
        corrupt_ui
    )
    with pytest.raises(importer_detail.PreventUpdate):
        importer_detail.export_importer_diversion_summary_to_excel(
            1,
            corrupt_reference,
            cached_ui[1],
        )


@pytest.mark.parametrize("workers", [1, 4, 8])
def test_diversion_source_is_single_flight_under_concurrency(
    workers,
    monkeypatch,
    persistent_importer_detail_cache,
):
    counter = {"count": 0}
    counter_lock = threading.Lock()

    def build_payload(*_args, **_kwargs):
        with counter_lock:
            counter["count"] += 1
        time.sleep(0.04)
        return {
            "destination_label": "China",
            "destination_countries": ["China"],
            "main_data": [],
            "charts_data": [],
        }

    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_diversion_source_version",
        lambda: f"diversion-single-flight-{workers}",
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_diversion_payload",
        build_payload,
    )
    with ThreadPoolExecutor(max_workers=workers) as pool:
        references = list(
            pool.map(
                lambda _index: (
                    importer_detail.refresh_importer_diversion_source(
                        0,
                        "country",
                        "China",
                        _catalog(),
                    )
                ),
                range(workers),
            )
        )

    assert counter["count"] == 1
    assert all(
        snapshots.snapshot_is_resolvable(reference)
        for reference in references
    )
    assert len({
        (reference["source_key"], reference["revision"])
        for reference in references
    }) == 1


@pytest.mark.parametrize("mode", ["missing", "corrupt"])
def test_allocation_reference_failure_is_visible_without_sql_fallback(
    mode,
    monkeypatch,
    persistent_importer_detail_cache,
):
    source_data = _allocation_source()
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_data["run_metadata"]),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        lambda *_args, **_kwargs: copy.deepcopy(source_data),
    )
    reference = importer_detail.refresh_origin_forecast_source(
        0,
        "country",
        "China",
        _catalog(),
    )
    _delete_or_corrupt_reference(reference, mode)
    monkeypatch.setattr(
        importer_detail,
        "fetch_origin_forecast_summary_data",
        lambda *_args, **_kwargs: pytest.fail(
            "corrupt allocation reference queried PostgreSQL"
        ),
    )

    output = importer_detail.update_origin_forecast_summary_table(
        "country",
        "China",
        [],
        _catalog(),
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        reference,
    )

    assert (
        importer_detail.IMPORTER_ALLOCATION_RECOVERY_MESSAGE
        in to_json(output)
    )


def test_nonempty_forecast_round_trips_exactly_and_chart_does_not_query_sql(
    monkeypatch,
    persistent_importer_detail_cache,
):
    monkeypatch.setattr(
        importer_detail,
        "fetch_woodmac_country_import_forecast_data",
        lambda _countries: _forecast_frame(),
    )
    _, reference = _load_reference(monkeypatch)

    restarted = importer_detail._load_import_analysis_forecast_data(
        reference
    )
    pd.testing.assert_frame_equal(
        restarted,
        _forecast_frame(),
        check_dtype=True,
        check_exact=True,
    )

    monkeypatch.setattr(
        importer_detail,
        "fetch_woodmac_country_import_forecast_data",
        lambda _countries: pytest.fail("chart callback queried WoodMac"),
    )
    importer_detail.update_import_analysis_charts(
        reference,
        30,
        "mcm_d",
        ["2025", "2026"],
    )


def test_forecast_month_is_part_of_source_key(
    monkeypatch,
    persistent_importer_detail_cache,
):
    payload = _base_payload()
    payload["woodmac_import_forecast"] = _forecast_frame()
    build_calls = []
    month = {"value": "2026-07-01"}
    monkeypatch.setattr(
        importer_detail,
        "_importer_detail_forecast_month_token",
        lambda: month["value"],
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        lambda aggregation, value, context: (
            build_calls.append((aggregation, value))
            or copy.deepcopy(payload)
        ),
    )

    first = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog(),
    )
    month["value"] = "2026-08-01"
    second = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog(),
    )

    assert first["source_key"] != second["source_key"]
    assert build_calls == [("country", "China"), ("country", "China")]


def test_reference_matches_legacy_for_year_charts_kpis_mix_and_base_export(
    monkeypatch,
    persistent_importer_detail_cache,
):
    payload, reference = _load_reference(monkeypatch)
    monkeypatch.setattr(
        importer_detail,
        "fetch_woodmac_country_import_forecast_data",
        lambda _countries: pd.DataFrame(),
    )

    legacy_years = importer_detail.update_import_analysis_year_selector(
        payload,
        ["2025"],
    )
    reference_years = importer_detail.update_import_analysis_year_selector(
        reference,
        ["2025"],
    )
    assert to_json(reference_years) == to_json(legacy_years)

    legacy_charts = importer_detail.update_import_analysis_charts(
        payload,
        30,
        "mcm_d",
        ["2025", "2026"],
    )
    reference_charts = importer_detail.update_import_analysis_charts(
        reference,
        30,
        "mcm_d",
        ["2025", "2026"],
    )
    assert to_json(reference_charts) == to_json(legacy_charts)

    scoped_fetches = []

    def fetch_scoped(_engine, _countries, **kwargs):
        scoped_fetches.append(dict(kwargs))
        return _scoped_frame().copy()

    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        fetch_scoped,
    )
    monkeypatch.setattr(
        importer_detail,
        "_load_importer_country_mapping_lookup",
        lambda _engine: pd.DataFrame(),
    )
    legacy_export = importer_detail.export_import_analysis_to_excel(
        1,
        "country",
        "China",
        _catalog(),
        30,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        payload,
    )
    reference_export = importer_detail.export_import_analysis_to_excel(
        1,
        "country",
        "China",
        _catalog(),
        30,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        reference,
    )
    assert _workbook_cells(reference_export) == _workbook_cells(legacy_export)
    assert len(scoped_fetches) == 2
    assert all(call.get("delivered_only") is True for call in scoped_fetches)


def test_status_source_drives_summary_and_export_without_database_reads(
    monkeypatch,
    persistent_importer_detail_cache,
):
    destination_context = {
        "display_label": "China",
        "destination_countries": ("China",),
    }
    status_payload = importer_detail._import_analysis_store_payload(
        "country",
        "China",
        destination_context,
        _scoped_frame_with_status(),
        forecast_df=pd.DataFrame(),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        lambda *_args, **_kwargs: copy.deepcopy(status_payload),
    )
    reference = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog(),
    )

    delivered_frame = _scoped_frame_with_status()
    delivered_frame = delivered_frame[
        delivered_frame["status"].eq("Delivered")
    ].drop(columns="status")
    monkeypatch.setattr(
        importer_detail,
        "_load_importer_country_mapping_lookup",
        lambda _engine: pd.DataFrame(),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: delivered_frame.copy(),
    )
    legacy_export = importer_detail.export_import_analysis_to_excel(
        1,
        "country",
        "China",
        _catalog(),
        30,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        _base_payload(),
    )
    legacy_summary = importer_detail.update_origin_summary_table(
        "country",
        "China",
        30,
        [],
        _catalog(),
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        "levels",
        5,
        3,
        3,
        _base_payload(),
    )

    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: pytest.fail(
            "summary or export queried scoped trades"
        ),
    )
    monkeypatch.setattr(
        importer_detail,
        "_load_importer_country_mapping_lookup",
        lambda *_args, **_kwargs: pytest.fail(
            "export queried country mappings"
        ),
    )

    summary = importer_detail.update_origin_summary_table(
        "country",
        "China",
        30,
        [],
        _catalog(),
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        "levels",
        5,
        3,
        3,
        reference,
    )
    assert "Error loading data" not in to_json(summary)
    assert to_json(summary) == to_json(legacy_summary)

    cached_export = importer_detail.export_import_analysis_to_excel(
        1,
        "country",
        "China",
        _catalog(),
        30,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        reference,
    )
    assert _workbook_cells(cached_export) == _workbook_cells(legacy_export)


def test_origin_summary_rejects_cached_rows_when_hierarchy_membership_changes(
    monkeypatch,
):
    cached_context = {
        "display_label": "Asia",
        "destination_countries": ("China",),
    }
    cached_payload = importer_detail._import_analysis_store_payload(
        "continent",
        "Asia",
        cached_context,
        _scoped_frame_with_status(),
        forecast_df=pd.DataFrame(),
    )
    refreshed_catalog = _catalog()
    refreshed_catalog.append(
        {
            **_catalog("Japan")[0],
            "continent": "Asia",
        }
    )
    fetched_scopes = []

    def fetch_summary(_engine, destination_countries, *_args, **kwargs):
        fetched_scopes.append(
            (
                tuple(destination_countries),
                kwargs.get("scoped_trades_df"),
            )
        )
        return pd.DataFrame()

    monkeypatch.setattr(
        importer_detail,
        "fetch_origin_summary_data",
        fetch_summary,
    )

    importer_detail.update_origin_summary_table(
        "continent",
        "Asia",
        30,
        [],
        refreshed_catalog,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        "levels",
        5,
        3,
        3,
        cached_payload,
    )

    assert fetched_scopes == [(("China", "Japan"), None)]


@pytest.mark.parametrize("mode", ["missing", "corrupt"])
def test_missing_or_corrupt_reference_has_explicit_recovery_and_no_raw_fallback(
    mode,
    monkeypatch,
    persistent_importer_detail_cache,
):
    _, reference = _load_reference(monkeypatch)
    _delete_or_corrupt_reference(reference, mode)

    expected_selector = (
        [{
            "label": importer_detail.IMPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE,
            "value": "__snapshot_unavailable__",
            "disabled": True,
        }],
        [],
    )
    assert (
        importer_detail.update_import_analysis_year_selector(reference, [])
        == expected_selector
    )

    charts = importer_detail.update_import_analysis_charts(
        reference,
        30,
        "mcm_d",
        [],
    )
    assert (
        charts[-1].children
        == importer_detail.IMPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE
    )
    assert charts[-1].role == "alert"

    fallback_fetches = []
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **kwargs: (
            fallback_fetches.append(kwargs) or _scoped_frame()
        ),
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Cached importer-detail data is unavailable",
    ):
        importer_detail.export_import_analysis_to_excel(
            1,
            "country",
            "China",
            _catalog(),
            30,
            importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
            "mcm_d",
            reference,
        )
    assert fallback_fetches == []


def test_wrong_namespace_nonresolvable_and_malformed_refs_are_rejected(
    monkeypatch,
    persistent_importer_detail_cache,
):
    wrong_reference = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": "another-page-v1",
        "source_key": "source",
        "revision": "00000000-0000-4000-8000-000000000000",
        "shared": True,
    }
    with pytest.raises(snapshots.SnapshotUnavailable):
        importer_detail._resolve_import_analysis_base_data(wrong_reference)

    payload = _base_payload()
    local_reference = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": importer_detail.IMPORTER_DETAIL_BASE_NAMESPACE,
        "source_key": "source",
        "revision": 1,
        "shared": False,
    }
    monkeypatch.setattr(
        importer_detail,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (local_reference, payload),
    )
    with pytest.raises(snapshots.SnapshotUnavailable):
        importer_detail.refresh_import_analysis_base_data(
            0,
            "country",
            "China",
            _catalog(),
        )

    malformed_reference, _ = snapshots.get_or_build_snapshot(
        importer_detail.engine,
        namespace=importer_detail.IMPORTER_DETAIL_BASE_NAMESPACE,
        source_key="malformed-payload",
        builder=lambda: {},
        force=True,
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Cached importer-detail data is unavailable",
    ):
        importer_detail._resolve_import_analysis_base_data(
            malformed_reference
        )


@pytest.mark.parametrize("workers", [1, 2, 4])
def test_base_loader_is_single_flight_for_one_two_and_four_callers(
    workers,
    monkeypatch,
    persistent_importer_detail_cache,
):
    counter = {"count": 0}
    counter_lock = threading.Lock()

    def fetch_scoped(*_args, **_kwargs):
        with counter_lock:
            counter["count"] += 1
        time.sleep(0.04)
        return _scoped_frame()

    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        fetch_scoped,
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: f"single-flight-{workers}",
    )

    with ThreadPoolExecutor(max_workers=workers) as pool:
        references = list(
            pool.map(
                lambda _index: (
                    importer_detail.refresh_import_analysis_base_data(
                        0,
                        "country",
                        "China",
                        _catalog(),
                    )
                ),
                range(workers),
            )
        )

    assert counter["count"] == 1
    assert all(
        snapshots.snapshot_is_resolvable(reference)
        for reference in references
    )
    assert len(
        {
            (
                reference["source_key"],
                reference["revision"],
            )
            for reference in references
        }
    ) == 1


def test_global_refresh_is_checked_once_and_forces_one_rebuild(
    monkeypatch,
    persistent_importer_detail_cache,
):
    refresh_calls = []
    fetch_calls = []
    monkeypatch.setattr(
        importer_detail,
        "_was_global_refresh_triggered",
        lambda: refresh_calls.append(True) or True,
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: (
            fetch_calls.append(True) or _scoped_frame()
        ),
    )

    reference = importer_detail.refresh_import_analysis_base_data(
        1,
        "country",
        "China",
        _catalog(),
    )

    assert refresh_calls == [True]
    assert fetch_calls == [True]
    assert snapshots.snapshot_is_resolvable(reference)


def test_country_hierarchy_route_diversion_and_maintenance_remain_independent_and_exports_are_stable(
    monkeypatch,
):
    independent_callbacks = (
        importer_detail.initialize_country_dropdown,
        importer_detail.update_origin_forecast_summary_table,
        importer_detail.update_route_analysis_charts_and_tables,
        importer_detail.export_importer_route_analysis_to_excel,
        importer_detail.process_diversion_data,
        importer_detail.update_diversion_ui,
        importer_detail.export_importer_diversion_summary_to_excel,
        importer_detail.update_maintenance_table,
    )
    assert all(
        "_resolve_import_analysis_base_data"
        not in callback_function.__code__.co_names
        for callback_function in independent_callbacks
    )
    assert (
        "_resolve_import_analysis_base_data"
        in importer_detail.update_origin_summary_table.__code__.co_names
    )

    route_frame = pd.DataFrame(
        {
            "year": [2025, 2025, 2026, 2026],
            "voyage_id": ["v1", "v2", "v3", "v4"],
            "origin_country_name": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "origin_classification_level1": [
                "Middle East",
                "Pacific",
                "North America",
                "Southeast Asia",
            ],
            "distanceDirect": [100.0, 100.0, 100.0, 100.0],
            "distanceViaSuez": [120.0, np.nan, 120.0, np.nan],
            "distanceViaPanama": [np.nan, 130.0, 130.0, np.nan],
        }
    )
    monkeypatch.setattr(
        importer_detail,
        "process_trade_and_distance_data",
        lambda *_args, **_kwargs: route_frame.copy(),
    )
    route_first = importer_detail.export_importer_route_analysis_to_excel(
        1,
        "Year",
        "origin_classification_level1",
        "country",
        "China",
        _catalog(),
    )
    route_second = importer_detail.export_importer_route_analysis_to_excel(
        1,
        "Year",
        "origin_classification_level1",
        "country",
        "China",
        _catalog(),
    )
    assert _workbook_cells(route_first) == _workbook_cells(route_second)

    diversion_store = {
        "main_data": [{
            "Diversion date": "2026-07-20",
            "Vessel": "Example LNG",
            "State": "Loaded",
            "Cubic Meters": 174500.25,
            "Added shipping days": 2.5,
        }]
    }
    diversion_columns = [
        {"field": "Diversion date"},
        {"field": "Vessel"},
        {"field": "State"},
        {"field": "Cubic Meters"},
        {"field": "Added shipping days"},
    ]
    diversion_first = importer_detail.export_importer_diversion_summary_to_excel(
        1,
        diversion_store,
        diversion_columns,
    )
    diversion_second = importer_detail.export_importer_diversion_summary_to_excel(
        1,
        diversion_store,
        diversion_columns,
    )
    assert _workbook_cells(diversion_first) == _workbook_cells(
        diversion_second
    )


def test_maintenance_controls_reuse_current_raw_source_with_exact_output(
    monkeypatch,
):
    raw_data = pd.DataFrame(
        {
            "plant_name": ["Ras Laffan"],
            "country_name": ["Qatar"],
            "lng_train_name_short": ["Train 1"],
            "year": [2026],
            "month": [7],
            "year_actual_forecast": ["Forecast"],
            "total_mtpa": [0.75],
            "metric_comment": ["Planned"],
            "date": pd.to_datetime(["2026-07-01"]),
        }
    )
    fetches = []
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or raw_data.copy()
        ),
    )
    baseline = importer_detail.update_maintenance_table(
        "country",
        "China",
        _catalog(),
        "mt",
        [],
    )
    assert fetches == [True]

    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "maintenance controls reread the source"
        ),
    )
    cached = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mt",
        [],
        maintenance_raw_data=baseline[2],
    )

    assert to_json(cached[:2]) == to_json(baseline[:2])
    assert cached[2] == baseline[2]


@pytest.mark.parametrize("error_message", ["source down", ""])
def test_maintenance_source_error_retries_on_next_control_change(
    monkeypatch,
    error_message,
):
    attempts = []

    def fail_source(*_args, **_kwargs):
        attempts.append("failed")
        raise RuntimeError(error_message)

    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        fail_source,
    )
    failed = importer_detail.update_maintenance_table(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
    )
    assert failed[2]["error"] == error_message

    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            attempts.append("retried") or pd.DataFrame()
        ),
    )
    retried = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mt",
        [],
        maintenance_raw_data=failed[2],
    )

    assert attempts == ["failed", "retried"]
    assert "No supplier maintenance data available" in to_json(retried[0])
    assert "error" not in retried[2]


def test_maintenance_database_failure_reaches_retryable_error_store(
    monkeypatch,
):
    monkeypatch.setattr(
        importer_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("maintenance database down")
        ),
    )

    failed = importer_detail.update_maintenance_table(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
    )

    assert failed[2]["error"] == "maintenance database down"
    assert "Error loading maintenance data" in to_json(failed[0])


def test_source_context_refresh_generation_is_not_sticky_for_later_selections(
    monkeypatch,
    persistent_importer_detail_cache,
):
    build_calls = []
    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        lambda aggregation, value, _context: (
            build_calls.append((aggregation, value))
            or _base_payload(value)
        ),
    )
    watermark = {
        "kpler_watermark": "2026-07-24T00:00:00Z",
        "woodmac_watermark": "2026-07-23T00:00:00Z",
        "distance_watermark": "2026-07-22T00:00:00Z",
    }
    initial_context = importer_detail._build_importer_detail_source_context(
        watermark
    )
    refresh_context = importer_detail._build_importer_detail_source_context(
        watermark,
        force_refresh=True,
    )

    importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=initial_context,
    )
    importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=refresh_context,
    )
    importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "Kuwait",
        _catalog("Kuwait"),
        source_context=refresh_context,
    )
    importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=refresh_context,
    )

    assert build_calls == [
        ("country", "China"),
        ("country", "China"),
        ("country", "Kuwait"),
    ]


def test_mapping_fingerprint_invalidates_base_and_route_snapshots(
    monkeypatch,
    persistent_importer_detail_cache,
):
    base_builds = []
    route_builds = []
    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        lambda aggregation, value, _context: (
            base_builds.append((aggregation, value))
            or _base_payload(value)
        ),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_route_source_payload",
        lambda countries: (
            route_builds.append(tuple(countries))
            or {
                "destination_countries": list(countries),
                "processed_df": pd.DataFrame(),
                "mapping_df": pd.DataFrame(),
            }
        ),
    )
    source_watermark = {
        "kpler_watermark": "kpler-v1",
        "woodmac_watermark": "woodmac-v1",
        "distance_watermark": "distance-v1",
        "mapping_fingerprint": "mapping-v1",
    }
    initial_context = (
        importer_detail._build_importer_detail_source_context(
            source_watermark
        )
    )
    changed_context = (
        importer_detail._build_importer_detail_source_context(
            {
                **source_watermark,
                "mapping_fingerprint": "mapping-v2",
            }
        )
    )

    first_base = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=initial_context,
    )
    changed_base = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=changed_context,
    )
    first_route = importer_detail.refresh_importer_route_analysis_source(
        "country",
        "China",
        _catalog("China"),
        initial_context,
    )
    changed_route = importer_detail.refresh_importer_route_analysis_source(
        "country",
        "China",
        _catalog("China"),
        changed_context,
    )

    assert base_builds == [("country", "China"), ("country", "China")]
    assert route_builds == [("China",), ("China",)]
    assert first_base["source_key"] != changed_base["source_key"]
    assert first_route["source_key"] != changed_route["source_key"]


def test_destination_catalog_exposes_unavailable_source_watermark_failure(
    monkeypatch,
):
    monkeypatch.setattr(
        importer_detail,
        "build_destination_catalog",
        lambda _engine: _catalog("China"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: (_ for _ in ()).throw(RuntimeError("watermark down")),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_maintenance_source_version",
        lambda: (_ for _ in ()).throw(RuntimeError("maintenance down")),
    )

    catalog, options, selected, source_context = (
        importer_detail.initialize_country_dropdown(
            0,
            "country",
            None,
            None,
        )
    )

    assert catalog == _catalog("China")
    assert options == [{"label": "China", "value": "China"}]
    assert selected == "China"
    assert source_context["source_watermark"] is None
    assert source_context["source_revision"]["status"] == "unavailable"
    assert "unavailable" in source_context["source_revision"]["message"]


def test_destination_catalog_reuses_last_good_revision_as_stale(
    monkeypatch,
):
    monkeypatch.setattr(
        importer_detail,
        "build_destination_catalog",
        lambda _engine: _catalog("China"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: (_ for _ in ()).throw(RuntimeError("watermark down")),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_maintenance_source_version",
        lambda: {"revision": "maintenance-v1"},
    )
    previous_context = (
        importer_detail._build_importer_detail_source_context(
            {"kpler_watermark": "kpler-v1"}
        )
    )

    catalog, options, selected, source_context = (
        importer_detail.initialize_country_dropdown(
            0,
            "country",
            _catalog("China"),
            {"aggregation": "country", "value": "China"},
            previous_context,
        )
    )

    assert catalog == _catalog("China")
    assert options == [{"label": "China", "value": "China"}]
    assert selected == "China"
    assert source_context["source_watermark"] == {
        "kpler_watermark": "kpler-v1"
    }
    assert source_context["source_revision"]["status"] == "stale"


def test_importer_initializer_failure_is_explicitly_unavailable(
    monkeypatch,
):
    monkeypatch.setattr(
        importer_detail,
        "_was_global_refresh_triggered",
        lambda: True,
    )
    monkeypatch.setattr(
        importer_detail,
        "build_destination_catalog",
        lambda _engine: _catalog("China"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: {"kpler_watermark": "kpler-v1"},
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_maintenance_source_version",
        lambda: None,
    )
    monkeypatch.setattr(
        importer_detail,
        "build_destination_value_options",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("option construction failed")
        ),
    )

    _catalog_output, options, selected, source_context = (
        importer_detail.initialize_country_dropdown(
            1,
            "country",
            None,
            None,
        )
    )

    assert options == []
    assert selected is None
    assert source_context["source_watermark"] is None
    assert source_context["source_revision"]["status"] == "unavailable"


def test_importer_maintenance_reference_is_small_exact_and_refreshes_once(
    monkeypatch,
    persistent_importer_detail_cache,
):
    raw_data = pd.DataFrame(
        {
            "plant_name": ["Ras Laffan"],
            "country_name": ["Qatar"],
            "lng_train_name_short": ["Train 1"],
            "year": [2026],
            "month": [7],
            "year_actual_forecast": ["Forecast"],
            "total_mtpa": [0.75],
            "metric_comment": ["Planned"],
            "date": pd.to_datetime(["2026-07-01"]),
        }
    )
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: raw_data.copy(),
    )
    legacy = importer_detail.update_maintenance_table(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
    )
    version = {
        "kpler_snapshot_id": "k1",
        "kpler_watermark": "2026-07-24T00:00:00Z",
        "kpler_content_sha256": "sha",
        "kpler_row_count": 10,
        "unplanned_watermark": "u1",
        "unplanned_row_count": 1,
        "planned_watermark": "p1",
        "planned_row_count": 1,
    }
    context = importer_detail._build_importer_detail_source_context(
        "base-v1",
        maintenance_source_version=version,
    )
    snapshotted = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
        source_context=context,
    )

    assert to_json(snapshotted[:2]) == to_json(legacy[:2])
    assert snapshots.is_snapshot_reference(
        snapshotted[2],
        importer_detail.IMPORTER_MAINTENANCE_SOURCE_NAMESPACE,
    )
    assert len(to_json(snapshotted[2]).encode("utf-8")) < 1_000
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "maintenance reference reread SQL"
        ),
    )
    cached = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
        maintenance_raw_data=snapshotted[2],
        source_context=context,
    )
    assert to_json(cached[:2]) == to_json(legacy[:2])

    refreshed_fetches = []
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            refreshed_fetches.append(True) or raw_data.copy()
        ),
    )
    refresh_context = importer_detail._build_importer_detail_source_context(
        "base-v1",
        force_refresh=True,
        maintenance_source_version=version,
    )
    refreshed = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
        maintenance_raw_data=snapshotted[2],
        source_context=refresh_context,
    )
    importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mt",
        [],
        maintenance_raw_data=refreshed[2],
        source_context=refresh_context,
    )
    assert refreshed_fetches == [True]


def test_importer_maintenance_metadata_failure_does_not_run_unversioned_query(
    monkeypatch,
):
    fetches = []
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or pd.DataFrame()
        ),
    )
    initial_context = (
        importer_detail._build_importer_detail_source_context(
            "metadata-unavailable",
            maintenance_source_version=None,
        )
    )

    first = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog("China"),
        "mcm_d",
        [],
        source_context=initial_context,
    )
    importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog("China"),
        "mt",
        [],
        maintenance_raw_data=first[2],
        source_context=initial_context,
    )

    refresh_context = (
        importer_detail._build_importer_detail_source_context(
            "metadata-unavailable",
            force_refresh=True,
            maintenance_source_version=None,
        )
    )
    refreshed = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog("China"),
        "mcm_d",
        [],
        maintenance_raw_data=first[2],
        source_context=refresh_context,
    )
    importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog("China"),
        "mt",
        [],
        maintenance_raw_data=refreshed[2],
        source_context=refresh_context,
    )

    assert fetches == []
    assert first[2]["error"]
    assert refreshed[2]["error"]


def test_importer_persistent_sources_track_mapping_and_refresh_generations(
    monkeypatch,
    persistent_importer_detail_cache,
):
    allocation_source = _allocation_source()
    allocation_metadata = []
    for fingerprint in (
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ):
        metadata = copy.deepcopy(allocation_source["run_metadata"])
        metadata["mapping_fingerprint"] = fingerprint
        allocation_metadata.append(metadata)
    allocation_builds = []
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: allocation_metadata.pop(0),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        lambda *_args, **kwargs: (
            allocation_builds.append(
                kwargs["run_metadata"]["mapping_fingerprint"]
            )
            or copy.deepcopy(allocation_source)
        ),
    )

    initial_context = {"refresh_generation": None}
    first_refresh_context = {
        "refresh_generation": "refresh-generation-1"
    }
    second_refresh_context = {
        "refresh_generation": "refresh-generation-2"
    }
    source_contexts = (
        initial_context,
        initial_context,
        first_refresh_context,
        first_refresh_context,
        second_refresh_context,
    )
    allocation_refs = [
        importer_detail.refresh_origin_forecast_source(
            source_context,
            "country",
            "China",
            _catalog(),
        )
        for source_context in source_contexts
    ]

    assert allocation_builds == [
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ]
    assert allocation_refs[0]["source_key"] != allocation_refs[1][
        "source_key"
    ]
    assert allocation_refs[1]["source_key"] != allocation_refs[2][
        "source_key"
    ]
    assert allocation_refs[2]["source_key"] == allocation_refs[3][
        "source_key"
    ]
    assert allocation_refs[3]["source_key"] != allocation_refs[4][
        "source_key"
    ]

    diversion_versions = [
        {
            "diversion_watermark": "diversion-v1",
            "country_mapping_fingerprint": fingerprint,
            "location_mapping_fingerprint": "location-v1",
        }
        for fingerprint in (
            "mapping-v1",
            "mapping-v2",
            "mapping-v2",
            "mapping-v2",
            "mapping-v2",
        )
    ]
    diversion_versions_for_fetch = copy.deepcopy(diversion_versions)
    diversion_builds = []
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_diversion_source_version",
        lambda: diversion_versions.pop(0),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_diversion_payload",
        lambda aggregation, value, catalog, **kwargs: (
            diversion_builds.append(
                kwargs["source_version"][
                    "country_mapping_fingerprint"
                ]
            )
            or {
                "destination_label": value,
                "destination_countries": [value],
                "main_data": [],
                "charts_data": [],
            }
        ),
    )

    diversion_refs = [
        importer_detail.refresh_importer_diversion_source(
            source_context,
            "country",
            "China",
            _catalog(),
        )
        for source_context in source_contexts
    ]

    assert diversion_builds == [
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ]
    assert diversion_refs[0]["source_key"] != diversion_refs[1][
        "source_key"
    ]
    assert diversion_refs[1]["source_key"] != diversion_refs[2][
        "source_key"
    ]
    assert diversion_refs[2]["source_key"] == diversion_refs[3][
        "source_key"
    ]
    assert diversion_refs[3]["source_key"] != diversion_refs[4][
        "source_key"
    ]

    captured_params = {}
    monkeypatch.setattr(
        importer_detail.pd,
        "read_sql",
        lambda _query, _engine, params: (
            captured_params.update(params) or pd.DataFrame()
        ),
    )
    importer_detail._fetch_importer_diversion_rows(
        ["China"],
        diversion_versions_for_fetch[0],
    )
    assert captured_params["source_version"] == "diversion-v1"

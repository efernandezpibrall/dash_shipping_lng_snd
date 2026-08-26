# Consolidated from test_importer_detail_snapshot_refs.py.

import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import threading
import time

from dash._utils import to_json
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import pytest

from pages import importer_detail
from utils import dashboard_snapshot_cache as snapshots


def _catalog(country="China"):
    return [{
        "destination_country_name": country,
        "country": country,
        "country_display": country,
        "continent": "Asia",
        "subcontinent": "East Asia",
        "basin": "Pacific",
        "country_classification_level1": "Asia",
        "country_classification": "Importer",
        "shipping_region": "North Asia",
    }]


def _scoped_frame():
    return pd.DataFrame(
        {
            "end_date": pd.to_datetime(
                [
                    "2025-01-02",
                    "2025-02-03",
                    "2026-01-04",
                    "2026-02-05",
                ]
            ),
            "cargo_mcm": np.array([62.25, np.nan, 74.5, 83.75]),
            "origin_country": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "origin_continent_chart": [
                "Asia",
                "Oceania",
                "North America",
                "Asia",
            ],
            "origin_continent": [
                "Asia",
                "Oceania",
                "North America",
                "Asia",
            ],
            "origin_shipping_region": [
                "Arabian Gulf",
                "Australia",
                "US Gulf",
                "Southeast Asia",
            ],
            "origin_basin": [
                "Pacific",
                "Pacific",
                "Atlantic",
                "Pacific",
            ],
            "origin_subcontinent": [
                "Middle East",
                "Oceania",
                "North America",
                None,
            ],
            "origin_classification_level1": [
                "Middle East",
                "Pacific",
                "North America",
                "Southeast Asia",
            ],
            "origin_classification": [
                "Producer",
                "Producer",
                "Producer",
                "Producer",
            ],
        }
    )


def _scoped_frame_with_status():
    frame = _scoped_frame()
    frame["status"] = [
        "Delivered",
        "Scheduled",
        "Delivered",
        "In Transit",
    ]
    return frame


def _forecast_frame():
    dates = pd.to_datetime(["2026-08-01", "2026-08-02"])
    return pd.DataFrame(
        {
            "date": dates,
            "year": dates.year,
            "day_of_year": dates.dayofyear,
            "month_day": dates.strftime("%b %d"),
            "mcmd": [334.85634408602147, 333.48172043010754],
            "is_forecast": [True, True],
            "source": ["Short Term", "Short Term"],
        }
    )


def _allocation_source():
    run_metadata = {
        "run_id": "allocation-run-importer",
        "analysis_date": pd.Timestamp("2026-07-20 12:00:00"),
        "forecast_start": pd.Timestamp("2026-07-01"),
        "forecast_end": pd.Timestamp("2028-12-01"),
        "supply_scenario": "base_view",
        "split_by_contract": True,
        "woodmac_short_term_outlook": "July 2026",
        "woodmac_long_term_outlook": "June 2026",
    }
    mapping_df = pd.DataFrame(
        {
            "country": ["Qatar", "Australia"],
            "country_name": ["Qatar", "Australia"],
            "continent": ["Asia", "Oceania"],
            "shipping_region": ["Arabian Gulf", "Australia"],
            "basin": ["Pacific", "Pacific"],
            "subcontinent": ["Middle East", "Oceania"],
            "country_classification_level1": [
                "Middle East",
                "Pacific",
            ],
            "country_classification": ["Producer", "Producer"],
        }
    )
    allocation_df = pd.DataFrame(
        {
            "date": pd.to_datetime(
                [
                    "2026-07-01",
                    "2026-08-01",
                    "2027-01-01",
                    "2028-01-01",
                ]
            ),
            "origin_country": [
                "Qatar",
                "Australia",
                "Qatar",
                "Australia",
            ],
            "allocated_volume_bcm": [1.25, 1.5, 3.0, 3.5],
            "alias": ["Qatar", "Australia", "Qatar", "Australia"],
            "country_display": [
                "Qatar",
                "Australia",
                "Qatar",
                "Australia",
            ],
            "continent": ["Asia", "Oceania", "Asia", "Oceania"],
            "country": ["Qatar", "Australia", "Qatar", "Australia"],
        }
    )
    demand_totals_df = pd.DataFrame(
        {
            "date": pd.to_datetime(
                [
                    "2026-07-01",
                    "2026-08-01",
                    "2027-01-01",
                    "2028-01-01",
                ]
            ),
            "forecast_demand_bcm": [2.8, 3.0, 6.5, 7.0],
        }
    )
    return {
        "destination_countries": ["China"],
        "selected_destination_aggregation": "country",
        "run_metadata": run_metadata,
        "mapping_df": mapping_df,
        "allocation_df": allocation_df,
        "internal_allocation_df": pd.DataFrame(
            columns=["date", "internal_allocation_bcm"]
        ),
        "demand_totals_df": demand_totals_df,
    }


def _base_payload(country="China"):
    context = {
        "display_label": country,
        "destination_countries": (country,),
    }
    payload = importer_detail._import_analysis_store_payload(
        "country",
        country,
        context,
        _scoped_frame(),
        forecast_df=pd.DataFrame(
            columns=[
                "date",
                "year",
                "day_of_year",
                "month_day",
                "mcmd",
                "is_forecast",
                "source",
            ]
        ),
    )
    payload["loaded_at"] = "2026-07-24T12:00:00"
    return payload


@pytest.mark.parametrize(
    ("aggregation", "mapping_column"),
    list(
        importer_detail.IMPORTER_SELECTION_TO_MAPPING_COLUMN.items()
    ),
)
def test_importer_net_scope_uses_only_whitelisted_mapping_columns(
    aggregation,
    mapping_column,
):
    assert (
        importer_detail._importer_net_scope_mapping_column(
            aggregation
        )
        == mapping_column
    )
    assert (
        importer_detail._importer_net_scope_mapping_column(
            "untrusted_column"
        )
        == "country_name"
    )


def test_scoped_trade_sql_maps_both_sides_nets_and_aggregates(
    monkeypatch,
):
    captured = {}
    sql_result = pd.DataFrame([{
        "end_date": pd.Timestamp("2026-07-01"),
        "cargo_mcm": 12.5,
        "origin_country": "Qatar",
        "origin_continent_chart": "Asia",
        "origin_continent": "Asia",
        "origin_shipping_region": "Arabian Gulf",
        "origin_basin": "Atlantic Basin",
        "origin_subcontinent": "Middle East",
        "origin_classification_level1": "Middle East",
        "origin_classification": "Producer",
        "destination_country_name": "Russian Federation",
        "status": "Delivered",
    }])

    def read_sql(query, _engine, params=None):
        captured["statement"] = str(query)
        captured["params"] = params
        return sql_result.copy()

    monkeypatch.setattr(importer_detail.pd, "read_sql", read_sql)
    result = importer_detail._fetch_importer_scoped_trades(
        importer_detail.engine,
        ["Russian Federation"],
        delivered_only=True,
        include_destination_context=True,
        include_status=True,
        selected_destination_aggregation=(
            "country_classification_level1"
        ),
    )

    statement = captured["statement"]
    assert (
        "INNER JOIN at_lng.mappings_country destination_map"
        in statement
    )
    assert (
        "LEFT JOIN at_lng.mappings_country origin_map"
        in statement
    )
    assert "origin_map.country_name" in statement
    assert (
        "origin_map.country_classification_level1"
        in statement
    )
    assert (
        "destination_map.country_classification_level1"
        in statement
    )
    assert "SUM(COALESCE(kt.cargo_destination_cubic_meters" in statement
    assert "GROUP BY" in statement
    assert captured["params"]["destination_countries"] == (
        "Russian Federation",
    )
    assert result["origin_country"].tolist() == ["Qatar"]
    assert result["cargo_mcm"].tolist() == [12.5]


def test_allocation_sql_excludes_internal_aliases_and_nets_footer(
    monkeypatch,
):
    statements = []
    mapping_df = pd.DataFrame([
        {
            "country": "Russian Federation",
            "country_name": "Russia",
            "continent": "Europe",
            "shipping_region": "Russia",
            "basin": "Atlantic Basin",
            "subcontinent": "Eastern Europe",
            "country_classification_level1": "Russia",
            "country_classification": "Russia",
        },
        {
            "country": "Qatar",
            "country_name": "Qatar",
            "continent": "Asia",
            "shipping_region": "Arabian Gulf",
            "basin": "Atlantic Basin",
            "subcontinent": "Middle East",
            "country_classification_level1": "Middle East",
            "country_classification": "Producer",
        },
    ])
    allocation_rows = pd.DataFrame([
        {
            "date": pd.Timestamp("2026-07-01"),
            "alias": "Russian Federation",
            "origin_country": "Russia",
            "country_display": "Russia",
            "country": "Russia",
            "continent": "Europe",
            "shipping_region": "Russia",
            "basin": "Atlantic Basin",
            "subcontinent": "Eastern Europe",
            "country_classification_level1": "Russia",
            "country_classification": "Russia",
            "is_internal_flow": True,
            "allocated_volume_bcm": 0.5,
        },
        {
            "date": pd.Timestamp("2026-07-01"),
            "alias": "Qatar",
            "origin_country": "Qatar",
            "country_display": "Qatar",
            "country": "Qatar",
            "continent": "Asia",
            "shipping_region": "Arabian Gulf",
            "basin": "Atlantic Basin",
            "subcontinent": "Middle East",
            "country_classification_level1": "Middle East",
            "country_classification": "Producer",
            "is_internal_flow": False,
            "allocated_volume_bcm": 1.5,
        },
    ])
    demand_df = pd.DataFrame({
        "date": [pd.Timestamp("2026-07-01")],
        "forecast_demand_bcm": [2.0],
    })

    def read_sql(query, _engine, **_kwargs):
        statement = str(query)
        statements.append(statement)
        if "mappings_country" in statement and (
            "fundamentals_supply_allocation" not in statement
        ):
            return mapping_df.copy()
        if "fundamentals_supply_allocation_demand_detail" in statement:
            return allocation_rows.copy()
        if "fundamentals_supply_allocation_demand_summary" in statement:
            return demand_df.copy()
        raise AssertionError(statement)

    monkeypatch.setattr(importer_detail.pd, "read_sql", read_sql)
    source = importer_detail._fetch_origin_forecast_source_data(
        importer_detail.engine,
        ["Russian Federation"],
        current_date="2026-07-28",
        run_metadata={"run_id": "test-run"},
        selected_destination_aggregation="country",
    )

    assert source["allocation_df"]["country"].tolist() == ["Qatar"]
    assert source["internal_allocation_df"].to_dict("records") == [{
        "date": pd.Timestamp("2026-07-01"),
        "internal_allocation_bcm": 0.5,
    }]
    allocation_sql = next(
        statement for statement in statements
        if "fundamentals_supply_allocation_demand_detail"
        in statement
    )
    assert "destination_map.country" in allocation_sql
    assert "origin_map.country" in allocation_sql
    assert "origin_map.country_name" in allocation_sql

    _summary, footer_rows, _metadata = (
        importer_detail._build_origin_forecast_summary_from_source(
            source,
            current_date="2026-07-28",
            origin_level="origin_country_name",
        )
    )
    assert footer_rows[0]["Continent"] == (
        "NET WOODMAC DEMAND TOTAL"
    )
    assert footer_rows[1]["Continent"] == (
        "NET ALLOCATED SUPPLY TOTAL"
    )
    assert footer_rows[0]["Jul'26"] == footer_rows[1]["Jul'26"]
    assert footer_rows[2]["Jul'26"] == 0.0


@pytest.fixture
def persistent_importer_detail_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "importer-detail-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(snapshots.LOCAL_CACHE_DIR_ENV, str(cache_directory))
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    monkeypatch.setattr(
        importer_detail,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: "2026-07-24T00:00:00Z",
    )
    monkeypatch.setattr(
        importer_detail,
        "fetch_woodmac_country_import_forecast_data",
        lambda _countries: pd.DataFrame(),
    )
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_reference(monkeypatch, country="China", frame=None):
    scoped_frame = frame.copy() if frame is not None else _scoped_frame()
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: scoped_frame.copy(),
    )
    reference = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        country,
        _catalog(country),
    )
    return _base_payload(country), reference


def _delete_or_corrupt_reference(reference, mode):
    stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if mode == "missing":
        stores.cache.delete(record_key, retry=True)
    else:
        stores.cache.set(record_key, b"corrupt", retry=True)
    snapshots.clear_local_snapshots()


def _workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: [
            [
                (cell.value, cell.data_type, cell.number_format)
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]
        for worksheet in workbook.worksheets
    }


@pytest.mark.parametrize("country", ["China", "Kuwait"])
def test_base_loader_emits_small_resolvable_reference_and_preserves_payload(
    country,
    monkeypatch,
    persistent_importer_detail_cache,
):
    payload, reference = _load_reference(monkeypatch, country)

    assert snapshots.is_snapshot_reference(
        reference,
        importer_detail.IMPORTER_DETAIL_BASE_NAMESPACE,
    )
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert len(to_json(reference).encode("utf-8")) < 50_000

    legacy = importer_detail._resolve_import_analysis_base_data(payload)
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    restarted = importer_detail._resolve_import_analysis_base_data(reference)

    pd.testing.assert_frame_equal(
        restarted[0],
        legacy[0],
        check_dtype=True,
        check_exact=True,
    )
    assert list(restarted[0].columns) == list(legacy[0].columns)
    assert restarted[0].isna().equals(legacy[0].isna())
    assert restarted[1:] == legacy[1:]


def test_allocation_reference_is_small_exact_and_control_paths_are_sql_free(
    monkeypatch,
    persistent_importer_detail_cache,
):
    source_data = _allocation_source()
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_data["run_metadata"]),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        lambda *_args, **_kwargs: copy.deepcopy(source_data),
    )

    reference = importer_detail.refresh_origin_forecast_source(
        0,
        "country",
        "China",
        _catalog(),
    )
    assert snapshots.is_snapshot_reference(
        reference,
        importer_detail.IMPORTER_ALLOCATION_SOURCE_NAMESPACE,
    )
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    resolved = importer_detail._resolve_origin_forecast_source_data(
        reference
    )
    for frame_key in (
        "mapping_df",
        "allocation_df",
        "internal_allocation_df",
        "demand_totals_df",
    ):
        pd.testing.assert_frame_equal(
            resolved[frame_key],
            source_data[frame_key],
            check_exact=True,
        )

    monkeypatch.setattr(
        importer_detail,
        "fetch_origin_forecast_summary_data",
        lambda *_args, **_kwargs: pytest.fail(
            "allocation controls queried PostgreSQL"
        ),
    )
    levels = [
        "origin_country_name",
        "origin_basin",
        "origin_subcontinent",
        "origin_classification_level1",
        "origin_classification",
    ]
    for level in levels:
        direct = importer_detail.update_origin_forecast_summary_table(
            "country",
            "China",
            ["Asia"],
            _catalog(),
            level,
            "mcm_d",
            source_data,
        )
        cached = importer_detail.update_origin_forecast_summary_table(
            "country",
            "China",
            ["Asia"],
            _catalog(),
            level,
            "mcm_d",
            reference,
        )
        assert to_json(cached) == to_json(direct)


def test_allocation_demand_only_source_matches_legacy_for_every_origin_level(
    monkeypatch,
):
    source_fixture = _allocation_source()
    raw_empty_allocation = pd.DataFrame(
        columns=[
            "date",
            "origin_country",
            "destination",
            "allocated_volume_bcm",
        ]
    )

    def read_sql(query, _engine, **_kwargs):
        statement = str(query)
        if "mappings_country" in statement:
            return source_fixture["mapping_df"].copy()
        if "fundamentals_supply_allocation_demand_summary" in statement:
            return source_fixture["demand_totals_df"].copy()
        if "fundamentals_supply_allocation_demand_detail" in statement:
            return raw_empty_allocation.copy()
        raise AssertionError(statement)

    monkeypatch.setattr(importer_detail.pd, "read_sql", read_sql)
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_fixture["run_metadata"]),
    )
    source_data = importer_detail._fetch_origin_forecast_source_data(
        importer_detail.engine,
        ["China"],
        run_metadata=copy.deepcopy(source_fixture["run_metadata"]),
    )

    levels = [
        "origin_country_name",
        "origin_shipping_region",
        "continent_origin_name",
        "origin_basin",
        "origin_subcontinent",
        "origin_classification_level1",
        "origin_classification",
    ]
    for level in levels:
        legacy = importer_detail.fetch_origin_forecast_summary_data(
            importer_detail.engine,
            ["China"],
            origin_level=level,
        )
        cached = importer_detail._build_origin_forecast_summary_from_source(
            source_data,
            origin_level=level,
        )
        pd.testing.assert_frame_equal(
            cached[0],
            legacy[0],
            check_exact=True,
        )
        assert cached[1] == legacy[1]
        assert cached[2] == legacy[2]


def test_allocation_no_run_uses_one_metadata_read_and_consistent_identity(
    monkeypatch,
    persistent_importer_detail_cache,
):
    metadata_reads = []

    def no_run(_engine):
        metadata_reads.append(True)
        if len(metadata_reads) > 1:
            raise AssertionError("allocation metadata was queried twice")
        return None

    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        no_run,
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        lambda *_args, **_kwargs: pytest.fail(
            "no-run refresh invoked the source loader"
        ),
    )

    reference = importer_detail.refresh_origin_forecast_source(
        0,
        "country",
        "China",
        _catalog(),
    )
    current_month = pd.Timestamp.today().replace(day=1).date().isoformat()
    expected_key = snapshots.build_source_key(
        importer_detail.IMPORTER_ALLOCATION_SOURCE_NAMESPACE,
        None,
        None,
        "country",
        ["China"],
        current_month,
    )
    assert metadata_reads == [True]
    assert reference["source_key"] == expected_key
    resolved = importer_detail._resolve_origin_forecast_source_data(
        reference
    )
    assert resolved["run_metadata"] is None
    assert resolved["allocation_df"].empty
    assert resolved["demand_totals_df"].empty
    manifest = snapshots._LOCAL_MANIFESTS[
        (
            reference["namespace"],
            reference["source_key"],
            reference["revision"],
        )
    ]
    assert manifest["run_id"] is None
    output = importer_detail.update_origin_forecast_summary_table(
        "country",
        "China",
        [],
        _catalog(),
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        reference,
    )
    assert "No compatible WoodMac supply-allocation SQL run" in to_json(
        output
    )


def test_route_reference_is_small_exact_and_controls_and_export_are_sql_free(
    monkeypatch,
    persistent_importer_detail_cache,
):
    route_source_version = {
        "kpler_watermark": "kpler-v1",
        "distance_watermark": "distance-v1",
    }
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_route_source_version",
        lambda: copy.deepcopy(route_source_version),
    )
    processed_df = pd.DataFrame(
        {
            "year": [2025, 2025, 2026, 2026],
            "month": [1, 4, 7, 10],
            "season": ["W", "S", "S", "W"],
            "quarter": ["Q1", "Q2", "Q3", "Q4"],
            "voyage_id": ["v1", "v2", "v3", "v4"],
            "origin_country_name": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "distanceDirect": [100.0, 100.0, 100.0, 100.0],
            "distanceViaSuez": [120.0, np.nan, 120.0, np.nan],
            "distanceViaPanama": [np.nan, 130.0, 130.0, np.nan],
            "selected_route": [
                "Direct",
                "ViaPanama",
                "ViaSuez",
                "Direct",
            ],
        }
    )
    mapping_df = pd.DataFrame(
        {
            "country": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "country_name": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "continent": [
                "Asia",
                "Oceania",
                "North America",
                "Asia",
            ],
            "basin": ["Pacific", "Pacific", "Atlantic", "Pacific"],
            "shipping_region": [
                "Arabian Gulf",
                "Australia",
                "US Gulf",
                "Southeast Asia",
            ],
            "subcontinent": [
                "Middle East",
                "Oceania",
                "North America",
                "Southeast Asia",
            ],
            "country_classification_level1": [
                "Middle East",
                "Pacific",
                "North America",
                "Southeast Asia",
            ],
            "country_classification": [
                "Producer",
                "Producer",
                "Producer",
                "Producer",
            ],
        }
    )
    source_data = {
        "destination_countries": ["China"],
        "processed_df": processed_df,
        "mapping_df": mapping_df,
    }
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_route_source_payload",
        lambda _countries: copy.deepcopy(source_data),
    )
    reference = importer_detail.refresh_importer_route_analysis_source(
        "country",
        "China",
        _catalog(),
        0,
    )
    assert snapshots.is_snapshot_reference(
        reference,
        importer_detail.IMPORTER_ROUTE_SOURCE_NAMESPACE,
    )
    assert len(to_json(reference).encode("utf-8")) < 10_000

    monkeypatch.setattr(
        importer_detail,
        "process_trade_and_distance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "route controls or export queried PostgreSQL"
        ),
    )
    monkeypatch.setattr(
        importer_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pytest.fail(
            "route controls queried mappings"
        ),
    )
    direct = importer_detail.update_route_analysis_charts_and_tables(
        "Year",
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "country",
        "China",
        _catalog(),
        source_data,
    )
    cached = importer_detail.update_route_analysis_charts_and_tables(
        "Year",
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "country",
        "China",
        _catalog(),
        reference,
    )
    assert to_json(cached) == to_json(direct)

    direct_export = (
        importer_detail.export_importer_route_analysis_to_excel(
            1,
            "Year",
            importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
            "country",
            "China",
            _catalog(),
            source_data,
        )
    )
    cached_export = (
        importer_detail.export_importer_route_analysis_to_excel(
            1,
            "Year",
            importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
            "country",
            "China",
            _catalog(),
            reference,
        )
    )
    assert _workbook_cells(cached_export) == _workbook_cells(
        direct_export
    )


def test_diversion_reference_is_small_exact_and_sql_free(
    monkeypatch,
    persistent_importer_detail_cache,
):
    source_data = {
        "destination_label": "China",
        "destination_countries": ["China"],
        "main_data": [{
            "Diversion date": "2026-06-01",
            "Vessel": "Test Vessel",
            "State": "Loaded",
            "Charterer": "Test Charterer",
            "Cubic Meters": 174500.0,
            "Origin location": "Sabine Pass",
            "Origin country": "United States",
            "Origin date": "2026-05-20",
            "Diverted from location": "Gate",
            "Diverted from country": "Netherlands",
            "Diverted from date": "2026-06-10",
            "New destination location": "Tianjin",
            "New destination country": "China",
            "New destination date": "2026-06-20",
            "Added shipping days": 10,
        }],
        "charts_data": [{
            "Diversion_month": "2026-06-01 00:00:00",
            "basin_combo": "Atlantic -> Pacific",
            "region_combo": "Northwest Europe -> North Asia",
            "country_combo": "Netherlands -> China",
            "Added shipping days": 10,
            "Cubic Meters": 174500.0,
        }],
    }
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_diversion_source_version",
        lambda: "diversion-v1",
    )
    captured_versions = []

    def build_diversion(*_args, **kwargs):
        captured_versions.append(kwargs.get("source_version"))
        return copy.deepcopy(source_data)

    monkeypatch.setattr(
        importer_detail,
        "_build_importer_diversion_payload",
        build_diversion,
    )
    reference = importer_detail.refresh_importer_diversion_source(
        0,
        "country",
        "China",
        _catalog(),
    )
    assert snapshots.is_snapshot_reference(
        reference,
        importer_detail.IMPORTER_DIVERSION_SOURCE_NAMESPACE,
    )
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert captured_versions == ["diversion-v1"]
    assert to_json(importer_detail.update_diversion_ui(None, "basin_combo")) == to_json(
        importer_detail.update_diversion_ui(
            {"main_data": [], "charts_data": []},
            "basin_combo",
        )
    )

    monkeypatch.setattr(
        importer_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pytest.fail(
            "diversion controls and export must not query SQL"
        ),
    )
    direct_ui = importer_detail.update_diversion_ui(
        source_data,
        "country_combo",
    )
    cached_ui = importer_detail.update_diversion_ui(
        reference,
        "country_combo",
    )
    assert to_json(cached_ui) == to_json(direct_ui)
    direct_export = (
        importer_detail.export_importer_diversion_summary_to_excel(
            1,
            source_data,
            direct_ui[1],
        )
    )
    cached_export = (
        importer_detail.export_importer_diversion_summary_to_excel(
            1,
            reference,
            cached_ui[1],
        )
    )
    assert _workbook_cells(cached_export) == _workbook_cells(
        direct_export
    )

    corrupt_reference = dict(reference, revision="missing")
    corrupt_ui = importer_detail.update_diversion_ui(
        corrupt_reference,
        "basin_combo",
    )
    assert importer_detail.IMPORTER_DIVERSION_RECOVERY_MESSAGE in to_json(
        corrupt_ui
    )
    with pytest.raises(importer_detail.PreventUpdate):
        importer_detail.export_importer_diversion_summary_to_excel(
            1,
            corrupt_reference,
            cached_ui[1],
        )


@pytest.mark.parametrize("workers", [1, 4, 8])
def test_diversion_source_is_single_flight_under_concurrency(
    workers,
    monkeypatch,
    persistent_importer_detail_cache,
):
    counter = {"count": 0}
    counter_lock = threading.Lock()

    def build_payload(*_args, **_kwargs):
        with counter_lock:
            counter["count"] += 1
        time.sleep(0.04)
        return {
            "destination_label": "China",
            "destination_countries": ["China"],
            "main_data": [],
            "charts_data": [],
        }

    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_diversion_source_version",
        lambda: f"diversion-single-flight-{workers}",
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_diversion_payload",
        build_payload,
    )
    with ThreadPoolExecutor(max_workers=workers) as pool:
        references = list(
            pool.map(
                lambda _index: (
                    importer_detail.refresh_importer_diversion_source(
                        0,
                        "country",
                        "China",
                        _catalog(),
                    )
                ),
                range(workers),
            )
        )

    assert counter["count"] == 1
    assert all(
        snapshots.snapshot_is_resolvable(reference)
        for reference in references
    )
    assert len({
        (reference["source_key"], reference["revision"])
        for reference in references
    }) == 1


@pytest.mark.parametrize("mode", ["missing", "corrupt"])
def test_allocation_reference_failure_is_visible_without_sql_fallback(
    mode,
    monkeypatch,
    persistent_importer_detail_cache,
):
    source_data = _allocation_source()
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_data["run_metadata"]),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        lambda *_args, **_kwargs: copy.deepcopy(source_data),
    )
    reference = importer_detail.refresh_origin_forecast_source(
        0,
        "country",
        "China",
        _catalog(),
    )
    _delete_or_corrupt_reference(reference, mode)
    monkeypatch.setattr(
        importer_detail,
        "fetch_origin_forecast_summary_data",
        lambda *_args, **_kwargs: pytest.fail(
            "corrupt allocation reference queried PostgreSQL"
        ),
    )

    output = importer_detail.update_origin_forecast_summary_table(
        "country",
        "China",
        [],
        _catalog(),
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        reference,
    )

    assert (
        importer_detail.IMPORTER_ALLOCATION_RECOVERY_MESSAGE
        in to_json(output)
    )


def test_nonempty_forecast_round_trips_exactly_and_chart_does_not_query_sql(
    monkeypatch,
    persistent_importer_detail_cache,
):
    monkeypatch.setattr(
        importer_detail,
        "fetch_woodmac_country_import_forecast_data",
        lambda _countries: _forecast_frame(),
    )
    _, reference = _load_reference(monkeypatch)

    restarted = importer_detail._load_import_analysis_forecast_data(
        reference
    )
    pd.testing.assert_frame_equal(
        restarted,
        _forecast_frame(),
        check_dtype=True,
        check_exact=True,
    )

    monkeypatch.setattr(
        importer_detail,
        "fetch_woodmac_country_import_forecast_data",
        lambda _countries: pytest.fail("chart callback queried WoodMac"),
    )
    importer_detail.update_import_analysis_charts(
        reference,
        30,
        "mcm_d",
        ["2025", "2026"],
    )


def test_forecast_month_is_part_of_source_key(
    monkeypatch,
    persistent_importer_detail_cache,
):
    payload = _base_payload()
    payload["woodmac_import_forecast"] = _forecast_frame()
    build_calls = []
    month = {"value": "2026-07-01"}
    monkeypatch.setattr(
        importer_detail,
        "_importer_detail_forecast_month_token",
        lambda: month["value"],
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        lambda aggregation, value, context: (
            build_calls.append((aggregation, value))
            or copy.deepcopy(payload)
        ),
    )

    first = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog(),
    )
    month["value"] = "2026-08-01"
    second = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog(),
    )

    assert first["source_key"] != second["source_key"]
    assert build_calls == [("country", "China"), ("country", "China")]


def test_reference_matches_legacy_for_year_charts_kpis_mix_and_base_export(
    monkeypatch,
    persistent_importer_detail_cache,
):
    payload, reference = _load_reference(monkeypatch)
    monkeypatch.setattr(
        importer_detail,
        "fetch_woodmac_country_import_forecast_data",
        lambda _countries: pd.DataFrame(),
    )

    legacy_years = importer_detail.update_import_analysis_year_selector(
        payload,
        ["2025"],
    )
    reference_years = importer_detail.update_import_analysis_year_selector(
        reference,
        ["2025"],
    )
    assert to_json(reference_years) == to_json(legacy_years)

    legacy_charts = importer_detail.update_import_analysis_charts(
        payload,
        30,
        "mcm_d",
        ["2025", "2026"],
    )
    reference_charts = importer_detail.update_import_analysis_charts(
        reference,
        30,
        "mcm_d",
        ["2025", "2026"],
    )
    assert to_json(reference_charts) == to_json(legacy_charts)

    scoped_fetches = []

    def fetch_scoped(_engine, _countries, **kwargs):
        scoped_fetches.append(dict(kwargs))
        return _scoped_frame().copy()

    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        fetch_scoped,
    )
    monkeypatch.setattr(
        importer_detail,
        "_load_importer_country_mapping_lookup",
        lambda _engine: pd.DataFrame(),
    )
    legacy_export = importer_detail.export_import_analysis_to_excel(
        1,
        "country",
        "China",
        _catalog(),
        30,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        payload,
    )
    reference_export = importer_detail.export_import_analysis_to_excel(
        1,
        "country",
        "China",
        _catalog(),
        30,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        reference,
    )
    assert _workbook_cells(reference_export) == _workbook_cells(legacy_export)
    assert len(scoped_fetches) == 2
    assert all(call.get("delivered_only") is True for call in scoped_fetches)


def test_status_source_drives_summary_and_export_without_database_reads(
    monkeypatch,
    persistent_importer_detail_cache,
):
    destination_context = {
        "display_label": "China",
        "destination_countries": ("China",),
    }
    status_payload = importer_detail._import_analysis_store_payload(
        "country",
        "China",
        destination_context,
        _scoped_frame_with_status(),
        forecast_df=pd.DataFrame(),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        lambda *_args, **_kwargs: copy.deepcopy(status_payload),
    )
    reference = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog(),
    )

    delivered_frame = _scoped_frame_with_status()
    delivered_frame = delivered_frame[
        delivered_frame["status"].eq("Delivered")
    ].drop(columns="status")
    monkeypatch.setattr(
        importer_detail,
        "_load_importer_country_mapping_lookup",
        lambda _engine: pd.DataFrame(),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: delivered_frame.copy(),
    )
    legacy_export = importer_detail.export_import_analysis_to_excel(
        1,
        "country",
        "China",
        _catalog(),
        30,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        _base_payload(),
    )
    legacy_summary = importer_detail.update_origin_summary_table(
        "country",
        "China",
        30,
        [],
        _catalog(),
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        "levels",
        5,
        3,
        3,
        _base_payload(),
    )

    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: pytest.fail(
            "summary or export queried scoped trades"
        ),
    )
    monkeypatch.setattr(
        importer_detail,
        "_load_importer_country_mapping_lookup",
        lambda *_args, **_kwargs: pytest.fail(
            "export queried country mappings"
        ),
    )

    summary = importer_detail.update_origin_summary_table(
        "country",
        "China",
        30,
        [],
        _catalog(),
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        "levels",
        5,
        3,
        3,
        reference,
    )
    assert "Error loading data" not in to_json(summary)
    assert to_json(summary) == to_json(legacy_summary)

    cached_export = importer_detail.export_import_analysis_to_excel(
        1,
        "country",
        "China",
        _catalog(),
        30,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        reference,
    )
    assert _workbook_cells(cached_export) == _workbook_cells(legacy_export)


def test_origin_summary_rejects_cached_rows_when_hierarchy_membership_changes(
    monkeypatch,
):
    cached_context = {
        "display_label": "Asia",
        "destination_countries": ("China",),
    }
    cached_payload = importer_detail._import_analysis_store_payload(
        "continent",
        "Asia",
        cached_context,
        _scoped_frame_with_status(),
        forecast_df=pd.DataFrame(),
    )
    refreshed_catalog = _catalog()
    refreshed_catalog.append(
        {
            **_catalog("Japan")[0],
            "continent": "Asia",
        }
    )
    fetched_scopes = []

    def fetch_summary(_engine, destination_countries, *_args, **kwargs):
        fetched_scopes.append(
            (
                tuple(destination_countries),
                kwargs.get("scoped_trades_df"),
            )
        )
        return pd.DataFrame()

    monkeypatch.setattr(
        importer_detail,
        "fetch_origin_summary_data",
        fetch_summary,
    )

    importer_detail.update_origin_summary_table(
        "continent",
        "Asia",
        30,
        [],
        refreshed_catalog,
        importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
        "mcm_d",
        "levels",
        5,
        3,
        3,
        cached_payload,
    )

    assert fetched_scopes == [(("China", "Japan"), None)]


@pytest.mark.parametrize("mode", ["missing", "corrupt"])
def test_missing_or_corrupt_reference_has_explicit_recovery_and_no_raw_fallback(
    mode,
    monkeypatch,
    persistent_importer_detail_cache,
):
    _, reference = _load_reference(monkeypatch)
    _delete_or_corrupt_reference(reference, mode)

    expected_selector = (
        [{
            "label": importer_detail.IMPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE,
            "value": "__snapshot_unavailable__",
            "disabled": True,
        }],
        [],
    )
    assert (
        importer_detail.update_import_analysis_year_selector(reference, [])
        == expected_selector
    )

    charts = importer_detail.update_import_analysis_charts(
        reference,
        30,
        "mcm_d",
        [],
    )
    assert (
        charts[-1].children
        == importer_detail.IMPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE
    )
    assert charts[-1].role == "alert"

    fallback_fetches = []
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **kwargs: (
            fallback_fetches.append(kwargs) or _scoped_frame()
        ),
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Cached importer-detail data is unavailable",
    ):
        importer_detail.export_import_analysis_to_excel(
            1,
            "country",
            "China",
            _catalog(),
            30,
            importer_detail.DEFAULT_IMPORTER_ORIGIN_LEVEL,
            "mcm_d",
            reference,
        )
    assert fallback_fetches == []


def test_wrong_namespace_nonresolvable_and_malformed_refs_are_rejected(
    monkeypatch,
    persistent_importer_detail_cache,
):
    wrong_reference = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": "another-page-v1",
        "source_key": "source",
        "revision": "00000000-0000-4000-8000-000000000000",
        "shared": True,
    }
    with pytest.raises(snapshots.SnapshotUnavailable):
        importer_detail._resolve_import_analysis_base_data(wrong_reference)

    payload = _base_payload()
    local_reference = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": importer_detail.IMPORTER_DETAIL_BASE_NAMESPACE,
        "source_key": "source",
        "revision": 1,
        "shared": False,
    }
    monkeypatch.setattr(
        importer_detail,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (local_reference, payload),
    )
    with pytest.raises(snapshots.SnapshotUnavailable):
        importer_detail.refresh_import_analysis_base_data(
            0,
            "country",
            "China",
            _catalog(),
        )

    malformed_reference, _ = snapshots.get_or_build_snapshot(
        importer_detail.engine,
        namespace=importer_detail.IMPORTER_DETAIL_BASE_NAMESPACE,
        source_key="malformed-payload",
        builder=lambda: {},
        force=True,
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Cached importer-detail data is unavailable",
    ):
        importer_detail._resolve_import_analysis_base_data(
            malformed_reference
        )


@pytest.mark.parametrize("workers", [1, 2, 4])
def test_base_loader_is_single_flight_for_one_two_and_four_callers(
    workers,
    monkeypatch,
    persistent_importer_detail_cache,
):
    counter = {"count": 0}
    counter_lock = threading.Lock()

    def fetch_scoped(*_args, **_kwargs):
        with counter_lock:
            counter["count"] += 1
        time.sleep(0.04)
        return _scoped_frame()

    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        fetch_scoped,
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: f"single-flight-{workers}",
    )

    with ThreadPoolExecutor(max_workers=workers) as pool:
        references = list(
            pool.map(
                lambda _index: (
                    importer_detail.refresh_import_analysis_base_data(
                        0,
                        "country",
                        "China",
                        _catalog(),
                    )
                ),
                range(workers),
            )
        )

    assert counter["count"] == 1
    assert all(
        snapshots.snapshot_is_resolvable(reference)
        for reference in references
    )
    assert len(
        {
            (
                reference["source_key"],
                reference["revision"],
            )
            for reference in references
        }
    ) == 1


def test_global_refresh_is_checked_once_and_forces_one_rebuild(
    monkeypatch,
    persistent_importer_detail_cache,
):
    refresh_calls = []
    fetch_calls = []
    monkeypatch.setattr(
        importer_detail,
        "_was_global_refresh_triggered",
        lambda: refresh_calls.append(True) or True,
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: (
            fetch_calls.append(True) or _scoped_frame()
        ),
    )

    reference = importer_detail.refresh_import_analysis_base_data(
        1,
        "country",
        "China",
        _catalog(),
    )

    assert refresh_calls == [True]
    assert fetch_calls == [True]
    assert snapshots.snapshot_is_resolvable(reference)


def test_country_hierarchy_route_diversion_and_maintenance_remain_independent_and_exports_are_stable(
    monkeypatch,
):
    independent_callbacks = (
        importer_detail.initialize_country_dropdown,
        importer_detail.update_origin_forecast_summary_table,
        importer_detail.update_route_analysis_charts_and_tables,
        importer_detail.export_importer_route_analysis_to_excel,
        importer_detail.process_diversion_data,
        importer_detail.update_diversion_ui,
        importer_detail.export_importer_diversion_summary_to_excel,
        importer_detail.update_maintenance_table,
    )
    assert all(
        "_resolve_import_analysis_base_data"
        not in callback_function.__code__.co_names
        for callback_function in independent_callbacks
    )
    assert (
        "_resolve_import_analysis_base_data"
        in importer_detail.update_origin_summary_table.__code__.co_names
    )

    route_frame = pd.DataFrame(
        {
            "year": [2025, 2025, 2026, 2026],
            "voyage_id": ["v1", "v2", "v3", "v4"],
            "origin_country_name": [
                "Qatar",
                "Australia",
                "United States",
                "Malaysia",
            ],
            "origin_classification_level1": [
                "Middle East",
                "Pacific",
                "North America",
                "Southeast Asia",
            ],
            "distanceDirect": [100.0, 100.0, 100.0, 100.0],
            "distanceViaSuez": [120.0, np.nan, 120.0, np.nan],
            "distanceViaPanama": [np.nan, 130.0, 130.0, np.nan],
        }
    )
    monkeypatch.setattr(
        importer_detail,
        "process_trade_and_distance_data",
        lambda *_args, **_kwargs: route_frame.copy(),
    )
    route_first = importer_detail.export_importer_route_analysis_to_excel(
        1,
        "Year",
        "origin_classification_level1",
        "country",
        "China",
        _catalog(),
    )
    route_second = importer_detail.export_importer_route_analysis_to_excel(
        1,
        "Year",
        "origin_classification_level1",
        "country",
        "China",
        _catalog(),
    )
    assert _workbook_cells(route_first) == _workbook_cells(route_second)

    diversion_store = {
        "main_data": [{
            "Diversion date": "2026-07-20",
            "Vessel": "Example LNG",
            "State": "Loaded",
            "Cubic Meters": 174500.25,
            "Added shipping days": 2.5,
        }]
    }
    diversion_columns = [
        {"field": "Diversion date"},
        {"field": "Vessel"},
        {"field": "State"},
        {"field": "Cubic Meters"},
        {"field": "Added shipping days"},
    ]
    diversion_first = importer_detail.export_importer_diversion_summary_to_excel(
        1,
        diversion_store,
        diversion_columns,
    )
    diversion_second = importer_detail.export_importer_diversion_summary_to_excel(
        1,
        diversion_store,
        diversion_columns,
    )
    assert _workbook_cells(diversion_first) == _workbook_cells(
        diversion_second
    )


def test_maintenance_controls_reuse_current_raw_source_with_exact_output(
    monkeypatch,
):
    raw_data = pd.DataFrame(
        {
            "plant_name": ["Ras Laffan"],
            "country_name": ["Qatar"],
            "lng_train_name_short": ["Train 1"],
            "year": [2026],
            "month": [7],
            "year_actual_forecast": ["Forecast"],
            "total_mtpa": [0.75],
            "metric_comment": ["Planned"],
            "date": pd.to_datetime(["2026-07-01"]),
        }
    )
    fetches = []
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or raw_data.copy()
        ),
    )
    baseline = importer_detail.update_maintenance_table(
        "country",
        "China",
        _catalog(),
        "mt",
        [],
    )
    assert fetches == [True]

    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "maintenance controls reread the source"
        ),
    )
    cached = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mt",
        [],
        maintenance_raw_data=baseline[2],
    )

    assert to_json(cached[:2]) == to_json(baseline[:2])
    assert cached[2] == baseline[2]


@pytest.mark.parametrize("error_message", ["source down", ""])
def test_maintenance_source_error_retries_on_next_control_change(
    monkeypatch,
    error_message,
):
    attempts = []

    def fail_source(*_args, **_kwargs):
        attempts.append("failed")
        raise RuntimeError(error_message)

    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        fail_source,
    )
    failed = importer_detail.update_maintenance_table(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
    )
    assert failed[2]["error"] == error_message

    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            attempts.append("retried") or pd.DataFrame()
        ),
    )
    retried = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mt",
        [],
        maintenance_raw_data=failed[2],
    )

    assert attempts == ["failed", "retried"]
    assert "No supplier maintenance data available" in to_json(retried[0])
    assert "error" not in retried[2]


def test_maintenance_database_failure_reaches_retryable_error_store(
    monkeypatch,
):
    monkeypatch.setattr(
        importer_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("maintenance database down")
        ),
    )

    failed = importer_detail.update_maintenance_table(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
    )

    assert failed[2]["error"] == "maintenance database down"
    assert "Error loading maintenance data" in to_json(failed[0])


def test_source_context_refresh_generation_is_not_sticky_for_later_selections(
    monkeypatch,
    persistent_importer_detail_cache,
):
    build_calls = []
    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        lambda aggregation, value, _context: (
            build_calls.append((aggregation, value))
            or _base_payload(value)
        ),
    )
    watermark = {
        "kpler_watermark": "2026-07-24T00:00:00Z",
        "woodmac_watermark": "2026-07-23T00:00:00Z",
        "distance_watermark": "2026-07-22T00:00:00Z",
    }
    initial_context = importer_detail._build_importer_detail_source_context(
        watermark
    )
    refresh_context = importer_detail._build_importer_detail_source_context(
        watermark,
        force_refresh=True,
    )

    importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=initial_context,
    )
    importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=refresh_context,
    )
    importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "Kuwait",
        _catalog("Kuwait"),
        source_context=refresh_context,
    )
    importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=refresh_context,
    )

    assert build_calls == [
        ("country", "China"),
        ("country", "China"),
        ("country", "Kuwait"),
    ]


def test_mapping_fingerprint_invalidates_base_and_route_snapshots(
    monkeypatch,
    persistent_importer_detail_cache,
):
    base_builds = []
    route_builds = []
    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        lambda aggregation, value, _context: (
            base_builds.append((aggregation, value))
            or _base_payload(value)
        ),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_route_source_payload",
        lambda countries: (
            route_builds.append(tuple(countries))
            or {
                "destination_countries": list(countries),
                "processed_df": pd.DataFrame(),
                "mapping_df": pd.DataFrame(),
            }
        ),
    )
    source_watermark = {
        "kpler_watermark": "kpler-v1",
        "woodmac_watermark": "woodmac-v1",
        "distance_watermark": "distance-v1",
        "mapping_fingerprint": "mapping-v1",
    }
    initial_context = (
        importer_detail._build_importer_detail_source_context(
            source_watermark
        )
    )
    changed_context = (
        importer_detail._build_importer_detail_source_context(
            {
                **source_watermark,
                "mapping_fingerprint": "mapping-v2",
            }
        )
    )

    first_base = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=initial_context,
    )
    changed_base = importer_detail.refresh_import_analysis_base_data(
        0,
        "country",
        "China",
        _catalog("China"),
        source_context=changed_context,
    )
    first_route = importer_detail.refresh_importer_route_analysis_source(
        "country",
        "China",
        _catalog("China"),
        initial_context,
    )
    changed_route = importer_detail.refresh_importer_route_analysis_source(
        "country",
        "China",
        _catalog("China"),
        changed_context,
    )

    assert base_builds == [("country", "China"), ("country", "China")]
    assert route_builds == [("China",), ("China",)]
    assert first_base["source_key"] != changed_base["source_key"]
    assert first_route["source_key"] != changed_route["source_key"]


def test_destination_catalog_exposes_unavailable_source_watermark_failure(
    monkeypatch,
):
    monkeypatch.setattr(
        importer_detail,
        "build_destination_catalog",
        lambda _engine: _catalog("China"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: (_ for _ in ()).throw(RuntimeError("watermark down")),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_maintenance_source_version",
        lambda: (_ for _ in ()).throw(RuntimeError("maintenance down")),
    )

    catalog, options, selected, source_context = (
        importer_detail.initialize_country_dropdown(
            0,
            "country",
            None,
            None,
        )
    )

    assert catalog == _catalog("China")
    assert options == [{"label": "China", "value": "China"}]
    assert selected == "China"
    assert source_context["source_watermark"] is None
    assert source_context["source_revision"]["status"] == "unavailable"
    assert "unavailable" in source_context["source_revision"]["message"]


def test_destination_catalog_reuses_last_good_revision_as_stale(
    monkeypatch,
):
    monkeypatch.setattr(
        importer_detail,
        "build_destination_catalog",
        lambda _engine: _catalog("China"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: (_ for _ in ()).throw(RuntimeError("watermark down")),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_maintenance_source_version",
        lambda: {"revision": "maintenance-v1"},
    )
    previous_context = (
        importer_detail._build_importer_detail_source_context(
            {"kpler_watermark": "kpler-v1"}
        )
    )

    catalog, options, selected, source_context = (
        importer_detail.initialize_country_dropdown(
            0,
            "country",
            _catalog("China"),
            {"aggregation": "country", "value": "China"},
            previous_context,
        )
    )

    assert catalog == _catalog("China")
    assert options == [{"label": "China", "value": "China"}]
    assert selected == "China"
    assert source_context["source_watermark"] == {
        "kpler_watermark": "kpler-v1"
    }
    assert source_context["source_revision"]["status"] == "stale"


def test_importer_initializer_failure_is_explicitly_unavailable(
    monkeypatch,
):
    monkeypatch.setattr(
        importer_detail,
        "_was_global_refresh_triggered",
        lambda: True,
    )
    monkeypatch.setattr(
        importer_detail,
        "build_destination_catalog",
        lambda _engine: _catalog("China"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: {"kpler_watermark": "kpler-v1"},
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_maintenance_source_version",
        lambda: None,
    )
    monkeypatch.setattr(
        importer_detail,
        "build_destination_value_options",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("option construction failed")
        ),
    )

    _catalog_output, options, selected, source_context = (
        importer_detail.initialize_country_dropdown(
            1,
            "country",
            None,
            None,
        )
    )

    assert options == []
    assert selected is None
    assert source_context["source_watermark"] is None
    assert source_context["source_revision"]["status"] == "unavailable"


def test_importer_maintenance_reference_is_small_exact_and_refreshes_once(
    monkeypatch,
    persistent_importer_detail_cache,
):
    raw_data = pd.DataFrame(
        {
            "plant_name": ["Ras Laffan"],
            "country_name": ["Qatar"],
            "lng_train_name_short": ["Train 1"],
            "year": [2026],
            "month": [7],
            "year_actual_forecast": ["Forecast"],
            "total_mtpa": [0.75],
            "metric_comment": ["Planned"],
            "date": pd.to_datetime(["2026-07-01"]),
        }
    )
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: raw_data.copy(),
    )
    legacy = importer_detail.update_maintenance_table(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
    )
    version = {
        "kpler_snapshot_id": "k1",
        "kpler_watermark": "2026-07-24T00:00:00Z",
        "kpler_content_sha256": "sha",
        "kpler_row_count": 10,
        "unplanned_watermark": "u1",
        "unplanned_row_count": 1,
        "planned_watermark": "p1",
        "planned_row_count": 1,
    }
    context = importer_detail._build_importer_detail_source_context(
        "base-v1",
        maintenance_source_version=version,
    )
    snapshotted = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
        source_context=context,
    )

    assert to_json(snapshotted[:2]) == to_json(legacy[:2])
    assert snapshots.is_snapshot_reference(
        snapshotted[2],
        importer_detail.IMPORTER_MAINTENANCE_SOURCE_NAMESPACE,
    )
    assert len(to_json(snapshotted[2]).encode("utf-8")) < 1_000
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "maintenance reference reread SQL"
        ),
    )
    cached = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
        maintenance_raw_data=snapshotted[2],
        source_context=context,
    )
    assert to_json(cached[:2]) == to_json(legacy[:2])

    refreshed_fetches = []
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            refreshed_fetches.append(True) or raw_data.copy()
        ),
    )
    refresh_context = importer_detail._build_importer_detail_source_context(
        "base-v1",
        force_refresh=True,
        maintenance_source_version=version,
    )
    refreshed = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mcm_d",
        [],
        maintenance_raw_data=snapshotted[2],
        source_context=refresh_context,
    )
    importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog(),
        "mt",
        [],
        maintenance_raw_data=refreshed[2],
        source_context=refresh_context,
    )
    assert refreshed_fetches == [True]


def test_importer_maintenance_metadata_failure_does_not_run_unversioned_query(
    monkeypatch,
):
    fetches = []
    monkeypatch.setattr(
        importer_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or pd.DataFrame()
        ),
    )
    initial_context = (
        importer_detail._build_importer_detail_source_context(
            "metadata-unavailable",
            maintenance_source_version=None,
        )
    )

    first = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog("China"),
        "mcm_d",
        [],
        source_context=initial_context,
    )
    importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog("China"),
        "mt",
        [],
        maintenance_raw_data=first[2],
        source_context=initial_context,
    )

    refresh_context = (
        importer_detail._build_importer_detail_source_context(
            "metadata-unavailable",
            force_refresh=True,
            maintenance_source_version=None,
        )
    )
    refreshed = importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog("China"),
        "mcm_d",
        [],
        maintenance_raw_data=first[2],
        source_context=refresh_context,
    )
    importer_detail._update_maintenance_table_from_source(
        "country",
        "China",
        _catalog("China"),
        "mt",
        [],
        maintenance_raw_data=refreshed[2],
        source_context=refresh_context,
    )

    assert fetches == []
    assert first[2]["error"]
    assert refreshed[2]["error"]


def test_importer_persistent_sources_track_mapping_and_refresh_generations(
    monkeypatch,
    persistent_importer_detail_cache,
):
    allocation_source = _allocation_source()
    allocation_metadata = []
    for fingerprint in (
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ):
        metadata = copy.deepcopy(allocation_source["run_metadata"])
        metadata["mapping_fingerprint"] = fingerprint
        allocation_metadata.append(metadata)
    allocation_builds = []
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: allocation_metadata.pop(0),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        lambda *_args, **kwargs: (
            allocation_builds.append(
                kwargs["run_metadata"]["mapping_fingerprint"]
            )
            or copy.deepcopy(allocation_source)
        ),
    )

    initial_context = {"refresh_generation": None}
    first_refresh_context = {
        "refresh_generation": "refresh-generation-1"
    }
    second_refresh_context = {
        "refresh_generation": "refresh-generation-2"
    }
    source_contexts = (
        initial_context,
        initial_context,
        first_refresh_context,
        first_refresh_context,
        second_refresh_context,
    )
    allocation_refs = [
        importer_detail.refresh_origin_forecast_source(
            source_context,
            "country",
            "China",
            _catalog(),
        )
        for source_context in source_contexts
    ]

    assert allocation_builds == [
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ]
    assert allocation_refs[0]["source_key"] != allocation_refs[1][
        "source_key"
    ]
    assert allocation_refs[1]["source_key"] != allocation_refs[2][
        "source_key"
    ]
    assert allocation_refs[2]["source_key"] == allocation_refs[3][
        "source_key"
    ]
    assert allocation_refs[3]["source_key"] != allocation_refs[4][
        "source_key"
    ]

    diversion_versions = [
        {
            "diversion_watermark": "diversion-v1",
            "country_mapping_fingerprint": fingerprint,
            "location_mapping_fingerprint": "location-v1",
        }
        for fingerprint in (
            "mapping-v1",
            "mapping-v2",
            "mapping-v2",
            "mapping-v2",
            "mapping-v2",
        )
    ]
    diversion_versions_for_fetch = copy.deepcopy(diversion_versions)
    diversion_builds = []
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_diversion_source_version",
        lambda: diversion_versions.pop(0),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_diversion_payload",
        lambda aggregation, value, catalog, **kwargs: (
            diversion_builds.append(
                kwargs["source_version"][
                    "country_mapping_fingerprint"
                ]
            )
            or {
                "destination_label": value,
                "destination_countries": [value],
                "main_data": [],
                "charts_data": [],
            }
        ),
    )

    diversion_refs = [
        importer_detail.refresh_importer_diversion_source(
            source_context,
            "country",
            "China",
            _catalog(),
        )
        for source_context in source_contexts
    ]

    assert diversion_builds == [
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ]
    assert diversion_refs[0]["source_key"] != diversion_refs[1][
        "source_key"
    ]
    assert diversion_refs[1]["source_key"] != diversion_refs[2][
        "source_key"
    ]
    assert diversion_refs[2]["source_key"] == diversion_refs[3][
        "source_key"
    ]
    assert diversion_refs[3]["source_key"] != diversion_refs[4][
        "source_key"
    ]

    captured_params = {}
    monkeypatch.setattr(
        importer_detail.pd,
        "read_sql",
        lambda _query, _engine, params: (
            captured_params.update(params) or pd.DataFrame()
        ),
    )
    importer_detail._fetch_importer_diversion_rows(
        ["China"],
        diversion_versions_for_fetch[0],
    )
    assert captured_params["source_version"] == "diversion-v1"


# Consolidated from test_importers_pbd_changes.py.

import base64
from datetime import date
from io import BytesIO

from openpyxl import load_workbook
import pandas as pd
import pytest
from dash import html

from pages import importer_detail, importers


def _snapshot_pair(current_date, baseline_date=None):
    current_date = pd.Timestamp(current_date)
    pair = {
        'current_snapshot_id': 200,
        'current_snapshot_date_utc': current_date.date(),
        'current_snapshot_timestamp_utc': current_date.replace(
            hour=5,
            minute=34,
            second=12,
            microsecond=778724,
        ),
        'current_facts_retained': True,
        'baseline_snapshot_id': None,
        'baseline_snapshot_date_utc': None,
        'baseline_snapshot_timestamp_utc': None,
        'baseline_facts_retained': None,
    }
    if baseline_date is not None:
        baseline_date = pd.Timestamp(baseline_date)
        pair.update({
            'baseline_snapshot_id': 199,
            'baseline_snapshot_date_utc': baseline_date.date(),
            'baseline_snapshot_timestamp_utc': baseline_date.replace(
                hour=5,
                minute=36,
                second=57,
                microsecond=998945,
            ),
            'baseline_facts_retained': True,
        })
    return pair


@pytest.mark.parametrize(
    (
        'current_date',
        'baseline_date',
        'expected_status',
        'expected_gap',
    ),
    [
        ('2026-07-30', '2026-07-29', 'exact', 1),
        ('2026-07-27', '2026-07-24', 'exact', 1),
        ('2026-07-26', '2026-07-24', 'exact', 1),
        ('2026-07-30', '2026-07-28', 'fallback', 2),
        ('2026-07-30', None, 'unavailable', None),
    ],
)
def test_importer_source_state_selects_previous_weekday_or_fallback(
    current_date,
    baseline_date,
    expected_status,
    expected_gap,
):
    state = importers._build_importers_source_state(
        _snapshot_pair(current_date, baseline_date),
        refresh_token='refresh-1',
    )

    assert state['format'] == importers.IMPORTERS_SOURCE_STATE_FORMAT
    assert state['current_snapshot']['snapshot_id'] == 200
    assert state['current_snapshot']['snapshot_date_utc'] == current_date
    assert state['watermark'].endswith('05:34:12.778724')
    assert state['baseline_status'] == expected_status
    assert state['business_day_gap'] == expected_gap
    if baseline_date is None:
        assert state['baseline_snapshot'] is None
    else:
        assert state['baseline_snapshot']['snapshot_id'] == 199
        assert (
            state['baseline_snapshot']['snapshot_date_utc']
            == baseline_date
        )


def test_source_cache_key_is_versioned_and_includes_exact_pair():
    source_state = importers._build_importers_source_state(
        _snapshot_pair('2026-07-30', '2026-07-29')
    )
    changed_pair = _snapshot_pair('2026-07-30', '2026-07-29')
    changed_pair['baseline_snapshot_timestamp_utc'] = (
        pd.Timestamp('2026-07-29T06:00:00')
    )
    changed_state = importers._build_importers_source_state(
        changed_pair
    )

    assert importers.IMPORTERS_SOURCE_NAMESPACE.endswith('-v4')
    assert importers.IMPORTERS_PERIOD_NAMESPACE.endswith('-v4')
    assert (
        importers._importers_source_snapshot_key(source_state)
        != importers._importers_source_snapshot_key(changed_state)
    )
    assert (
        importers._importers_source_snapshot_key(source_state)
        == importers._importers_source_snapshot_key({
            **source_state,
            'refresh_token': 'ignored',
        })
    )


def test_scoped_trade_query_binds_exact_snapshot_and_bounded_window(
    monkeypatch,
):
    captured = {}

    def read_sql(query, _engine, params=None):
        captured['statement'] = str(query)
        captured['params'] = params
        return pd.DataFrame(columns=[
            'end_date',
            'cargo_mcm',
            'origin_country',
            'origin_continent_chart',
            'origin_continent',
            'origin_shipping_region',
            'origin_basin',
            'origin_subcontinent',
            'origin_classification_level1',
            'origin_classification',
            'destination_country_name',
        ])

    monkeypatch.setattr(importer_detail.pd, 'read_sql', read_sql)
    importer_detail._fetch_importer_scoped_trades(
        importer_detail.engine,
        ['France'],
        min_end_date=date(2026, 6, 30),
        max_end_date=date(2026, 7, 29),
        snapshot_timestamp_utc='2026-07-29T05:36:57.998945',
        delivered_only=True,
        include_destination_context=True,
    )

    assert (
        'kt.upload_timestamp_utc = :snapshot_timestamp_utc'
        in captured['statement']
    )
    assert 'kt."end"::date <= :max_end_date' in captured['statement']
    assert captured['params']['min_end_date'] == date(2026, 6, 30)
    assert captured['params']['max_end_date'] == date(2026, 7, 29)
    assert (
        captured['params']['snapshot_timestamp_utc']
        == '2026-07-29T05:36:57.998945'
    )


def test_catalog_ranking_query_uses_atomic_current_snapshot(
    monkeypatch,
):
    captured = {}

    def read_sql(query, _engine, params=None):
        captured['statement'] = str(query)
        captured['params'] = params
        return pd.DataFrame(columns=[
            'destination_country_name',
            'avg_30d_mcmd',
        ])

    monkeypatch.setattr(importers.pd, 'read_sql', read_sql)
    importers._fetch_importers_catalog_ranking_source_df(
        '2026-07-30T05:34:12.778724',
        '2026-07-30',
    )

    assert (
        'CAST(:snapshot_timestamp_utc AS timestamptz)'
        in captured['statement']
    )
    assert captured['params'] == {
        'snapshot_timestamp_utc': (
            '2026-07-30T05:34:12.778724'
        ),
        'ranking_start_date': date(2026, 7, 1),
        'ranking_end_date': date(2026, 7, 31),
    }


def _trade_frame(as_of_date, daily_by_origin):
    rows = []
    for flow_date in pd.date_range(
        pd.Timestamp(as_of_date) - pd.Timedelta(days=29),
        pd.Timestamp(as_of_date),
    ):
        for origin_country, daily_mcmd in daily_by_origin.items():
            rows.append({
                'end_date': flow_date,
                'cargo_mcm': float(daily_mcmd),
                'origin_country': origin_country,
                'destination_country_name': 'France',
            })
    return pd.DataFrame(rows)


def test_period_payload_uses_exact_pair_and_outer_joins_changes(
    monkeypatch,
):
    current_df = _trade_frame(
        '2026-07-30',
        {'A': 10, 'B': 5},
    )
    baseline_df = _trade_frame(
        '2026-07-29',
        {'A': 8, 'C': 4},
    )
    query_calls = []

    def fetch_scoped_trades(*_args, **kwargs):
        query_calls.append(kwargs)
        if kwargs.get('snapshot_timestamp_utc', '').startswith(
            '2026-07-29'
        ):
            return baseline_df.copy()
        return current_df.copy()

    monkeypatch.setattr(
        importers,
        '_fetch_importer_scoped_trades',
        fetch_scoped_trades,
    )
    source_state = importers._build_importers_source_state(
        _snapshot_pair('2026-07-30', '2026-07-29')
    )
    payload = importers._build_period_payload(
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        'Country',
        'origin_country_name',
        'show_all',
        45,
        source_state,
    )

    assert payload['format'] == importers.IMPORTERS_PERIOD_PAYLOAD_FORMAT
    assert payload['snapshot_comparison']['status'] == 'exact'
    assert len(query_calls) == 2
    assert query_calls[0]['snapshot_timestamp_utc'].startswith(
        '2026-07-30'
    )
    assert query_calls[0]['max_end_date'] == '2026-07-30'
    assert query_calls[1]['min_end_date'] == date(2026, 6, 30)
    assert query_calls[1]['max_end_date'] == '2026-07-29'

    records_by_origin = {
        record['country']: record
        for record in payload['show_all'][0]['records']
    }
    assert records_by_origin['A']['Δ 30D vs PBD'] == pytest.approx(2)
    assert records_by_origin['A']['Δ 7D vs PBD'] == pytest.approx(2)
    assert records_by_origin['B']['Δ 30D vs PBD'] == pytest.approx(5)
    assert records_by_origin['B']['Δ 7D vs PBD'] == pytest.approx(5)
    assert records_by_origin['C']['Δ 30D vs PBD'] == pytest.approx(-4)
    assert records_by_origin['C']['Δ 7D vs PBD'] == pytest.approx(-4)
    assert '45D' in records_by_origin['A']


def test_small_origin_grouping_reuses_current_vintage_taxonomy():
    current_df = pd.concat([
        _trade_frame('2026-07-30', {'Small': 1}),
        _trade_frame('2026-07-30', {'Large': 20}),
    ], ignore_index=True)
    current_df['origin_shipping_region'] = 'Atlantic'
    baseline_df = _trade_frame(
        '2026-07-29',
        {'Small': 50, 'Large': 20},
    )
    baseline_df['origin_shipping_region'] = 'Atlantic'

    grouped_current, grouping_config = (
        importers.group_small_importer_origin_countries(
            current_df,
            'origin_shipping_region',
            as_of_date='2026-07-30',
            return_grouping_config=True,
        )
    )
    grouped_baseline = (
        importers.group_small_importer_origin_countries(
            baseline_df,
            'origin_shipping_region',
            grouping_config=grouping_config,
        )
    )

    assert 'Rest of countries' in set(
        grouped_current['origin_country']
    )
    assert 'Rest of countries' in set(
        grouped_baseline['origin_country']
    )
    assert 'Small' not in set(grouped_baseline['origin_country'])
    assert 'Large' in set(grouped_baseline['origin_country'])


def test_rolling_windows_use_inclusive_snapshot_boundaries():
    rows = []
    for flow_date in pd.date_range('2026-06-30', '2026-07-30'):
        rows.append({
            'end_date': flow_date,
            'cargo_mcm': (
                310.0
                if flow_date == pd.Timestamp('2026-06-30')
                else 10.0
            ),
            'origin_country': 'A',
            'destination_country_name': 'France',
        })
    summary_df = (
        importers.build_importer_origin_summary_from_scoped_trades(
            pd.DataFrame(rows),
            rolling_window_days=30,
            origin_level='origin_country_name',
            current_date='2026-07-30',
        )
    )

    assert summary_df.loc[0, '30D'] == 10
    assert summary_df.loc[0, '7D'] == 10


@pytest.mark.parametrize(
    ('rolling_avg_days', 'expected_query_start'),
    [
        (1, '2021-01-01'),
        (7, '2020-12-26'),
        (30, '2020-12-03'),
        (180, '2020-07-06'),
    ],
)
def test_importer_chart_query_start_uses_exact_rolling_warmup(
    rolling_avg_days,
    expected_query_start,
):
    assert importers._get_importer_chart_query_start_date(
        rolling_avg_days
    ) == expected_query_start


def test_first_visible_importer_chart_point_has_complete_180_day_window():
    query_start = importers._get_importer_chart_query_start_date(180)
    dates = pd.date_range(query_start, '2021-01-01', freq='D')
    assert len(dates) == 180

    scoped_df = pd.DataFrame({
        'end_date': dates,
        'cargo_mcm': [180.0, *([0.0] * 179)],
        'origin_continent': ['Asia'] * 180,
    })
    demand_df = importer_detail._build_importer_total_import_df(
        scoped_df,
        rolling_window_days=180,
        current_date='2021-01-01',
        chart_start_date=query_start,
        display_start_date=importers.IMPORTER_CHART_DISPLAY_START_DATE,
    )
    continent_df = importer_detail._build_importer_continent_chart_df(
        scoped_df,
        rolling_window_days=180,
        current_date='2021-01-01',
        chart_start_date=query_start,
        display_start_date=importers.IMPORTER_CHART_DISPLAY_START_DATE,
    )

    first_demand = demand_df.iloc[0]
    first_continent = continent_df[
        continent_df['continent_origin'] == 'Asia'
    ].iloc[0]
    assert first_demand['date'] == pd.Timestamp('2021-01-01')
    assert first_demand['rolling_avg'] == pytest.approx(1.0)
    assert first_continent['date'] == pd.Timestamp('2021-01-01')
    assert first_continent['rolling_avg'] == pytest.approx(1.0)


@pytest.mark.parametrize(
    ('volume_metric', 'quantity_kind', 'precision'),
    [
        ('mcm_d', 'rate', 0),
        ('bcm', 'period_volume', 1),
        ('mt', 'period_volume', 1),
        ('mtpa', 'rate', 1),
    ],
)
def test_importer_overview_volume_metric_policy(
    volume_metric,
    quantity_kind,
    precision,
):
    assert [
        option['value'] for option in importers.VOLUME_METRIC_OPTIONS
    ] == ['mcm_d', 'bcm', 'mt', 'mtpa']
    metric_info = importers._get_importer_volume_metric_info(
        volume_metric
    )
    assert metric_info['quantity_kind'] == quantity_kind
    assert metric_info['display_precision'] == precision
    assert importers._get_importer_volume_metric_display_precision(
        volume_metric
    ) == precision
    assert importers._get_importer_volume_metric_plotly_number_format(
        volume_metric
    ) == f',.{precision}f'


@pytest.mark.parametrize(
    (
        'volume_metric',
        'expected_measure',
        'expected_latest',
        'expected_delta',
        'expected_plotly_format',
    ),
    [
        ('mcm_d', 'Rolling Average', '120', '+20', ',.0f'),
        ('bcm', 'Rolling Volume', '5.4', '+0.9', ',.1f'),
        ('mt', 'Rolling Volume', '4.0', '+0.7', ',.1f'),
        ('mtpa', 'Rolling Average', '32.2', '+5.4', ',.1f'),
    ],
)
def test_importer_chart_precision_conversion_titles_hovers_and_kpis(
    volume_metric,
    expected_measure,
    expected_latest,
    expected_delta,
    expected_plotly_format,
):
    rolling_avg_days = 45
    demand_records = [
        {
            'date': f'{year}-01-15',
            'year': str(year),
            'month_day': 'Jan 15',
            'rolling_avg': value,
            'is_forecast': False,
        }
        for year, value in ((2025, 100.0), (2026, 120.0))
    ]
    origin_records = [
        {
            **record,
            'continent_origin': 'Asia',
            'percentage': 62.5,
        }
        for record in demand_records
    ]

    demand_chart = importers.create_importer_demand_chart(
        demand_records,
        volume_metric,
        ['2025', '2026'],
        rolling_avg_days,
    )
    demand_metrics = importers.get_importer_demand_chart_header_metrics(
        demand_records,
        volume_metric,
        ['2025', '2026'],
        rolling_avg_days,
    )
    demand_current_text = importers._format_importer_chart_current_value(
        demand_metrics,
        importers.get_volume_metric_info(volume_metric)['label'],
        volume_metric,
    )
    demand_delta_text = _component_text(
        importers._build_importer_chart_delta_pill(
            'YoY',
            demand_metrics['delta_value'],
            demand_metrics['delta_pct'],
            volume_metric,
        )
    )

    origin_chart = importers.create_importer_origin_continent_chart(
        origin_records,
        'absolute',
        volume_metric,
        ['2025', '2026'],
        rolling_avg_days,
    )
    origin_kpis = importers._calculate_origin_continent_kpis(
        origin_records,
        'absolute',
        volume_metric,
        ['2025', '2026'],
        rolling_avg_days,
    )
    titles = importers.update_importer_rolling_section_titles(
        rolling_avg_days,
        volume_metric,
    )

    assert demand_chart.layout.yaxis.tickformat == expected_plotly_format
    assert all(
        f':{expected_plotly_format}' in trace.hovertemplate
        for trace in demand_chart.data
        if trace.hovertemplate
    )
    assert demand_current_text.endswith(
        f'{expected_latest} '
        f"{importers.get_volume_metric_info(volume_metric)['label']}"
    )
    assert expected_delta in demand_delta_text
    assert origin_chart.layout.yaxis.tickformat == expected_plotly_format
    assert all(
        f':{expected_plotly_format}' in trace.hovertemplate
        for trace in origin_chart.data
        if trace.hovertemplate
    )
    assert origin_kpis[0]['latest_text'] == expected_latest
    assert origin_kpis[0]['yoy_value_text'] == expected_delta
    assert titles == (
        f'LNG Demand - 45-Day {expected_measure}',
        'LNG Demand by Origin Continent - '
        f'45-Day {expected_measure}',
    )

    export_df = importers._build_chart_export_df(
        {'Global': demand_records},
        volume_metric,
        ['2026'],
        rolling_avg_days=rolling_avg_days,
    )
    export_column = importers._get_importer_rolling_metric_export_column_name(
        rolling_avg_days,
        volume_metric,
    )
    assert export_column in export_df.columns
    assert export_df.loc[0, export_column] == pytest.approx(
        demand_metrics['latest_value']
    )


@pytest.mark.parametrize('volume_metric', ['mcm_d', 'bcm', 'mt', 'mtpa'])
def test_importer_percentage_chart_is_metric_independent_and_one_decimal(
    volume_metric,
):
    records = [{
        'date': '2026-01-15',
        'year': '2026',
        'month_day': 'Jan 15',
        'continent_origin': 'Asia',
        'rolling_avg': 120.0,
        'percentage': 62.54,
        'is_forecast': False,
    }]
    chart = importers.create_importer_origin_continent_chart(
        records,
        'percentage',
        volume_metric,
        ['2026'],
        45,
    )
    kpis = importers._calculate_origin_continent_kpis(
        records,
        'percentage',
        volume_metric,
        ['2026'],
        45,
    )

    assert chart.layout.yaxis.tickformat == ',.1f'
    assert all(
        '%{y:,.1f}%' in trace.hovertemplate
        for trace in chart.data
        if trace.hovertemplate
    )
    assert kpis[0]['latest_text'] == '62.5%'


def _period_payload(status='exact'):
    baseline_snapshot = (
        {
            'snapshot_id': 199,
            'snapshot_date_utc': '2026-07-29',
            'snapshot_timestamp_utc': (
                '2026-07-29T05:36:57.998945'
            ),
        }
        if status in {'exact', 'fallback'}
        else None
    )
    records = [
        {
            'continent': 'A',
            'country': 'A',
            "Q2'26": 70,
            "Q2'25": 50,
            "Jun'26": 70,
            "Jun'25": 50,
            '30D': 75,
            '30D_PP': 60,
            '30D_Y1': 50,
            'Δ 7D-30D': -5,
            'Δ 30D Y/Y': 25,
            "W29'26": 75,
            "W28'26": 65,
            "W29'25": 50,
            '7D': 70,
            '7D_PP': 60,
            '7D_Y1': 50,
            '30D_PBD_CURRENT': 75,
            '7D_PBD_CURRENT': 70,
            '30D_PBD': 50,
            '7D_PBD': 90,
            'Δ 30D vs PBD': 25,
            'Δ 7D vs PBD': -20,
        },
        {
            'continent': 'B',
            'country': 'B',
            "Q2'26": 30,
            "Q2'25": 50,
            "Jun'26": 30,
            "Jun'25": 50,
            '30D': 25,
            '30D_PP': 40,
            '30D_Y1': 50,
            'Δ 7D-30D': 5,
            'Δ 30D Y/Y': -25,
            "W29'26": 25,
            "W28'26": 35,
            "W29'25": 50,
            '7D': 30,
            '7D_PP': 40,
            '7D_Y1': 50,
            '30D_PBD_CURRENT': 25,
            '7D_PBD_CURRENT': 30,
            '30D_PBD': 50,
            '7D_PBD': 10,
            'Δ 30D vs PBD': -25,
            'Δ 7D vs PBD': 20,
        },
    ]
    if status == 'unavailable':
        for record in records:
            for column_name in (
                *importers.IMPORTER_PERIOD_PBD_CURRENT_COLUMNS,
                *importers.IMPORTER_PERIOD_PBD_REFERENCE_COLUMNS,
                *importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS,
            ):
                record[column_name] = float('nan')
    return {
        'format': importers.IMPORTERS_PERIOD_PAYLOAD_FORMAT,
        'active_grouping_mode': 'show_all',
        'show_all': [{
            'label': 'France',
            'key': 'France',
            'records': records,
        }],
        'group_small_countries': [],
        'snapshot_comparison': {
            'status': status,
            'current_snapshot': {
                'snapshot_id': 200,
                'snapshot_date_utc': '2026-07-30',
                'snapshot_timestamp_utc': (
                    '2026-07-30T05:34:12.778724'
                ),
            },
            'baseline_snapshot': baseline_snapshot,
            'business_day_gap': 2 if status == 'fallback' else 1,
        },
    }


@pytest.mark.parametrize(
    'comparison_basis',
    ['levels', 'previous_period', 'same_period_last_year'],
)
def test_pbd_columns_remain_last_and_market_share_is_pp(
    comparison_basis,
):
    component, records = importers.update_period_analysis_table(
        _period_payload(),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        'mcm_d',
        30,
        'show_all',
        'percentage',
        comparison_basis,
        5,
        3,
        3,
    )

    assert isinstance(component, html.Div)
    assert list(records[0])[-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )
    assert '30D_PBD' not in records[0]
    child_records = {
        record['Aggregation'].strip(): record
        for record in records
        if record['Aggregation'].strip() in {'A', 'B'}
    }
    assert child_records['A']['Δ 30D vs PBD'] == 25
    assert child_records['A']['Δ 7D vs PBD'] == -20
    assert child_records['B']['Δ 30D vs PBD'] == -25
    assert child_records['B']['Δ 7D vs PBD'] == 20

    grid = component.children[0]
    assert [column['field'] for column in grid.columnDefs][-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )
    grid_children = {
        record['Aggregation'].strip(): record
        for record in grid.rowData
        if record['Aggregation'].strip() in {'A', 'B'}
    }
    assert grid_children['A']['Δ 30D vs PBD'] == '+25.0 pp'
    assert grid_children['A']['Δ 7D vs PBD'] == '-20.0 pp'


def _component_text(component):
    if component is None:
        return ''
    if isinstance(component, str):
        return component
    if isinstance(component, (list, tuple)):
        return ' '.join(_component_text(item) for item in component)
    return _component_text(getattr(component, 'children', None))


def test_unavailable_baseline_shows_dash_and_explicit_warning():
    component, records = importers.update_period_analysis_table(
        _period_payload(status='unavailable'),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        'mcm_d',
        30,
        'show_all',
        'absolute',
        'levels',
        5,
        3,
        3,
    )

    assert list(records[0])[-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )
    grid = component.children[0]
    assert all(
        row['Δ 30D vs PBD'] == '—'
        and row['Δ 7D vs PBD'] == '—'
        for row in grid.rowData
    )
    assert 'PBD baseline unavailable' in _component_text(component)


@pytest.mark.parametrize('volume_metric', ['mcm_d', 'bcm', 'mt', 'mtpa'])
def test_pbd_volume_delta_uses_selected_unit(volume_metric):
    _component, records = importers.update_period_analysis_table(
        _period_payload(),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        volume_metric,
        30,
        'show_all',
        'absolute',
        'levels',
        5,
        3,
        3,
    )
    child_a = next(
        record
        for record in records
        if record['Aggregation'].strip() == 'A'
    )
    expected = {
        'mcm_d': 25.0,
        'bcm': round(25.0 * 30 / importer_detail.MCM_PER_BCM, 1),
        'mt': round(25.0 * 30 / importer_detail.MCM_PER_MT, 1),
        'mtpa': round(25.0 * importer_detail.MMTPA_PER_MCM_D, 1),
    }[volume_metric]
    assert child_a['Δ 30D vs PBD'] == pytest.approx(expected)


@pytest.mark.parametrize(
    ('volume_metric', 'expected_quarter', 'expected_30d', 'expected_7d'),
    [
        ('mcm_d', '70', '75', '70'),
        ('bcm', '6.4', '2.2', '0.5'),
        ('mt', '4.7', '1.7', '0.4'),
        ('mtpa', '18.8', '20.1', '18.8'),
    ],
)
def test_importer_period_grid_uses_metric_precision_and_period_days(
    volume_metric,
    expected_quarter,
    expected_30d,
    expected_7d,
):
    component, _records = importers.update_period_analysis_table(
        _period_payload(),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        volume_metric,
        30,
        'show_all',
        'absolute',
        'levels',
        5,
        3,
        3,
    )
    grid = component.children[0]
    child_a = next(
        record
        for record in grid.rowData
        if record['Aggregation'].strip() == 'A'
    )

    assert child_a["Q2'26"] == expected_quarter
    assert child_a['30D'] == expected_30d
    assert child_a['7D'] == expected_7d
    if volume_metric in {'bcm', 'mt'}:
        assert child_a['Δ 7D-30D'] == '—'
        assert 'different horizons' in _component_text(component)
    else:
        assert child_a['Δ 7D-30D'] != '—'


def test_importer_period_delta_uses_underlying_precision_before_rounding():
    converted = importers._convert_importer_period_absolute_volume_metric(
        pd.DataFrame([{
            'Importer': 'Global',
            'Aggregation': '',
            '30D_PBD_CURRENT': 2.0,
            '30D_PBD': 4 / 3,
            'Δ 30D vs PBD': 999.0,
        }]),
        'bcm',
        30,
    )

    # The converted levels are 0.06 and 0.04 bcm. Preserve their true 0.02
    # difference until the display boundary; pre-rounding operands yields 0.1.
    raw_delta = converted.loc[0, 'Δ 30D vs PBD']
    assert raw_delta == pytest.approx(0.02)
    assert importers._format_importer_period_grid_value(
        raw_delta,
        view_type='absolute',
        is_delta=True,
        is_pbd_delta=True,
        volume_metric='bcm',
    ) == '0.0'


@pytest.mark.parametrize('volume_metric', ['mcm_d', 'bcm', 'mt', 'mtpa'])
def test_importer_period_percentage_is_one_decimal_and_metric_independent(
    volume_metric,
):
    component, _records = importers.update_period_analysis_table(
        _period_payload(),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        volume_metric,
        30,
        'show_all',
        'percentage',
        'levels',
        5,
        3,
        3,
    )
    grid = component.children[0]
    child_a = next(
        record
        for record in grid.rowData
        if record['Aggregation'].strip() == 'A'
    )

    assert child_a['30D'] == '75.0%'
    assert child_a['7D'] == '70.0%'
    assert child_a['Δ 30D vs PBD'] == '+25.0 pp'


def test_seven_day_period_display_has_one_window_column_and_no_self_delta():
    component, records = importers.update_period_analysis_table(
        _period_payload(),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        'bcm',
        7,
        'show_all',
        'absolute',
        'levels',
        5,
        3,
        3,
    )
    grid = component.children[0]
    fields = [column['field'] for column in grid.columnDefs]
    child_a = next(
        record
        for record in grid.rowData
        if record['Aggregation'].strip() == 'A'
    )

    assert fields.count('7D') == 1
    assert 'Δ 7D-7D' not in fields
    assert 'Δ 7D-7D' not in records[0]
    assert child_a['7D'] == '0.5'


@pytest.mark.parametrize(
    ('volume_metric', 'input_value', 'expected'),
    [
        ('mcm_d', -0.4, '0'),
        ('bcm', -0.04, '0.0'),
        ('mt', -0.04, '0.0'),
        ('mtpa', -0.04, '0.0'),
    ],
)
def test_importer_metric_formatting_never_emits_signed_zero(
    volume_metric,
    input_value,
    expected,
):
    label = importers.get_volume_metric_info(volume_metric)['label']
    current_text = importers._format_importer_chart_current_value(
        {'latest_label': 'Jan 01', 'latest_value': input_value},
        label,
        volume_metric,
    )
    delta_text = _component_text(
        importers._build_importer_chart_delta_pill(
            'YoY',
            input_value,
            -0.4,
            volume_metric,
        )
    )

    assert current_text == f'Jan 01: {expected} {label}'
    assert '-0' not in delta_text
    assert '+0' not in delta_text
    assert importers._format_origin_kpi_value(
        input_value,
        'absolute',
        volume_metric,
    ) == expected
    assert importers._format_importer_period_grid_value(
        input_value,
        view_type='absolute',
        is_delta=True,
        volume_metric=volume_metric,
    ) == expected


def test_pbd_market_share_handles_zero_denominators_as_add_remove():
    display_df = pd.DataFrame([
        {
            'Importer': '▶ France',
            'Aggregation': 'Total',
            '30D_PBD_CURRENT': 100,
            '7D_PBD_CURRENT': 0,
            '30D_PBD': 0,
            '7D_PBD': 100,
        },
        {
            'Importer': '',
            'Aggregation': '    A',
            '30D_PBD_CURRENT': 100,
            '7D_PBD_CURRENT': 0,
            '30D_PBD': 0,
            '7D_PBD': 100,
        },
    ])
    percentage_df = (
        importers._apply_importer_period_pbd_percentage_view(
            display_df
        )
    )
    percentage_df = (
        importers._recalculate_importer_period_pbd_deltas(
            percentage_df
        )
    )

    assert percentage_df.loc[1, 'Δ 30D vs PBD'] == 100
    assert percentage_df.loc[1, 'Δ 7D vs PBD'] == -100


def test_period_export_keeps_pbd_columns_at_end(monkeypatch):
    captured = {}

    def send_export(export_df, filename_prefix, sheet_name):
        captured['columns'] = list(export_df.columns)
        captured['sheet_name'] = sheet_name
        return {'ok': True}

    monkeypatch.setattr(
        importers,
        '_send_export_dataframe',
        send_export,
    )
    display_records = [{
        'Importer': 'Global',
        'Aggregation': '',
        '30D': 10,
        '7D': 12,
        'Δ 30D vs PBD': 2,
        'Δ 7D vs PBD': -1,
    }]
    result = importers.export_period_analysis_to_excel(
        1,
        display_records,
        'origin_shipping_region',
        30,
        'absolute',
        'levels',
    )

    assert result == {'ok': True}
    assert captured['sheet_name'] == 'Period Analysis'
    assert captured['columns'][-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )


def test_period_excel_sheet_reconciles_rendered_pbd_values():
    display_records = [{
        'Importer': 'Global',
        'Aggregation': '',
        '30D': 10,
        '7D': 12,
        'Δ 30D vs PBD': 2,
        'Δ 7D vs PBD': -1,
    }]
    download = importers.export_period_analysis_to_excel(
        1,
        display_records,
        'origin_shipping_region',
        30,
        'absolute',
        'levels',
    )
    workbook = load_workbook(
        BytesIO(base64.b64decode(download['content']))
    )
    worksheet = workbook['Period Analysis']
    rows = list(worksheet.iter_rows(values_only=True))

    assert list(rows[0])[-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )
    assert rows[1][-2:] == ('+2', '-1')


# Consolidated from test_importers_snapshot_refs.py.

import base64
import copy
from concurrent.futures import ThreadPoolExecutor
import inspect
from io import BytesIO
import statistics
import threading
import time

from dash import html, no_update
from dash._callback import GLOBAL_CALLBACK_MAP
from dash._utils import to_json
from flask import Flask, Response
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import pytest

from pages import importers
from utils import dashboard_snapshot_cache as snapshots


def _make_overview_payload(
    *,
    entity_count=2,
    years=(2025, 2026),
    points_per_year=6,
    continents=("Africa", "Asia"),
):
    chart_entities = [
        {
            "key": f"Importer {index}",
            "label": f"Importer {index}",
            "destination_countries": [f"Country {index}"],
            "avg_30d_mcmd": float(20 - index),
            "is_global": index == 0,
        }
        for index in range(entity_count)
    ]
    table_entities = [
        dict(entity)
        for entity in chart_entities
        if not entity.get("is_global")
    ] or [dict(chart_entities[0])]
    demand_records = {}
    origin_records = {}
    for entity_index, entity in enumerate(chart_entities):
        demand_rows = []
        origin_rows = []
        for year in years:
            dates = pd.date_range(
                f"{year}-01-01",
                periods=points_per_year,
                freq="D",
            )
            for day_index, date in enumerate(dates):
                base_value = float(
                    40 + entity_index * 3 + day_index
                )
                demand_rows.append({
                    "date": date.isoformat(),
                    "year": str(year),
                    "month_day": date.strftime("%b %d"),
                    "rolling_avg": base_value,
                    "is_forecast": year == max(years),
                })
                for continent_index, continent in enumerate(continents):
                    origin_rows.append({
                        "date": date.isoformat(),
                        "year": str(year),
                        "month_day": date.strftime("%b %d"),
                        "continent_origin": continent,
                        "rolling_avg": base_value / len(continents),
                        "percentage": float(
                            100 / len(continents)
                            + continent_index
                        ),
                        "is_forecast": year == max(years),
                    })
        demand_records[entity["label"]] = demand_rows
        origin_records[entity["label"]] = origin_rows
    return {
        "chart_entities": chart_entities,
        "table_entities": table_entities,
        "demand_cube": snapshots.pack_record_mapping(demand_records),
        "origin_cube": snapshots.pack_record_mapping(origin_records),
    }


def _make_source_payload():
    return {
        "catalog_df": pd.DataFrame({
            "destination_country_name": ["Country 0", "Country 1"],
            "country_display": ["Country 0", "Country 1"],
        }),
        "ranking_df": pd.DataFrame({
            "destination_country_name": ["Country 0", "Country 1"],
            "country_display": ["Country 0", "Country 1"],
            "avg_30d_mcmd": [20.0, 19.0],
        }),
        "scoped_trades_df": pd.DataFrame({"scope": [1]}),
    }


def _install_overview_builders(
    monkeypatch,
    overview_payload,
    source_payload=None,
):
    source_payload = source_payload or _make_source_payload()
    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        lambda: copy.deepcopy(source_payload),
    )
    monkeypatch.setattr(
        importers,
        "_build_importers_overview_payload_from_source",
        lambda *_args: copy.deepcopy(overview_payload),
    )


def _build_realistic_typed_overview_payload(
    monkeypatch,
    *,
    include_codec_fields,
    timezone_aware,
):
    entity = {
        "key": "Importer 1",
        "label": "Importer 1",
        "destination_countries": ["Country 1"],
        "avg_30d_mcmd": 20.0,
        "is_global": False,
    }
    dates = [
        pd.Timestamp(
            "2026-01-01 12:30:15.123456789",
            tz="Europe/London" if timezone_aware else None,
        ),
        pd.Timestamp(
            "2026-01-02 12:30:15.987654321",
            tz="Europe/London" if timezone_aware else None,
        ),
    ]
    demand_columns = {
        "date": pd.Series(dates, dtype=object),
        "year": pd.Series(
            [np.int16(2026), np.int16(2026)],
            dtype=object,
        ),
        "month_day": ["Jan 01", "Jan 02"],
        "rolling_avg": pd.Series(
            [np.float32(12.5), np.float32(13.25)],
            dtype=object,
        ),
        "is_forecast": pd.Series(
            [np.bool_(False), np.bool_(True)],
            dtype=object,
        ),
    }
    if include_codec_fields:
        demand_columns.update({
            "missing_timestamp": pd.Series(
                [pd.NaT, pd.NaT],
                dtype=object,
            ),
            "numpy_timestamp": pd.Series(
                [
                    np.datetime64(
                        "2026-01-01T12:30:15.123456789",
                        "ns",
                    ),
                    np.datetime64("NaT", "ns"),
                ],
                dtype=object,
            ),
            "pandas_missing": pd.Series(
                [pd.NA, pd.NA],
                dtype=object,
            ),
            "nullable_text": pd.Series(
                [None, "available"],
                dtype=object,
            ),
            "numpy_int": pd.Series(
                [np.int32(7), np.int32(8)],
                dtype=object,
            ),
            "numpy_float": pd.Series(
                [np.float64(1.25), np.float64(np.nan)],
                dtype=object,
            ),
            "numpy_bool": pd.Series(
                [np.bool_(True), np.bool_(False)],
                dtype=object,
            ),
        })
    demand_df = pd.DataFrame(demand_columns)
    origin_df = demand_df.assign(
        continent_origin=["Asia", "Americas"],
        percentage=[62.5, 37.5],
    )

    monkeypatch.setattr(
        importers,
        "_build_destination_entities",
        lambda *_args, **_kwargs: [copy.deepcopy(entity)],
    )
    monkeypatch.setattr(
        importers,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: pd.DataFrame({"scope": [1]}),
    )
    monkeypatch.setattr(
        importers,
        "_filter_scoped_trades_for_entity",
        lambda scoped_df, *_args: scoped_df,
    )
    monkeypatch.setattr(
        importers,
        "_build_importer_total_import_df",
        lambda *_args, **_kwargs: demand_df.copy(),
    )
    monkeypatch.setattr(
        importers,
        "_build_importer_continent_chart_df",
        lambda *_args, **_kwargs: origin_df.copy(),
    )
    payload = importers._build_importers_overview_payload_from_source(
        _make_source_payload(),
        "Country",
        30,
    )
    if include_codec_fields:
        cube = payload["demand_cube"]
        column_indexes = {
            column: cube["columns"].index(column)
            for column in (
                "year",
                "rolling_avg",
                "is_forecast",
                "numpy_timestamp",
                "pandas_missing",
                "numpy_int",
                "numpy_float",
                "numpy_bool",
            )
        }
        typed_values = [
            {
                "year": np.int16(2026),
                "rolling_avg": np.float32(12.5),
                "is_forecast": np.bool_(False),
                "numpy_timestamp": np.datetime64(
                    "2026-01-01T12:30:15.123456789",
                    "ns",
                ),
                "pandas_missing": pd.NA,
                "numpy_int": np.int32(7),
                "numpy_float": np.float64(1.25),
                "numpy_bool": np.bool_(True),
            },
            {
                "year": np.int16(2026),
                "rolling_avg": np.float32(13.25),
                "is_forecast": np.bool_(True),
                "numpy_timestamp": np.datetime64("NaT", "ns"),
                "pandas_missing": pd.NA,
                "numpy_int": np.int32(8),
                "numpy_float": np.float64(np.nan),
                "numpy_bool": np.bool_(False),
            },
        ]
        for row, values in zip(cube["rows"], typed_values):
            for column, value in values.items():
                row[column_indexes[column]] = value
    return payload


@pytest.fixture
def persistent_importer_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "importers-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_representative_references(monkeypatch):
    overview_payload = _make_overview_payload()
    _install_overview_builders(monkeypatch, overview_payload)
    overview_stores = importers.refresh_overview_data(
        {"watermark": "2026-07-24T00:00:00"},
        "Country",
        30,
    )
    return overview_payload, overview_stores


def _load_legacy_v1_overview_references(overview_payload):
    reference, _payload = snapshots.get_or_build_snapshot(
        importers.engine,
        namespace="importers-overview-v1",
        source_key="legacy-v1-migration-guard-fixture",
        builder=lambda: (
            importers._prepare_importers_overview_snapshot_payload(
                copy.deepcopy(overview_payload)
            )
        ),
    )
    return (
        snapshots.with_snapshot_slot(reference, "chart_entities"),
        snapshots.with_snapshot_slot(reference, "table_entities"),
        snapshots.with_snapshot_slot(reference, "demand_cube"),
        snapshots.with_snapshot_slot(reference, "origin_cube"),
    )


def test_importer_overview_emits_small_resolvable_references_and_survives_restart(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload, overview_stores = (
        _load_representative_references(monkeypatch)
    )

    for store in overview_stores:
        assert snapshots.is_snapshot_reference(store)
        assert snapshots.snapshot_is_resolvable(store)
        assert len(to_json(store).encode("utf-8")) < 10_000
    assert len(to_json(overview_stores).encode("utf-8")) < 50_000

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()

    assert importers._resolve_importers_entities_store(
        overview_stores[0],
        "chart_entities",
    ) == overview_payload["chart_entities"]
    assert importers._resolve_importers_entities_store(
        overview_stores[1],
        "table_entities",
    ) == overview_payload["table_entities"]
    assert importers._resolve_importers_chart_store(
        overview_stores[2]
    ) == snapshots.unpack_record_mapping(
        overview_payload["demand_cube"]
    )
    assert importers._resolve_importers_chart_store(
        overview_stores[3]
    ) == snapshots.unpack_record_mapping(
        overview_payload["origin_cube"]
    )


def test_legacy_v1_references_show_recovery_while_v2_references_work(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload, v2_stores = _load_representative_references(
        monkeypatch
    )
    v1_stores = _load_legacy_v1_overview_references(
        overview_payload
    )
    assert all(
        snapshots.snapshot_is_resolvable(store)
        for store in v1_stores
    )
    assert all(
        store["namespace"] == "importers-overview-v1"
        for store in v1_stores
    )

    options, selected = (
        importers.update_demand_year_selector_options(
            v1_stores[2],
            ["2026"],
        )
    )
    assert selected == []
    assert (
        options[0]["label"]
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )

    chart_notice = importers.update_demand_charts(
        v1_stores[2],
        v2_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(chart_notice, html.Div)
    assert (
        chart_notice.children
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )

    entity_notice = importers.update_demand_charts(
        v2_stores[2],
        v1_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(entity_notice, html.Div)
    assert (
        entity_notice.children
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )

    v2_options, v2_selected = (
        importers.update_demand_year_selector_options(
            v2_stores[2],
            ["2026"],
        )
    )
    assert v2_selected == ["2026"]
    assert any(
        option["value"] == "2026"
        for option in v2_options
    )
    v2_chart = importers.update_demand_charts(
        v2_stores[2],
        v2_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(v2_chart, html.Div)
    assert (
        v2_chart.children
        != importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )


def test_tagged_codec_preserves_real_builder_scalar_types_and_dataframe(
    monkeypatch,
):
    overview_payload = _build_realistic_typed_overview_payload(
        monkeypatch,
        include_codec_fields=True,
        timezone_aware=True,
    )
    prepared_payload = (
        importers._prepare_importers_overview_snapshot_payload(
            overview_payload
        )
    )
    decoded_cube = importers._decode_importers_json_payload(
        prepared_payload["demand_cube"],
        importers.IMPORTERS_RECORD_CUBE_FORMAT,
    )
    raw_records = snapshots.unpack_record_mapping(
        overview_payload["demand_cube"]
    )["Importer 1"]
    decoded_records = snapshots.unpack_record_mapping(
        decoded_cube
    )["Importer 1"]

    assert list(decoded_records[0]) == list(raw_records[0])
    assert isinstance(decoded_records[0]["date"], pd.Timestamp)
    assert (
        str(decoded_records[0]["date"].tz)
        == str(raw_records[0]["date"].tz)
        == "Europe/London"
    )
    assert (
        decoded_records[0]["date"].value
        == raw_records[0]["date"].value
    )
    assert decoded_records[0]["missing_timestamp"] is pd.NaT
    assert decoded_records[0]["pandas_missing"] is pd.NA
    assert decoded_records[0]["nullable_text"] is None
    assert type(decoded_records[0]["year"]) is np.int16
    assert type(decoded_records[0]["rolling_avg"]) is np.float32
    assert type(decoded_records[0]["is_forecast"]) is np.bool_
    assert type(decoded_records[0]["numpy_int"]) is np.int32
    assert type(decoded_records[0]["numpy_float"]) is np.float64
    assert type(decoded_records[0]["numpy_bool"]) is np.bool_
    assert (
        decoded_records[0]["numpy_timestamp"].dtype
        == raw_records[0]["numpy_timestamp"].dtype
    )
    assert (
        decoded_records[0]["numpy_timestamp"]
        == raw_records[0]["numpy_timestamp"]
    )
    assert np.isnat(decoded_records[1]["numpy_timestamp"])

    raw_frame = pd.DataFrame(raw_records)
    decoded_frame = pd.DataFrame(decoded_records)
    pd.testing.assert_frame_equal(
        decoded_frame,
        raw_frame,
        check_dtype=True,
        check_exact=True,
    )
    assert list(decoded_frame.columns) == list(raw_frame.columns)
    assert decoded_frame.isna().equals(raw_frame.isna())


def test_typed_reference_preserves_chart_csv_and_xlsx_exports(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload = _build_realistic_typed_overview_payload(
        monkeypatch,
        include_codec_fields=False,
        timezone_aware=False,
    )
    _install_overview_builders(monkeypatch, overview_payload)
    overview_stores = importers.refresh_overview_data(
        {"watermark": "typed-export-source"},
        "Country",
        30,
    )
    raw_demand = snapshots.unpack_record_mapping(
        overview_payload["demand_cube"]
    )
    reference_demand = importers._resolve_importers_chart_store(
        overview_stores[2]
    )

    raw_chart = importers.update_demand_charts(
        raw_demand,
        overview_payload["chart_entities"],
        "mcm_d",
        ["2026"],
        30,
    )
    reference_chart = importers.update_demand_charts(
        overview_stores[2],
        overview_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert to_json(reference_chart) == to_json(raw_chart)

    raw_export_frame = importers._build_chart_export_df(
        raw_demand,
        "mcm_d",
        ["2026"],
    )
    reference_export_frame = importers._build_chart_export_df(
        reference_demand,
        "mcm_d",
        ["2026"],
    )
    pd.testing.assert_frame_equal(
        reference_export_frame,
        raw_export_frame,
        check_dtype=True,
        check_exact=True,
    )
    assert (
        reference_export_frame.to_csv(index=False)
        == raw_export_frame.to_csv(index=False)
    )

    raw_download = importers.export_demand_to_excel(
        1,
        raw_demand,
        "mcm_d",
        ["2026"],
        30,
    )
    reference_download = importers.export_demand_to_excel(
        1,
        overview_stores[2],
        "mcm_d",
        ["2026"],
        30,
    )
    raw_workbook = load_workbook(
        BytesIO(base64.b64decode(raw_download["content"]))
    )
    reference_workbook = load_workbook(
        BytesIO(base64.b64decode(reference_download["content"]))
    )
    raw_sheet = raw_workbook["Demand"]
    reference_sheet = reference_workbook["Demand"]

    def worksheet_cells(worksheet):
        return [
            [
                (
                    cell.value,
                    cell.data_type,
                    cell.number_format,
                )
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]

    assert worksheet_cells(reference_sheet) == worksheet_cells(raw_sheet)
    headers = [cell.value for cell in raw_sheet[1]]
    date_column = headers.index("date") + 1
    raw_date_cell = raw_sheet.cell(row=2, column=date_column)
    reference_date_cell = reference_sheet.cell(
        row=2,
        column=date_column,
    )
    assert raw_date_cell.data_type == reference_date_cell.data_type == "d"
    assert (
        raw_date_cell.number_format
        == reference_date_cell.number_format
        != "General"
    )


def test_importer_overview_never_falls_back_to_raw_payloads(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload = _make_overview_payload()
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": importers.IMPORTERS_OVERVIEW_NAMESPACE,
        "source_key": "source",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        importers,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (
            non_resolvable,
            overview_payload,
        ),
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        importers.refresh_overview_data(
            {"watermark": "stable"},
            "Country",
            30,
        )


def test_legacy_and_reference_downstream_outputs_are_exactly_equal(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload, overview_stores = (
        _load_representative_references(monkeypatch)
    )
    raw_chart_entities = overview_payload["chart_entities"]
    raw_demand = snapshots.unpack_record_mapping(
        overview_payload["demand_cube"]
    )
    raw_origin = snapshots.unpack_record_mapping(
        overview_payload["origin_cube"]
    )
    selected_years = ["2025", "2026"]

    assert to_json(
        importers.update_demand_year_selector_options(
            raw_demand,
            selected_years,
        )
    ) == to_json(
        importers.update_demand_year_selector_options(
            overview_stores[2],
            selected_years,
        )
    )
    assert to_json(
        importers.update_origin_year_selector_options(
            raw_origin,
            selected_years,
        )
    ) == to_json(
        importers.update_origin_year_selector_options(
            overview_stores[3],
            selected_years,
        )
    )
    assert to_json(
        importers.update_demand_charts(
            raw_demand,
            raw_chart_entities,
            "mcm_d",
            selected_years,
            30,
        )
    ) == to_json(
        importers.update_demand_charts(
            overview_stores[2],
            overview_stores[0],
            "mcm_d",
            selected_years,
            30,
        )
    )
    assert to_json(
        importers.update_origin_continent_charts(
            raw_origin,
            raw_chart_entities,
            "mcm_d",
            selected_years,
            "absolute",
            30,
        )
    ) == to_json(
        importers.update_origin_continent_charts(
            overview_stores[3],
            overview_stores[0],
            "mcm_d",
            selected_years,
            "absolute",
            30,
        )
    )

    def capture_export(export_df, filename_prefix, sheet_name):
        return {
            "columns": list(export_df.columns),
            "records": export_df.to_dict("records"),
            "filename_prefix": filename_prefix,
            "sheet_name": sheet_name,
        }

    monkeypatch.setattr(
        importers,
        "_send_export_dataframe",
        capture_export,
    )
    assert importers.export_demand_to_excel(
        1,
        raw_demand,
        "mcm_d",
        selected_years,
        30,
    ) == importers.export_demand_to_excel(
        1,
        overview_stores[2],
        "mcm_d",
        selected_years,
        30,
    )
    assert importers.export_origin_continent_to_excel(
        1,
        raw_origin,
        "mcm_d",
        selected_years,
        "absolute",
        30,
    ) == importers.export_origin_continent_to_excel(
        1,
        overview_stores[3],
        "mcm_d",
        selected_years,
        "absolute",
        30,
    )


@pytest.mark.parametrize("corruption_mode", ["missing", "corrupt"])
def test_missing_or_corrupt_reference_shows_explicit_refresh_recovery(
    monkeypatch,
    persistent_importer_cache,
    corruption_mode,
):
    _overview_payload, overview_stores = (
        _load_representative_references(monkeypatch)
    )
    stores = snapshots._get_persistent_stores()
    reference = overview_stores[2]
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if corruption_mode == "missing":
        stores.cache.delete(record_key, retry=True)
    else:
        stores.cache.set(record_key, b"corrupt", retry=True)
    snapshots.clear_local_snapshots()

    options, selected = (
        importers.update_demand_year_selector_options(
            overview_stores[2],
            ["2026"],
        )
    )
    assert selected == []
    assert (
        options[0]["label"]
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    chart_notice = importers.update_demand_charts(
        overview_stores[2],
        overview_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(chart_notice, html.Div)
    assert (
        chart_notice.children
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        importers.export_demand_to_excel(
            1,
            overview_stores[2],
            "mcm_d",
            ["2026"],
            30,
        )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        importers.refresh_period_data(
            overview_stores[1],
            "Default",
            "Country",
            30,
            "shipping_region",
            {"watermark": "2026-07-24T00:00:00"},
        )


def test_source_cache_reuses_reads_until_watermark_or_date_changes(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload = _make_overview_payload()
    source_payload = _make_source_payload()
    source_calls = 0
    derived_calls = []

    def source_builder():
        nonlocal source_calls
        source_calls += 1
        return copy.deepcopy(source_payload)

    def derived_builder(
        _source,
        classification_mode,
        rolling_avg_days,
        _source_state=None,
    ):
        derived_calls.append((classification_mode, rolling_avg_days))
        return copy.deepcopy(overview_payload)

    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        source_builder,
    )
    monkeypatch.setattr(
        importers,
        "_build_importers_overview_payload_from_source",
        derived_builder,
    )

    source_state = {
        "watermark": "2026-07-24T00:00:00",
        "as_of_date": "2026-07-25",
    }
    first = importers.refresh_overview_data(
        source_state,
        "Country",
        30,
    )
    warm = importers.refresh_overview_data(
        source_state,
        "Country",
        30,
    )
    changed_rolling = importers.refresh_overview_data(
        source_state,
        "Country",
        45,
    )
    changed_classification = importers.refresh_overview_data(
        source_state,
        "Classification Level 1",
        45,
    )
    next_source = importers.refresh_overview_data(
        {
            "watermark": "2026-07-25T00:00:00",
            "as_of_date": "2026-07-25",
        },
        "Country",
        30,
    )

    assert source_calls == 2
    assert derived_calls == [
        ("Country", 30),
        ("Country", 45),
        ("Classification Level 1", 45),
        ("Country", 30),
    ]
    assert warm == first
    assert changed_rolling[2]["source_key"] != first[2]["source_key"]
    assert (
        changed_classification[2]["source_key"]
        != changed_rolling[2]["source_key"]
    )
    assert next_source[2]["source_key"] != first[2]["source_key"]
    assert all(
        snapshots.snapshot_is_resolvable(store)
        for store in next_source
    )


def test_explicit_refresh_forces_source_and_derived_once_per_new_state(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload = _make_overview_payload()
    source_payload = _make_source_payload()
    source_calls = 0
    derived_calls = []

    def source_builder():
        nonlocal source_calls
        source_calls += 1
        return copy.deepcopy(source_payload)

    def derived_builder(
        _source,
        classification_mode,
        rolling_avg_days,
        _source_state=None,
    ):
        derived_calls.append((classification_mode, rolling_avg_days))
        return copy.deepcopy(overview_payload)

    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        source_builder,
    )
    monkeypatch.setattr(
        importers,
        "_build_importers_overview_payload_from_source",
        derived_builder,
    )
    triggered_id = "imp-overview-source-state-store"
    monkeypatch.setattr(
        importers,
        "_importers_overview_triggered_id",
        lambda: triggered_id,
    )

    base_state = {
        "watermark": "same-watermark",
        "as_of_date": "2026-07-25",
    }
    initial = importers.refresh_overview_data(
        base_state,
        "Country",
        30,
    )
    warm = importers.refresh_overview_data(
        base_state,
        "Country",
        30,
    )
    assert warm == initial
    assert source_calls == 1
    assert derived_calls == [("Country", 30)]

    first_refresh = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-one"},
        "Country",
        30,
    )
    second_refresh = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-two"},
        "Country",
        30,
    )
    assert source_calls == 1
    assert derived_calls == [("Country", 30)]
    assert first_refresh == initial
    assert second_refresh == initial
    assert all(
        store["namespace"] == importers.IMPORTERS_OVERVIEW_NAMESPACE
        for store in first_refresh + second_refresh
    )

    triggered_id = "imp-overview-classification-mode"
    classification = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-two"},
        "Classification Level 1",
        30,
    )
    assert source_calls == 1
    assert derived_calls[-1] == ("Classification Level 1", 30)

    triggered_id = "imp-overview-rolling-window-days-input"
    rolling = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-two"},
        "Classification Level 1",
        45,
    )
    assert source_calls == 1
    assert derived_calls[-1] == ("Classification Level 1", 45)

    replay = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-two"},
        "Classification Level 1",
        45,
    )
    assert replay == rolling
    assert source_calls == 1
    assert len(derived_calls) == 3
    assert all(
        snapshots.snapshot_is_resolvable(store)
        for store in classification
    )


def test_source_state_refresh_status_is_operational_only(
    monkeypatch,
):
    monkeypatch.setattr(
        importers,
        "_fetch_importers_source_watermark",
        lambda: pd.Timestamp("2026-07-25T00:00:00"),
    )
    is_refresh = False
    monkeypatch.setattr(
        importers,
        "_was_global_refresh_triggered",
        lambda: is_refresh,
    )

    initial, initial_status = (
        importers.load_importers_overview_source_state(0)
    )
    assert initial["refresh_token"] is None
    assert initial_status["refresh_generation"] == 0

    is_refresh = True
    first, first_status = importers.load_importers_overview_source_state(
        1,
        initial,
    )
    second, second_status = importers.load_importers_overview_source_state(
        2,
        initial,
    )
    assert first is no_update
    assert second is no_update
    assert first_status["refresh_generation"] == 1
    assert second_status["refresh_generation"] == 2
    assert (
        importers._importers_source_snapshot_key(initial)
        == importers._importers_source_snapshot_key(initial)
    )


def test_importer_source_miss_revalidates_once_and_hit_does_not_requery(
    monkeypatch,
    persistent_importer_cache,
):
    source_pair = {
        "current_snapshot_id": 101,
        "current_snapshot_date_utc": "2026-07-25",
        "current_snapshot_timestamp_utc": "2026-07-25T12:00:00Z",
        "current_facts_retained": True,
    }
    source_state = importers._build_importers_source_state(
        source_pair,
        refresh_token=None,
    )
    source_calls = 0
    validation_calls = 0

    def build_source(_source_state):
        nonlocal source_calls
        source_calls += 1
        return copy.deepcopy(_make_source_payload())

    def fetch_state():
        nonlocal validation_calls
        validation_calls += 1
        return dict(source_pair)

    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        build_source,
    )
    monkeypatch.setattr(
        importers,
        "_fetch_importers_source_watermark",
        fetch_state,
    )

    first = importers._load_importers_source_snapshot(source_state)
    second = importers._load_importers_source_snapshot(source_state)

    assert first == second
    assert source_calls == 1
    assert validation_calls == 1


def test_importer_source_drift_does_not_publish_partial_snapshot(
    monkeypatch,
    persistent_importer_cache,
):
    source_pair = {
        "current_snapshot_id": 101,
        "current_snapshot_date_utc": "2026-07-25",
        "current_snapshot_timestamp_utc": "2026-07-25T12:00:00Z",
        "current_facts_retained": True,
    }
    source_state = importers._build_importers_source_state(
        source_pair,
        refresh_token=None,
    )
    changed_pair = {
        **source_pair,
        "current_snapshot_id": 102,
        "current_snapshot_timestamp_utc": "2026-07-25T13:00:00Z",
    }
    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        lambda _state: copy.deepcopy(_make_source_payload()),
    )
    monkeypatch.setattr(
        importers,
        "_fetch_importers_source_watermark",
        lambda: changed_pair,
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="changed during snapshot construction",
    ):
        importers._load_importers_source_snapshot(source_state)

    assert snapshots.get_snapshot_if_available(
        importers.engine,
        namespace=importers.IMPORTERS_SOURCE_NAMESPACE,
        source_key=importers._importers_source_snapshot_key(source_state),
    ) is None


def test_period_callback_builds_renders_and_exports_populated_payload(
    monkeypatch,
    persistent_importer_cache,
):
    period_callback = GLOBAL_CALLBACK_MAP[
        "imp-overview-period-data-store.data"
    ]
    period_inputs = {
        (item["id"], item["property"])
        for item in period_callback["inputs"]
    }
    assert (
        "imp-overview-table-entities-store",
        "data",
    ) in period_inputs
    assert (
        "global-refresh-button",
        "n_clicks",
    ) not in period_inputs
    assert period_callback["state"] == [{
        "id": "imp-overview-source-state-store",
        "property": "data",
    }]

    _overview_payload, overview_stores = (
        _load_representative_references(monkeypatch)
    )
    scoped_trade_calls = []

    def fetch_scoped_trades(*_args, **kwargs):
        scoped_trade_calls.append(kwargs)
        return pd.DataFrame({"scope": [1]})

    summary_df = pd.DataFrame([{
        "continent": "Asia",
        "country": "Qatar",
        "Q2'26": 90.0,
        "Jun'26": 30.0,
        "30D": 30.0,
        "W29'26": 7.0,
        "7D": 7.0,
    }])
    monkeypatch.setattr(
        importers,
        "_fetch_importer_scoped_trades",
        fetch_scoped_trades,
    )
    monkeypatch.setattr(
        importers,
        "_filter_scoped_trades_for_entity",
        lambda scoped_df, *_args: scoped_df,
    )
    monkeypatch.setattr(
        importers,
        "build_importer_origin_summary_from_scoped_trades",
        lambda *_args, **_kwargs: summary_df.copy(),
    )

    current_payload = importers.refresh_period_data(
        overview_stores[1],
        "Country",
        "origin_shipping_region",
            30,
            "show_all",
            {"watermark": "2026-07-24T00:00:00"},
        )
    assert snapshots.is_snapshot_reference(current_payload)
    resolved_payload = snapshots.resolve_snapshot(
        current_payload,
        importers.engine,
        expected_namespace=importers.IMPORTERS_PERIOD_NAMESPACE,
    )
    assert resolved_payload["active_grouping_mode"] == "show_all"
    assert len(resolved_payload["show_all"]) == 1
    assert resolved_payload["show_all"][0]["records"] == (
        summary_df.to_dict("records")
    )
    assert scoped_trade_calls == [{
        "delivered_only": True,
        "include_destination_context": True,
        "selected_destination_aggregation": "country",
    }]

    current_component, current_display = (
        importers.update_period_analysis_table(
            current_payload,
            [],
            overview_stores[1],
            "mcm_d",
            30,
            "show_all",
            "absolute",
            "levels",
            5,
            3,
            3,
        )
    )
    assert current_display
    assert isinstance(current_component, html.Div)
    assert current_component.className == (
        "importer-period-table-shell"
    )
    assert current_display[0]["Importer"] == importers.IMPORTER_GLOBAL_LABEL

    monkeypatch.setattr(
        importers,
        "_send_export_dataframe",
        lambda export_df, filename_prefix, sheet_name: {
            "records": export_df.to_dict("records"),
            "filename_prefix": filename_prefix,
            "sheet_name": sheet_name,
        },
    )
    export_result = importers.export_period_analysis_to_excel(
        1,
        current_display,
        "origin_shipping_region",
        30,
        "absolute",
        "levels",
    )
    page_rows = current_component.children[0].rowData
    visible_columns = list(current_display[0])
    assert export_result["records"] == [
        {
            column: page_row[column]
            for column in visible_columns
        }
        for page_row in page_rows
    ]
    assert export_result["sheet_name"] == "Period Analysis"


def test_source_build_uses_two_catalog_reads_and_defers_scoped_sql(
    monkeypatch,
):
    barrier = threading.Barrier(2)
    loader_threads = set()
    read_calls = []
    mapping_source = pd.DataFrame([
        {
            "country_name": "Alpha Display",
            "country": "Alpha",
            "continent": "Asia",
            "subcontinent": "Eastern Asia",
            "basin": "Pacific",
            "country_classification_level1": "Core",
            "country_classification": "Tier 1",
            "shipping_region": "North Asia",
        },
        {
            "country_name": "Beta Display",
            "country": "Beta",
            "continent": "Europe",
            "subcontinent": "Northern Europe",
            "basin": "Atlantic",
            "country_classification_level1": "Growth",
            "country_classification": "Tier 2",
            "shipping_region": "NWE",
        },
    ])
    catalog_ranking_source = pd.DataFrame({
        "destination_country_name": ["Alpha", "Beta"],
        "avg_30d_mcmd": [12.5, np.nan],
    })

    def load_catalog_ranking():
        read_calls.append("catalog_ranking")
        loader_threads.add(threading.get_ident())
        barrier.wait(timeout=2)
        return catalog_ranking_source.copy()

    def load_mappings():
        read_calls.append("mappings")
        loader_threads.add(threading.get_ident())
        barrier.wait(timeout=2)
        return mapping_source.copy()

    monkeypatch.setattr(
        importers,
        "_fetch_importers_catalog_ranking_source_df",
        load_catalog_ranking,
    )
    monkeypatch.setattr(
        importers,
        "_fetch_importers_mapping_source_df",
        load_mappings,
    )
    payload = importers._build_importers_source_payload()

    assert sorted(read_calls) == [
        "catalog_ranking",
        "mappings",
    ]
    assert len(loader_threads) == 2
    assert payload["scoped_trades_df"].empty
    assert payload["catalog_df"].to_dict("records") == [
        {
            "destination_country_name": "Alpha",
            "country": "Alpha",
            "country_display": "Alpha Display",
            "continent": "Asia",
            "subcontinent": "Eastern Asia",
            "basin": "Pacific",
            "country_classification_level1": "Core",
            "country_classification": "Tier 1",
            "shipping_region": "North Asia",
        },
        {
            "destination_country_name": "Beta",
            "country": "Beta",
            "country_display": "Beta Display",
            "continent": "Europe",
            "subcontinent": "Northern Europe",
            "basin": "Atlantic",
            "country_classification_level1": "Growth",
            "country_classification": "Tier 2",
            "shipping_region": "NWE",
        },
    ]
    pd.testing.assert_frame_equal(
        payload["ranking_df"],
        pd.DataFrame({
            "destination_country_name": ["Alpha"],
            "avg_30d_mcmd": [12.5],
            "country_display": ["Alpha Display"],
        }),
        check_dtype=True,
        check_exact=True,
    )


def test_mapping_lookup_preserves_all_russia_and_singapore_aliases():
    mapping_source = pd.DataFrame([
        {
            "country_name": "Russia",
            "country": "Russia",
            "continent": "Europe",
            "shipping_region": "Russia",
            "basin": "Atlantic Basin",
            "subcontinent": "Eastern Europe",
            "country_classification_level1": "Russia",
            "country_classification": "Russia",
        },
        {
            "country_name": "Russia",
            "country": "Russian Federation",
            "continent": "Europe",
            "shipping_region": "Russia",
            "basin": "Atlantic Basin",
            "subcontinent": "Eastern Europe",
            "country_classification_level1": "Russia",
            "country_classification": "Russia",
        },
        {
            "country_name": "Singapore",
            "country": "Singapore",
            "continent": "Asia",
            "shipping_region": "SE Asia",
            "basin": "Pacific Basin",
            "subcontinent": "Southeast Asia",
            "country_classification_level1": "South East Asia",
            "country_classification": "SE Asia",
        },
        {
            "country_name": "Singapore",
            "country": "Singapore Republic",
            "continent": "Asia",
            "shipping_region": "SE Asia",
            "basin": "Pacific Basin",
            "subcontinent": "Southeast Asia",
            "country_classification_level1": "South East Asia",
            "country_classification": "SE Asia",
        },
    ])

    lookup = importers._build_importer_mapping_lookup_from_source(
        mapping_source
    ).set_index("mapping_key")

    assert {
        "Russia",
        "Russian Federation",
        "Singapore",
        "Singapore Republic",
    }.issubset(lookup.index)
    assert (
        lookup.loc[
            "Russian Federation",
            "origin_classification_level1",
        ]
        == "Russia"
    )
    assert (
        lookup.loc[
            "Singapore Republic",
            "origin_classification_level1",
        ]
        == "South East Asia"
    )


def test_source_and_derived_keys_have_only_approved_dependencies():
    source_state = {
        "watermark": "revision-a",
        "as_of_date": "2026-07-25",
    }
    base_key = importers._importers_source_snapshot_key(source_state)
    assert base_key == importers._importers_source_snapshot_key({
        **source_state,
        "classification_mode": "Country",
        "rolling_avg_days": 30,
        "refresh_token": "excluded-from-key",
        "ignored": "value",
    })
    assert base_key != importers._importers_source_snapshot_key({
        **source_state,
        "watermark": "revision-b",
    })
    assert base_key != importers._importers_source_snapshot_key({
        **source_state,
        "as_of_date": "2026-07-26",
    })

    source_reference = {
        "namespace": importers.IMPORTERS_SOURCE_NAMESPACE,
        "source_key": base_key,
        "revision": "source-revision-a",
    }
    derived_key = importers._importers_overview_snapshot_key(
        source_reference,
        "Country",
        30,
    )
    assert derived_key != importers._importers_overview_snapshot_key(
        source_reference,
        "Country",
        45,
    )
    assert derived_key != importers._importers_overview_snapshot_key(
        source_reference,
        "Classification Level 1",
        30,
    )
    assert derived_key != importers._importers_overview_snapshot_key(
        {**source_reference, "revision": "source-revision-b"},
        "Country",
        30,
    )


def test_year_selectors_resolve_compact_metadata_without_decoding_cubes(
    monkeypatch,
    persistent_importer_cache,
):
    payload, stores = _load_representative_references(monkeypatch)
    expected_demand = importers.update_demand_year_selector_options(
        snapshots.unpack_record_mapping(payload["demand_cube"]),
        ["2026"],
    )
    expected_origin = importers.update_origin_year_selector_options(
        snapshots.unpack_record_mapping(payload["origin_cube"]),
        ["2026"],
    )

    def fail_if_cube_is_resolved(_charts_data):
        raise AssertionError("selector decompressed the chart cube")

    monkeypatch.setattr(
        importers,
        "_resolve_importers_chart_store",
        fail_if_cube_is_resolved,
    )
    assert to_json(importers.update_demand_year_selector_options(
        stores[2],
        ["2026"],
    )) == to_json(expected_demand)
    assert to_json(importers.update_origin_year_selector_options(
        stores[3],
        ["2026"],
    )) == to_json(expected_origin)


def test_chart_callbacks_prepare_each_entity_once_and_keep_public_signatures(
    monkeypatch,
):
    payload = _make_overview_payload(entity_count=4)
    entities = payload["chart_entities"]
    demand = snapshots.unpack_record_mapping(payload["demand_cube"])
    origin = snapshots.unpack_record_mapping(payload["origin_cube"])

    assert list(inspect.signature(
        importers.create_importer_demand_chart
    ).parameters) == [
        "data",
        "volume_metric",
        "selected_years",
        "rolling_avg_days",
    ]
    assert list(inspect.signature(
        importers.get_importer_demand_chart_header_metrics
    ).parameters) == [
        "data",
        "volume_metric",
        "selected_years",
        "rolling_avg_days",
    ]
    assert list(inspect.signature(
        importers.create_importer_origin_continent_chart
    ).parameters) == [
        "data",
        "chart_type",
        "volume_metric",
        "selected_years",
        "rolling_avg_days",
    ]
    assert list(inspect.signature(
        importers._calculate_origin_continent_kpis
    ).parameters) == [
        "data",
        "chart_type",
        "volume_metric",
        "selected_years",
        "rolling_avg_days",
    ]

    demand_calls = 0
    original_demand_prepare = (
        importers._prepare_importer_demand_chart_dataframe
    )

    def count_demand_prepare(*args, **kwargs):
        nonlocal demand_calls
        demand_calls += 1
        return original_demand_prepare(*args, **kwargs)

    monkeypatch.setattr(
        importers,
        "_prepare_importer_demand_chart_dataframe",
        count_demand_prepare,
    )
    monkeypatch.setattr(
        importers,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("metric render performed source I/O")
        ),
    )
    importers.update_demand_charts(
        demand,
        entities,
        "bcm",
        ["2025", "2026"],
        30,
    )
    assert demand_calls == len(entities)

    origin_calls = 0
    original_origin_prepare = (
        importers._prepare_importer_origin_chart_dataframe
    )

    def count_origin_prepare(*args, **kwargs):
        nonlocal origin_calls
        origin_calls += 1
        return original_origin_prepare(*args, **kwargs)

    monkeypatch.setattr(
        importers,
        "_prepare_importer_origin_chart_dataframe",
        count_origin_prepare,
    )
    importers.update_origin_continent_charts(
        origin,
        entities,
        "bcm",
        ["2025", "2026"],
        "absolute",
        30,
    )
    assert origin_calls == len(entities)


def test_overview_callback_is_triggered_once_through_source_state():
    assert importers.IMPORTERS_OVERVIEW_NAMESPACE == (
        "importers-overview-v5"
    )
    callback_key = next(
        key
        for key in GLOBAL_CALLBACK_MAP
        if "imp-overview-chart-entities-store.data" in key
    )
    callback_inputs = GLOBAL_CALLBACK_MAP[callback_key]["inputs"]
    assert callback_inputs == [
        {
            "id": "imp-overview-source-state-store",
            "property": "data",
        },
        {
            "id": "imp-overview-classification-mode",
            "property": "value",
        },
        {
            "id": "imp-overview-rolling-window-days-input",
            "property": "value",
        },
    ]


@pytest.mark.parametrize("pool_size", [1, 2, 4])
def test_importer_overview_loader_single_flight_at_pool_sizes(
    monkeypatch,
    persistent_importer_cache,
    pool_size,
):
    payload = _make_overview_payload()
    source_payload = _make_source_payload()
    source_build_calls = 0
    derived_build_calls = 0
    build_lock = threading.Lock()

    def source_builder():
        nonlocal source_build_calls
        with build_lock:
            source_build_calls += 1
        time.sleep(0.05)
        return copy.deepcopy(source_payload)

    def derived_builder(*_args):
        nonlocal derived_build_calls
        with build_lock:
            derived_build_calls += 1
        time.sleep(0.05)
        return copy.deepcopy(payload)

    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        source_builder,
    )
    monkeypatch.setattr(
        importers,
        "_build_importers_overview_payload_from_source",
        derived_builder,
    )

    def load():
        return importers.refresh_overview_data(
            {"watermark": "concurrency-stable-source"},
            "Country",
            30,
        )

    with ThreadPoolExecutor(max_workers=pool_size) as executor:
        results = list(executor.map(lambda _index: load(), range(pool_size)))

    assert source_build_calls == 1
    assert derived_build_calls == 1
    assert all(result == results[0] for result in results)
    assert all(
        snapshots.snapshot_is_resolvable(store)
        for store in results[0]
    )
    assert len(to_json(results[0]).encode("utf-8")) < 50_000


def test_importer_warm_loader_response_benchmark_gate(
    monkeypatch,
    persistent_importer_cache,
):
    template = _make_overview_payload(
        entity_count=12,
        years=(2022, 2023, 2024, 2025, 2026),
        points_per_year=120,
        continents=("Africa", "Americas", "Asia", "Europe"),
    )

    def build_payload(*_args):
        return copy.deepcopy(template)

    _install_overview_builders(monkeypatch, template)
    baseline_payload = build_payload()

    app = Flask(__name__)

    @app.get("/baseline")
    def baseline_response():
        response_payload = (
            baseline_payload["chart_entities"],
            baseline_payload["table_entities"],
            snapshots.unpack_record_mapping(
                baseline_payload["demand_cube"]
            ),
            snapshots.unpack_record_mapping(
                baseline_payload["origin_cube"]
            ),
        )
        return Response(
            to_json(response_payload),
            mimetype="application/json",
        )

    @app.get("/reference")
    def reference_response():
        return Response(
            to_json(
                importers.refresh_overview_data(
                    {"watermark": "stable-source"},
                    "Country",
                    30,
                )
            ),
            mimetype="application/json",
        )

    client = app.test_client()
    assert client.get("/baseline").status_code == 200
    assert client.get("/reference").status_code == 200

    timings = {"baseline": [], "reference": []}
    sizes = {"baseline": [], "reference": []}
    for _iteration in range(6):
        for mode in ("baseline", "reference"):
            started = time.perf_counter()
            response = client.get(f"/{mode}")
            elapsed = time.perf_counter() - started
            assert response.status_code == 200
            timings[mode].append(elapsed)
            sizes[mode].append(len(response.data))

    baseline_seconds = statistics.median(timings["baseline"])
    reference_seconds = statistics.median(timings["reference"])
    improvement = (
        baseline_seconds - reference_seconds
    ) / baseline_seconds
    reference_bytes = max(sizes["reference"])
    baseline_bytes = max(sizes["baseline"])
    print(
        "IMPORTERS_WARM_BENCHMARK "
        f"baseline_ms={baseline_seconds * 1000:.2f} "
        f"reference_ms={reference_seconds * 1000:.2f} "
        f"improvement_pct={improvement * 100:.2f} "
        f"baseline_bytes={baseline_bytes} "
        f"reference_bytes={reference_bytes}"
    )

    assert reference_bytes < 50_000
    assert improvement > 0.10

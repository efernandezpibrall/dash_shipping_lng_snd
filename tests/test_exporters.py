# Consolidated from test_exporter_detail_snapshot_refs.py.

import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import threading
import time

from dash._utils import to_json
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import pytest

from pages import exporter_detail
from utils import dashboard_snapshot_cache as snapshots


def _destination_frame(country="United States"):
    return pd.DataFrame(
        {
            "start_date": pd.to_datetime(
                [
                    "2025-01-02",
                    "2025-02-03",
                    "2026-01-04",
                    "2026-02-05",
                ]
            ),
            "cargo_mcm": np.array([60.25, np.nan, 72.5, 81.75]),
            "destination_country": [
                "Japan",
                "Spain",
                country,
                "United Kingdom",
            ],
            "destination_continent": [
                "Asia",
                "Europe",
                "North America",
                "Europe",
            ],
            "destination_shipping_region": [
                "North Asia",
                "Iberia",
                "US Gulf",
                "Northwest Europe",
            ],
            "destination_basin": [
                "Pacific",
                "Atlantic",
                "Atlantic",
                "Atlantic",
            ],
            "destination_subcontinent": [
                "East Asia",
                "Southern Europe",
                "North America",
                "Western Europe",
            ],
            "destination_classification_level1": [
                "Asia",
                "Europe",
                "North America",
                None,
            ],
            "destination_classification": [
                "Importer",
                "Importer",
                "Producer",
                "Importer",
            ],
        }
    )


def _origin_plant_frame():
    frame = _destination_frame().copy()
    frame.insert(2, "origin_zone", ["Gulf", "Gulf", "Atlantic", "Atlantic"])
    frame.insert(
        3,
        "origin_plant",
        ["Sabine Pass", "Corpus Christi", "Cove Point", "Cove Point"],
    )
    return frame


def _forecast_frame():
    dates = pd.to_datetime(["2026-08-01", "2026-08-02"])
    return pd.DataFrame(
        {
            "date": dates,
            "year": dates.year,
            "day_of_year": dates.dayofyear,
            "month_day": dates.strftime("%b %d"),
            "mcmd": [434.85634408602147, 433.48172043010754],
            "woodmac_mtpa": [116.75, 116.38],
            "is_forecast": [True, True],
            "source": ["Short Term", "Short Term"],
        }
    )


def _allocation_source(country="United States"):
    run_metadata = {
        "run_id": "allocation-run-exporter",
        "analysis_date": pd.Timestamp("2026-07-20 12:00:00"),
        "forecast_start": pd.Timestamp("2026-07-01"),
        "forecast_end": pd.Timestamp("2028-12-01"),
        "supply_scenario": "base_view",
        "split_by_contract": True,
        "woodmac_short_term_outlook": "July 2026",
        "woodmac_long_term_outlook": "June 2026",
    }
    mapping_df = pd.DataFrame(
        {
            "country": ["Spain", "Japan"],
            "country_name": ["Spain", "Japan"],
            "continent": ["Europe", "Asia"],
            "basin": ["Atlantic", "Pacific"],
            "shipping_region": ["Iberia", "North Asia"],
            "subcontinent": ["Southern Europe", "East Asia"],
            "country_classification_level1": ["Europe", "Asia"],
            "country_classification": ["Importer", "Importer"],
        }
    )
    allocation_df = pd.DataFrame(
        {
            "date": pd.to_datetime(
                [
                    "2026-07-01",
                    "2026-08-01",
                    "2027-01-01",
                    "2028-01-01",
                ]
            ),
            "destination_country": ["Spain", "Japan", "Spain", "Japan"],
            "allocated_volume_bcm": [1.25, 1.5, 3.0, 3.5],
            "alias": ["Spain", "Japan", "Spain", "Japan"],
            "country_display": ["Spain", "Japan", "Spain", "Japan"],
            "continent": ["Europe", "Asia", "Europe", "Asia"],
            "basin": ["Atlantic", "Pacific", "Atlantic", "Pacific"],
            "shipping_region": [
                "Iberia",
                "North Asia",
                "Iberia",
                "North Asia",
            ],
            "subcontinent": [
                "Southern Europe",
                "East Asia",
                "Southern Europe",
                "East Asia",
            ],
            "country_classification_level1": [
                "Europe",
                "Asia",
                "Europe",
                "Asia",
            ],
            "country_classification": [
                "Importer",
                "Importer",
                "Importer",
                "Importer",
            ],
            "country": ["Spain", "Japan", "Spain", "Japan"],
        }
    )
    return {
        "origin_country": country,
        "run_metadata": run_metadata,
        "mapping_df": mapping_df,
        "allocation_df": allocation_df,
    }


def _base_payload(country="United States"):
    origin_scope = {
        "country": country,
        "continent": "North America",
        "shipping_region": "US Gulf",
        "basin": "Atlantic",
        "subcontinent": None,
        "country_classification_level1": "North America",
        "country_classification": "Producer",
        "continent_conflict": False,
    }
    return {
        "origin_country": country,
        "normalized_destination_trades": exporter_detail._store_dataframe(
            _destination_frame(country)
        ),
        "origin_plant_destination_trades": exporter_detail._store_dataframe(
            _origin_plant_frame()
        ),
        "woodmac_export_forecast": exporter_detail._store_dataframe(
            pd.DataFrame(
                columns=[
                    "date",
                    "year",
                    "day_of_year",
                    "month_day",
                    "mcmd",
                    "woodmac_mtpa",
                    "is_forecast",
                    "source",
                ]
            )
        ),
        "origin_scope": origin_scope,
        "available_years": [2025, 2026],
        "loaded_at": "2026-07-24T12:00:00",
    }


@pytest.fixture
def persistent_exporter_detail_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "exporter-detail-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(snapshots.LOCAL_CACHE_DIR_ENV, str(cache_directory))
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    monkeypatch.setattr(
        exporter_detail,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: "2026-07-24T00:00:00Z",
    )
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_reference(monkeypatch, country="United States", payload=None):
    payload = payload or _base_payload(country)
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda selected_country: copy.deepcopy(payload),
    )
    reference = exporter_detail.refresh_exporter_detail_base_data(country, 0)
    return payload, reference


def _delete_or_corrupt_reference(reference, mode):
    stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if mode == "missing":
        stores.cache.delete(record_key, retry=True)
    else:
        stores.cache.set(record_key, b"corrupt", retry=True)
    snapshots.clear_local_snapshots()


def _exporter_detail_workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: [
            [
                (cell.value, cell.data_type, cell.number_format)
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]
        for worksheet in workbook.worksheets
    }


def _without_loaded_at(payload):
    return {
        key: value
        for key, value in payload.items()
        if key != "loaded_at"
    }


@pytest.mark.parametrize(
    "country",
    ["United States", "Equatorial Guinea"],
)
def test_base_loader_emits_small_resolvable_reference_and_preserves_payload(
    country,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    payload, reference = _load_reference(monkeypatch, country)

    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_DETAIL_BASE_NAMESPACE,
    )
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert len(to_json(reference).encode("utf-8")) < 50_000

    legacy_frames = exporter_detail._resolve_exporter_detail_base_data(payload)
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    restarted_frames = exporter_detail._resolve_exporter_detail_base_data(
        reference
    )

    pd.testing.assert_frame_equal(
        restarted_frames[0],
        legacy_frames[0],
        check_dtype=True,
        check_exact=True,
    )
    assert list(restarted_frames[0].columns) == list(legacy_frames[0].columns)
    assert restarted_frames[0].isna().equals(legacy_frames[0].isna())
    assert restarted_frames[1] == legacy_frames[1]
    pd.testing.assert_frame_equal(
        restarted_frames[2],
        legacy_frames[2],
        check_dtype=True,
        check_exact=True,
    )
    assert restarted_frames[3] == country


def test_allocation_reference_is_small_exact_and_control_paths_are_sql_free(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    source_data = _allocation_source()
    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_data["run_metadata"]),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        lambda *_args, **_kwargs: copy.deepcopy(source_data),
    )

    reference = exporter_detail.refresh_destination_forecast_source(
        "United States",
        0,
    )
    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_ALLOCATION_SOURCE_NAMESPACE,
    )
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    resolved = exporter_detail._resolve_destination_forecast_source_data(
        reference
    )
    pd.testing.assert_frame_equal(
        resolved["mapping_df"],
        source_data["mapping_df"],
        check_exact=True,
    )
    pd.testing.assert_frame_equal(
        resolved["allocation_df"],
        source_data["allocation_df"],
        check_exact=True,
    )

    monkeypatch.setattr(
        exporter_detail,
        "fetch_destination_forecast_summary_data",
        lambda *_args, **_kwargs: pytest.fail(
            "allocation controls queried PostgreSQL"
        ),
    )
    levels = [
        "destination_country_name",
        "destination_shipping_region",
        "destination_basin",
        "destination_subcontinent",
        "destination_classification_level1",
        "destination_classification",
    ]
    for level in levels:
        direct = exporter_detail.update_destination_forecast_summary_table(
            "United States",
            ["Europe"],
            level,
            "mcm_d",
            source_data,
        )
        cached = exporter_detail.update_destination_forecast_summary_table(
            "United States",
            ["Europe"],
            level,
            "mcm_d",
            reference,
        )
        assert to_json(cached) == to_json(direct)


def test_allocation_no_run_uses_one_metadata_read_and_consistent_identity(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    metadata_reads = []

    def no_run(_engine):
        metadata_reads.append(True)
        if len(metadata_reads) > 1:
            raise AssertionError("allocation metadata was queried twice")
        return None

    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        no_run,
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        lambda *_args, **_kwargs: pytest.fail(
            "no-run refresh invoked the source loader"
        ),
    )

    reference = exporter_detail.refresh_destination_forecast_source(
        "United States",
        0,
    )
    current_month = pd.Timestamp.today().replace(day=1).date().isoformat()
    expected_key = snapshots.build_source_key(
        exporter_detail.EXPORTER_ALLOCATION_SOURCE_NAMESPACE,
        None,
        None,
        "United States",
        current_month,
    )
    assert metadata_reads == [True]
    assert reference["source_key"] == expected_key
    resolved = exporter_detail._resolve_destination_forecast_source_data(
        reference
    )
    assert resolved["run_metadata"] is None
    assert resolved["allocation_df"].empty
    manifest = snapshots._LOCAL_MANIFESTS[
        (
            reference["namespace"],
            reference["source_key"],
            reference["revision"],
        )
    ]
    assert manifest["run_id"] is None
    output = exporter_detail.update_destination_forecast_summary_table(
        "United States",
        [],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        reference,
    )
    assert "No compatible WoodMac supply-allocation SQL run" in to_json(
        output
    )


def test_route_reference_is_small_exact_and_controls_and_export_are_sql_free(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    _, base_reference = _load_reference(monkeypatch)
    base_reference["route_source_version"] = {
        "kpler_watermark": "kpler-v1",
        "distance_watermark": "distance-v1",
    }
    processed_df = pd.DataFrame(
        {
            "year": [2025, 2025, 2026, 2026],
            "month": [1, 4, 7, 10],
            "season": ["W", "S", "S", "W"],
            "quarter": ["Q1", "Q2", "Q3", "Q4"],
            "voyage_id": ["v1", "v2", "v3", "v4"],
            "destination_country_name": [
                "Spain",
                "Japan",
                "France",
                "Brazil",
            ],
            "distanceDirect": [100.0, 100.0, 100.0, 100.0],
            "distanceViaSuez": [120.0, np.nan, 120.0, np.nan],
            "distanceViaPanama": [np.nan, 130.0, 130.0, np.nan],
            "selected_route": [
                "Direct",
                "ViaPanama",
                "ViaSuez",
                "Direct",
            ],
        }
    )
    mapping_df = pd.DataFrame(
        {
            "country": ["Spain", "Japan", "France", "Brazil"],
            "country_name": ["Spain", "Japan", "France", "Brazil"],
            "continent": ["Europe", "Asia", "Europe", "South America"],
            "basin": ["Atlantic", "Pacific", "Atlantic", "Atlantic"],
            "shipping_region": [
                "Iberia",
                "North Asia",
                "Northwest Europe",
                "South America",
            ],
            "subcontinent": [
                "Southern Europe",
                "East Asia",
                "Western Europe",
                "South America",
            ],
            "country_classification_level1": [
                "Europe",
                "Asia",
                "Europe",
                "South America",
            ],
            "country_classification": [
                "Importer",
                "Importer",
                "Importer",
                "Importer",
            ],
        }
    )
    source_data = {
        "origin_country": "United States",
        "processed_df": processed_df,
        "mapping_df": mapping_df,
    }
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_route_source_payload",
        lambda _country: copy.deepcopy(source_data),
    )
    reference = exporter_detail.refresh_exporter_route_analysis_source(
        "United States",
        base_reference,
    )
    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_ROUTE_SOURCE_NAMESPACE,
    )
    assert len(to_json(reference).encode("utf-8")) < 10_000

    monkeypatch.setattr(
        exporter_detail,
        "process_trade_and_distance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "route controls or export queried PostgreSQL"
        ),
    )
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pytest.fail(
            "route controls queried mappings"
        ),
    )
    direct = exporter_detail.update_route_analysis_charts_and_tables(
        "Year",
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "United States",
        source_data,
    )
    cached = exporter_detail.update_route_analysis_charts_and_tables(
        "Year",
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "United States",
        reference,
    )
    assert to_json(cached) == to_json(direct)

    direct_export = exporter_detail.export_route_analysis_to_excel(
        1,
        "Year",
        "destination_country_name",
        "United States",
        source_data,
    )
    cached_export = exporter_detail.export_route_analysis_to_excel(
        1,
        "Year",
        "destination_country_name",
        "United States",
        reference,
    )
    assert _exporter_detail_workbook_cells(cached_export) == _exporter_detail_workbook_cells(
        direct_export
    )


def test_diversion_reference_is_small_exact_and_sql_free(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    source_data = {
        "origin_country": "United States",
        "main_data": [{
            "Diversion date": "2026-06-01",
            "Vessel": "Test Vessel",
            "State": "Loaded",
            "Charterer": "Test Charterer",
            "Cubic Meters": 174500.0,
            "Origin location": "Sabine Pass",
            "Origin country": "United States",
            "Origin date": "2026-05-20",
            "Diverted from location": "Gate",
            "Diverted from country": "Netherlands",
            "Diverted from date": "2026-06-10",
            "New destination location": "Futtsu",
            "New destination country": "Japan",
            "New destination date": "2026-06-20",
            "Added shipping days": 10,
        }],
        "charts_data": [{
            "Diversion_month": "2026-06-01 00:00:00",
            "basin_combo": "Atlantic -> Pacific",
            "region_combo": "Northwest Europe -> North Asia",
            "country_combo": "Netherlands -> Japan",
            "Added shipping days": 10,
            "Cubic Meters": 174500.0,
        }],
    }
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_diversion_source_version",
        lambda: "diversion-v1",
    )
    captured_versions = []

    def build_diversion(*_args, **kwargs):
        captured_versions.append(kwargs.get("source_version"))
        return copy.deepcopy(source_data)

    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_diversion_payload",
        build_diversion,
    )
    reference = exporter_detail.refresh_exporter_diversion_source(
        0,
        "United States",
    )
    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_DIVERSION_SOURCE_NAMESPACE,
    )
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert captured_versions == ["diversion-v1"]
    assert to_json(exporter_detail.update_diversion_ui(None, "basin_combo")) == to_json(
        exporter_detail.update_diversion_ui(
            {"main_data": [], "charts_data": []},
            "basin_combo",
        )
    )

    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pytest.fail(
            "diversion controls and export must not query SQL"
        ),
    )
    direct_ui = exporter_detail.update_diversion_ui(
        source_data,
        "region_combo",
    )
    cached_ui = exporter_detail.update_diversion_ui(
        reference,
        "region_combo",
    )
    assert to_json(cached_ui) == to_json(direct_ui)
    direct_export = exporter_detail.export_diversion_summary_to_excel(
        1,
        source_data,
        direct_ui[1],
    )
    cached_export = exporter_detail.export_diversion_summary_to_excel(
        1,
        reference,
        cached_ui[1],
    )
    assert _exporter_detail_workbook_cells(cached_export) == _exporter_detail_workbook_cells(
        direct_export
    )

    corrupt_reference = dict(reference, revision="missing")
    corrupt_ui = exporter_detail.update_diversion_ui(
        corrupt_reference,
        "basin_combo",
    )
    assert exporter_detail.EXPORTER_DIVERSION_RECOVERY_MESSAGE in to_json(
        corrupt_ui
    )
    with pytest.raises(exporter_detail.PreventUpdate):
        exporter_detail.export_diversion_summary_to_excel(
            1,
            corrupt_reference,
            cached_ui[1],
        )


@pytest.mark.parametrize("workers", [1, 4, 8])
def test_diversion_source_is_single_flight_under_concurrency(
    workers,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    counter = {"count": 0}
    counter_lock = threading.Lock()

    def build_payload(*_args, **_kwargs):
        with counter_lock:
            counter["count"] += 1
        time.sleep(0.04)
        return {
            "origin_country": "United States",
            "main_data": [],
            "charts_data": [],
        }

    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_diversion_source_version",
        lambda: f"diversion-single-flight-{workers}",
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_diversion_payload",
        build_payload,
    )
    with ThreadPoolExecutor(max_workers=workers) as pool:
        references = list(
            pool.map(
                lambda _index: (
                    exporter_detail.refresh_exporter_diversion_source(
                        0,
                        "United States",
                    )
                ),
                range(workers),
            )
        )

    assert counter["count"] == 1
    assert all(
        snapshots.snapshot_is_resolvable(reference)
        for reference in references
    )
    assert len({
        (reference["source_key"], reference["revision"])
        for reference in references
    }) == 1


@pytest.mark.parametrize("mode", ["missing", "corrupt"])
def test_allocation_reference_failure_is_visible_without_sql_fallback(
    mode,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    source_data = _allocation_source()
    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_data["run_metadata"]),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        lambda *_args, **_kwargs: copy.deepcopy(source_data),
    )
    reference = exporter_detail.refresh_destination_forecast_source(
        "United States",
        0,
    )
    _delete_or_corrupt_reference(reference, mode)
    monkeypatch.setattr(
        exporter_detail,
        "fetch_destination_forecast_summary_data",
        lambda *_args, **_kwargs: pytest.fail(
            "corrupt allocation reference queried PostgreSQL"
        ),
    )

    output = exporter_detail.update_destination_forecast_summary_table(
        "United States",
        [],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        reference,
    )

    assert (
        exporter_detail.EXPORTER_ALLOCATION_RECOVERY_MESSAGE
        in to_json(output)
    )


def test_nonempty_forecast_round_trips_exactly_and_chart_does_not_query_sql(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    payload = _base_payload()
    payload["woodmac_export_forecast"] = _forecast_frame()
    _, reference = _load_reference(monkeypatch, payload=payload)

    restarted = exporter_detail._load_exporter_detail_forecast_data(reference)
    pd.testing.assert_frame_equal(
        restarted,
        _forecast_frame(),
        check_dtype=True,
        check_exact=True,
    )

    monkeypatch.setattr(
        exporter_detail,
        "fetch_woodmac_country_export_forecast_data",
        lambda _country: pytest.fail("chart callback queried WoodMac"),
    )
    exporter_detail.update_exporter_detail_supply_charts(
        reference,
        30,
        "mcm_d",
        ["2025", "2026"],
    )


def test_forecast_month_is_part_of_source_key(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    payload = _base_payload()
    payload["woodmac_export_forecast"] = _forecast_frame()
    build_calls = []
    month = {"value": "2026-07-01"}
    monkeypatch.setattr(
        exporter_detail,
        "_exporter_detail_forecast_month_token",
        lambda: month["value"],
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda country: build_calls.append(country) or copy.deepcopy(payload),
    )

    first = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        0,
    )
    month["value"] = "2026-08-01"
    second = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        0,
    )

    assert first["source_key"] != second["source_key"]
    assert build_calls == ["United States", "United States"]


def test_reference_matches_legacy_for_selectors_charts_summaries_tables_and_supply_export(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    payload, reference = _load_reference(monkeypatch)
    monkeypatch.setattr(
        exporter_detail,
        "fetch_woodmac_country_export_forecast_data",
        lambda _country: pd.DataFrame(),
    )

    legacy_selectors = (
        exporter_detail.update_exporter_detail_supply_year_selector(
            payload,
            ["2025"],
        ),
        exporter_detail.update_exporter_detail_continent_year_selector(
            payload,
            ["2025"],
        ),
    )
    reference_selectors = (
        exporter_detail.update_exporter_detail_supply_year_selector(
            reference,
            ["2025"],
        ),
        exporter_detail.update_exporter_detail_continent_year_selector(
            reference,
            ["2025"],
        ),
    )
    assert to_json(reference_selectors) == to_json(legacy_selectors)

    legacy_charts = exporter_detail.update_exporter_detail_supply_charts(
        payload,
        30,
        "mcm_d",
        ["2025", "2026"],
    )
    reference_charts = exporter_detail.update_exporter_detail_supply_charts(
        reference,
        30,
        "mcm_d",
        ["2025", "2026"],
    )
    assert to_json(reference_charts) == to_json(legacy_charts)

    legacy_summaries = (
        exporter_detail.refresh_exporter_detail_summary_data_stores(
            "United States",
            30,
            exporter_detail.DEFAULT_DESTINATION_LEVEL,
            payload,
        )
    )
    reference_summaries = (
        exporter_detail.refresh_exporter_detail_summary_data_stores(
            "United States",
            30,
            exporter_detail.DEFAULT_DESTINATION_LEVEL,
            reference,
        )
    )
    assert _without_loaded_at(reference_summaries[0]) == _without_loaded_at(
        legacy_summaries[0]
    )
    assert _without_loaded_at(reference_summaries[1]) == _without_loaded_at(
        legacy_summaries[1]
    )

    destination_legacy = exporter_detail.update_destination_summary_table(
        "United States",
        30,
        ["Europe"],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        "previous",
        5,
        3,
        3,
        legacy_summaries[0],
    )
    destination_reference = exporter_detail.update_destination_summary_table(
        "United States",
        30,
        ["Europe"],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        "previous",
        5,
        3,
        3,
        reference_summaries[0],
    )
    assert to_json(destination_reference) == to_json(destination_legacy)

    origin_legacy = exporter_detail.update_origin_plant_summary_table(
        "United States",
        30,
        ["Gulf"],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        "previous",
        5,
        3,
        3,
        legacy_summaries[1],
    )
    origin_reference = exporter_detail.update_origin_plant_summary_table(
        "United States",
        30,
        ["Gulf"],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        "previous",
        5,
        3,
        3,
        reference_summaries[1],
    )
    assert to_json(origin_reference) == to_json(origin_legacy)

    legacy_export = exporter_detail.export_supply_analysis_to_excel(
        1,
        "United States",
        30,
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        payload,
    )
    reference_export = exporter_detail.export_supply_analysis_to_excel(
        1,
        "United States",
        30,
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        reference,
    )
    assert _exporter_detail_workbook_cells(reference_export) == _exporter_detail_workbook_cells(legacy_export)


@pytest.mark.parametrize("mode", ["missing", "corrupt"])
def test_missing_or_corrupt_reference_has_explicit_recovery_and_no_raw_fallback(
    mode,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    _, reference = _load_reference(monkeypatch)
    _delete_or_corrupt_reference(reference, mode)

    expected_selector = (
        [{
            "label": exporter_detail.EXPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE,
            "value": "__snapshot_unavailable__",
            "disabled": True,
        }],
        [],
    )
    assert (
        exporter_detail.update_exporter_detail_supply_year_selector(
            reference,
            [],
        )
        == expected_selector
    )
    assert (
        exporter_detail.update_exporter_detail_continent_year_selector(
            reference,
            [],
        )
        == expected_selector
    )

    chart_result = exporter_detail.update_exporter_detail_supply_charts(
        reference,
        30,
        "mcm_d",
        [],
    )
    assert (
        chart_result[-1].children
        == exporter_detail.EXPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE
    )
    assert chart_result[-1].role == "alert"

    summary_result = (
        exporter_detail.refresh_exporter_detail_summary_data_stores(
            "United States",
            30,
            exporter_detail.DEFAULT_DESTINATION_LEVEL,
            reference,
        )
    )
    assert all(
        item["error"]
        == exporter_detail.EXPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE
        for item in summary_result
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Cached exporter-detail data is unavailable",
    ):
        exporter_detail.export_supply_analysis_to_excel(
            1,
            "United States",
            30,
            exporter_detail.DEFAULT_DESTINATION_LEVEL,
            "mcm_d",
            reference,
        )


def test_wrong_namespace_and_nonresolvable_reference_are_rejected(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    wrong_reference = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": "another-page-v1",
        "source_key": "source",
        "revision": "00000000-0000-4000-8000-000000000000",
        "shared": True,
    }
    with pytest.raises(snapshots.SnapshotUnavailable):
        exporter_detail._resolve_exporter_detail_base_data(wrong_reference)

    payload = _base_payload()
    local_reference = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": exporter_detail.EXPORTER_DETAIL_BASE_NAMESPACE,
        "source_key": "source",
        "revision": 1,
        "shared": False,
    }
    monkeypatch.setattr(
        exporter_detail,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (local_reference, payload),
    )
    with pytest.raises(snapshots.SnapshotUnavailable):
        exporter_detail.refresh_exporter_detail_base_data(
            "United States",
            0,
        )

    malformed_reference, _ = snapshots.get_or_build_snapshot(
        exporter_detail.engine,
        namespace=exporter_detail.EXPORTER_DETAIL_BASE_NAMESPACE,
        source_key="malformed-payload",
        builder=lambda: {},
        force=True,
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Cached exporter-detail data is unavailable",
    ):
        exporter_detail._resolve_exporter_detail_base_data(
            malformed_reference
        )


@pytest.mark.parametrize("workers", [1, 2, 4])
def test_base_loader_is_single_flight_for_one_two_and_four_callers(
    workers,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    counter = {"count": 0}
    counter_lock = threading.Lock()

    def build_payload(_country):
        with counter_lock:
            counter["count"] += 1
        time.sleep(0.04)
        return _base_payload()

    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        build_payload,
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: f"single-flight-{workers}",
    )

    with ThreadPoolExecutor(max_workers=workers) as pool:
        references = list(
            pool.map(
                lambda _index: (
                    exporter_detail.refresh_exporter_detail_base_data(
                        "United States",
                        0,
                    )
                ),
                range(workers),
            )
        )

    assert counter["count"] == 1
    assert all(
        snapshots.snapshot_is_resolvable(reference)
        for reference in references
    )
    assert len(
        {
            (
                reference["source_key"],
                reference["revision"],
            )
            for reference in references
        }
    ) == 1


def test_global_refresh_is_checked_once_and_forces_one_rebuild(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    refresh_calls = []
    build_calls = []
    monkeypatch.setattr(
        exporter_detail,
        "_was_global_refresh_triggered",
        lambda: refresh_calls.append(True) or True,
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda country: build_calls.append(country) or _base_payload(country),
    )

    reference = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        1,
    )

    assert refresh_calls == [True]
    assert build_calls == ["United States"]
    assert snapshots.snapshot_is_resolvable(reference)


def test_route_diversion_and_maintenance_remain_outside_base_snapshot_and_exports_are_stable(
    monkeypatch,
):
    independent_callbacks = (
        exporter_detail.update_route_analysis_charts_and_tables,
        exporter_detail.export_route_analysis_to_excel,
        exporter_detail.process_diversion_data,
        exporter_detail.update_diversion_ui,
        exporter_detail.export_diversion_summary_to_excel,
        exporter_detail.update_maintenance_table,
        exporter_detail.update_maintenance_seasonal_chart,
    )
    assert all(
        "_resolve_exporter_detail_base_data"
        not in callback_function.__code__.co_names
        for callback_function in independent_callbacks
    )

    route_frame = pd.DataFrame(
        {
            "year": [2025, 2025, 2026, 2026],
            "voyage_id": ["v1", "v2", "v3", "v4"],
            "destination_country_name": [
                "Spain",
                "Japan",
                "France",
                "Brazil",
            ],
            "destination_shipping_region": [
                "Iberia",
                "North Asia",
                "Northwest Europe",
                "South America",
            ],
            "distanceDirect": [100.0, 100.0, 100.0, 100.0],
            "distanceViaSuez": [120.0, np.nan, 120.0, np.nan],
            "distanceViaPanama": [np.nan, 130.0, 130.0, np.nan],
        }
    )
    monkeypatch.setattr(
        exporter_detail,
        "process_trade_and_distance_data",
        lambda *_args, **_kwargs: route_frame.copy(),
    )
    route_first = exporter_detail.export_route_analysis_to_excel(
        1,
        "Year",
        "destination_country_name",
        "United States",
    )
    route_second = exporter_detail.export_route_analysis_to_excel(
        1,
        "Year",
        "destination_country_name",
        "United States",
    )
    assert _exporter_detail_workbook_cells(route_first) == _exporter_detail_workbook_cells(route_second)

    diversion_store = {
        "main_data": [
            {
                "Diversion date": "2026-07-20",
                "Vessel": "Example LNG",
                "State": "Loaded",
                "Cubic Meters": 174500.25,
                "Added shipping days": 2.5,
            }
        ]
    }
    diversion_columns = [
        {"field": "Diversion date"},
        {"field": "Vessel"},
        {"field": "State"},
        {"field": "Cubic Meters"},
        {"field": "Added shipping days"},
    ]
    diversion_first = exporter_detail.export_diversion_summary_to_excel(
        1,
        diversion_store,
        diversion_columns,
    )
    diversion_second = exporter_detail.export_diversion_summary_to_excel(
        1,
        diversion_store,
        diversion_columns,
    )
    assert _exporter_detail_workbook_cells(diversion_first) == _exporter_detail_workbook_cells(
        diversion_second
    )


def test_maintenance_controls_reuse_current_raw_source_with_exact_output(
    monkeypatch,
):
    raw_data = pd.DataFrame(
        {
            "id_plant": [1],
            "plant_name": ["Sabine Pass"],
            "country_name": ["United States"],
            "id_lng_train": [11],
            "lng_train_name_short": ["Train 1"],
            "year": [2026],
            "month": [7],
            "year_actual_forecast": ["Forecast"],
            "total_mtpa": [0.75],
            "metric_comment": ["Planned"],
            "train_capacity_mmtpa": [4.5],
            "train_capacity_mcmd": [16.77],
            "plant_capacity_mmtpa": [27.0],
            "plant_capacity_mcmd": [100.62],
            "date": pd.to_datetime(["2026-07-01"]),
        }
    )
    fetches = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or raw_data.copy()
        ),
    )
    baseline = exporter_detail.update_maintenance_table(
        "United States",
        "mt",
        3,
        3,
        [],
    )
    assert fetches == [True]

    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "maintenance controls reread the source"
        ),
    )
    cached = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=baseline[2],
    )

    assert to_json(cached[:2]) == to_json(baseline[:2])
    assert cached[2] == baseline[2]


@pytest.mark.parametrize("error_message", ["source down", ""])
def test_maintenance_source_error_retries_on_next_control_change(
    monkeypatch,
    error_message,
):
    attempts = []

    def fail_source(*_args, **_kwargs):
        attempts.append("failed")
        raise RuntimeError(error_message)

    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        fail_source,
    )
    failed = exporter_detail.update_maintenance_table(
        "United States",
        "mcm_d",
        3,
        3,
        [],
    )
    assert failed[2]["error"] == error_message

    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            attempts.append("retried") or pd.DataFrame()
        ),
    )
    retried = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=failed[2],
    )

    assert attempts == ["failed", "retried"]
    assert "No maintenance data available" in to_json(retried[0])
    assert "error" not in retried[2]


def test_maintenance_database_failure_reaches_retryable_error_store(
    monkeypatch,
):
    monkeypatch.setattr(
        exporter_detail.engine,
        "connect",
        lambda: (_ for _ in ()).throw(RuntimeError("alias database down")),
    )
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("maintenance database down")
        ),
    )

    failed = exporter_detail.update_maintenance_table(
        "United States",
        "mcm_d",
        3,
        3,
        [],
    )

    assert failed[2]["error"] == "maintenance database down"
    assert "Error loading maintenance data" in to_json(failed[0])


def test_source_context_refresh_generation_is_not_sticky_for_later_selections(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    build_calls = []
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda country: (
            build_calls.append(country) or _base_payload(country)
        ),
    )
    watermark = {
        "kpler_watermark": "2026-07-24T00:00:00Z",
        "woodmac_watermark": "2026-07-23T00:00:00Z",
        "distance_watermark": "2026-07-22T00:00:00Z",
    }
    initial_context = exporter_detail._build_exporter_detail_source_context(
        watermark
    )
    refresh_context = exporter_detail._build_exporter_detail_source_context(
        watermark,
        force_refresh=True,
    )

    exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=initial_context,
    )
    exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=refresh_context,
    )
    exporter_detail.refresh_exporter_detail_base_data(
        "Qatar",
        source_context=refresh_context,
    )
    exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=refresh_context,
    )

    assert build_calls == ["United States", "United States", "Qatar"]


def test_mapping_fingerprint_invalidates_base_and_route_snapshots(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    base_builds = []
    route_builds = []
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda country: (
            base_builds.append(country) or _base_payload(country)
        ),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_route_source_payload",
        lambda country: (
            route_builds.append(country)
            or {
                "origin_country": country,
                "processed_df": pd.DataFrame(),
                "mapping_df": pd.DataFrame(),
            }
        ),
    )
    source_watermark = {
        "kpler_watermark": "kpler-v1",
        "woodmac_watermark": "woodmac-v1",
        "distance_watermark": "distance-v1",
        "mapping_fingerprint": "mapping-v1",
    }
    initial_context = (
        exporter_detail._build_exporter_detail_source_context(
            source_watermark
        )
    )
    changed_context = (
        exporter_detail._build_exporter_detail_source_context(
            {
                **source_watermark,
                "mapping_fingerprint": "mapping-v2",
            }
        )
    )

    first_base = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=initial_context,
    )
    changed_base = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=changed_context,
    )
    first_route = exporter_detail.refresh_exporter_route_analysis_source(
        "United States",
        initial_context,
    )
    changed_route = exporter_detail.refresh_exporter_route_analysis_source(
        "United States",
        changed_context,
    )

    assert base_builds == ["United States", "United States"]
    assert route_builds == ["United States", "United States"]
    assert first_base["source_key"] != changed_base["source_key"]
    assert first_route["source_key"] != changed_route["source_key"]


def test_exporter_route_falls_back_when_context_watermark_is_unavailable(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    metadata_reads = []
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_route_source_version",
        lambda: (
            metadata_reads.append(True)
            or {
                "kpler_watermark": "kpler-v1",
                "distance_watermark": "distance-v1",
                "mapping_fingerprint": "mapping-v1",
            }
        ),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_route_source_payload",
        lambda country: {
            "origin_country": country,
            "processed_df": pd.DataFrame(),
            "mapping_df": pd.DataFrame(),
        },
    )
    degraded_context = (
        exporter_detail._build_exporter_detail_source_context(
            "metadata-unavailable"
        )
    )

    reference = exporter_detail.refresh_exporter_route_analysis_source(
        "United States",
        degraded_context,
    )

    assert metadata_reads == [True]
    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_ROUTE_SOURCE_NAMESPACE,
    )


def test_country_catalog_exposes_unavailable_source_watermark_failure(monkeypatch):
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pd.DataFrame(
            {"origin_country_name": ["Qatar", "United States"]}
        ),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: (_ for _ in ()).throw(RuntimeError("watermark down")),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_maintenance_source_version",
        lambda: (_ for _ in ()).throw(RuntimeError("maintenance down")),
    )

    options, selected, source_context = (
        exporter_detail.initialize_country_dropdown(0)
    )

    assert options == [
        {"label": "Qatar", "value": "Qatar"},
        {"label": "United States", "value": "United States"},
    ]
    assert selected == "United States"
    assert source_context["source_watermark"] is None
    assert source_context["source_revision"]["status"] == "unavailable"
    assert "unavailable" in source_context["source_revision"]["message"]


def test_country_catalog_reuses_last_good_revision_as_stale(monkeypatch):
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pd.DataFrame(
            {"origin_country_name": ["Qatar", "United States"]}
        ),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: (_ for _ in ()).throw(RuntimeError("watermark down")),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_maintenance_source_version",
        lambda: {"revision": "maintenance-v1"},
    )
    previous_context = (
        exporter_detail._build_exporter_detail_source_context(
            {"kpler_watermark": "kpler-v1"}
        )
    )
    existing_options = [
        {"label": "Qatar", "value": "Qatar"},
        {"label": "United States", "value": "United States"},
    ]

    options, selected, source_context = (
        exporter_detail.initialize_country_dropdown(
            0,
            existing_options,
            "Qatar",
            previous_context,
        )
    )

    assert options == existing_options
    assert selected == "Qatar"
    assert source_context["source_watermark"] == {
        "kpler_watermark": "kpler-v1"
    }
    assert source_context["source_revision"]["status"] == "stale"


def test_exporter_maintenance_reference_is_small_exact_and_refreshes_once(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    raw_data = pd.DataFrame(
        {
            "id_plant": [1],
            "plant_name": ["Sabine Pass"],
            "country_name": ["United States"],
            "id_lng_train": [11],
            "lng_train_name_short": ["Train 1"],
            "year": [2026],
            "month": [7],
            "year_actual_forecast": ["Forecast"],
            "total_mtpa": [0.75],
            "metric_comment": ["Planned"],
            "train_capacity_mmtpa": [4.5],
            "train_capacity_mcmd": [16.77],
            "plant_capacity_mmtpa": [27.0],
            "plant_capacity_mcmd": [100.62],
            "date": pd.to_datetime(["2026-07-01"]),
        }
    )
    fetches = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or raw_data.copy()
        ),
    )
    legacy = exporter_detail.update_maintenance_table(
        "United States",
        "mcm_d",
        3,
        3,
        [],
    )
    version = {
        "unplanned_watermark": "u1",
        "unplanned_row_count": 1,
        "planned_watermark": "p1",
        "planned_row_count": 1,
        "train_capacity_watermark": "t1",
        "train_capacity_row_count": 1,
        "plant_capacity_watermark": "s1",
        "plant_capacity_row_count": 1,
        "mapping_fingerprint": "m1",
    }
    context = exporter_detail._build_exporter_detail_source_context(
        "base-v1",
        maintenance_source_version=version,
    )
    snapshotted = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        source_context=context,
    )

    assert to_json(snapshotted[:2]) == to_json(legacy[:2])
    assert snapshots.is_snapshot_reference(
        snapshotted[2],
        exporter_detail.EXPORTER_MAINTENANCE_SOURCE_NAMESPACE,
    )
    assert len(to_json(snapshotted[2]).encode("utf-8")) < 1_000
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "maintenance reference reread SQL"
        ),
    )
    cached = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        maintenance_raw_data=snapshotted[2],
        source_context=context,
    )
    assert to_json(cached[:2]) == to_json(legacy[:2])

    refreshed_fetches = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            refreshed_fetches.append(True) or raw_data.copy()
        ),
    )
    refresh_context = exporter_detail._build_exporter_detail_source_context(
        "base-v1",
        force_refresh=True,
        maintenance_source_version=version,
    )
    refreshed = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        maintenance_raw_data=snapshotted[2],
        source_context=refresh_context,
    )
    exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=refreshed[2],
        source_context=refresh_context,
    )
    assert refreshed_fetches == [True]


def test_exporter_maintenance_metadata_failure_does_not_run_unversioned_query(
    monkeypatch,
):
    fetches = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or pd.DataFrame()
        ),
    )
    initial_context = (
        exporter_detail._build_exporter_detail_source_context(
            "metadata-unavailable",
            maintenance_source_version=None,
        )
    )

    first = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        source_context=initial_context,
    )
    exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=first[2],
        source_context=initial_context,
    )

    refresh_context = (
        exporter_detail._build_exporter_detail_source_context(
            "metadata-unavailable",
            force_refresh=True,
            maintenance_source_version=None,
        )
    )
    refreshed = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        maintenance_raw_data=first[2],
        source_context=refresh_context,
    )
    exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=refreshed[2],
        source_context=refresh_context,
    )

    assert fetches == []
    assert first[2]["error"]
    assert refreshed[2]["error"]


def test_exporter_persistent_sources_track_mapping_and_refresh_generations(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    allocation_source = _allocation_source()
    allocation_metadata = []
    for fingerprint in (
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ):
        metadata = copy.deepcopy(allocation_source["run_metadata"])
        metadata["mapping_fingerprint"] = fingerprint
        allocation_metadata.append(metadata)
    allocation_builds = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: allocation_metadata.pop(0),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        lambda *_args, **kwargs: (
            allocation_builds.append(
                kwargs["run_metadata"]["mapping_fingerprint"]
            )
            or copy.deepcopy(allocation_source)
        ),
    )

    initial_context = {"refresh_generation": None}
    first_refresh_context = {
        "refresh_generation": "refresh-generation-1"
    }
    second_refresh_context = {
        "refresh_generation": "refresh-generation-2"
    }
    source_contexts = (
        initial_context,
        initial_context,
        first_refresh_context,
        first_refresh_context,
        second_refresh_context,
    )
    allocation_refs = [
        exporter_detail.refresh_destination_forecast_source(
            "United States",
            source_context,
        )
        for source_context in source_contexts
    ]

    assert allocation_builds == [
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ]
    assert allocation_refs[0]["source_key"] != allocation_refs[1][
        "source_key"
    ]
    assert allocation_refs[1]["source_key"] != allocation_refs[2][
        "source_key"
    ]
    assert allocation_refs[2]["source_key"] == allocation_refs[3][
        "source_key"
    ]
    assert allocation_refs[3]["source_key"] != allocation_refs[4][
        "source_key"
    ]

    diversion_versions = [
        {
            "diversion_watermark": "diversion-v1",
            "country_mapping_fingerprint": fingerprint,
            "location_mapping_fingerprint": "location-v1",
        }
        for fingerprint in (
            "mapping-v1",
            "mapping-v2",
            "mapping-v2",
            "mapping-v2",
            "mapping-v2",
        )
    ]
    diversion_versions_for_fetch = copy.deepcopy(diversion_versions)
    diversion_builds = []
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_diversion_source_version",
        lambda: diversion_versions.pop(0),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_diversion_payload",
        lambda country, **kwargs: (
            diversion_builds.append(
                kwargs["source_version"][
                    "country_mapping_fingerprint"
                ]
            )
            or {
                "origin_country": country,
                "main_data": [],
                "charts_data": [],
            }
        ),
    )

    diversion_refs = [
        exporter_detail.refresh_exporter_diversion_source(
            source_context,
            "United States",
        )
        for source_context in source_contexts
    ]

    assert diversion_builds == [
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ]
    assert diversion_refs[0]["source_key"] != diversion_refs[1][
        "source_key"
    ]
    assert diversion_refs[1]["source_key"] != diversion_refs[2][
        "source_key"
    ]
    assert diversion_refs[2]["source_key"] == diversion_refs[3][
        "source_key"
    ]
    assert diversion_refs[3]["source_key"] != diversion_refs[4][
        "source_key"
    ]

    captured_params = {}
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda _query, _engine, params: (
            captured_params.update(params) or pd.DataFrame()
        ),
    )
    exporter_detail._fetch_exporter_diversion_rows(
        "United States",
        diversion_versions_for_fetch[0],
    )
    assert captured_params["source_version"] == "diversion-v1"


# Consolidated from test_exporters_pbd_changes.py.

import base64
from io import BytesIO

from openpyxl import load_workbook
import numpy as np
import pandas as pd
import pytest

from pages import exporters


def _snapshot_pair(
    current_date,
    baseline_date=None,
    *,
    current_timestamp=None,
    baseline_timestamp=None,
):
    current_date = pd.Timestamp(current_date)
    pair = {
        'current_snapshot_id': 200,
        'current_snapshot_date_utc': current_date.date(),
        'current_snapshot_timestamp_utc': (
            current_timestamp
            or current_date.replace(
                hour=5,
                minute=34,
                second=12,
                microsecond=778724,
            )
        ),
        'current_facts_retained': True,
        'baseline_snapshot_id': None,
        'baseline_snapshot_date_utc': None,
        'baseline_snapshot_timestamp_utc': None,
        'baseline_facts_retained': None,
    }
    if baseline_date is not None:
        baseline_date = pd.Timestamp(baseline_date)
        pair.update({
            'baseline_snapshot_id': 199,
            'baseline_snapshot_date_utc': baseline_date.date(),
            'baseline_snapshot_timestamp_utc': (
                baseline_timestamp
                or baseline_date.replace(
                    hour=5,
                    minute=36,
                    second=57,
                    microsecond=998945,
                )
            ),
            'baseline_facts_retained': True,
        })
    return pair


@pytest.mark.parametrize(
    ('current_date', 'baseline_date', 'expected_status', 'expected_gap'),
    [
        ('2026-07-30', '2026-07-29', 'exact', 1),
        ('2026-07-27', '2026-07-24', 'exact', 1),
        ('2026-07-26', '2026-07-24', 'exact', 1),
        ('2026-07-30', '2026-07-28', 'fallback', 2),
        ('2026-07-30', None, 'unavailable', None),
    ],
)
def test_source_state_selects_previous_weekday_or_labels_fallback(
    current_date,
    baseline_date,
    expected_status,
    expected_gap,
):
    state = exporters._build_exporters_source_state(
        _snapshot_pair(current_date, baseline_date),
        refresh_token='refresh-1',
    )

    assert state['format'] == exporters.EXPORTERS_SOURCE_STATE_FORMAT
    assert state['format'].endswith('-v2')
    assert state['current_snapshot']['snapshot_id'] == 200
    assert state['current_snapshot']['snapshot_date_utc'] == current_date
    assert state['source_watermark'].endswith('05:34:12.778724')
    assert state['baseline_status'] == expected_status
    assert state['business_day_gap'] == expected_gap
    if baseline_date is None:
        assert state['baseline_snapshot'] is None
    else:
        assert state['baseline_snapshot']['snapshot_id'] == 199
        assert (
            state['baseline_snapshot']['snapshot_date_utc']
            == baseline_date
        )
        assert state['baseline_snapshot'][
            'snapshot_timestamp_utc'
        ].endswith('05:36:57.998945')


def _flow_row(
    flow_date,
    supply_country,
    mcmd,
    *,
    demand_country='Destination',
    supply_classification='Supply class',
    demand_classification='Demand class',
):
    return {
        'supply_classification': supply_classification,
        'supply_country': supply_country,
        'supply_installation': f'{supply_country} terminal',
        'demand_classification': demand_classification,
        'demand_country': demand_country,
        'flow_date': pd.Timestamp(flow_date),
        'mcmd': float(mcmd),
    }


def _constant_window(
    as_of_date,
    supply_country,
    mcmd,
    **row_kwargs,
):
    return [
        _flow_row(
            flow_date,
            supply_country,
            mcmd,
            **row_kwargs,
        )
        for flow_date in pd.date_range(
            end=pd.Timestamp(as_of_date),
            periods=30,
            freq='D',
        )
    ]


def test_rolling_windows_use_inclusive_7d_and_30d_boundaries():
    as_of_date = pd.Timestamp('2026-07-30')
    frame = pd.DataFrame([
        _flow_row(as_of_date - pd.Timedelta(days=6), 'A', 7),
        _flow_row(as_of_date - pd.Timedelta(days=7), 'A', 70),
        _flow_row(as_of_date - pd.Timedelta(days=29), 'A', 30),
        _flow_row(as_of_date - pd.Timedelta(days=30), 'A', 300),
    ])

    result = exporters._build_supply_dest_rolling_windows_from_df(
        frame,
        'Country',
        'Installation',
        as_of_date,
    )
    row = result.loc[result['supply_country'] == 'A'].iloc[0]

    assert row['7D'] == pytest.approx(1.0)
    assert row['30D'] == pytest.approx(3.6)


@pytest.mark.parametrize(
    ('classification_mode', 'aggregation_mode'),
    [
        ('Country', 'Installation'),
        ('Country', 'Country'),
        ('Country', 'Classification Level 1'),
        ('Classification Level 1', 'Installation'),
        ('Classification Level 1', 'Country'),
        ('Classification Level 1', 'Classification Level 1'),
    ],
)
def test_pbd_outer_join_keeps_additions_removals_and_all_aggregations(
    classification_mode,
    aggregation_mode,
):
    current_date = pd.Timestamp('2026-07-30')
    baseline_date = pd.Timestamp('2026-07-29')
    current_rows = (
        _constant_window(current_date, 'A', 10)
        + _constant_window(
            current_date,
            'B',
            5,
            demand_country='Destination B',
            demand_classification='Demand class B',
        )
        + _constant_window(
            current_date,
            'Internal',
            100,
            demand_country='Internal',
            supply_classification='Internal class',
            demand_classification='Internal class',
        )
        + [_flow_row(current_date - pd.Timedelta(days=30), 'A', 300)]
    )
    baseline_rows = (
        _constant_window(baseline_date, 'A', 8)
        + _constant_window(
            baseline_date,
            'C',
            4,
            demand_country='Destination C',
            demand_classification='Demand class C',
        )
        + _constant_window(
            baseline_date,
            'Internal',
            100,
            demand_country='Internal',
            supply_classification='Internal class',
            demand_classification='Internal class',
        )
        + [_flow_row(baseline_date - pd.Timedelta(days=30), 'A', 300)]
    )

    current = exporters._build_supply_dest_rolling_windows_from_df(
        pd.DataFrame(current_rows),
        classification_mode,
        aggregation_mode,
        current_date,
    )
    baseline = exporters._build_supply_dest_rolling_windows_from_df(
        pd.DataFrame(baseline_rows),
        classification_mode,
        aggregation_mode,
        baseline_date,
    )
    merged = exporters._merge_supply_dest_pbd_rolling_windows(
        current,
        baseline,
        classification_mode,
        aggregation_mode,
        baseline_available=True,
    )
    detail = merged[
        ~merged['supply_country'].isin(['Total', 'Internal'])
    ]

    a_row = detail.loc[detail['supply_country'] == 'A'].iloc[0]
    b_row = detail.loc[detail['supply_country'] == 'B'].iloc[0]
    c_row = detail.loc[detail['supply_country'] == 'C'].iloc[0]

    assert a_row['Δ 30D vs PBD'] == pytest.approx(2)
    assert a_row['Δ 7D vs PBD'] == pytest.approx(2)
    assert b_row['Δ 30D vs PBD'] == pytest.approx(5)
    assert b_row['Δ 7D vs PBD'] == pytest.approx(5)
    assert c_row['Δ 30D vs PBD'] == pytest.approx(-4)
    assert c_row['Δ 7D vs PBD'] == pytest.approx(-4)
    assert 'Internal' not in detail['supply_country'].tolist()


def test_outer_join_preserves_a_fully_removed_current_vintage():
    baseline = exporters._build_supply_dest_rolling_windows_from_df(
        pd.DataFrame(
            _constant_window('2026-07-29', 'Removed', 4)
        ),
        'Country',
        'Installation',
        '2026-07-29',
    )

    merged = exporters._merge_supply_dest_pbd_rolling_windows(
        pd.DataFrame(),
        baseline,
        'Country',
        'Installation',
        baseline_available=True,
    )

    row = merged.iloc[0]
    assert row['30D'] == 0
    assert row['7D'] == 0
    assert row['Δ 30D vs PBD'] == -4
    assert row['Δ 7D vs PBD'] == -4


def test_small_country_taxonomy_is_frozen_from_current_vintage():
    current = pd.DataFrame(
        _constant_window('2026-07-30', 'Small', 1)
        + _constant_window('2026-07-30', 'Large', 100)
    )
    baseline = pd.DataFrame(
        _constant_window('2026-07-29', 'Small', 50)
        + _constant_window('2026-07-29', 'Large', 100)
    )

    grouped_current, grouping_config = (
        exporters.group_small_supply_dest_countries(
            current,
            'Country',
            'Installation',
            as_of_date='2026-07-30',
            return_grouping_config=True,
        )
    )
    grouped_baseline = exporters.group_small_supply_dest_countries(
        baseline,
        'Country',
        'Installation',
        grouping_config=grouping_config,
    )

    assert set(grouped_current['supply_country']) == {
        'Large',
        'Rest of countries',
    }
    assert set(grouped_baseline['supply_country']) == {
        'Large',
        'Rest of countries',
    }


class _ConnectionContext:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False


class _Engine:
    def connect(self):
        return _ConnectionContext()


@pytest.mark.parametrize(
    ('rolling_avg_days', 'expected_query_start'),
    [
        (1, '2021-01-01'),
        (7, '2020-12-26'),
        (30, '2020-12-03'),
        (180, '2020-07-06'),
    ],
)
def test_rolling_query_start_uses_exact_preceding_days(
    rolling_avg_days,
    expected_query_start,
):
    assert exporters._get_rolling_query_start_date(
        '2021-01-01',
        rolling_avg_days,
    ) == expected_query_start


@pytest.mark.parametrize('rolling_avg_days', [1, 7, 30, 180])
def test_continent_selected_window_uses_exact_rolling_warmup(
    rolling_avg_days,
):
    _years, query_start, display_start = (
        exporters._get_continent_chart_selected_window(
            ['2025'],
            rolling_avg_days,
        )
    )

    assert (
        pd.Timestamp(display_start) - pd.Timestamp(query_start)
    ).days == rolling_avg_days - 1


def test_supply_queries_gate_on_complete_180_day_windows(monkeypatch):
    captured_queries = []

    def fake_read_sql(query, _connection, params=None):
        captured_queries.append(str(query))
        return pd.DataFrame()

    monkeypatch.setattr(exporters.pd, 'read_sql', fake_read_sql)
    engine = _Engine()

    exporters.fetch_global_supply_data(engine, 'at_lng', rolling_avg_days=180)
    exporters.fetch_country_supply_data(
        engine,
        'at_lng',
        'United States',
        'Country',
        180,
    )
    exporters.fetch_country_supply_data(
        engine,
        'at_lng',
        'Atlantic',
        'Classification Level 1',
        180,
    )
    exporters._fetch_country_supply_chart_batch(engine, 'at_lng', 180)
    exporters._fetch_classification_supply_chart_batch(
        engine,
        'at_lng',
        ['Atlantic'],
        180,
    )

    assert len(captured_queries) == 5
    for query in captured_queries:
        assert "'2020-07-06'::date" in query
        assert '179 PRECEDING' in query
        assert 'COUNT(*) OVER rolling_window' in query
        assert 'rolling_window_day_count = 180' in query
        assert "WHERE date >= '2021-01-01'" in query


@pytest.mark.parametrize('classification_mode', ['Country', 'Classification Level 1'])
def test_continent_queries_gate_on_complete_180_day_windows(
    monkeypatch,
    classification_mode,
):
    captured = {}

    def fake_read_sql(query, _connection, params=None):
        captured['query'] = str(query)
        return pd.DataFrame()

    monkeypatch.setattr(exporters.pd, 'read_sql', fake_read_sql)
    exporters.fetch_continent_chart_data_batch(
        _Engine(),
        'at_lng',
        ['Global'],
        classification_mode,
        selected_years=['2025'],
        rolling_avg_days=180,
    )

    query = captured['query']
    assert "'2023-07-06'::date" in query
    assert '179 PRECEDING' in query
    assert 'COUNT(*) OVER rolling_window' in query
    assert 'rolling_window_day_count = 180' in query
    assert "WHERE date >= '2024-01-01'" in query


def test_pbd_query_binds_exact_snapshot_and_only_30_calendar_days(
    monkeypatch,
):
    captured = {}

    def fake_read_sql(query, connection, params):
        captured['query'] = str(query)
        captured['params'] = params
        return pd.DataFrame([{
            'supply_classification': 'Supply class',
            'supply_country': 'A',
            'supply_installation': 'A terminal',
            'demand_classification': 'Demand class',
            'demand_country': 'Destination',
            'flow_date': pd.Timestamp('2026-07-29'),
            'year': 2026,
            'quarter': 3,
            'month': 7,
            'week': 31,
            'volume': 1000,
        }])

    monkeypatch.setattr(exporters.pd, 'read_sql', fake_read_sql)
    result = exporters.fetch_supply_destination_pbd_base_data(
        _Engine(),
        'at_lng',
        '2026-07-29T05:36:57.998945',
        '2026-07-29',
    )

    assert not result.empty
    assert 'latest_data' not in captured['query']
    assert 'kt.start >= :window_start' in captured['query']
    assert 'kt.start < :end_exclusive' in captured['query']
    assert captured['params']['snapshot_timestamp_utc'] == pd.Timestamp(
        '2026-07-29T05:36:57.998945'
    )
    assert captured['params']['window_start'] == pd.Timestamp('2026-06-30')
    assert captured['params']['end_exclusive'] == pd.Timestamp('2026-07-30')


@pytest.mark.parametrize('volume_metric', ['mcm_d', 'bcm', 'mt', 'mtpa'])
def test_pbd_absolute_changes_convert_in_selected_unit(volume_metric):
    frame = pd.DataFrame([{
        'Supply Country': 'A',
        '30D': 136.0,
        '30D_PBD': 68.0,
        '7D': 136.0,
        '7D_PBD': 68.0,
        'Δ 30D vs PBD': 999.0,
        'Δ 7D vs PBD': 999.0,
    }])

    converted = exporters._convert_supply_dest_absolute_volume_metric(
        frame,
        volume_metric,
    )
    assert converted.loc[0, 'Δ 30D vs PBD'] == pytest.approx(
        round(
            converted.loc[0, '30D']
            - converted.loc[0, '30D_PBD'],
            1,
        )
    )
    assert converted.loc[0, 'Δ 7D vs PBD'] == pytest.approx(
        round(
            converted.loc[0, '7D']
            - converted.loc[0, '7D_PBD'],
            1,
        )
    )


def test_bcm_metric_uses_each_destination_column_period_length():
    frame = pd.DataFrame([{
        'Supply Country': 'A',
        '2024': 100.0,
        '30D': 100.0,
        '7D': 100.0,
        'Δ 7D-30D': 0.0,
    }])

    converted = exporters._convert_supply_dest_absolute_volume_metric(
        frame,
        'bcm',
    )

    assert converted.loc[0, '2024'] == pytest.approx(36.6)
    assert converted.loc[0, '30D'] == pytest.approx(3.0)
    assert converted.loc[0, '7D'] == pytest.approx(0.7)
    assert pd.isna(converted.loc[0, 'Δ 7D-30D'])


@pytest.mark.parametrize('volume_metric', ['mcm_d', 'mtpa'])
def test_rate_metrics_keep_comparable_7d_30d_delta(volume_metric):
    converted = exporters._convert_supply_dest_absolute_volume_metric(
        pd.DataFrame([{
            'Supply Country': 'A',
            '30D': 100.0,
            '7D': 120.0,
            'Δ 7D-30D': 20.0,
        }]),
        volume_metric,
    )

    assert pd.notna(converted.loc[0, 'Δ 7D-30D'])
    assert converted.loc[0, 'Δ 7D-30D'] == pytest.approx(
        round(converted.loc[0, '7D'] - converted.loc[0, '30D'], 1)
    )


def test_bcm_metric_is_available_and_exports_converted_rolling_volume():
    assert {'label': 'bcm', 'value': 'bcm'} in exporters.VOLUME_METRIC_OPTIONS

    charts_data = {
        'Global': [{
            'date': '2026-07-30',
            'year': '2026',
            'month_day': 'Jul 30',
            'rolling_avg': 100.0,
            'is_forecast': False,
        }]
    }
    download = exporters.export_supply_charts_to_excel(
        1,
        charts_data,
        'bcm',
        30,
    )
    workbook = load_workbook(
        BytesIO(base64.b64decode(download['content']))
    )
    worksheet = workbook['All Data']
    headers = [cell.value for cell in worksheet[1]]
    bcm_column = headers.index('rolling_volume_30d (bcm)') + 1

    assert worksheet.cell(row=2, column=bcm_column).value == pytest.approx(3.0)


def test_continent_bcm_export_uses_volume_header_and_value(monkeypatch):
    continent_frame = pd.DataFrame([{
        'entity_name': 'Global',
        'date': pd.Timestamp('2026-07-30'),
        'continent_destination': 'Asia',
        'year': 2026,
        'day_of_year': 211,
        'month_day': 'Jul 30',
        'rolling_avg': 100.0,
        'percentage': 100.0,
        'is_forecast': False,
    }])
    monkeypatch.setattr(
        exporters,
        '_resolve_or_load_exporters_continent_export_payload',
        lambda *_args: {
            'entities': ['Global'],
            'data': continent_frame,
        },
    )

    download = exporters.export_continent_charts_to_excel(
        1,
        {'placeholder': True},
        'absolute',
        'Country',
        'bcm',
        ['2026'],
        30,
    )
    workbook = load_workbook(BytesIO(base64.b64decode(download['content'])))
    worksheet = workbook['All Data']
    headers = [cell.value for cell in worksheet[1]]
    bcm_column = headers.index('rolling_volume_30d (bcm)') + 1

    assert worksheet.cell(row=2, column=bcm_column).value == pytest.approx(3.0)


def _summary_payload(*, status='exact', baseline_total=100):
    rows = [
        {
            'supply_country': 'A',
            'supply_installation': 'A terminal',
            '2024': 10.0,
            '2025': 20.0,
            '30D': 25.0,
            '30D_PP': 20.0,
            '30D_Y1': 15.0,
            '7D': 30.0,
            '7D_PP': 25.0,
            '7D_Y1': 20.0,
            'Δ 7D-30D': 5.0,
            'Δ 30D Y/Y': 10.0,
            '30D_PBD': baseline_total / 2,
            '7D_PBD': baseline_total / 10,
            'Δ 30D vs PBD': 25.0 - baseline_total / 2,
            'Δ 7D vs PBD': 30.0 - baseline_total / 10,
        },
        {
            'supply_country': 'B',
            'supply_installation': 'B terminal',
            '2024': 30.0,
            '2025': 40.0,
            '30D': 75.0,
            '30D_PP': 80.0,
            '30D_Y1': 85.0,
            '7D': 70.0,
            '7D_PP': 75.0,
            '7D_Y1': 80.0,
            'Δ 7D-30D': -5.0,
            'Δ 30D Y/Y': -10.0,
            '30D_PBD': baseline_total / 2,
            '7D_PBD': baseline_total * 0.9,
            'Δ 30D vs PBD': 75.0 - baseline_total / 2,
            'Δ 7D vs PBD': 70.0 - baseline_total * 0.9,
        },
    ]
    if status == 'unavailable':
        for row in rows:
            for column_name in (
                *exporters.SUPPLY_DEST_PBD_REFERENCE_COLUMNS,
                *exporters.SUPPLY_DEST_PBD_DELTA_COLUMNS,
            ):
                row[column_name] = np.nan
    current_snapshot = {
        'snapshot_id': 200,
        'snapshot_date_utc': '2026-07-30',
        'snapshot_timestamp_utc': '2026-07-30T05:34:12.778724',
        'facts_retained': True,
    }
    baseline_snapshot = (
        {
            'snapshot_id': 199,
            'snapshot_date_utc': '2026-07-29',
            'snapshot_timestamp_utc': '2026-07-29T05:36:57.998945',
            'facts_retained': True,
        }
        if status != 'unavailable'
        else None
    )
    return {
        'format': exporters.EXPORTERS_SUPPLY_DEST_SUMMARY_FORMAT,
        'show_all': rows,
        'group_small_countries': rows,
        'comparison': {
            'status': status,
            'current_snapshot': current_snapshot,
            'baseline_snapshot': baseline_snapshot,
            'business_day_gap': 1 if baseline_snapshot else None,
        },
    }


def _walk_exporters_pbd_components(component):
    if component is None:
        return
    if isinstance(component, (list, tuple)):
        for child in component:
            yield from _walk_exporters_pbd_components(child)
        return
    yield component
    yield from _walk_exporters_pbd_components(getattr(component, 'children', None))


def _first_exporters_pbd_grid(component):
    return next(
        item
        for item in _walk_exporters_pbd_components(component)
        if hasattr(item, 'rowData')
    )


def _column_fields(column_defs):
    fields = []
    for column in column_defs:
        children = column.get('children') or []
        if children:
            fields.extend(_column_fields(children))
            continue
        field = column.get('field') or column.get('id')
        if field:
            fields.append(field)
    return fields


def _component_text(component):
    values = []
    for item in _walk_exporters_pbd_components(component):
        if isinstance(item, str):
            values.append(item)
    return ' '.join(values)


def _render_table(
    payload,
    *,
    view_type='absolute',
    comparison_basis='levels',
    volume_metric='mcm_d',
):
    return exporters.update_supply_dest_table(
        payload,
        [],
        [],
        [],
        'Country',
        view_type,
        'Installation',
        comparison_basis,
        'show_all',
        volume_metric,
        2,
        0,
        0,
        0,
    )


@pytest.mark.parametrize(
    ('volume_metric', 'expected_measure'),
    [
        ('mcm_d', 'Rolling Average'),
        ('mtpa', 'Rolling Average'),
        ('bcm', 'Rolling Volume'),
        ('mt', 'Rolling Volume'),
    ],
)
def test_rolling_section_titles_follow_metric_semantics(
    volume_metric,
    expected_measure,
):
    titles = exporters.update_rolling_average_section_titles(
        30,
        volume_metric,
    )

    assert titles == (
        f'LNG Supply - 30-Day {expected_measure}',
        f'LNG Supply by Destination Continent - 30-Day {expected_measure}',
    )


@pytest.mark.parametrize(
    ('volume_metric', 'expected_precision', 'expected_plotly_format'),
    [
        ('mcm_d', 0, ',.0f'),
        ('mtpa', 1, ',.1f'),
        ('bcm', 1, ',.1f'),
        ('mt', 1, ',.1f'),
    ],
)
def test_volume_metric_chart_precision_policy(
    volume_metric,
    expected_precision,
    expected_plotly_format,
):
    assert (
        exporters._get_volume_metric_chart_precision(volume_metric)
        == expected_precision
    )
    assert (
        exporters._get_volume_metric_plotly_number_format(volume_metric)
        == expected_plotly_format
    )
    assert exporters._get_volume_metric_zero_tolerance(volume_metric) == pytest.approx(
        0.5 * (10 ** -expected_precision)
    )


@pytest.mark.parametrize(
    (
        'volume_metric',
        'volume_label',
        'expected_value',
        'expected_plotly_format',
    ),
    [
        ('mcm_d', 'mcm/d', '10', ':,.0f'),
        ('bcm', 'bcm', '0.3', ':,.1f'),
        ('mt', 'MT', '0.2', ':,.1f'),
        ('mtpa', 'MMTPA', '2.7', ':,.1f'),
    ],
)
def test_volume_metric_page_formatting_uses_selected_precision(
    volume_metric,
    volume_label,
    expected_value,
    expected_plotly_format,
):
    supply_records = [
        {
            'date': f'{year}-01-01',
            'year': str(year),
            'month_day': 'Jan 01',
            'rolling_avg': value,
            'is_forecast': False,
        }
        for year, value in ((2025, 0.0), (2026, 10.0))
    ]

    supply_chart = exporters.create_supply_chart(
        supply_records,
        volume_metric=volume_metric,
        selected_years=['2025', '2026'],
        rolling_avg_days=30,
    )
    supply_card = exporters.update_supply_charts(
        {'Global': supply_records},
        volume_metric,
        ['2025', '2026'],
        30,
    )
    supply_text = _component_text(supply_card)

    supply_hovertemplates = [
        trace.hovertemplate
        for trace in supply_chart.data
        if trace.hovertemplate
    ]
    assert supply_hovertemplates
    assert all(
        expected_plotly_format in template
        for template in supply_hovertemplates
    )
    assert supply_chart.layout.yaxis.tickformat == expected_plotly_format[1:]
    assert f'Jan 01: {expected_value} {volume_label}' in supply_text
    assert f'+{expected_value}' in supply_text

    continent_frame = pd.DataFrame([
        {
            'entity_name': 'Global',
            'date': pd.Timestamp(f'{year}-01-01'),
            'continent_destination': 'Asia',
            'year': year,
            'day_of_year': 1,
            'month_day': 'Jan 01',
            'rolling_avg': value,
            'percentage': 100.0 if value else 0.0,
            'is_forecast': False,
        }
        for year, value in ((2025, 0.0), (2026, 10.0))
    ])
    continent_chart = exporters._create_continent_destination_chart_from_df(
        continent_frame,
        volume_metric=volume_metric,
        selected_years=['2025', '2026'],
        rolling_avg_days=30,
    )
    continent_metrics = continent_chart.layout.meta['continent_kpis']

    continent_hovertemplates = [
        trace.hovertemplate
        for trace in continent_chart.data
        if trace.hovertemplate
    ]
    assert continent_hovertemplates
    assert all(
        expected_plotly_format in template
        for template in continent_hovertemplates
    )
    assert continent_chart.layout.yaxis.tickformat == expected_plotly_format[1:]
    assert continent_metrics[0]['latest_text'] == expected_value
    assert continent_metrics[0]['yoy_value_text'] == f'+{expected_value}'

    zero_frame = continent_frame.copy()
    zero_frame['rolling_avg'] = 0.0
    zero_chart = exporters._create_continent_destination_chart_from_df(
        zero_frame,
        volume_metric=volume_metric,
        selected_years=['2025', '2026'],
        rolling_avg_days=30,
    )
    assert zero_chart.layout.meta['continent_kpis'] == []


@pytest.mark.parametrize(
    ('volume_metric', 'volume_label', 'input_value', 'expected_value'),
    [
        ('mcm_d', 'mcm/d', -0.4, '0'),
        ('bcm', 'bcm', -0.04, '0.0'),
        ('mt', 'MT', -0.04, '0.0'),
        ('mtpa', 'MMTPA', -0.04, '0.0'),
    ],
)
def test_fractional_metric_formatting_never_emits_signed_zero(
    volume_metric,
    volume_label,
    input_value,
    expected_value,
):
    current_text = exporters._format_supply_chart_current_value(
        {'latest_label': 'Jan 01', 'latest_value': input_value},
        volume_metric,
    )
    delta_text = _component_text(
        exporters._build_supply_chart_delta_pill(
            'YoY',
            input_value,
            -0.4,
            volume_metric,
        )
    )

    assert current_text == f'Jan 01: {expected_value} {volume_label}'
    assert '-0' not in delta_text
    assert exporters._format_continent_kpi_value(
        input_value,
        'absolute',
        volume_label,
        volume_metric,
    ) == f'{expected_value} {volume_label}'
    assert exporters._format_continent_kpi_pct(-0.4) == ' (0%)'


@pytest.mark.parametrize(
    ('volume_metric', 'zero_value', 'visible_value'),
    [
        ('mcm_d', 0.5, 0.6),
        ('mtpa', 0.04, 0.06),
    ],
)
def test_metric_zero_threshold_preserves_visible_values_and_direction(
    volume_metric,
    zero_value,
    visible_value,
):
    assert exporters._continent_kpi_value_displays_zero(
        zero_value,
        'absolute',
        volume_metric,
    )
    assert not exporters._continent_kpi_value_displays_zero(
        visible_value,
        'absolute',
        volume_metric,
    )
    assert exporters._continent_kpi_direction_class(
        zero_value,
        'absolute',
        volume_metric,
    ) == 'continent-kpi-delta-neutral'
    assert exporters._continent_kpi_direction_class(
        visible_value,
        'absolute',
        volume_metric,
    ) == 'continent-kpi-delta-positive'


def test_mt_destination_grid_uses_one_decimal_for_converted_values():
    grid = _first_exporters_pbd_grid(
        _render_table(
            _summary_payload(),
            volume_metric='mt',
        )
    )
    a_row = next(
        row
        for row in grid.rowData
        if 'A' in row.get('Supply Country', '')
    )

    assert a_row['2024'] == '2.7'
    assert a_row['2025'] == '5.4'
    assert a_row['30D'] == '0.6'
    assert a_row['7D'] == '0.2'
    assert a_row['Δ 7D-30D'] == '—'
    assert a_row['Δ 30D Y/Y'] == '0.3'
    assert a_row['Δ 30D vs PBD'] == '-0.5'
    assert a_row['Δ 7D vs PBD'] == '+0.1'


def test_mtpa_destination_grid_uses_one_decimal_for_converted_values():
    grid = _first_exporters_pbd_grid(
        _render_table(
            _summary_payload(),
            volume_metric='mtpa',
        )
    )
    a_row = next(
        row
        for row in grid.rowData
        if 'A' in row.get('Supply Country', '')
    )

    assert a_row['2024'] == '2.7'
    assert a_row['2025'] == '5.4'
    assert a_row['30D'] == '6.7'
    assert a_row['7D'] == '8.1'
    assert a_row['Δ 7D-30D'] == '1.4'
    assert a_row['Δ 30D Y/Y'] == '2.7'
    assert a_row['Δ 30D vs PBD'] == '-6.7'
    assert a_row['Δ 7D vs PBD'] == '+5.4'


def test_mcm_d_destination_grid_normalizes_display_zero_and_raw_delta():
    display_df = pd.DataFrame({
        'Supply Country': ['A', 'B'],
        'Δ 30D vs PBD': [0.5, -0.5],
    })
    columns = exporters._build_supply_dest_columns(
        display_df,
        volume_metric='mcm_d',
    )
    grid_df, _ = exporters._build_supply_dest_summary_grid_display(
        display_df,
        columns,
        volume_metric='mcm_d',
    )

    assert grid_df['Δ 30D vs PBD'].tolist() == ['0', '0']
    assert grid_df['__supply_dest_delta_30d_pbd_raw'].tolist() == [0.0, 0.0]


@pytest.mark.parametrize('comparison_basis', ['levels', 'previous_period', 'same_period_last_year'])
@pytest.mark.parametrize('volume_metric', ['bcm', 'mt'])
def test_period_volume_delta_is_unavailable_across_comparison_modes(
    comparison_basis,
    volume_metric,
):
    rendered = _render_table(
        _summary_payload(),
        comparison_basis=comparison_basis,
        volume_metric=volume_metric,
    )
    grid = _first_exporters_pbd_grid(rendered)
    a_row = next(
        row
        for row in grid.rowData
        if 'A' in row.get('Supply Country', '')
    )

    assert a_row['Δ 7D-30D'] == '—'
    assert 'period totals cover different horizons' in _component_text(rendered)


@pytest.mark.parametrize('volume_metric', ['mcm_d', 'bcm', 'mt', 'mtpa'])
def test_volume_metric_precision_does_not_change_percentage_view(volume_metric):
    grid = _first_exporters_pbd_grid(
        _render_table(
            _summary_payload(),
            view_type='percentage',
            volume_metric=volume_metric,
        )
    )
    a_row = next(
        row
        for row in grid.rowData
        if 'A' in row.get('Supply Country', '')
    )

    assert a_row['30D'] == '25.0%'
    assert a_row['Δ 7D-30D'] != '—'
    assert a_row['Δ 7D-30D'].endswith(' pp')


def test_destination_export_preserves_unavailable_period_volume_delta():
    grid = _first_exporters_pbd_grid(
        _render_table(
            _summary_payload(),
            volume_metric='bcm',
        )
    )
    download = exporters.export_supply_dest_table_to_excel(
        1,
        [grid.rowData],
        [grid.rowData],
        [grid.columnDefs],
        'bcm',
        'absolute',
    )
    workbook = load_workbook(BytesIO(base64.b64decode(download['content'])))
    worksheet = workbook['Supply by Destination']
    headers = [cell.value for cell in worksheet[1]]
    delta_column = headers.index('Δ 7D-30D') + 1

    assert all(
        worksheet.cell(row=row, column=delta_column).value == '—'
        for row in range(2, worksheet.max_row + 1)
    )
    assert 'different horizons' in worksheet.cell(
        row=1,
        column=delta_column,
    ).comment.text


@pytest.mark.parametrize(
    'comparison_basis',
    ['levels', 'previous_period', 'same_period_last_year'],
)
def test_pbd_columns_remain_last_and_unchanged_across_comparisons(
    comparison_basis,
):
    grid = _first_exporters_pbd_grid(
        _render_table(
            _summary_payload(),
            comparison_basis=comparison_basis,
        )
    )
    fields = _column_fields(grid.columnDefs)
    a_row = next(
        row
        for row in grid.rowData
        if 'A' in row.get('Supply Country', '')
    )

    assert fields[-2:] == list(exporters.SUPPLY_DEST_PBD_DELTA_COLUMNS)
    assert '30D_PBD' not in fields
    assert '7D_PBD' not in fields
    assert a_row['Δ 30D vs PBD'] == '-25'
    assert a_row['Δ 7D vs PBD'] == '+20'


def test_mcm_d_absolute_summary_levels_display_no_decimals():
    grid = _first_exporters_pbd_grid(_render_table(_summary_payload()))
    a_row = next(
        row
        for row in grid.rowData
        if 'A' in row.get('Supply Country', '')
    )

    assert a_row['2024'] == '10'
    assert a_row['30D'] == '25'


def test_market_share_pbd_changes_are_percentage_points_and_zero_safe():
    grid = _first_exporters_pbd_grid(
        _render_table(
            _summary_payload(),
            view_type='percentage',
        )
    )
    rows_by_country = {
        row['Supply Country'].replace('▶', '').strip(): row
        for row in grid.rowData
    }

    assert rows_by_country['A']['30D'] == '25.0%'
    assert rows_by_country['A']['Δ 30D vs PBD'] == '-25.0 pp'
    assert rows_by_country['A']['Δ 7D vs PBD'] == '+20.0 pp'
    assert rows_by_country['B']['Δ 30D vs PBD'] == '+25.0 pp'
    assert rows_by_country['B']['Δ 7D vs PBD'] == '-20.0 pp'

    zero_baseline_grid = _first_exporters_pbd_grid(
        _render_table(
            _summary_payload(baseline_total=0),
            view_type='percentage',
        )
    )
    zero_rows = {
        row['Supply Country'].replace('▶', '').strip(): row
        for row in zero_baseline_grid.rowData
    }
    assert zero_rows['A']['Δ 30D vs PBD'] == '+25.0 pp'
    assert zero_rows['B']['Δ 30D vs PBD'] == '+75.0 pp'


def test_unavailable_baseline_shows_dashes_and_explicit_warning():
    table = _render_table(_summary_payload(status='unavailable'))
    grid = _first_exporters_pbd_grid(table)
    text = _component_text(table)

    assert all(
        row['Δ 30D vs PBD'] == '—'
        and row['Δ 7D vs PBD'] == '—'
        for row in grid.rowData
    )
    assert 'PBD baseline unavailable' in text


def test_snapshot_pair_lineage_displays_exact_timestamps():
    text = _component_text(_render_table(_summary_payload()))

    assert 'Jul 30, 2026 05:34:12.778724 UTC' in text
    assert 'Jul 29, 2026 05:36:57.998945 UTC' in text
    assert 'window roll plus Kpler revisions' in text


def test_excel_uses_rendered_order_and_pbd_values():
    grid = _first_exporters_pbd_grid(_render_table(_summary_payload()))
    download = exporters.export_supply_dest_table_to_excel(
        1,
        [grid.rowData],
        [grid.rowData],
        [grid.columnDefs],
    )
    workbook = load_workbook(
        BytesIO(base64.b64decode(download['content']))
    )
    worksheet = workbook['Supply by Destination']
    rows = list(worksheet.iter_rows(values_only=True))
    headers = list(rows[0])

    assert headers[-2:] == list(exporters.SUPPLY_DEST_PBD_DELTA_COLUMNS)
    assert rows[1][-2:] == (
        grid.rowData[0]['Δ 30D vs PBD'],
        grid.rowData[0]['Δ 7D vs PBD'],
    )


def test_snapshot_cache_keys_are_versioned_and_include_exact_pair():
    current_reference = {'namespace': 'current', 'key': 'current-key'}
    baseline_reference = {'namespace': 'baseline', 'key': 'baseline-key'}
    source_state = exporters._build_exporters_source_state(
        _snapshot_pair('2026-07-30', '2026-07-29')
    )

    first_key = exporters._exporters_destination_summary_source_key(
        current_reference,
        'Country',
        'Installation',
        baseline_reference,
        source_state,
    )
    changed_state = exporters._build_exporters_source_state(
        _snapshot_pair(
            '2026-07-30',
            '2026-07-29',
            baseline_timestamp='2026-07-29T06:00:00',
        )
    )
    changed_key = exporters._exporters_destination_summary_source_key(
        current_reference,
        'Country',
        'Installation',
        baseline_reference,
        changed_state,
    )

    assert exporters.EXPORTERS_DESTINATION_SUMMARY_NAMESPACE.endswith('-v2')
    assert exporters.EXPORTERS_SUPPLY_DEST_SUMMARY_FORMAT.endswith('-v2')
    assert first_key != changed_key


# Consolidated from test_exporters_snapshot_refs.py.

import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import statistics
import threading
import time

from dash import html, no_update
from dash._utils import to_json
from flask import Flask, Response
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import pytest

from pages import exporters
from utils import dashboard_snapshot_cache as snapshots


def _make_exporters_payload(
    *,
    entity_count=2,
    years=(2025, 2026),
    points_per_year=6,
    typed=False,
    timezone_aware=False,
    supply_dest_rows=3,
):
    charts_data = {}
    for entity_index in range(entity_count):
        entity_name = (
            "Global"
            if entity_index == 0
            else f"Exporter {entity_index}"
        )
        records = []
        for year in years:
            dates = pd.date_range(
                f"{year}-01-01",
                periods=points_per_year,
                freq="D",
            )
            for day_index, date in enumerate(dates):
                if timezone_aware:
                    date = date.tz_localize("Europe/London")
                record = {
                    "date": date,
                    "year": (
                        np.int16(year)
                        if typed
                        else str(year)
                    ),
                    "month_day": date.strftime("%b %d"),
                    "rolling_avg": (
                        np.float32(40 + entity_index + day_index)
                        if typed
                        else float(40 + entity_index + day_index)
                    ),
                    "is_forecast": (
                        np.bool_(year == max(years))
                        if typed
                        else year == max(years)
                    ),
                }
                if typed:
                    record.update({
                        "missing_timestamp": pd.NaT,
                        "numpy_timestamp": (
                            np.datetime64("NaT", "ns")
                            if day_index == points_per_year - 1
                            else np.datetime64(
                                date.tz_localize(None)
                                if timezone_aware
                                else date,
                                "ns",
                            )
                        ),
                        "pandas_missing": pd.NA,
                        "nullable_text": None,
                        "numpy_int": np.int32(day_index),
                        "numpy_float": np.float64(day_index + 0.25),
                        "numpy_bool": np.bool_(day_index % 2 == 0),
                    })
                records.append(record)
        charts_data[entity_name] = records

    destination_records = []
    for row_index in range(supply_dest_rows):
        destination_records.append({
            "supply_country": f"Country {row_index}",
            "supply_installation": f"Installation {row_index}",
            "30D": (
                np.float32(20 + row_index)
                if typed
                else float(20 + row_index)
            ),
            "7D": (
                np.float64(21 + row_index)
                if typed
                else float(21 + row_index)
            ),
            "30D_Y1": float(18 + row_index),
            "Q1'26": float(19 + row_index),
            "Jun'26": float(20 + row_index),
            "W25'26": float(21 + row_index),
            "nullable_note": pd.NA if typed else None,
        })
    destination_records.append({
        "supply_country": "GRAND TOTAL",
        "supply_installation": "",
        "30D": float(20 * supply_dest_rows),
        "7D": float(21 * supply_dest_rows),
        "30D_Y1": float(18 * supply_dest_rows),
        "Q1'26": float(19 * supply_dest_rows),
        "Jun'26": float(20 * supply_dest_rows),
        "W25'26": float(21 * supply_dest_rows),
        "nullable_note": None,
    })
    return {
        "charts_cube": snapshots.pack_record_mapping(charts_data),
        "continent_entities": list(charts_data),
        "supply_dest": {
            "show_all": copy.deepcopy(destination_records),
            "group_small_countries": copy.deepcopy(
                destination_records
            ),
        },
    }


def _make_continent_frame(entities, years=(2025, 2026)):
    records = []
    for entity_index, entity_name in enumerate(entities):
        for year in years:
            for day_index, date in enumerate(
                pd.date_range(f"{year}-01-01", periods=3, freq="D")
            ):
                rolling_avg = float(10 + entity_index + day_index)
                records.append({
                    "entity_name": entity_name,
                    "date": date,
                    "continent_destination": (
                        "Asia" if day_index % 2 == 0 else "Europe"
                    ),
                    "year": np.int16(year),
                    "day_of_year": np.int16(day_index + 1),
                    "month_day": date.strftime("%b %d"),
                    "rolling_avg": np.float32(rolling_avg),
                    "percentage": np.float32(
                        60 if day_index % 2 == 0 else 40
                    ),
                    "is_forecast": np.bool_(year == max(years)),
                })
    return pd.DataFrame.from_records(records)


def _source_state(refresh_token=None):
    return {
        "format": exporters.EXPORTERS_SOURCE_STATE_FORMAT,
        "source_watermark": "2026-07-24T00:00:00",
        "as_of_date": "2026-07-24",
        "refresh_token": refresh_token,
    }


def _raw_continent_store(
    payload,
    continent_frame,
    *,
    selected_years=None,
    classification_mode="Country",
    rolling_avg_days=30,
):
    selected_years, _query_start_date, _display_start_date = (
        exporters._get_continent_chart_selected_window(
            selected_years
        )
    )
    return {
        "entities": list(payload["continent_entities"]),
        "data": continent_frame.copy(),
        "source_state": _source_state(),
        "classification_mode": classification_mode,
        "rolling_avg_days": rolling_avg_days,
        "selected_years": selected_years,
    }


def _install_exporters_data_sources(
    monkeypatch,
    payload,
    *,
    continent_frame=None,
):
    charts_data = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )
    entities = list(payload["continent_entities"])
    continent_frame = (
        continent_frame.copy()
        if continent_frame is not None
        else _make_continent_frame(entities)
    )
    monkeypatch.setattr(
        exporters,
        "_get_exporter_entity_names",
        lambda *_args: list(entities),
    )
    monkeypatch.setattr(
        exporters,
        "_fetch_supply_chart_data_for_entities",
        lambda *_args: {
            entity_name: pd.DataFrame(
                copy.deepcopy(charts_data[entity_name])
            )
            for entity_name in entities
        },
    )
    monkeypatch.setattr(
        exporters,
        "fetch_supply_destination_base_data",
        lambda *_args: pd.DataFrame({
            "base_row": [1],
        }),
    )
    monkeypatch.setattr(
        exporters,
        "build_supply_dest_summary_store_payload",
        lambda *_args: copy.deepcopy(payload["supply_dest"]),
    )
    def fetch_continent(
        _engine,
        _schema,
        _entity_names,
        _classification_mode,
        selected_years=None,
        rolling_avg_days=30,
    ):
        _active_years, _query_start_date, display_start_date = (
            exporters._get_continent_chart_selected_window(
                selected_years
            )
        )
        if not _active_years:
            return pd.DataFrame()
        dates = pd.to_datetime(
            continent_frame.get("date"),
            errors="coerce",
        )
        return continent_frame[
            dates >= pd.Timestamp(display_start_date)
        ].copy()

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_continent,
    )
    return continent_frame


@pytest.fixture
def persistent_exporters_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "exporters-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    monkeypatch.setattr(
        exporters,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    monkeypatch.setattr(
        exporters,
        "_fetch_exporters_source_watermark",
        lambda: pd.Timestamp("2026-07-24T00:00:00"),
    )
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_representative_references(monkeypatch, payload=None):
    payload = payload or _make_exporters_payload()
    continent_frame = _install_exporters_data_sources(
        monkeypatch,
        payload,
    )
    stores = exporters.refresh_all_data(
        _source_state(),
        "Country",
        "Installation",
        30,
    )
    return payload, continent_frame, stores


def _walk_exporters_snapshot_components(component):
    if component is None:
        return
    if isinstance(component, (list, tuple)):
        for child in component:
            yield from _walk_exporters_snapshot_components(child)
        return
    yield component
    yield from _walk_exporters_snapshot_components(
        getattr(component, "children", None)
    )


def _first_exporters_snapshot_grid(component):
    return next(
        item
        for item in _walk_exporters_snapshot_components(component)
        if hasattr(item, "rowData")
    )


def _exporters_workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: {
            "cells": [
                [
                    (
                        cell.value,
                        cell.data_type,
                        cell.number_format,
                    )
                    for cell in row
                ]
                for row in worksheet.iter_rows()
            ],
            "widths": {
                key: dimension.width
                for key, dimension
                in worksheet.column_dimensions.items()
            },
        }
        for worksheet in workbook.worksheets
    }


def _assert_workbook_date_number_format(
    download,
    expected_format="YYYY-MM-DD",
):
    workbook = _exporters_workbook_cells(download)
    for worksheet in workbook.values():
        cells = worksheet["cells"]
        headers = [cell[0] for cell in cells[0]]
        if "date" not in headers:
            continue
        date_index = headers.index("date")
        date_formats = {
            row[date_index][2]
            for row in cells[1:]
            if row[date_index][0] is not None
        }
        assert date_formats == {expected_format}


def test_exporters_loader_emits_small_resolvable_refs_and_survives_restart(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    charts_store, continent_store, supply_dest_store = stores

    for store in (charts_store, continent_store, supply_dest_store):
        assert snapshots.is_snapshot_reference(store)
        assert snapshots.snapshot_is_resolvable(store)
        assert len(to_json(store).encode("utf-8")) < 10_000
    assert len(to_json(stores).encode("utf-8")) < 50_000

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()

    assert exporters._resolve_exporters_store(
        charts_store
    ) == snapshots.unpack_record_mapping(payload["charts_cube"])
    assert (
        exporters._resolve_exporters_store(supply_dest_store)
        == payload["supply_dest"]
    )
    continent_payload = exporters._resolve_exporters_continent_payload(
        continent_store
    )
    assert continent_payload["entities"] == payload["continent_entities"]
    assert continent_payload["selected_years"] == (
        exporters._default_continent_chart_selected_years(
            exporters._get_continent_chart_available_years()
        )
    )
    pd.testing.assert_frame_equal(
        continent_payload["data"],
        continent_frame,
        check_dtype=True,
        check_exact=True,
    )


def test_exporters_tagged_codec_preserves_typed_dataframe_and_nulls():
    payload = _make_exporters_payload(
        typed=True,
        timezone_aware=True,
    )
    prepared = exporters._prepare_exporters_overview_snapshot_payload(
        payload
    )
    decoded_cube = exporters._decode_exporters_json_payload(
        prepared["charts_cube"]
    )
    raw_records = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )["Global"]
    decoded_records = snapshots.unpack_record_mapping(
        decoded_cube
    )["Global"]

    assert list(decoded_records[0]) == list(raw_records[0])
    assert isinstance(decoded_records[0]["date"], pd.Timestamp)
    assert str(decoded_records[0]["date"].tz) == "Europe/London"
    assert (
        decoded_records[0]["date"].value
        == raw_records[0]["date"].value
    )
    assert decoded_records[0]["missing_timestamp"] is pd.NaT
    assert decoded_records[0]["pandas_missing"] is pd.NA
    assert decoded_records[0]["nullable_text"] is None
    assert type(decoded_records[0]["year"]) is np.int16
    assert type(decoded_records[0]["rolling_avg"]) is np.float32
    assert type(decoded_records[0]["is_forecast"]) is np.bool_
    assert type(decoded_records[0]["numpy_int"]) is np.int32
    assert type(decoded_records[0]["numpy_float"]) is np.float64
    assert type(decoded_records[0]["numpy_bool"]) is np.bool_
    assert np.isnat(decoded_records[-1]["numpy_timestamp"])

    pd.testing.assert_frame_equal(
        pd.DataFrame(decoded_records),
        pd.DataFrame(raw_records),
        check_dtype=True,
        check_exact=True,
    )
    decoded_supply = exporters._decode_exporters_json_payload(
        prepared["supply_dest"]
    )
    assert (
        type(decoded_supply["show_all"][0]["30D"])
        is np.float32
    )
    assert (
        type(decoded_supply["show_all"][0]["7D"])
        is np.float64
    )
    assert decoded_supply["show_all"][0]["nullable_note"] is pd.NA


def test_exporters_loader_never_falls_back_to_raw_payloads(
    monkeypatch,
    persistent_exporters_cache,
):
    payload = _make_exporters_payload()
    _install_exporters_data_sources(monkeypatch, payload)
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": exporters.EXPORTERS_SUPPLY_CHARTS_NAMESPACE,
        "source_key": "source",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        exporters,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (
            non_resolvable,
            payload,
        ),
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        exporters.refresh_all_data(
            _source_state(),
            "Country",
            "Installation",
            30,
        )


def test_exporters_legacy_and_reference_controls_charts_tables_equal(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    charts_store, _continent_store, supply_dest_store = stores
    raw_charts = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )
    selected_years = ["2025", "2026"]

    assert to_json(
        exporters.update_supply_year_selector_options(
            raw_charts,
            selected_years,
        )
    ) == to_json(
        exporters.update_supply_year_selector_options(
            charts_store,
            selected_years,
        )
    )
    assert to_json(
        exporters.update_supply_charts(
            raw_charts,
            "mcm_d",
            selected_years,
            30,
        )
    ) == to_json(
        exporters.update_supply_charts(
            charts_store,
            "mcm_d",
            selected_years,
            30,
        )
    )
    raw_table = exporters.update_supply_dest_table(
        payload["supply_dest"],
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    reference_table = exporters.update_supply_dest_table(
        supply_dest_store,
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    assert to_json(reference_table) == to_json(raw_table)


def test_exporters_continent_reference_controls_and_charts_equal(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]
    selected_years = ["2025", "2026"]
    raw_continent_store = _raw_continent_store(
        payload,
        continent_frame,
        selected_years=selected_years,
    )

    assert to_json(
        exporters.update_continent_year_selector_options(
            raw_continent_store,
            selected_years,
        )
    ) == to_json(
        exporters.update_continent_year_selector_options(
            continent_store,
            selected_years,
        )
    )
    for chart_type in ("absolute", "percentage"):
        assert to_json(
            exporters.update_continent_charts(
                raw_continent_store,
                chart_type,
                "Country",
                "mcm_d",
                selected_years,
                30,
            )
        ) == to_json(
            exporters.update_continent_charts(
                continent_store,
                chart_type,
                "Country",
                "mcm_d",
                selected_years,
                30,
            )
        )


def test_exporters_continent_interactions_and_export_do_not_read_database(
    monkeypatch,
    persistent_exporters_cache,
):
    _payload, continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]
    custom_fetches = []

    def fetch_custom_selection(
        _engine,
        _schema,
        _entity_names,
        _classification_mode,
        selected_years=None,
        rolling_avg_days=30,
    ):
        custom_fetches.append(list(selected_years or []))
        return continent_frame.copy()

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_custom_selection,
    )
    first_render = exporters.update_continent_charts(
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025"],
        30,
    )
    assert isinstance(first_render, html.Div)
    assert custom_fetches == [["2025"]]
    first_export = exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025"],
        30,
    )
    assert first_export is not None
    assert custom_fetches == [
        ["2025"],
        exporters._get_continent_chart_available_years(),
    ]

    def unexpected_database_read(*_args, **_kwargs):
        raise AssertionError("continent interaction queried the database")

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        unexpected_database_read,
    )
    monkeypatch.setattr(
        exporters.pd,
        "read_sql",
        unexpected_database_read,
    )

    for chart_type, metric, years in (
        ("absolute", "mcm_d", ["2025"]),
        ("absolute", "bcm", ["2025", "2026"]),
        ("absolute", "mtpa", ["2025", "2026"]),
        ("percentage", "mcm_d", ["2025"]),
    ):
        rendered = exporters.update_continent_charts(
            continent_store,
            chart_type,
            "Country",
            metric,
            years,
            30,
        )
        assert isinstance(rendered, html.Div)

    download = exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025"],
        30,
    )
    assert download is not None


@pytest.mark.parametrize(
    ("requested_years", "expected_years"),
    [
        (None, None),
        (["2026"], ["2026"]),
        (["2022", "2025"], ["2022", "2025"]),
    ],
)
def test_exporters_continent_snapshot_uses_exact_selected_year_window(
    monkeypatch,
    requested_years,
    expected_years,
):
    if expected_years is None:
        expected_years = (
            exporters._default_continent_chart_selected_years(
                exporters._get_continent_chart_available_years()
            )
        )
    captured = {}

    def fetch_batch(
        _engine,
        _schema,
        entity_names,
        classification_mode,
        selected_years,
        rolling_avg_days,
    ):
        captured.update({
            "entity_names": entity_names,
            "classification_mode": classification_mode,
            "selected_years": selected_years,
            "rolling_avg_days": rolling_avg_days,
        })
        return pd.DataFrame()

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_batch,
    )

    payload = exporters._build_exporters_continent_snapshot_payload(
        object(),
        "at_lng",
        _source_state(),
        "Country",
        30,
        ["Global"],
        requested_years,
    )

    assert captured["selected_years"] == expected_years
    assert captured["entity_names"] == ["Global"]
    assert payload["entities"] == ["Global"]
    assert payload["data"].empty
    assert payload["selected_years"] == expected_years
    assert payload["classification_mode"] == "Country"
    assert payload["rolling_avg_days"] == 30
    assert payload["source_state"] == _source_state()


def test_exporters_default_window_excludes_historical_only_zero_continents(
    monkeypatch,
):
    entities = ["Angola", "Russian Federation"]
    category_observations = {
        entity_name: {
            "Asia": pd.Timestamp("2025-01-10"),
            "Europe": pd.Timestamp("2025-01-11"),
            "Unknown": pd.Timestamp("2025-01-12"),
            "Americas": pd.Timestamp("2022-06-01"),
        }
        for entity_name in entities
    }

    def fetch_windowed_fixture(
        _engine,
        _schema,
        entity_names,
        _classification_mode,
        selected_years=None,
        rolling_avg_days=30,
    ):
        _active_years, query_start_date, display_start_date = (
            exporters._get_continent_chart_selected_window(
                selected_years
            )
        )
        query_start = pd.Timestamp(query_start_date)
        return pd.DataFrame.from_records([
            {
                "entity_name": entity_name,
                "date": pd.Timestamp(display_start_date),
                "continent_destination": continent,
                "year": int(display_start_date[:4]),
                "day_of_year": 1,
                "month_day": "Jan 01",
                "rolling_avg": 0.0,
                "percentage": 0.0,
                "is_forecast": False,
            }
            for entity_name in entity_names
            for continent, observed_at
            in category_observations[entity_name].items()
            if observed_at >= query_start
        ])

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_windowed_fixture,
    )

    default_payload = (
        exporters._build_exporters_continent_snapshot_payload(
            object(),
            "at_lng",
            _source_state(),
            "Country",
            30,
            entities,
            None,
        )
    )
    historical_payload = (
        exporters._build_exporters_continent_snapshot_payload(
            object(),
            "at_lng",
            _source_state(),
            "Country",
            30,
            entities,
            ["2022"],
        )
    )

    for entity_name in entities:
        default_categories = set(
            default_payload["data"].loc[
                default_payload["data"]["entity_name"] == entity_name,
                "continent_destination",
            ]
        )
        historical_categories = set(
            historical_payload["data"].loc[
                historical_payload["data"]["entity_name"] == entity_name,
                "continent_destination",
            ]
        )
        assert default_categories == {"Asia", "Europe", "Unknown"}
        assert "Americas" not in default_categories
        assert "Americas" in historical_categories


def test_exporters_excel_uses_lazy_full_history_category_universe(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]
    default_payload = exporters._resolve_exporters_continent_payload(
        continent_store
    )
    assert "Americas" not in set(
        default_payload["data"]["continent_destination"]
    )

    export_fetches = []

    def fetch_full_history(
        _engine,
        _schema,
        entity_names,
        _classification_mode,
        selected_years=None,
        rolling_avg_days=30,
    ):
        export_fetches.append(list(selected_years or []))
        records = []
        for entity_name in entity_names:
            for year in (2025, 2026):
                for continent_index, continent in enumerate(
                    ("Americas", "Asia", "Europe", "Unknown")
                ):
                    date = pd.Timestamp(
                        year=year,
                        month=1,
                        day=continent_index + 1,
                    )
                    records.append({
                        "entity_name": entity_name,
                        "date": date,
                        "continent_destination": continent,
                        "year": year,
                        "day_of_year": continent_index + 1,
                        "month_day": date.strftime("%b %d"),
                        "rolling_avg": (
                            0.0 if continent == "Americas" else 10.0
                        ),
                        "percentage": (
                            0.0 if continent == "Americas" else 100 / 3
                        ),
                        "is_forecast": False,
                    })
        return pd.DataFrame.from_records(records)

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_full_history,
    )

    absolute_download = exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025", "2026"],
        30,
    )
    percentage_download = exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "percentage",
        "Country",
        "mcm_d",
        ["2025", "2026"],
        30,
    )

    assert export_fetches == [
        exporters._get_continent_chart_available_years()
    ]
    for download in (absolute_download, percentage_download):
        workbook = _exporters_workbook_cells(download)
        cells = workbook["All Data"]["cells"]
        headers = [cell[0] for cell in cells[0]]
        rows = [
            dict(zip(headers, [cell[0] for cell in row]))
            for row in cells[1:]
        ]
        assert {row["continent_destination"] for row in rows} == {
            "Americas",
            "Asia",
            "Europe",
            "Unknown",
        }
        assert {int(row["year"]) for row in rows} == {2025, 2026}
        assert any(
            row["continent_destination"] == "Americas"
            for row in rows
        )


def test_exporters_continent_runtime_failure_keeps_legacy_fallbacks(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]

    def fail_build(*_args, **_kwargs):
        raise RuntimeError("transient continent query failure")

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fail_build,
    )

    rendered = exporters.update_continent_charts(
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2022"],
        30,
    )
    assert isinstance(rendered, html.Div)
    assert (
        rendered.children
        != exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    graphs = [
        component
        for component in _walk_exporters_snapshot_components(rendered)
        if hasattr(component, "figure")
    ]
    assert len(graphs) == len(payload["continent_entities"])
    assert all(len(graph.figure.data) == 0 for graph in graphs)

    assert exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025", "2026"],
        30,
    ) is None


def test_exporters_three_excel_exports_preserve_workbook_cells(
    monkeypatch,
    persistent_exporters_cache,
):
    payload = _make_exporters_payload(typed=True)
    payload, continent_frame, stores = _load_representative_references(
        monkeypatch,
        payload,
    )
    charts_store, continent_store, supply_dest_store = stores
    raw_charts = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )

    raw_supply_download = exporters.export_supply_charts_to_excel(
        1,
        raw_charts,
        "mcm_d",
        30,
    )
    reference_supply_download = (
        exporters.export_supply_charts_to_excel(
            1,
            charts_store,
            "mcm_d",
            30,
        )
    )
    raw_supply_cells = _exporters_workbook_cells(raw_supply_download)
    reference_supply_cells = _exporters_workbook_cells(
        reference_supply_download
    )
    assert reference_supply_cells == raw_supply_cells
    supply_date_cell = raw_supply_cells["All Data"]["cells"][1][1]
    assert supply_date_cell[1] == "d"
    assert supply_date_cell[2] != "General"

    raw_table = exporters.update_supply_dest_table(
        payload["supply_dest"],
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    reference_table = exporters.update_supply_dest_table(
        supply_dest_store,
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    raw_grid = _first_exporters_snapshot_grid(raw_table)
    reference_grid = _first_exporters_snapshot_grid(reference_table)
    raw_table_download = (
        exporters.export_supply_dest_table_to_excel(
            1,
            [raw_grid.rowData],
            [raw_grid.rowData],
            [raw_grid.columnDefs],
        )
    )
    reference_table_download = (
        exporters.export_supply_dest_table_to_excel(
            1,
            [reference_grid.rowData],
            [reference_grid.rowData],
            [reference_grid.columnDefs],
        )
    )
    assert _exporters_workbook_cells(
        reference_table_download
    ) == _exporters_workbook_cells(raw_table_download)

    raw_continent_store = _raw_continent_store(
        payload,
        continent_frame,
    )
    raw_continent_download = (
        exporters.export_continent_charts_to_excel(
            1,
            raw_continent_store,
            "absolute",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )
    )
    reference_continent_download = (
        exporters.export_continent_charts_to_excel(
            1,
            continent_store,
            "absolute",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )
    )
    assert _exporters_workbook_cells(
        reference_continent_download
    ) == _exporters_workbook_cells(raw_continent_download)
    _assert_workbook_date_number_format(
        reference_continent_download
    )

    raw_percentage_download = (
        exporters.export_continent_charts_to_excel(
            1,
            raw_continent_store,
            "percentage",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )
    )
    reference_percentage_download = (
        exporters.export_continent_charts_to_excel(
            1,
            continent_store,
            "percentage",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )
    )
    assert _exporters_workbook_cells(
        reference_percentage_download
    ) == _exporters_workbook_cells(raw_percentage_download)
    _assert_workbook_date_number_format(
        reference_percentage_download
    )


@pytest.mark.parametrize("corruption_mode", ["missing", "corrupt"])
def test_exporters_missing_or_corrupt_refs_show_explicit_recovery(
    monkeypatch,
    persistent_exporters_cache,
    corruption_mode,
):
    _payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    charts_store, _continent_store, supply_dest_store = stores
    persistent_stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        charts_store["namespace"],
        charts_store["source_key"],
        charts_store["revision"],
    )
    if corruption_mode == "missing":
        persistent_stores.cache.delete(record_key, retry=True)
    else:
        persistent_stores.cache.set(
            record_key,
            b"corrupt",
            retry=True,
        )
    snapshots.clear_local_snapshots()

    options, selected = (
        exporters.update_supply_year_selector_options(
            charts_store,
            ["2026"],
        )
    )
    assert selected == []
    assert (
        options[0]["label"]
        == exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    chart_notice = exporters.update_supply_charts(
        charts_store,
        "mcm_d",
        ["2026"],
        30,
    )
    table_notice = exporters.update_supply_dest_table(
        supply_dest_store,
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    assert isinstance(chart_notice, html.Div)
    assert (
        chart_notice.children
        == exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    assert _first_exporters_snapshot_grid(table_notice).rowData
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        exporters.export_supply_charts_to_excel(
            1,
            charts_store,
            "mcm_d",
            30,
        )


@pytest.mark.parametrize("corruption_mode", ["missing", "corrupt"])
def test_exporters_missing_or_corrupt_continent_ref_recovers_explicitly(
    monkeypatch,
    persistent_exporters_cache,
    corruption_mode,
):
    _payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]
    persistent_stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        continent_store["namespace"],
        continent_store["source_key"],
        continent_store["revision"],
    )
    if corruption_mode == "missing":
        persistent_stores.cache.delete(record_key, retry=True)
    else:
        persistent_stores.cache.set(
            record_key,
            b"corrupt",
            retry=True,
        )
    snapshots.clear_local_snapshots()

    options, selected = (
        exporters.update_continent_year_selector_options(
            continent_store,
            ["2026"],
        )
    )
    assert selected == []
    assert (
        options[0]["label"]
        == exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    notice = exporters.update_continent_charts(
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(notice, html.Div)
    assert notice.children == exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        exporters.export_continent_charts_to_excel(
            1,
            continent_store,
            "absolute",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )


def test_exporters_source_watermark_only_runs_on_initial_or_global_refresh(
    monkeypatch,
    persistent_exporters_cache,
):
    watermark_calls = []
    global_refresh = False

    monkeypatch.setattr(
        exporters,
        "_fetch_exporters_source_watermark",
        lambda: (
            watermark_calls.append("read")
            or pd.Timestamp("2026-07-24T00:00:00")
        ),
    )
    monkeypatch.setattr(
        exporters,
        "_was_global_refresh_triggered",
        lambda: global_refresh,
    )

    initial_state, initial_status = (
        exporters.refresh_exporters_source_state(0, 0)
    )
    assert initial_state["refresh_token"] is None
    assert initial_status["status"] == "checked"
    assert watermark_calls == ["read"]

    payload = _make_exporters_payload()
    _install_exporters_data_sources(monkeypatch, payload)
    exporters.refresh_all_data(
        initial_state,
        "Country",
        "Installation",
        30,
    )
    exporters.refresh_all_data(
        initial_state,
        "Country",
        "Country",
        30,
    )
    assert watermark_calls == ["read"]

    global_refresh = True
    refreshed_state, refreshed_status = (
        exporters.refresh_exporters_source_state(
            0,
            1,
            initial_state,
        )
    )
    assert refreshed_state is no_update
    assert refreshed_status["refresh_generation"] == 1
    assert watermark_calls == ["read", "read"]


def test_exporters_source_keys_follow_only_their_dependencies():
    source_state = _source_state()
    changed_source_state = _source_state("forced")
    entity_names = ["Global", "Exporter 1"]

    destination_key = exporters._exporters_destination_base_source_key(
        source_state
    )
    assert destination_key == (
        exporters._exporters_destination_base_source_key(source_state)
    )
    assert destination_key == (
        exporters._exporters_destination_base_source_key(
            changed_source_state
        )
    )

    supply_key = exporters._exporters_supply_charts_source_key(
        source_state,
        "Country",
        30,
        entity_names,
    )
    assert supply_key != exporters._exporters_supply_charts_source_key(
        source_state,
        "Country",
        14,
        entity_names,
    )
    assert supply_key != exporters._exporters_supply_charts_source_key(
        source_state,
        "Classification Level 1",
        30,
        entity_names,
    )

    continent_key = exporters._exporters_continent_data_source_key(
        source_state,
        "Country",
        30,
        entity_names,
        ["2025", "2026"],
    )
    assert continent_key != (
        exporters._exporters_continent_data_source_key(
            source_state,
            "Country",
            30,
            entity_names + ["Exporter 2"],
            ["2025", "2026"],
        )
    )
    assert continent_key != (
        exporters._exporters_continent_data_source_key(
            source_state,
            "Country",
            30,
            entity_names,
            ["2026"],
        )
    )

    base_reference = {
        "namespace": exporters.EXPORTERS_DESTINATION_BASE_NAMESPACE,
        "source_key": destination_key,
        "revision": "revision-a",
    }
    summary_key = exporters._exporters_destination_summary_source_key(
        base_reference,
        "Country",
        "Installation",
    )
    assert summary_key != (
        exporters._exporters_destination_summary_source_key(
            base_reference,
            "Country",
            "Country",
        )
    )
    changed_reference = dict(
        base_reference,
        revision="revision-b",
    )
    assert summary_key != (
        exporters._exporters_destination_summary_source_key(
            changed_reference,
            "Country",
            "Installation",
        )
    )


def test_exporters_global_refresh_token_reuses_all_snapshot_builds(
    monkeypatch,
    persistent_exporters_cache,
):
    payload = _make_exporters_payload()
    call_counts = {
        "supply": 0,
        "destination": 0,
        "continent": 0,
        "summary": 0,
    }
    charts_data = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )
    entities = list(payload["continent_entities"])

    monkeypatch.setattr(
        exporters,
        "_get_exporter_entity_names",
        lambda *_args: entities,
    )

    def fetch_supply(*_args):
        call_counts["supply"] += 1
        return {
            entity: pd.DataFrame(copy.deepcopy(charts_data[entity]))
            for entity in entities
        }

    def fetch_destination(*_args):
        call_counts["destination"] += 1
        return pd.DataFrame({"base_row": [1]})

    def fetch_continent(*_args, **_kwargs):
        call_counts["continent"] += 1
        return _make_continent_frame(entities)

    def build_summary(*_args):
        call_counts["summary"] += 1
        return copy.deepcopy(payload["supply_dest"])

    monkeypatch.setattr(
        exporters,
        "_fetch_supply_chart_data_for_entities",
        fetch_supply,
    )
    monkeypatch.setattr(
        exporters,
        "fetch_supply_destination_base_data",
        fetch_destination,
    )
    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_continent,
    )
    monkeypatch.setattr(
        exporters,
        "build_supply_dest_summary_store_payload",
        build_summary,
    )

    original = exporters.refresh_all_data(
        _source_state(),
        "Country",
        "Installation",
        30,
    )
    warm = exporters.refresh_all_data(
        _source_state(),
        "Country",
        "Installation",
        30,
    )
    rebuilt = exporters.refresh_all_data(
        _source_state("forced-refresh"),
        "Country",
        "Installation",
        30,
    )

    assert call_counts == {
        "supply": 1,
        "destination": 1,
        "continent": 1,
        "summary": 1,
    }
    assert warm == original
    assert rebuilt == original


def test_exporters_source_drift_keeps_prior_atomic_bundle_active(
    monkeypatch,
    persistent_exporters_cache,
):
    payload = _make_exporters_payload()
    _install_exporters_data_sources(monkeypatch, payload)

    def source_pair(snapshot_id, timestamp):
        return {
            "current_snapshot_id": snapshot_id,
            "current_snapshot_date_utc": "2026-07-25",
            "current_snapshot_timestamp_utc": timestamp,
            "current_facts_retained": True,
        }

    pair_a = source_pair(101, "2026-07-25T12:00:00Z")
    state_a = exporters._build_exporters_source_state(pair_a)
    monkeypatch.setattr(
        exporters,
        "_fetch_exporters_source_watermark",
        lambda: pair_a,
    )
    exporters.refresh_all_data(
        state_a,
        "Country",
        "Installation",
        30,
    )
    entities = list(payload["continent_entities"])
    bundle_key_a = exporters._build_source_key(
        exporters.EXPORTERS_REFRESH_BUNDLE_NAMESPACE,
        exporters._exporters_semantic_source_state(state_a),
        "Country",
        "Installation",
        30,
        entities,
    )
    prior_bundle = snapshots.get_snapshot_if_available(
        exporters.engine,
        namespace=exporters.EXPORTERS_REFRESH_BUNDLE_NAMESPACE,
        source_key=bundle_key_a,
    )
    assert prior_bundle is not None

    pair_b = source_pair(102, "2026-07-25T13:00:00Z")
    state_b = exporters._build_exporters_source_state(pair_b)
    pair_c = source_pair(103, "2026-07-25T14:00:00Z")
    monkeypatch.setattr(
        exporters,
        "_fetch_exporters_source_watermark",
        lambda: pair_c,
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="changed during snapshot construction",
    ):
        exporters.refresh_all_data(
            state_b,
            "Country",
            "Installation",
            30,
        )

    assert snapshots.get_snapshot_if_available(
        exporters.engine,
        namespace=exporters.EXPORTERS_REFRESH_BUNDLE_NAMESPACE,
        source_key=bundle_key_a,
    ) == prior_bundle
    bundle_key_b = exporters._build_source_key(
        exporters.EXPORTERS_REFRESH_BUNDLE_NAMESPACE,
        exporters._exporters_semantic_source_state(state_b),
        "Country",
        "Installation",
        30,
        entities,
    )
    assert snapshots.get_snapshot_if_available(
        exporters.engine,
        namespace=exporters.EXPORTERS_REFRESH_BUNDLE_NAMESPACE,
        source_key=bundle_key_b,
    ) is None


@pytest.mark.parametrize("pool_size", [1, 2, 4])
def test_exporters_loader_single_flight_at_pool_sizes(
    monkeypatch,
    persistent_exporters_cache,
    pool_size,
):
    payload = _make_exporters_payload()
    charts_data = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )
    entities = list(payload["continent_entities"])
    call_counts = {
        "supply": 0,
        "destination": 0,
        "continent": 0,
        "summary": 0,
    }
    build_lock = threading.Lock()

    def mark_call(name):
        with build_lock:
            call_counts[name] += 1
        time.sleep(0.05)

    monkeypatch.setattr(
        exporters,
        "_get_exporter_entity_names",
        lambda *_args: entities,
    )

    def fetch_supply(*_args):
        mark_call("supply")
        return {
            entity: pd.DataFrame(copy.deepcopy(charts_data[entity]))
            for entity in entities
        }

    def fetch_destination(*_args):
        mark_call("destination")
        return pd.DataFrame({"base_row": [1]})

    def fetch_continent(*_args, **_kwargs):
        mark_call("continent")
        return _make_continent_frame(entities)

    def build_summary(*_args):
        mark_call("summary")
        return copy.deepcopy(payload["supply_dest"])

    monkeypatch.setattr(
        exporters,
        "_fetch_supply_chart_data_for_entities",
        fetch_supply,
    )
    monkeypatch.setattr(
        exporters,
        "fetch_supply_destination_base_data",
        fetch_destination,
    )
    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_continent,
    )
    monkeypatch.setattr(
        exporters,
        "build_supply_dest_summary_store_payload",
        build_summary,
    )

    def load():
        return exporters.refresh_all_data(
            _source_state(),
            "Country",
            "Installation",
            30,
        )

    with ThreadPoolExecutor(max_workers=pool_size) as executor:
        results = list(
            executor.map(lambda _index: load(), range(pool_size))
        )

    assert call_counts == {
        "supply": 1,
        "destination": 1,
        "continent": 1,
        "summary": 1,
    }
    assert all(result == results[0] for result in results)
    assert snapshots.snapshot_is_resolvable(results[0][0])
    assert snapshots.snapshot_is_resolvable(results[0][1])
    assert snapshots.snapshot_is_resolvable(results[0][2])
    assert len(to_json(results[0]).encode("utf-8")) < 50_000


def test_exporters_independent_cold_loads_overlap_and_keep_named_order(
    monkeypatch,
):
    barrier = threading.Barrier(3)

    def loader(name):
        barrier.wait(timeout=1)
        return ({"name": name}, name)

    monkeypatch.setattr(
        exporters,
        "_load_exporters_destination_base_snapshot",
        lambda *_args: loader("destination"),
    )
    monkeypatch.setattr(
        exporters,
        "_load_exporters_supply_charts_snapshot",
        lambda *_args: loader("supply"),
    )
    monkeypatch.setattr(
        exporters,
        "_load_exporters_continent_snapshot",
        lambda *_args: loader("continent"),
    )

    loaded = exporters._load_exporters_independent_snapshots(
        object(),
        "at_lng",
        _source_state(),
        "Country",
        30,
        ["Global"],
    )

    assert list(loaded) == [
        "destination_base",
        "supply_charts",
        "continent_data",
    ]
    assert loaded["destination_base"][1] == "destination"
    assert loaded["supply_charts"][1] == "supply"
    assert loaded["continent_data"][1] == "continent"


def test_exporters_independent_loader_propagates_failure_without_partial_result(
    monkeypatch,
):
    monkeypatch.setattr(
        exporters,
        "_load_exporters_destination_base_snapshot",
        lambda *_args: ({"name": "destination"}, "destination"),
    )
    monkeypatch.setattr(
        exporters,
        "_load_exporters_supply_charts_snapshot",
        lambda *_args: (_ for _ in ()).throw(
            RuntimeError("supply failed")
        ),
    )
    monkeypatch.setattr(
        exporters,
        "_load_exporters_continent_snapshot",
        lambda *_args: ({"name": "continent"}, "continent"),
    )

    with pytest.raises(RuntimeError, match="supply failed"):
        exporters._load_exporters_independent_snapshots(
            object(),
            "at_lng",
            _source_state(),
            "Country",
            30,
            ["Global"],
        )


def test_exporters_warm_loader_response_benchmark_gate(
    monkeypatch,
    persistent_exporters_cache,
):
    template = _make_exporters_payload(
        entity_count=12,
        years=(2022, 2023, 2024, 2025, 2026),
        points_per_year=120,
        supply_dest_rows=1500,
    )
    _install_exporters_data_sources(monkeypatch, template)

    app = Flask(__name__)

    @app.get("/baseline")
    def baseline_response():
        response_payload = (
            snapshots.unpack_record_mapping(
                template["charts_cube"]
            ),
            template["continent_entities"],
            template["supply_dest"],
        )
        return Response(
            to_json(response_payload),
            mimetype="application/json",
        )

    @app.get("/reference")
    def reference_response():
        return Response(
            to_json(
                exporters.refresh_all_data(
                    _source_state(),
                    "Country",
                    "Installation",
                    30,
                )
            ),
            mimetype="application/json",
        )

    client = app.test_client()
    assert client.get("/baseline").status_code == 200
    assert client.get("/reference").status_code == 200

    timings = {"baseline": [], "reference": []}
    sizes = {"baseline": [], "reference": []}
    for _iteration in range(6):
        for mode in ("baseline", "reference"):
            started = time.perf_counter()
            response = client.get(f"/{mode}")
            elapsed = time.perf_counter() - started
            assert response.status_code == 200
            timings[mode].append(elapsed)
            sizes[mode].append(len(response.data))

    baseline_seconds = statistics.median(timings["baseline"])
    reference_seconds = statistics.median(timings["reference"])
    improvement = (
        baseline_seconds - reference_seconds
    ) / baseline_seconds
    reference_bytes = max(sizes["reference"])
    print(
        "EXPORTERS_WARM_BENCHMARK "
        f"baseline_ms={baseline_seconds * 1000:.2f} "
        f"reference_ms={reference_seconds * 1000:.2f} "
        f"improvement_pct={improvement * 100:.2f} "
        f"baseline_bytes={max(sizes['baseline'])} "
        f"reference_bytes={reference_bytes}"
    )

    assert reference_bytes < 50_000
    assert improvement > 0.10

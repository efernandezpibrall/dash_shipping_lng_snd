import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import statistics
import threading
import time

from dash import html
from dash._utils import to_json
from flask import Flask, Response
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import pytest

from pages import exporters
from utils import dashboard_snapshot_cache as snapshots


def _make_exporters_payload(
    *,
    entity_count=2,
    years=(2025, 2026),
    points_per_year=6,
    typed=False,
    timezone_aware=False,
    supply_dest_rows=3,
):
    charts_data = {}
    for entity_index in range(entity_count):
        entity_name = (
            "Global"
            if entity_index == 0
            else f"Exporter {entity_index}"
        )
        records = []
        for year in years:
            dates = pd.date_range(
                f"{year}-01-01",
                periods=points_per_year,
                freq="D",
            )
            for day_index, date in enumerate(dates):
                if timezone_aware:
                    date = date.tz_localize("Europe/London")
                record = {
                    "date": date,
                    "year": (
                        np.int16(year)
                        if typed
                        else str(year)
                    ),
                    "month_day": date.strftime("%b %d"),
                    "rolling_avg": (
                        np.float32(40 + entity_index + day_index)
                        if typed
                        else float(40 + entity_index + day_index)
                    ),
                    "is_forecast": (
                        np.bool_(year == max(years))
                        if typed
                        else year == max(years)
                    ),
                }
                if typed:
                    record.update({
                        "missing_timestamp": pd.NaT,
                        "numpy_timestamp": (
                            np.datetime64("NaT", "ns")
                            if day_index == points_per_year - 1
                            else np.datetime64(
                                date.tz_localize(None)
                                if timezone_aware
                                else date,
                                "ns",
                            )
                        ),
                        "pandas_missing": pd.NA,
                        "nullable_text": None,
                        "numpy_int": np.int32(day_index),
                        "numpy_float": np.float64(day_index + 0.25),
                        "numpy_bool": np.bool_(day_index % 2 == 0),
                    })
                records.append(record)
        charts_data[entity_name] = records

    destination_records = []
    for row_index in range(supply_dest_rows):
        destination_records.append({
            "supply_country": f"Country {row_index}",
            "supply_installation": f"Installation {row_index}",
            "30D": (
                np.float32(20 + row_index)
                if typed
                else float(20 + row_index)
            ),
            "7D": (
                np.float64(21 + row_index)
                if typed
                else float(21 + row_index)
            ),
            "30D_Y1": float(18 + row_index),
            "Q1'26": float(19 + row_index),
            "Jun'26": float(20 + row_index),
            "W25'26": float(21 + row_index),
            "nullable_note": pd.NA if typed else None,
        })
    destination_records.append({
        "supply_country": "GRAND TOTAL",
        "supply_installation": "",
        "30D": float(20 * supply_dest_rows),
        "7D": float(21 * supply_dest_rows),
        "30D_Y1": float(18 * supply_dest_rows),
        "Q1'26": float(19 * supply_dest_rows),
        "Jun'26": float(20 * supply_dest_rows),
        "W25'26": float(21 * supply_dest_rows),
        "nullable_note": None,
    })
    return {
        "charts_cube": snapshots.pack_record_mapping(charts_data),
        "continent_entities": list(charts_data),
        "supply_dest": {
            "show_all": copy.deepcopy(destination_records),
            "group_small_countries": copy.deepcopy(
                destination_records
            ),
        },
    }


def _make_continent_frame(entities, years=(2025, 2026)):
    records = []
    for entity_index, entity_name in enumerate(entities):
        for year in years:
            for day_index, date in enumerate(
                pd.date_range(f"{year}-01-01", periods=3, freq="D")
            ):
                rolling_avg = float(10 + entity_index + day_index)
                records.append({
                    "entity_name": entity_name,
                    "date": date,
                    "continent_destination": (
                        "Asia" if day_index % 2 == 0 else "Europe"
                    ),
                    "year": np.int16(year),
                    "day_of_year": np.int16(day_index + 1),
                    "month_day": date.strftime("%b %d"),
                    "rolling_avg": np.float32(rolling_avg),
                    "percentage": np.float32(
                        60 if day_index % 2 == 0 else 40
                    ),
                    "is_forecast": np.bool_(year == max(years)),
                })
    return pd.DataFrame.from_records(records)


def _source_state(refresh_token=None):
    return {
        "format": exporters.EXPORTERS_SOURCE_STATE_FORMAT,
        "source_watermark": "2026-07-24T00:00:00",
        "as_of_date": "2026-07-24",
        "refresh_token": refresh_token,
    }


def _raw_continent_store(
    payload,
    continent_frame,
    *,
    selected_years=None,
    classification_mode="Country",
    rolling_avg_days=30,
):
    selected_years, _query_start_date, _display_start_date = (
        exporters._get_continent_chart_selected_window(
            selected_years
        )
    )
    return {
        "entities": list(payload["continent_entities"]),
        "data": continent_frame.copy(),
        "source_state": _source_state(),
        "classification_mode": classification_mode,
        "rolling_avg_days": rolling_avg_days,
        "selected_years": selected_years,
    }


def _install_exporters_data_sources(
    monkeypatch,
    payload,
    *,
    continent_frame=None,
):
    charts_data = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )
    entities = list(payload["continent_entities"])
    continent_frame = (
        continent_frame.copy()
        if continent_frame is not None
        else _make_continent_frame(entities)
    )
    monkeypatch.setattr(
        exporters,
        "_get_exporter_entity_names",
        lambda *_args: list(entities),
    )
    monkeypatch.setattr(
        exporters,
        "_fetch_supply_chart_data_for_entities",
        lambda *_args: {
            entity_name: pd.DataFrame(
                copy.deepcopy(charts_data[entity_name])
            )
            for entity_name in entities
        },
    )
    monkeypatch.setattr(
        exporters,
        "fetch_supply_destination_base_data",
        lambda *_args: pd.DataFrame({
            "base_row": [1],
        }),
    )
    monkeypatch.setattr(
        exporters,
        "build_supply_dest_summary_store_payload",
        lambda *_args: copy.deepcopy(payload["supply_dest"]),
    )
    def fetch_continent(
        _engine,
        _schema,
        _entity_names,
        _classification_mode,
        selected_years=None,
        rolling_avg_days=30,
    ):
        _active_years, _query_start_date, display_start_date = (
            exporters._get_continent_chart_selected_window(
                selected_years
            )
        )
        if not _active_years:
            return pd.DataFrame()
        dates = pd.to_datetime(
            continent_frame.get("date"),
            errors="coerce",
        )
        return continent_frame[
            dates >= pd.Timestamp(display_start_date)
        ].copy()

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_continent,
    )
    return continent_frame


@pytest.fixture
def persistent_exporters_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "exporters-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    monkeypatch.setattr(
        exporters,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    monkeypatch.setattr(
        exporters,
        "_fetch_exporters_source_watermark",
        lambda: pd.Timestamp("2026-07-24T00:00:00"),
    )
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_representative_references(monkeypatch, payload=None):
    payload = payload or _make_exporters_payload()
    continent_frame = _install_exporters_data_sources(
        monkeypatch,
        payload,
    )
    stores = exporters.refresh_all_data(
        _source_state(),
        "Country",
        "Installation",
        30,
    )
    return payload, continent_frame, stores


def _walk_components(component):
    if component is None:
        return
    if isinstance(component, (list, tuple)):
        for child in component:
            yield from _walk_components(child)
        return
    yield component
    yield from _walk_components(
        getattr(component, "children", None)
    )


def _first_grid(component):
    return next(
        item
        for item in _walk_components(component)
        if hasattr(item, "rowData")
    )


def _workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: {
            "cells": [
                [
                    (
                        cell.value,
                        cell.data_type,
                        cell.number_format,
                    )
                    for cell in row
                ]
                for row in worksheet.iter_rows()
            ],
            "widths": {
                key: dimension.width
                for key, dimension
                in worksheet.column_dimensions.items()
            },
        }
        for worksheet in workbook.worksheets
    }


def _assert_workbook_date_number_format(
    download,
    expected_format="YYYY-MM-DD",
):
    workbook = _workbook_cells(download)
    for worksheet in workbook.values():
        cells = worksheet["cells"]
        headers = [cell[0] for cell in cells[0]]
        if "date" not in headers:
            continue
        date_index = headers.index("date")
        date_formats = {
            row[date_index][2]
            for row in cells[1:]
            if row[date_index][0] is not None
        }
        assert date_formats == {expected_format}


def test_exporters_loader_emits_small_resolvable_refs_and_survives_restart(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    charts_store, continent_store, supply_dest_store = stores

    for store in (charts_store, continent_store, supply_dest_store):
        assert snapshots.is_snapshot_reference(store)
        assert snapshots.snapshot_is_resolvable(store)
        assert len(to_json(store).encode("utf-8")) < 10_000
    assert len(to_json(stores).encode("utf-8")) < 50_000

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()

    assert exporters._resolve_exporters_store(
        charts_store
    ) == snapshots.unpack_record_mapping(payload["charts_cube"])
    assert (
        exporters._resolve_exporters_store(supply_dest_store)
        == payload["supply_dest"]
    )
    continent_payload = exporters._resolve_exporters_continent_payload(
        continent_store
    )
    assert continent_payload["entities"] == payload["continent_entities"]
    assert continent_payload["selected_years"] == (
        exporters._default_continent_chart_selected_years(
            exporters._get_continent_chart_available_years()
        )
    )
    pd.testing.assert_frame_equal(
        continent_payload["data"],
        continent_frame,
        check_dtype=True,
        check_exact=True,
    )


def test_exporters_tagged_codec_preserves_typed_dataframe_and_nulls():
    payload = _make_exporters_payload(
        typed=True,
        timezone_aware=True,
    )
    prepared = exporters._prepare_exporters_overview_snapshot_payload(
        payload
    )
    decoded_cube = exporters._decode_exporters_json_payload(
        prepared["charts_cube"]
    )
    raw_records = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )["Global"]
    decoded_records = snapshots.unpack_record_mapping(
        decoded_cube
    )["Global"]

    assert list(decoded_records[0]) == list(raw_records[0])
    assert isinstance(decoded_records[0]["date"], pd.Timestamp)
    assert str(decoded_records[0]["date"].tz) == "Europe/London"
    assert (
        decoded_records[0]["date"].value
        == raw_records[0]["date"].value
    )
    assert decoded_records[0]["missing_timestamp"] is pd.NaT
    assert decoded_records[0]["pandas_missing"] is pd.NA
    assert decoded_records[0]["nullable_text"] is None
    assert type(decoded_records[0]["year"]) is np.int16
    assert type(decoded_records[0]["rolling_avg"]) is np.float32
    assert type(decoded_records[0]["is_forecast"]) is np.bool_
    assert type(decoded_records[0]["numpy_int"]) is np.int32
    assert type(decoded_records[0]["numpy_float"]) is np.float64
    assert type(decoded_records[0]["numpy_bool"]) is np.bool_
    assert np.isnat(decoded_records[-1]["numpy_timestamp"])

    pd.testing.assert_frame_equal(
        pd.DataFrame(decoded_records),
        pd.DataFrame(raw_records),
        check_dtype=True,
        check_exact=True,
    )
    decoded_supply = exporters._decode_exporters_json_payload(
        prepared["supply_dest"]
    )
    assert (
        type(decoded_supply["show_all"][0]["30D"])
        is np.float32
    )
    assert (
        type(decoded_supply["show_all"][0]["7D"])
        is np.float64
    )
    assert decoded_supply["show_all"][0]["nullable_note"] is pd.NA


def test_exporters_loader_never_falls_back_to_raw_payloads(
    monkeypatch,
    persistent_exporters_cache,
):
    payload = _make_exporters_payload()
    _install_exporters_data_sources(monkeypatch, payload)
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": exporters.EXPORTERS_SUPPLY_CHARTS_NAMESPACE,
        "source_key": "source",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        exporters,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (
            non_resolvable,
            payload,
        ),
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        exporters.refresh_all_data(
            _source_state(),
            "Country",
            "Installation",
            30,
        )


def test_exporters_legacy_and_reference_controls_charts_tables_equal(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    charts_store, _continent_store, supply_dest_store = stores
    raw_charts = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )
    selected_years = ["2025", "2026"]

    assert to_json(
        exporters.update_supply_year_selector_options(
            raw_charts,
            selected_years,
        )
    ) == to_json(
        exporters.update_supply_year_selector_options(
            charts_store,
            selected_years,
        )
    )
    assert to_json(
        exporters.update_supply_charts(
            raw_charts,
            "mcm_d",
            selected_years,
            30,
        )
    ) == to_json(
        exporters.update_supply_charts(
            charts_store,
            "mcm_d",
            selected_years,
            30,
        )
    )
    raw_table = exporters.update_supply_dest_table(
        payload["supply_dest"],
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    reference_table = exporters.update_supply_dest_table(
        supply_dest_store,
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    assert to_json(reference_table) == to_json(raw_table)


def test_exporters_continent_reference_controls_and_charts_equal(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]
    selected_years = ["2025", "2026"]
    raw_continent_store = _raw_continent_store(
        payload,
        continent_frame,
        selected_years=selected_years,
    )

    assert to_json(
        exporters.update_continent_year_selector_options(
            raw_continent_store,
            selected_years,
        )
    ) == to_json(
        exporters.update_continent_year_selector_options(
            continent_store,
            selected_years,
        )
    )
    for chart_type in ("absolute", "percentage"):
        assert to_json(
            exporters.update_continent_charts(
                raw_continent_store,
                chart_type,
                "Country",
                "mcm_d",
                selected_years,
                30,
            )
        ) == to_json(
            exporters.update_continent_charts(
                continent_store,
                chart_type,
                "Country",
                "mcm_d",
                selected_years,
                30,
            )
        )


def test_exporters_continent_interactions_and_export_do_not_read_database(
    monkeypatch,
    persistent_exporters_cache,
):
    _payload, continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]
    custom_fetches = []

    def fetch_custom_selection(
        _engine,
        _schema,
        _entity_names,
        _classification_mode,
        selected_years=None,
        rolling_avg_days=30,
    ):
        custom_fetches.append(list(selected_years or []))
        return continent_frame.copy()

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_custom_selection,
    )
    first_render = exporters.update_continent_charts(
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025"],
        30,
    )
    assert isinstance(first_render, html.Div)
    assert custom_fetches == [["2025"]]
    first_export = exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025"],
        30,
    )
    assert first_export is not None
    assert custom_fetches == [
        ["2025"],
        exporters._get_continent_chart_available_years(),
    ]

    def unexpected_database_read(*_args, **_kwargs):
        raise AssertionError("continent interaction queried the database")

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        unexpected_database_read,
    )
    monkeypatch.setattr(
        exporters.pd,
        "read_sql",
        unexpected_database_read,
    )

    for chart_type, metric, years in (
        ("absolute", "mcm_d", ["2025"]),
        ("absolute", "mtpa", ["2025", "2026"]),
        ("percentage", "mcm_d", ["2025"]),
    ):
        rendered = exporters.update_continent_charts(
            continent_store,
            chart_type,
            "Country",
            metric,
            years,
            30,
        )
        assert isinstance(rendered, html.Div)

    download = exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025"],
        30,
    )
    assert download is not None


@pytest.mark.parametrize(
    ("requested_years", "expected_years"),
    [
        (None, None),
        (["2026"], ["2026"]),
        (["2022", "2025"], ["2022", "2025"]),
    ],
)
def test_exporters_continent_snapshot_uses_exact_selected_year_window(
    monkeypatch,
    requested_years,
    expected_years,
):
    if expected_years is None:
        expected_years = (
            exporters._default_continent_chart_selected_years(
                exporters._get_continent_chart_available_years()
            )
        )
    captured = {}

    def fetch_batch(
        _engine,
        _schema,
        entity_names,
        classification_mode,
        selected_years,
        rolling_avg_days,
    ):
        captured.update({
            "entity_names": entity_names,
            "classification_mode": classification_mode,
            "selected_years": selected_years,
            "rolling_avg_days": rolling_avg_days,
        })
        return pd.DataFrame()

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_batch,
    )

    payload = exporters._build_exporters_continent_snapshot_payload(
        object(),
        "at_lng",
        _source_state(),
        "Country",
        30,
        ["Global"],
        requested_years,
    )

    assert captured["selected_years"] == expected_years
    assert captured["entity_names"] == ["Global"]
    assert payload["entities"] == ["Global"]
    assert payload["data"].empty
    assert payload["selected_years"] == expected_years
    assert payload["classification_mode"] == "Country"
    assert payload["rolling_avg_days"] == 30
    assert payload["source_state"] == _source_state()


def test_exporters_default_window_excludes_historical_only_zero_continents(
    monkeypatch,
):
    entities = ["Angola", "Russian Federation"]
    category_observations = {
        entity_name: {
            "Asia": pd.Timestamp("2025-01-10"),
            "Europe": pd.Timestamp("2025-01-11"),
            "Unknown": pd.Timestamp("2025-01-12"),
            "Americas": pd.Timestamp("2022-06-01"),
        }
        for entity_name in entities
    }

    def fetch_windowed_fixture(
        _engine,
        _schema,
        entity_names,
        _classification_mode,
        selected_years=None,
        rolling_avg_days=30,
    ):
        _active_years, query_start_date, display_start_date = (
            exporters._get_continent_chart_selected_window(
                selected_years
            )
        )
        query_start = pd.Timestamp(query_start_date)
        return pd.DataFrame.from_records([
            {
                "entity_name": entity_name,
                "date": pd.Timestamp(display_start_date),
                "continent_destination": continent,
                "year": int(display_start_date[:4]),
                "day_of_year": 1,
                "month_day": "Jan 01",
                "rolling_avg": 0.0,
                "percentage": 0.0,
                "is_forecast": False,
            }
            for entity_name in entity_names
            for continent, observed_at
            in category_observations[entity_name].items()
            if observed_at >= query_start
        ])

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_windowed_fixture,
    )

    default_payload = (
        exporters._build_exporters_continent_snapshot_payload(
            object(),
            "at_lng",
            _source_state(),
            "Country",
            30,
            entities,
            None,
        )
    )
    historical_payload = (
        exporters._build_exporters_continent_snapshot_payload(
            object(),
            "at_lng",
            _source_state(),
            "Country",
            30,
            entities,
            ["2022"],
        )
    )

    for entity_name in entities:
        default_categories = set(
            default_payload["data"].loc[
                default_payload["data"]["entity_name"] == entity_name,
                "continent_destination",
            ]
        )
        historical_categories = set(
            historical_payload["data"].loc[
                historical_payload["data"]["entity_name"] == entity_name,
                "continent_destination",
            ]
        )
        assert default_categories == {"Asia", "Europe", "Unknown"}
        assert "Americas" not in default_categories
        assert "Americas" in historical_categories


def test_exporters_excel_uses_lazy_full_history_category_universe(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]
    default_payload = exporters._resolve_exporters_continent_payload(
        continent_store
    )
    assert "Americas" not in set(
        default_payload["data"]["continent_destination"]
    )

    export_fetches = []

    def fetch_full_history(
        _engine,
        _schema,
        entity_names,
        _classification_mode,
        selected_years=None,
        rolling_avg_days=30,
    ):
        export_fetches.append(list(selected_years or []))
        records = []
        for entity_name in entity_names:
            for year in (2025, 2026):
                for continent_index, continent in enumerate(
                    ("Americas", "Asia", "Europe", "Unknown")
                ):
                    date = pd.Timestamp(
                        year=year,
                        month=1,
                        day=continent_index + 1,
                    )
                    records.append({
                        "entity_name": entity_name,
                        "date": date,
                        "continent_destination": continent,
                        "year": year,
                        "day_of_year": continent_index + 1,
                        "month_day": date.strftime("%b %d"),
                        "rolling_avg": (
                            0.0 if continent == "Americas" else 10.0
                        ),
                        "percentage": (
                            0.0 if continent == "Americas" else 100 / 3
                        ),
                        "is_forecast": False,
                    })
        return pd.DataFrame.from_records(records)

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_full_history,
    )

    absolute_download = exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025", "2026"],
        30,
    )
    percentage_download = exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "percentage",
        "Country",
        "mcm_d",
        ["2025", "2026"],
        30,
    )

    assert export_fetches == [
        exporters._get_continent_chart_available_years()
    ]
    for download in (absolute_download, percentage_download):
        workbook = _workbook_cells(download)
        cells = workbook["All Data"]["cells"]
        headers = [cell[0] for cell in cells[0]]
        rows = [
            dict(zip(headers, [cell[0] for cell in row]))
            for row in cells[1:]
        ]
        assert {row["continent_destination"] for row in rows} == {
            "Americas",
            "Asia",
            "Europe",
            "Unknown",
        }
        assert {int(row["year"]) for row in rows} == {2025, 2026}
        assert any(
            row["continent_destination"] == "Americas"
            for row in rows
        )


def test_exporters_continent_runtime_failure_keeps_legacy_fallbacks(
    monkeypatch,
    persistent_exporters_cache,
):
    payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]

    def fail_build(*_args, **_kwargs):
        raise RuntimeError("transient continent query failure")

    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fail_build,
    )

    rendered = exporters.update_continent_charts(
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2022"],
        30,
    )
    assert isinstance(rendered, html.Div)
    assert (
        rendered.children
        != exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    graphs = [
        component
        for component in _walk_components(rendered)
        if hasattr(component, "figure")
    ]
    assert len(graphs) == len(payload["continent_entities"])
    assert all(len(graph.figure.data) == 0 for graph in graphs)

    assert exporters.export_continent_charts_to_excel(
        1,
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2025", "2026"],
        30,
    ) is None


def test_exporters_three_excel_exports_preserve_workbook_cells(
    monkeypatch,
    persistent_exporters_cache,
):
    payload = _make_exporters_payload(typed=True)
    payload, continent_frame, stores = _load_representative_references(
        monkeypatch,
        payload,
    )
    charts_store, continent_store, supply_dest_store = stores
    raw_charts = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )

    raw_supply_download = exporters.export_supply_charts_to_excel(
        1,
        raw_charts,
        "mcm_d",
        30,
    )
    reference_supply_download = (
        exporters.export_supply_charts_to_excel(
            1,
            charts_store,
            "mcm_d",
            30,
        )
    )
    raw_supply_cells = _workbook_cells(raw_supply_download)
    reference_supply_cells = _workbook_cells(
        reference_supply_download
    )
    assert reference_supply_cells == raw_supply_cells
    supply_date_cell = raw_supply_cells["All Data"]["cells"][1][1]
    assert supply_date_cell[1] == "d"
    assert supply_date_cell[2] != "General"

    raw_table = exporters.update_supply_dest_table(
        payload["supply_dest"],
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    reference_table = exporters.update_supply_dest_table(
        supply_dest_store,
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    raw_grid = _first_grid(raw_table)
    reference_grid = _first_grid(reference_table)
    raw_table_download = (
        exporters.export_supply_dest_table_to_excel(
            1,
            [raw_grid.rowData],
            [raw_grid.rowData],
            [raw_grid.columnDefs],
        )
    )
    reference_table_download = (
        exporters.export_supply_dest_table_to_excel(
            1,
            [reference_grid.rowData],
            [reference_grid.rowData],
            [reference_grid.columnDefs],
        )
    )
    assert _workbook_cells(
        reference_table_download
    ) == _workbook_cells(raw_table_download)

    raw_continent_store = _raw_continent_store(
        payload,
        continent_frame,
    )
    raw_continent_download = (
        exporters.export_continent_charts_to_excel(
            1,
            raw_continent_store,
            "absolute",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )
    )
    reference_continent_download = (
        exporters.export_continent_charts_to_excel(
            1,
            continent_store,
            "absolute",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )
    )
    assert _workbook_cells(
        reference_continent_download
    ) == _workbook_cells(raw_continent_download)
    _assert_workbook_date_number_format(
        reference_continent_download
    )

    raw_percentage_download = (
        exporters.export_continent_charts_to_excel(
            1,
            raw_continent_store,
            "percentage",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )
    )
    reference_percentage_download = (
        exporters.export_continent_charts_to_excel(
            1,
            continent_store,
            "percentage",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )
    )
    assert _workbook_cells(
        reference_percentage_download
    ) == _workbook_cells(raw_percentage_download)
    _assert_workbook_date_number_format(
        reference_percentage_download
    )


@pytest.mark.parametrize("corruption_mode", ["missing", "corrupt"])
def test_exporters_missing_or_corrupt_refs_show_explicit_recovery(
    monkeypatch,
    persistent_exporters_cache,
    corruption_mode,
):
    _payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    charts_store, _continent_store, supply_dest_store = stores
    persistent_stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        charts_store["namespace"],
        charts_store["source_key"],
        charts_store["revision"],
    )
    if corruption_mode == "missing":
        persistent_stores.cache.delete(record_key, retry=True)
    else:
        persistent_stores.cache.set(
            record_key,
            b"corrupt",
            retry=True,
        )
    snapshots.clear_local_snapshots()

    options, selected = (
        exporters.update_supply_year_selector_options(
            charts_store,
            ["2026"],
        )
    )
    assert selected == []
    assert (
        options[0]["label"]
        == exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    chart_notice = exporters.update_supply_charts(
        charts_store,
        "mcm_d",
        ["2026"],
        30,
    )
    table_notice = exporters.update_supply_dest_table(
        supply_dest_store,
        [],
        [],
        [],
        "Country",
        "absolute",
        "Installation",
        "levels",
        "show_all",
        "mcm_d",
        0,
        5,
        3,
        3,
    )
    assert isinstance(chart_notice, html.Div)
    assert (
        chart_notice.children
        == exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    assert _first_grid(table_notice).rowData
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        exporters.export_supply_charts_to_excel(
            1,
            charts_store,
            "mcm_d",
            30,
        )


@pytest.mark.parametrize("corruption_mode", ["missing", "corrupt"])
def test_exporters_missing_or_corrupt_continent_ref_recovers_explicitly(
    monkeypatch,
    persistent_exporters_cache,
    corruption_mode,
):
    _payload, _continent_frame, stores = (
        _load_representative_references(monkeypatch)
    )
    continent_store = stores[1]
    persistent_stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        continent_store["namespace"],
        continent_store["source_key"],
        continent_store["revision"],
    )
    if corruption_mode == "missing":
        persistent_stores.cache.delete(record_key, retry=True)
    else:
        persistent_stores.cache.set(
            record_key,
            b"corrupt",
            retry=True,
        )
    snapshots.clear_local_snapshots()

    options, selected = (
        exporters.update_continent_year_selector_options(
            continent_store,
            ["2026"],
        )
    )
    assert selected == []
    assert (
        options[0]["label"]
        == exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    notice = exporters.update_continent_charts(
        continent_store,
        "absolute",
        "Country",
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(notice, html.Div)
    assert notice.children == exporters.EXPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        exporters.export_continent_charts_to_excel(
            1,
            continent_store,
            "absolute",
            "Country",
            "mcm_d",
            ["2026"],
            30,
        )


def test_exporters_source_watermark_only_runs_on_initial_or_global_refresh(
    monkeypatch,
    persistent_exporters_cache,
):
    watermark_calls = []
    global_refresh = False

    monkeypatch.setattr(
        exporters,
        "_fetch_exporters_source_watermark",
        lambda: (
            watermark_calls.append("read")
            or pd.Timestamp("2026-07-24T00:00:00")
        ),
    )
    monkeypatch.setattr(
        exporters,
        "_was_global_refresh_triggered",
        lambda: global_refresh,
    )

    initial_state = exporters.refresh_exporters_source_state(0, 0)
    assert initial_state["refresh_token"] is None
    assert watermark_calls == ["read"]

    payload = _make_exporters_payload()
    _install_exporters_data_sources(monkeypatch, payload)
    exporters.refresh_all_data(
        initial_state,
        "Country",
        "Installation",
        30,
    )
    exporters.refresh_all_data(
        initial_state,
        "Country",
        "Country",
        30,
    )
    assert watermark_calls == ["read"]

    global_refresh = True
    refreshed_state = exporters.refresh_exporters_source_state(0, 1)
    assert refreshed_state["refresh_token"]
    assert watermark_calls == ["read", "read"]


def test_exporters_source_keys_follow_only_their_dependencies():
    source_state = _source_state()
    changed_source_state = _source_state("forced")
    entity_names = ["Global", "Exporter 1"]

    destination_key = exporters._exporters_destination_base_source_key(
        source_state
    )
    assert destination_key == (
        exporters._exporters_destination_base_source_key(source_state)
    )
    assert destination_key != (
        exporters._exporters_destination_base_source_key(
            changed_source_state
        )
    )

    supply_key = exporters._exporters_supply_charts_source_key(
        source_state,
        "Country",
        30,
        entity_names,
    )
    assert supply_key != exporters._exporters_supply_charts_source_key(
        source_state,
        "Country",
        14,
        entity_names,
    )
    assert supply_key != exporters._exporters_supply_charts_source_key(
        source_state,
        "Classification Level 1",
        30,
        entity_names,
    )

    continent_key = exporters._exporters_continent_data_source_key(
        source_state,
        "Country",
        30,
        entity_names,
        ["2025", "2026"],
    )
    assert continent_key != (
        exporters._exporters_continent_data_source_key(
            source_state,
            "Country",
            30,
            entity_names + ["Exporter 2"],
            ["2025", "2026"],
        )
    )
    assert continent_key != (
        exporters._exporters_continent_data_source_key(
            source_state,
            "Country",
            30,
            entity_names,
            ["2026"],
        )
    )

    base_reference = {
        "namespace": exporters.EXPORTERS_DESTINATION_BASE_NAMESPACE,
        "source_key": destination_key,
        "revision": "revision-a",
    }
    summary_key = exporters._exporters_destination_summary_source_key(
        base_reference,
        "Country",
        "Installation",
    )
    assert summary_key != (
        exporters._exporters_destination_summary_source_key(
            base_reference,
            "Country",
            "Country",
        )
    )
    changed_reference = dict(
        base_reference,
        revision="revision-b",
    )
    assert summary_key != (
        exporters._exporters_destination_summary_source_key(
            changed_reference,
            "Country",
            "Installation",
        )
    )


def test_exporters_global_refresh_token_forces_all_snapshot_rebuilds(
    monkeypatch,
    persistent_exporters_cache,
):
    payload = _make_exporters_payload()
    call_counts = {
        "supply": 0,
        "destination": 0,
        "continent": 0,
        "summary": 0,
    }
    charts_data = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )
    entities = list(payload["continent_entities"])

    monkeypatch.setattr(
        exporters,
        "_get_exporter_entity_names",
        lambda *_args: entities,
    )

    def fetch_supply(*_args):
        call_counts["supply"] += 1
        return {
            entity: pd.DataFrame(copy.deepcopy(charts_data[entity]))
            for entity in entities
        }

    def fetch_destination(*_args):
        call_counts["destination"] += 1
        return pd.DataFrame({"base_row": [1]})

    def fetch_continent(*_args, **_kwargs):
        call_counts["continent"] += 1
        return _make_continent_frame(entities)

    def build_summary(*_args):
        call_counts["summary"] += 1
        return copy.deepcopy(payload["supply_dest"])

    monkeypatch.setattr(
        exporters,
        "_fetch_supply_chart_data_for_entities",
        fetch_supply,
    )
    monkeypatch.setattr(
        exporters,
        "fetch_supply_destination_base_data",
        fetch_destination,
    )
    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_continent,
    )
    monkeypatch.setattr(
        exporters,
        "build_supply_dest_summary_store_payload",
        build_summary,
    )

    original = exporters.refresh_all_data(
        _source_state(),
        "Country",
        "Installation",
        30,
    )
    warm = exporters.refresh_all_data(
        _source_state(),
        "Country",
        "Installation",
        30,
    )
    rebuilt = exporters.refresh_all_data(
        _source_state("forced-refresh"),
        "Country",
        "Installation",
        30,
    )

    assert call_counts == {
        "supply": 2,
        "destination": 2,
        "continent": 2,
        "summary": 2,
    }
    assert warm == original
    assert all(
        current["revision"] != previous["revision"]
        for current, previous in zip(rebuilt, original)
    )


@pytest.mark.parametrize("pool_size", [1, 2, 4])
def test_exporters_loader_single_flight_at_pool_sizes(
    monkeypatch,
    persistent_exporters_cache,
    pool_size,
):
    payload = _make_exporters_payload()
    charts_data = snapshots.unpack_record_mapping(
        payload["charts_cube"]
    )
    entities = list(payload["continent_entities"])
    call_counts = {
        "supply": 0,
        "destination": 0,
        "continent": 0,
        "summary": 0,
    }
    build_lock = threading.Lock()

    def mark_call(name):
        with build_lock:
            call_counts[name] += 1
        time.sleep(0.05)

    monkeypatch.setattr(
        exporters,
        "_get_exporter_entity_names",
        lambda *_args: entities,
    )

    def fetch_supply(*_args):
        mark_call("supply")
        return {
            entity: pd.DataFrame(copy.deepcopy(charts_data[entity]))
            for entity in entities
        }

    def fetch_destination(*_args):
        mark_call("destination")
        return pd.DataFrame({"base_row": [1]})

    def fetch_continent(*_args, **_kwargs):
        mark_call("continent")
        return _make_continent_frame(entities)

    def build_summary(*_args):
        mark_call("summary")
        return copy.deepcopy(payload["supply_dest"])

    monkeypatch.setattr(
        exporters,
        "_fetch_supply_chart_data_for_entities",
        fetch_supply,
    )
    monkeypatch.setattr(
        exporters,
        "fetch_supply_destination_base_data",
        fetch_destination,
    )
    monkeypatch.setattr(
        exporters,
        "fetch_continent_chart_data_batch",
        fetch_continent,
    )
    monkeypatch.setattr(
        exporters,
        "build_supply_dest_summary_store_payload",
        build_summary,
    )

    def load():
        return exporters.refresh_all_data(
            _source_state(),
            "Country",
            "Installation",
            30,
        )

    with ThreadPoolExecutor(max_workers=pool_size) as executor:
        results = list(
            executor.map(lambda _index: load(), range(pool_size))
        )

    assert call_counts == {
        "supply": 1,
        "destination": 1,
        "continent": 1,
        "summary": 1,
    }
    assert all(result == results[0] for result in results)
    assert snapshots.snapshot_is_resolvable(results[0][0])
    assert snapshots.snapshot_is_resolvable(results[0][1])
    assert snapshots.snapshot_is_resolvable(results[0][2])
    assert len(to_json(results[0]).encode("utf-8")) < 50_000


def test_exporters_independent_cold_loads_overlap_and_keep_named_order(
    monkeypatch,
):
    barrier = threading.Barrier(3)

    def loader(name):
        barrier.wait(timeout=1)
        return ({"name": name}, name)

    monkeypatch.setattr(
        exporters,
        "_load_exporters_destination_base_snapshot",
        lambda *_args: loader("destination"),
    )
    monkeypatch.setattr(
        exporters,
        "_load_exporters_supply_charts_snapshot",
        lambda *_args: loader("supply"),
    )
    monkeypatch.setattr(
        exporters,
        "_load_exporters_continent_snapshot",
        lambda *_args: loader("continent"),
    )

    loaded = exporters._load_exporters_independent_snapshots(
        object(),
        "at_lng",
        _source_state(),
        "Country",
        30,
        ["Global"],
    )

    assert list(loaded) == [
        "destination_base",
        "supply_charts",
        "continent_data",
    ]
    assert loaded["destination_base"][1] == "destination"
    assert loaded["supply_charts"][1] == "supply"
    assert loaded["continent_data"][1] == "continent"


def test_exporters_independent_loader_propagates_failure_without_partial_result(
    monkeypatch,
):
    monkeypatch.setattr(
        exporters,
        "_load_exporters_destination_base_snapshot",
        lambda *_args: ({"name": "destination"}, "destination"),
    )
    monkeypatch.setattr(
        exporters,
        "_load_exporters_supply_charts_snapshot",
        lambda *_args: (_ for _ in ()).throw(
            RuntimeError("supply failed")
        ),
    )
    monkeypatch.setattr(
        exporters,
        "_load_exporters_continent_snapshot",
        lambda *_args: ({"name": "continent"}, "continent"),
    )

    with pytest.raises(RuntimeError, match="supply failed"):
        exporters._load_exporters_independent_snapshots(
            object(),
            "at_lng",
            _source_state(),
            "Country",
            30,
            ["Global"],
        )


def test_exporters_warm_loader_response_benchmark_gate(
    monkeypatch,
    persistent_exporters_cache,
):
    template = _make_exporters_payload(
        entity_count=12,
        years=(2022, 2023, 2024, 2025, 2026),
        points_per_year=120,
        supply_dest_rows=1500,
    )
    _install_exporters_data_sources(monkeypatch, template)

    app = Flask(__name__)

    @app.get("/baseline")
    def baseline_response():
        response_payload = (
            snapshots.unpack_record_mapping(
                template["charts_cube"]
            ),
            template["continent_entities"],
            template["supply_dest"],
        )
        return Response(
            to_json(response_payload),
            mimetype="application/json",
        )

    @app.get("/reference")
    def reference_response():
        return Response(
            to_json(
                exporters.refresh_all_data(
                    _source_state(),
                    "Country",
                    "Installation",
                    30,
                )
            ),
            mimetype="application/json",
        )

    client = app.test_client()
    assert client.get("/baseline").status_code == 200
    assert client.get("/reference").status_code == 200

    timings = {"baseline": [], "reference": []}
    sizes = {"baseline": [], "reference": []}
    for _iteration in range(6):
        for mode in ("baseline", "reference"):
            started = time.perf_counter()
            response = client.get(f"/{mode}")
            elapsed = time.perf_counter() - started
            assert response.status_code == 200
            timings[mode].append(elapsed)
            sizes[mode].append(len(response.data))

    baseline_seconds = statistics.median(timings["baseline"])
    reference_seconds = statistics.median(timings["reference"])
    improvement = (
        baseline_seconds - reference_seconds
    ) / baseline_seconds
    reference_bytes = max(sizes["reference"])
    print(
        "EXPORTERS_WARM_BENCHMARK "
        f"baseline_ms={baseline_seconds * 1000:.2f} "
        f"reference_ms={reference_seconds * 1000:.2f} "
        f"improvement_pct={improvement * 100:.2f} "
        f"baseline_bytes={max(sizes['baseline'])} "
        f"reference_bytes={reference_bytes}"
    )

    assert reference_bytes < 50_000
    assert improvement > 0.10

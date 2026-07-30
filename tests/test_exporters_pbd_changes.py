import base64
from io import BytesIO

from openpyxl import load_workbook
import numpy as np
import pandas as pd
import pytest

from pages import exporters


def _snapshot_pair(
    current_date,
    baseline_date=None,
    *,
    current_timestamp=None,
    baseline_timestamp=None,
):
    current_date = pd.Timestamp(current_date)
    pair = {
        'current_snapshot_id': 200,
        'current_snapshot_date_utc': current_date.date(),
        'current_snapshot_timestamp_utc': (
            current_timestamp
            or current_date.replace(
                hour=5,
                minute=34,
                second=12,
                microsecond=778724,
            )
        ),
        'current_facts_retained': True,
        'baseline_snapshot_id': None,
        'baseline_snapshot_date_utc': None,
        'baseline_snapshot_timestamp_utc': None,
        'baseline_facts_retained': None,
    }
    if baseline_date is not None:
        baseline_date = pd.Timestamp(baseline_date)
        pair.update({
            'baseline_snapshot_id': 199,
            'baseline_snapshot_date_utc': baseline_date.date(),
            'baseline_snapshot_timestamp_utc': (
                baseline_timestamp
                or baseline_date.replace(
                    hour=5,
                    minute=36,
                    second=57,
                    microsecond=998945,
                )
            ),
            'baseline_facts_retained': True,
        })
    return pair


@pytest.mark.parametrize(
    ('current_date', 'baseline_date', 'expected_status', 'expected_gap'),
    [
        ('2026-07-30', '2026-07-29', 'exact', 1),
        ('2026-07-27', '2026-07-24', 'exact', 1),
        ('2026-07-26', '2026-07-24', 'exact', 1),
        ('2026-07-30', '2026-07-28', 'fallback', 2),
        ('2026-07-30', None, 'unavailable', None),
    ],
)
def test_source_state_selects_previous_weekday_or_labels_fallback(
    current_date,
    baseline_date,
    expected_status,
    expected_gap,
):
    state = exporters._build_exporters_source_state(
        _snapshot_pair(current_date, baseline_date),
        refresh_token='refresh-1',
    )

    assert state['format'] == exporters.EXPORTERS_SOURCE_STATE_FORMAT
    assert state['format'].endswith('-v2')
    assert state['current_snapshot']['snapshot_id'] == 200
    assert state['current_snapshot']['snapshot_date_utc'] == current_date
    assert state['source_watermark'].endswith('05:34:12.778724')
    assert state['baseline_status'] == expected_status
    assert state['business_day_gap'] == expected_gap
    if baseline_date is None:
        assert state['baseline_snapshot'] is None
    else:
        assert state['baseline_snapshot']['snapshot_id'] == 199
        assert (
            state['baseline_snapshot']['snapshot_date_utc']
            == baseline_date
        )
        assert state['baseline_snapshot'][
            'snapshot_timestamp_utc'
        ].endswith('05:36:57.998945')


def _flow_row(
    flow_date,
    supply_country,
    mcmd,
    *,
    demand_country='Destination',
    supply_classification='Supply class',
    demand_classification='Demand class',
):
    return {
        'supply_classification': supply_classification,
        'supply_country': supply_country,
        'supply_installation': f'{supply_country} terminal',
        'demand_classification': demand_classification,
        'demand_country': demand_country,
        'flow_date': pd.Timestamp(flow_date),
        'mcmd': float(mcmd),
    }


def _constant_window(
    as_of_date,
    supply_country,
    mcmd,
    **row_kwargs,
):
    return [
        _flow_row(
            flow_date,
            supply_country,
            mcmd,
            **row_kwargs,
        )
        for flow_date in pd.date_range(
            end=pd.Timestamp(as_of_date),
            periods=30,
            freq='D',
        )
    ]


def test_rolling_windows_use_inclusive_7d_and_30d_boundaries():
    as_of_date = pd.Timestamp('2026-07-30')
    frame = pd.DataFrame([
        _flow_row(as_of_date - pd.Timedelta(days=6), 'A', 7),
        _flow_row(as_of_date - pd.Timedelta(days=7), 'A', 70),
        _flow_row(as_of_date - pd.Timedelta(days=29), 'A', 30),
        _flow_row(as_of_date - pd.Timedelta(days=30), 'A', 300),
    ])

    result = exporters._build_supply_dest_rolling_windows_from_df(
        frame,
        'Country',
        'Installation',
        as_of_date,
    )
    row = result.loc[result['supply_country'] == 'A'].iloc[0]

    assert row['7D'] == pytest.approx(1.0)
    assert row['30D'] == pytest.approx(3.6)


@pytest.mark.parametrize(
    ('classification_mode', 'aggregation_mode'),
    [
        ('Country', 'Installation'),
        ('Country', 'Country'),
        ('Country', 'Classification Level 1'),
        ('Classification Level 1', 'Installation'),
        ('Classification Level 1', 'Country'),
        ('Classification Level 1', 'Classification Level 1'),
    ],
)
def test_pbd_outer_join_keeps_additions_removals_and_all_aggregations(
    classification_mode,
    aggregation_mode,
):
    current_date = pd.Timestamp('2026-07-30')
    baseline_date = pd.Timestamp('2026-07-29')
    current_rows = (
        _constant_window(current_date, 'A', 10)
        + _constant_window(
            current_date,
            'B',
            5,
            demand_country='Destination B',
            demand_classification='Demand class B',
        )
        + _constant_window(
            current_date,
            'Internal',
            100,
            demand_country='Internal',
            supply_classification='Internal class',
            demand_classification='Internal class',
        )
        + [_flow_row(current_date - pd.Timedelta(days=30), 'A', 300)]
    )
    baseline_rows = (
        _constant_window(baseline_date, 'A', 8)
        + _constant_window(
            baseline_date,
            'C',
            4,
            demand_country='Destination C',
            demand_classification='Demand class C',
        )
        + _constant_window(
            baseline_date,
            'Internal',
            100,
            demand_country='Internal',
            supply_classification='Internal class',
            demand_classification='Internal class',
        )
        + [_flow_row(baseline_date - pd.Timedelta(days=30), 'A', 300)]
    )

    current = exporters._build_supply_dest_rolling_windows_from_df(
        pd.DataFrame(current_rows),
        classification_mode,
        aggregation_mode,
        current_date,
    )
    baseline = exporters._build_supply_dest_rolling_windows_from_df(
        pd.DataFrame(baseline_rows),
        classification_mode,
        aggregation_mode,
        baseline_date,
    )
    merged = exporters._merge_supply_dest_pbd_rolling_windows(
        current,
        baseline,
        classification_mode,
        aggregation_mode,
        baseline_available=True,
    )
    detail = merged[
        ~merged['supply_country'].isin(['Total', 'Internal'])
    ]

    a_row = detail.loc[detail['supply_country'] == 'A'].iloc[0]
    b_row = detail.loc[detail['supply_country'] == 'B'].iloc[0]
    c_row = detail.loc[detail['supply_country'] == 'C'].iloc[0]

    assert a_row['Δ 30D vs PBD'] == pytest.approx(2)
    assert a_row['Δ 7D vs PBD'] == pytest.approx(2)
    assert b_row['Δ 30D vs PBD'] == pytest.approx(5)
    assert b_row['Δ 7D vs PBD'] == pytest.approx(5)
    assert c_row['Δ 30D vs PBD'] == pytest.approx(-4)
    assert c_row['Δ 7D vs PBD'] == pytest.approx(-4)
    assert 'Internal' not in detail['supply_country'].tolist()


def test_outer_join_preserves_a_fully_removed_current_vintage():
    baseline = exporters._build_supply_dest_rolling_windows_from_df(
        pd.DataFrame(
            _constant_window('2026-07-29', 'Removed', 4)
        ),
        'Country',
        'Installation',
        '2026-07-29',
    )

    merged = exporters._merge_supply_dest_pbd_rolling_windows(
        pd.DataFrame(),
        baseline,
        'Country',
        'Installation',
        baseline_available=True,
    )

    row = merged.iloc[0]
    assert row['30D'] == 0
    assert row['7D'] == 0
    assert row['Δ 30D vs PBD'] == -4
    assert row['Δ 7D vs PBD'] == -4


def test_small_country_taxonomy_is_frozen_from_current_vintage():
    current = pd.DataFrame(
        _constant_window('2026-07-30', 'Small', 1)
        + _constant_window('2026-07-30', 'Large', 100)
    )
    baseline = pd.DataFrame(
        _constant_window('2026-07-29', 'Small', 50)
        + _constant_window('2026-07-29', 'Large', 100)
    )

    grouped_current, grouping_config = (
        exporters.group_small_supply_dest_countries(
            current,
            'Country',
            'Installation',
            as_of_date='2026-07-30',
            return_grouping_config=True,
        )
    )
    grouped_baseline = exporters.group_small_supply_dest_countries(
        baseline,
        'Country',
        'Installation',
        grouping_config=grouping_config,
    )

    assert set(grouped_current['supply_country']) == {
        'Large',
        'Rest of countries',
    }
    assert set(grouped_baseline['supply_country']) == {
        'Large',
        'Rest of countries',
    }


class _ConnectionContext:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False


class _Engine:
    def connect(self):
        return _ConnectionContext()


def test_pbd_query_binds_exact_snapshot_and_only_30_calendar_days(
    monkeypatch,
):
    captured = {}

    def fake_read_sql(query, connection, params):
        captured['query'] = str(query)
        captured['params'] = params
        return pd.DataFrame([{
            'supply_classification': 'Supply class',
            'supply_country': 'A',
            'supply_installation': 'A terminal',
            'demand_classification': 'Demand class',
            'demand_country': 'Destination',
            'flow_date': pd.Timestamp('2026-07-29'),
            'year': 2026,
            'quarter': 3,
            'month': 7,
            'week': 31,
            'volume': 1000,
        }])

    monkeypatch.setattr(exporters.pd, 'read_sql', fake_read_sql)
    result = exporters.fetch_supply_destination_pbd_base_data(
        _Engine(),
        'at_lng',
        '2026-07-29T05:36:57.998945',
        '2026-07-29',
    )

    assert not result.empty
    assert 'latest_data' not in captured['query']
    assert 'kt.start >= :window_start' in captured['query']
    assert 'kt.start < :end_exclusive' in captured['query']
    assert captured['params']['snapshot_timestamp_utc'] == pd.Timestamp(
        '2026-07-29T05:36:57.998945'
    )
    assert captured['params']['window_start'] == pd.Timestamp('2026-06-30')
    assert captured['params']['end_exclusive'] == pd.Timestamp('2026-07-30')


@pytest.mark.parametrize('volume_metric', ['mcm_d', 'mt', 'mtpa'])
def test_pbd_absolute_changes_convert_in_selected_unit(volume_metric):
    frame = pd.DataFrame([{
        'Supply Country': 'A',
        '30D': 136.0,
        '30D_PBD': 68.0,
        '7D': 136.0,
        '7D_PBD': 68.0,
        'Δ 30D vs PBD': 999.0,
        'Δ 7D vs PBD': 999.0,
    }])

    converted = exporters._convert_supply_dest_absolute_volume_metric(
        frame,
        volume_metric,
    )
    assert converted.loc[0, 'Δ 30D vs PBD'] == pytest.approx(
        round(
            converted.loc[0, '30D']
            - converted.loc[0, '30D_PBD'],
            1,
        )
    )
    assert converted.loc[0, 'Δ 7D vs PBD'] == pytest.approx(
        round(
            converted.loc[0, '7D']
            - converted.loc[0, '7D_PBD'],
            1,
        )
    )


def _summary_payload(*, status='exact', baseline_total=100):
    rows = [
        {
            'supply_country': 'A',
            'supply_installation': 'A terminal',
            '2024': 10.0,
            '2025': 20.0,
            '30D': 25.0,
            '30D_PP': 20.0,
            '30D_Y1': 15.0,
            '7D': 30.0,
            '7D_PP': 25.0,
            '7D_Y1': 20.0,
            'Δ 7D-30D': 5.0,
            'Δ 30D Y/Y': 10.0,
            '30D_PBD': baseline_total / 2,
            '7D_PBD': baseline_total / 10,
            'Δ 30D vs PBD': 25.0 - baseline_total / 2,
            'Δ 7D vs PBD': 30.0 - baseline_total / 10,
        },
        {
            'supply_country': 'B',
            'supply_installation': 'B terminal',
            '2024': 30.0,
            '2025': 40.0,
            '30D': 75.0,
            '30D_PP': 80.0,
            '30D_Y1': 85.0,
            '7D': 70.0,
            '7D_PP': 75.0,
            '7D_Y1': 80.0,
            'Δ 7D-30D': -5.0,
            'Δ 30D Y/Y': -10.0,
            '30D_PBD': baseline_total / 2,
            '7D_PBD': baseline_total * 0.9,
            'Δ 30D vs PBD': 75.0 - baseline_total / 2,
            'Δ 7D vs PBD': 70.0 - baseline_total * 0.9,
        },
    ]
    if status == 'unavailable':
        for row in rows:
            for column_name in (
                *exporters.SUPPLY_DEST_PBD_REFERENCE_COLUMNS,
                *exporters.SUPPLY_DEST_PBD_DELTA_COLUMNS,
            ):
                row[column_name] = np.nan
    current_snapshot = {
        'snapshot_id': 200,
        'snapshot_date_utc': '2026-07-30',
        'snapshot_timestamp_utc': '2026-07-30T05:34:12.778724',
        'facts_retained': True,
    }
    baseline_snapshot = (
        {
            'snapshot_id': 199,
            'snapshot_date_utc': '2026-07-29',
            'snapshot_timestamp_utc': '2026-07-29T05:36:57.998945',
            'facts_retained': True,
        }
        if status != 'unavailable'
        else None
    )
    return {
        'format': exporters.EXPORTERS_SUPPLY_DEST_SUMMARY_FORMAT,
        'show_all': rows,
        'group_small_countries': rows,
        'comparison': {
            'status': status,
            'current_snapshot': current_snapshot,
            'baseline_snapshot': baseline_snapshot,
            'business_day_gap': 1 if baseline_snapshot else None,
        },
    }


def _walk_components(component):
    if component is None:
        return
    if isinstance(component, (list, tuple)):
        for child in component:
            yield from _walk_components(child)
        return
    yield component
    yield from _walk_components(getattr(component, 'children', None))


def _first_grid(component):
    return next(
        item
        for item in _walk_components(component)
        if hasattr(item, 'rowData')
    )


def _column_fields(column_defs):
    fields = []
    for column in column_defs:
        children = column.get('children') or []
        if children:
            fields.extend(_column_fields(children))
            continue
        field = column.get('field') or column.get('id')
        if field:
            fields.append(field)
    return fields


def _component_text(component):
    values = []
    for item in _walk_components(component):
        if isinstance(item, str):
            values.append(item)
    return ' '.join(values)


def _render_table(
    payload,
    *,
    view_type='absolute',
    comparison_basis='levels',
    volume_metric='mcm_d',
):
    return exporters.update_supply_dest_table(
        payload,
        [],
        [],
        [],
        'Country',
        view_type,
        'Installation',
        comparison_basis,
        'show_all',
        volume_metric,
        2,
        0,
        0,
        0,
    )


@pytest.mark.parametrize(
    'comparison_basis',
    ['levels', 'previous_period', 'same_period_last_year'],
)
def test_pbd_columns_remain_last_and_unchanged_across_comparisons(
    comparison_basis,
):
    grid = _first_grid(
        _render_table(
            _summary_payload(),
            comparison_basis=comparison_basis,
        )
    )
    fields = _column_fields(grid.columnDefs)
    a_row = next(
        row
        for row in grid.rowData
        if 'A' in row.get('Supply Country', '')
    )

    assert fields[-2:] == list(exporters.SUPPLY_DEST_PBD_DELTA_COLUMNS)
    assert '30D_PBD' not in fields
    assert '7D_PBD' not in fields
    assert a_row['Δ 30D vs PBD'] == '-25'
    assert a_row['Δ 7D vs PBD'] == '+20'


def test_market_share_pbd_changes_are_percentage_points_and_zero_safe():
    grid = _first_grid(
        _render_table(
            _summary_payload(),
            view_type='percentage',
        )
    )
    rows_by_country = {
        row['Supply Country'].replace('▶', '').strip(): row
        for row in grid.rowData
    }

    assert rows_by_country['A']['Δ 30D vs PBD'] == '-25 pp'
    assert rows_by_country['A']['Δ 7D vs PBD'] == '+20 pp'
    assert rows_by_country['B']['Δ 30D vs PBD'] == '+25 pp'
    assert rows_by_country['B']['Δ 7D vs PBD'] == '-20 pp'

    zero_baseline_grid = _first_grid(
        _render_table(
            _summary_payload(baseline_total=0),
            view_type='percentage',
        )
    )
    zero_rows = {
        row['Supply Country'].replace('▶', '').strip(): row
        for row in zero_baseline_grid.rowData
    }
    assert zero_rows['A']['Δ 30D vs PBD'] == '+25 pp'
    assert zero_rows['B']['Δ 30D vs PBD'] == '+75 pp'


def test_unavailable_baseline_shows_dashes_and_explicit_warning():
    table = _render_table(_summary_payload(status='unavailable'))
    grid = _first_grid(table)
    text = _component_text(table)

    assert all(
        row['Δ 30D vs PBD'] == '—'
        and row['Δ 7D vs PBD'] == '—'
        for row in grid.rowData
    )
    assert 'PBD baseline unavailable' in text


def test_snapshot_pair_lineage_displays_exact_timestamps():
    text = _component_text(_render_table(_summary_payload()))

    assert 'Jul 30, 2026 05:34:12.778724 UTC' in text
    assert 'Jul 29, 2026 05:36:57.998945 UTC' in text
    assert 'window roll plus Kpler revisions' in text


def test_excel_uses_rendered_order_and_pbd_values():
    grid = _first_grid(_render_table(_summary_payload()))
    download = exporters.export_supply_dest_table_to_excel(
        1,
        [grid.rowData],
        [grid.rowData],
        [grid.columnDefs],
    )
    workbook = load_workbook(
        BytesIO(base64.b64decode(download['content']))
    )
    worksheet = workbook['Supply by Destination']
    rows = list(worksheet.iter_rows(values_only=True))
    headers = list(rows[0])

    assert headers[-2:] == list(exporters.SUPPLY_DEST_PBD_DELTA_COLUMNS)
    assert rows[1][-2:] == (
        grid.rowData[0]['Δ 30D vs PBD'],
        grid.rowData[0]['Δ 7D vs PBD'],
    )


def test_snapshot_cache_keys_are_versioned_and_include_exact_pair():
    current_reference = {'namespace': 'current', 'key': 'current-key'}
    baseline_reference = {'namespace': 'baseline', 'key': 'baseline-key'}
    source_state = exporters._build_exporters_source_state(
        _snapshot_pair('2026-07-30', '2026-07-29')
    )

    first_key = exporters._exporters_destination_summary_source_key(
        current_reference,
        'Country',
        'Installation',
        baseline_reference,
        source_state,
    )
    changed_state = exporters._build_exporters_source_state(
        _snapshot_pair(
            '2026-07-30',
            '2026-07-29',
            baseline_timestamp='2026-07-29T06:00:00',
        )
    )
    changed_key = exporters._exporters_destination_summary_source_key(
        current_reference,
        'Country',
        'Installation',
        baseline_reference,
        changed_state,
    )

    assert exporters.EXPORTERS_DESTINATION_SUMMARY_NAMESPACE.endswith('-v2')
    assert exporters.EXPORTERS_SUPPLY_DEST_SUMMARY_FORMAT.endswith('-v2')
    assert first_key != changed_key

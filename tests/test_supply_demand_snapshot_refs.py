import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import threading
import time

from dash import html
from dash._utils import to_json
from openpyxl import load_workbook
import pandas as pd
import pytest

from pages import demand, supply
from utils import dashboard_snapshot_cache as snapshots
from utils import provider_flow_snapshot


PAGES = (
    (
        "supply",
        supply,
        "woodmac_export",
        "ea_export",
        supply.SUPPLY_SNAPSHOT_RECOVERY_MESSAGE,
    ),
    (
        "demand",
        demand,
        "woodmac_import",
        "ea_import",
        demand.DEMAND_SNAPSHOT_RECOVERY_MESSAGE,
    ),
)


def _flow_frame(multiplier=1.0):
    return pd.DataFrame(
        {
            "month": pd.to_datetime(
                [
                    "2025-01-01",
                    "2025-02-01",
                    "2025-01-01",
                    "2025-02-01",
                ]
            ),
            "country_name": [
                "Qatar",
                "Qatar",
                "United States",
                "United States",
            ],
            "total_mmtpa": [
                10.25 * multiplier,
                11.5 * multiplier,
                8.75 * multiplier,
                9.5 * multiplier,
            ],
        }
    )


def _mapping_frame():
    return pd.DataFrame(
        [
            {
                "country": "Qatar",
                "country_name": "Qatar",
                "continent": "Asia",
                "subcontinent": "Middle East",
                "basin": "Pacific",
                "country_classification_level1": "Middle East",
                "country_classification": "Producer",
                "shipping_region": "Arabian Gulf",
            },
            {
                "country": "USA",
                "country_name": "United States",
                "continent": "North America",
                "subcontinent": "North America",
                "basin": "Atlantic",
                "country_classification_level1": "North America",
                "country_classification": "Producer",
                "shipping_region": "US Gulf",
            },
        ]
    )


def _provider_payload():
    return {
        "woodmac_export": _flow_frame(1.0),
        "ea_export": _flow_frame(1.1),
        "woodmac_import": _flow_frame(0.9),
        "ea_import": _flow_frame(0.8),
        "mapping": _mapping_frame(),
        "woodmac_export_options": {
            "short_term": [],
            "long_term": [],
        },
        "woodmac_import_options": {
            "short_term": [],
            "long_term": [],
        },
        "ea_comparison_runs": [],
        "ea_export_options": [],
        "ea_import_options": [],
        "current_ea": {
            "run_id": 154,
            "snapshot_at": "2026-07-24T14:36:47.113486Z",
            "change_count": 0,
            "delete_count": 0,
        },
        "errors": {},
    }


@pytest.fixture
def persistent_provider_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "provider-flow-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _install_snapshot_getter(
    monkeypatch,
    page_module,
    payload,
    *,
    source_key,
    build_counter=None,
    build_delay=0.0,
    force_calls=None,
):
    def builder():
        if build_counter is not None:
            with build_counter["lock"]:
                build_counter["count"] += 1
        if build_delay:
            time.sleep(build_delay)
        return copy.deepcopy(payload)

    def getter(*, force=False):
        if force_calls is not None:
            force_calls.append(force)
        return snapshots.get_or_build_snapshot(
            page_module.engine,
            namespace=provider_flow_snapshot.NAMESPACE,
            source_key=source_key,
            builder=builder,
            force=force,
            manifest={"source_state": "stable"},
        )

    monkeypatch.setattr(
        page_module,
        "_get_provider_flow_snapshot",
        getter,
    )
    return getter


def _load_page_references(
    monkeypatch,
    page_name,
    page_module,
    payload,
    *,
    source_key_suffix="representative",
):
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    _install_snapshot_getter(
        monkeypatch,
        page_module,
        payload,
        source_key=f"{page_name}-{source_key_suffix}",
    )
    return page_module.load_balance_source_data(0)


def _workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: [
            [
                (
                    cell.value,
                    cell.data_type,
                    cell.number_format,
                )
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]
        for worksheet in workbook.worksheets
    }


def _delete_reference(reference, corruption_mode):
    stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if corruption_mode == "missing":
        stores.cache.delete(record_key, retry=True)
    else:
        stores.cache.set(record_key, b"corrupt", retry=True)
    snapshots.clear_local_snapshots()


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_loader_emits_ordered_small_refs_and_survives_restart(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    result = _load_page_references(
        monkeypatch,
        page_name,
        page_module,
        payload,
    )
    woodmac_ref, ea_ref = result[:2]

    assert woodmac_ref["slot"] == woodmac_slot
    assert ea_ref["slot"] == ea_slot
    for reference in (woodmac_ref, ea_ref):
        assert snapshots.is_snapshot_reference(
            reference,
            provider_flow_snapshot.NAMESPACE,
        )
        assert snapshots.snapshot_is_resolvable(reference)
        assert len(to_json(reference).encode("utf-8")) < 10_000
    assert len(to_json(result).encode("utf-8")) < 50_000
    assert result[2] == page_module.get_available_countries(
        [payload[woodmac_slot], payload[ea_slot]]
    )
    assert result[4] is None

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    pd.testing.assert_frame_equal(
        page_module._deserialize_dataframe(woodmac_ref),
        payload[woodmac_slot],
        check_dtype=True,
        check_exact=True,
    )
    pd.testing.assert_frame_equal(
        page_module._deserialize_dataframe(ea_ref),
        payload[ea_slot],
        check_dtype=True,
        check_exact=True,
    )


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_reference_controls_tables_deltas_and_two_exports_match_legacy(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    result = _load_page_references(
        monkeypatch,
        page_name,
        page_module,
        payload,
        source_key_suffix="parity",
    )
    woodmac_ref, ea_ref = result[:2]
    raw_woodmac = page_module._serialize_dataframe(
        payload[woodmac_slot]
    )
    raw_ea = page_module._serialize_dataframe(payload[ea_slot])
    countries = result[2]
    lookup = result[3]
    woodmac_metadata = result[5]
    ea_metadata = result[6]
    selected = ["Qatar"]
    start_date = "2025-01-01"
    end_date = "2025-02-28"
    monkeypatch.setattr(
        page_module,
        "_fetch_comparison_raw_df",
        lambda *_args, **_kwargs: (
            payload[woodmac_slot].copy(),
            None,
        ),
    )

    callback_pairs = [
        (
            page_module.update_balance_country_options(
                countries,
                "continent",
                lookup,
                raw_woodmac,
                raw_ea,
                start_date,
                end_date,
                selected,
                selected,
            ),
            page_module.update_balance_country_options(
                countries,
                "continent",
                lookup,
                woodmac_ref,
                ea_ref,
                start_date,
                end_date,
                selected,
                selected,
            ),
        ),
        (
            page_module.update_balance_date_range(
                raw_woodmac,
                raw_ea,
                start_date,
                end_date,
            ),
            page_module.update_balance_date_range(
                woodmac_ref,
                ea_ref,
                start_date,
                end_date,
            ),
        ),
        (
            page_module.render_balance_tables(
                raw_woodmac,
                raw_ea,
                woodmac_metadata,
                ea_metadata,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
            page_module.render_balance_tables(
                woodmac_ref,
                ea_ref,
                woodmac_metadata,
                ea_metadata,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
        ),
        (
            page_module.render_comparison_delta_table(
                raw_woodmac,
                raw_ea,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
                "ea",
                None,
                None,
                153,
            ),
            page_module.render_comparison_delta_table(
                woodmac_ref,
                ea_ref,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
                "ea",
                None,
                None,
                153,
            ),
        ),
        (
            page_module.render_ea_comparison_delta_table(
                raw_woodmac,
                raw_ea,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
                "woodmac",
                None,
                None,
                None,
            ),
            page_module.render_ea_comparison_delta_table(
                woodmac_ref,
                ea_ref,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
                "woodmac",
                None,
                None,
                None,
            ),
        ),
    ]
    for raw_result, reference_result in callback_pairs:
        assert to_json(raw_result) == to_json(reference_result)

    export_pairs = [
        (
            page_module.export_woodmac_balance_excel(
                1,
                raw_woodmac,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
            page_module.export_woodmac_balance_excel(
                1,
                woodmac_ref,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
        ),
        (
            page_module.export_ea_balance_excel(
                1,
                raw_ea,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
            page_module.export_ea_balance_excel(
                1,
                ea_ref,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
        ),
    ]
    for raw_download, reference_download in export_pairs:
        assert _workbook_cells(raw_download) == _workbook_cells(
            reference_download
        )


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
@pytest.mark.parametrize("worker_count", (1, 2, 4))
def test_page_provider_loader_single_flight_at_one_two_and_four_callers(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    worker_count,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    build_counter = {
        "count": 0,
        "lock": threading.Lock(),
    }
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    _install_snapshot_getter(
        monkeypatch,
        page_module,
        payload,
        source_key=f"{page_name}-single-flight-{worker_count}",
        build_counter=build_counter,
        build_delay=0.04,
    )

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        results = list(
            executor.map(
                lambda _index: page_module.load_balance_source_data(0),
                range(worker_count),
            )
        )

    assert build_counter["count"] == 1
    assert all(result == results[0] for result in results)
    assert snapshots.snapshot_is_resolvable(results[0][0])
    assert snapshots.snapshot_is_resolvable(results[0][1])
    assert len(to_json(results[0]).encode("utf-8")) < 50_000


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_loader_never_uses_non_resolvable_or_legacy_fallback(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
):
    payload = _provider_payload()
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": provider_flow_snapshot.NAMESPACE,
        "source_key": f"{page_name}-non-resolvable",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        page_module,
        "_get_provider_flow_snapshot",
        lambda **_kwargs: (non_resolvable, payload),
    )
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        page_module.load_balance_source_data(0)


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
@pytest.mark.parametrize("corruption_mode", ("missing", "corrupt"))
def test_missing_or_corrupt_page_ref_has_explicit_recovery_and_blocks_exports(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    corruption_mode,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    result = _load_page_references(
        monkeypatch,
        page_name,
        page_module,
        payload,
        source_key_suffix=f"recovery-{corruption_mode}",
    )
    woodmac_ref, ea_ref = result[:2]
    countries = result[2]
    lookup = result[3]
    _delete_reference(woodmac_ref, corruption_mode)

    rendered = page_module.render_balance_tables(
        woodmac_ref,
        ea_ref,
        result[5],
        result[6],
        ["Qatar"],
        "rest_of_world",
        "2025-01-01",
        "2025-02-28",
        "monthly",
        "country",
        lookup,
    )
    assert isinstance(rendered[0], html.Div)
    assert rendered[0].children == recovery_message
    assert isinstance(rendered[2], html.Div)
    assert rendered[2].children == recovery_message

    selector = page_module.update_balance_country_options(
        countries,
        "continent",
        lookup,
        woodmac_ref,
        ea_ref,
        "2025-01-01",
        "2025-02-28",
        ["Qatar"],
        ["Qatar"],
    )
    assert selector[0][0]["label"] == recovery_message
    assert selector[1] == []
    assert selector[2] is True

    export_calls = [
        lambda: page_module.export_woodmac_balance_excel(
            1,
            woodmac_ref,
            ["Qatar"],
            "rest_of_world",
            "2025-01-01",
            "2025-02-28",
            "monthly",
            "country",
            lookup,
        ),
        lambda: page_module.export_ea_balance_excel(
            1,
            ea_ref,
            ["Qatar"],
            "rest_of_world",
            "2025-01-01",
            "2025-02-28",
            "monthly",
            "country",
            lookup,
        ),
    ]
    for export_call in export_calls:
        with pytest.raises(
            snapshots.SnapshotUnavailable,
            match="Click the global Refresh",
        ):
            export_call()


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_global_refresh_is_checked_once_and_forces_one_rebuild(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    build_counter = {
        "count": 0,
        "lock": threading.Lock(),
    }
    force_calls = []
    _install_snapshot_getter(
        monkeypatch,
        page_module,
        payload,
        source_key=f"{page_name}-global-refresh",
        build_counter=build_counter,
        force_calls=force_calls,
    )
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    old_result = page_module.load_balance_source_data(0)
    refresh_checks = 0

    def refresh_triggered():
        nonlocal refresh_checks
        refresh_checks += 1
        return True

    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        refresh_triggered,
    )
    refreshed_result = page_module.load_balance_source_data(1)

    assert refresh_checks == 1
    assert force_calls == [False, True]
    assert build_counter["count"] == 2
    assert refreshed_result[0] != old_result[0]
    assert refreshed_result[1] != old_result[1]


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_provider_failure_is_fail_closed(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
):
    monkeypatch.setattr(
        page_module,
        "_get_provider_flow_snapshot",
        lambda **_kwargs: (_ for _ in ()).throw(
            RuntimeError("source state changed")
        ),
    )
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        page_module.load_balance_source_data(0)

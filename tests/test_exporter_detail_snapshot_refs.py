import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import threading
import time

from dash._utils import to_json
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import pytest

from pages import exporter_detail
from utils import dashboard_snapshot_cache as snapshots


def _destination_frame(country="United States"):
    return pd.DataFrame(
        {
            "start_date": pd.to_datetime(
                [
                    "2025-01-02",
                    "2025-02-03",
                    "2026-01-04",
                    "2026-02-05",
                ]
            ),
            "cargo_mcm": np.array([60.25, np.nan, 72.5, 81.75]),
            "destination_country": [
                "Japan",
                "Spain",
                country,
                "United Kingdom",
            ],
            "destination_continent": [
                "Asia",
                "Europe",
                "North America",
                "Europe",
            ],
            "destination_shipping_region": [
                "North Asia",
                "Iberia",
                "US Gulf",
                "Northwest Europe",
            ],
            "destination_basin": [
                "Pacific",
                "Atlantic",
                "Atlantic",
                "Atlantic",
            ],
            "destination_subcontinent": [
                "East Asia",
                "Southern Europe",
                "North America",
                "Western Europe",
            ],
            "destination_classification_level1": [
                "Asia",
                "Europe",
                "North America",
                None,
            ],
            "destination_classification": [
                "Importer",
                "Importer",
                "Producer",
                "Importer",
            ],
        }
    )


def _origin_plant_frame():
    frame = _destination_frame().copy()
    frame.insert(2, "origin_zone", ["Gulf", "Gulf", "Atlantic", "Atlantic"])
    frame.insert(
        3,
        "origin_plant",
        ["Sabine Pass", "Corpus Christi", "Cove Point", "Cove Point"],
    )
    return frame


def _forecast_frame():
    dates = pd.to_datetime(["2026-08-01", "2026-08-02"])
    return pd.DataFrame(
        {
            "date": dates,
            "year": dates.year,
            "day_of_year": dates.dayofyear,
            "month_day": dates.strftime("%b %d"),
            "mcmd": [434.85634408602147, 433.48172043010754],
            "woodmac_mtpa": [116.75, 116.38],
            "is_forecast": [True, True],
            "source": ["Short Term", "Short Term"],
        }
    )


def _allocation_source(country="United States"):
    run_metadata = {
        "run_id": "allocation-run-exporter",
        "analysis_date": pd.Timestamp("2026-07-20 12:00:00"),
        "forecast_start": pd.Timestamp("2026-07-01"),
        "forecast_end": pd.Timestamp("2028-12-01"),
        "supply_scenario": "base_view",
        "split_by_contract": True,
        "woodmac_short_term_outlook": "July 2026",
        "woodmac_long_term_outlook": "June 2026",
    }
    mapping_df = pd.DataFrame(
        {
            "country": ["Spain", "Japan"],
            "country_name": ["Spain", "Japan"],
            "continent": ["Europe", "Asia"],
            "basin": ["Atlantic", "Pacific"],
            "shipping_region": ["Iberia", "North Asia"],
            "subcontinent": ["Southern Europe", "East Asia"],
            "country_classification_level1": ["Europe", "Asia"],
            "country_classification": ["Importer", "Importer"],
        }
    )
    allocation_df = pd.DataFrame(
        {
            "date": pd.to_datetime(
                [
                    "2026-07-01",
                    "2026-08-01",
                    "2027-01-01",
                    "2028-01-01",
                ]
            ),
            "destination_country": ["Spain", "Japan", "Spain", "Japan"],
            "allocated_volume_bcm": [1.25, 1.5, 3.0, 3.5],
            "alias": ["Spain", "Japan", "Spain", "Japan"],
            "country_display": ["Spain", "Japan", "Spain", "Japan"],
            "continent": ["Europe", "Asia", "Europe", "Asia"],
            "basin": ["Atlantic", "Pacific", "Atlantic", "Pacific"],
            "shipping_region": [
                "Iberia",
                "North Asia",
                "Iberia",
                "North Asia",
            ],
            "subcontinent": [
                "Southern Europe",
                "East Asia",
                "Southern Europe",
                "East Asia",
            ],
            "country_classification_level1": [
                "Europe",
                "Asia",
                "Europe",
                "Asia",
            ],
            "country_classification": [
                "Importer",
                "Importer",
                "Importer",
                "Importer",
            ],
            "country": ["Spain", "Japan", "Spain", "Japan"],
        }
    )
    return {
        "origin_country": country,
        "run_metadata": run_metadata,
        "mapping_df": mapping_df,
        "allocation_df": allocation_df,
    }


def _base_payload(country="United States"):
    origin_scope = {
        "country": country,
        "continent": "North America",
        "shipping_region": "US Gulf",
        "basin": "Atlantic",
        "subcontinent": None,
        "country_classification_level1": "North America",
        "country_classification": "Producer",
        "continent_conflict": False,
    }
    return {
        "origin_country": country,
        "normalized_destination_trades": exporter_detail._store_dataframe(
            _destination_frame(country)
        ),
        "origin_plant_destination_trades": exporter_detail._store_dataframe(
            _origin_plant_frame()
        ),
        "woodmac_export_forecast": exporter_detail._store_dataframe(
            pd.DataFrame(
                columns=[
                    "date",
                    "year",
                    "day_of_year",
                    "month_day",
                    "mcmd",
                    "woodmac_mtpa",
                    "is_forecast",
                    "source",
                ]
            )
        ),
        "origin_scope": origin_scope,
        "available_years": [2025, 2026],
        "loaded_at": "2026-07-24T12:00:00",
    }


@pytest.fixture
def persistent_exporter_detail_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "exporter-detail-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(snapshots.LOCAL_CACHE_DIR_ENV, str(cache_directory))
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    monkeypatch.setattr(
        exporter_detail,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: "2026-07-24T00:00:00Z",
    )
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_reference(monkeypatch, country="United States", payload=None):
    payload = payload or _base_payload(country)
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda selected_country: copy.deepcopy(payload),
    )
    reference = exporter_detail.refresh_exporter_detail_base_data(country, 0)
    return payload, reference


def _delete_or_corrupt_reference(reference, mode):
    stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if mode == "missing":
        stores.cache.delete(record_key, retry=True)
    else:
        stores.cache.set(record_key, b"corrupt", retry=True)
    snapshots.clear_local_snapshots()


def _workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: [
            [
                (cell.value, cell.data_type, cell.number_format)
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]
        for worksheet in workbook.worksheets
    }


def _without_loaded_at(payload):
    return {
        key: value
        for key, value in payload.items()
        if key != "loaded_at"
    }


@pytest.mark.parametrize(
    "country",
    ["United States", "Equatorial Guinea"],
)
def test_base_loader_emits_small_resolvable_reference_and_preserves_payload(
    country,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    payload, reference = _load_reference(monkeypatch, country)

    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_DETAIL_BASE_NAMESPACE,
    )
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert len(to_json(reference).encode("utf-8")) < 50_000

    legacy_frames = exporter_detail._resolve_exporter_detail_base_data(payload)
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    restarted_frames = exporter_detail._resolve_exporter_detail_base_data(
        reference
    )

    pd.testing.assert_frame_equal(
        restarted_frames[0],
        legacy_frames[0],
        check_dtype=True,
        check_exact=True,
    )
    assert list(restarted_frames[0].columns) == list(legacy_frames[0].columns)
    assert restarted_frames[0].isna().equals(legacy_frames[0].isna())
    assert restarted_frames[1] == legacy_frames[1]
    pd.testing.assert_frame_equal(
        restarted_frames[2],
        legacy_frames[2],
        check_dtype=True,
        check_exact=True,
    )
    assert restarted_frames[3] == country


def test_allocation_reference_is_small_exact_and_control_paths_are_sql_free(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    source_data = _allocation_source()
    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_data["run_metadata"]),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        lambda *_args, **_kwargs: copy.deepcopy(source_data),
    )

    reference = exporter_detail.refresh_destination_forecast_source(
        "United States",
        0,
    )
    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_ALLOCATION_SOURCE_NAMESPACE,
    )
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    resolved = exporter_detail._resolve_destination_forecast_source_data(
        reference
    )
    pd.testing.assert_frame_equal(
        resolved["mapping_df"],
        source_data["mapping_df"],
        check_exact=True,
    )
    pd.testing.assert_frame_equal(
        resolved["allocation_df"],
        source_data["allocation_df"],
        check_exact=True,
    )

    monkeypatch.setattr(
        exporter_detail,
        "fetch_destination_forecast_summary_data",
        lambda *_args, **_kwargs: pytest.fail(
            "allocation controls queried PostgreSQL"
        ),
    )
    levels = [
        "destination_country_name",
        "destination_shipping_region",
        "destination_basin",
        "destination_subcontinent",
        "destination_classification_level1",
        "destination_classification",
    ]
    for level in levels:
        direct = exporter_detail.update_destination_forecast_summary_table(
            "United States",
            ["Europe"],
            level,
            "mcm_d",
            source_data,
        )
        cached = exporter_detail.update_destination_forecast_summary_table(
            "United States",
            ["Europe"],
            level,
            "mcm_d",
            reference,
        )
        assert to_json(cached) == to_json(direct)


def test_allocation_no_run_uses_one_metadata_read_and_consistent_identity(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    metadata_reads = []

    def no_run(_engine):
        metadata_reads.append(True)
        if len(metadata_reads) > 1:
            raise AssertionError("allocation metadata was queried twice")
        return None

    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        no_run,
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        lambda *_args, **_kwargs: pytest.fail(
            "no-run refresh invoked the source loader"
        ),
    )

    reference = exporter_detail.refresh_destination_forecast_source(
        "United States",
        0,
    )
    current_month = pd.Timestamp.today().replace(day=1).date().isoformat()
    expected_key = snapshots.build_source_key(
        exporter_detail.EXPORTER_ALLOCATION_SOURCE_NAMESPACE,
        None,
        None,
        "United States",
        current_month,
    )
    assert metadata_reads == [True]
    assert reference["source_key"] == expected_key
    resolved = exporter_detail._resolve_destination_forecast_source_data(
        reference
    )
    assert resolved["run_metadata"] is None
    assert resolved["allocation_df"].empty
    manifest = snapshots._LOCAL_MANIFESTS[
        (
            reference["namespace"],
            reference["source_key"],
            reference["revision"],
        )
    ]
    assert manifest["run_id"] is None
    output = exporter_detail.update_destination_forecast_summary_table(
        "United States",
        [],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        reference,
    )
    assert "No compatible WoodMac supply-allocation SQL run" in to_json(
        output
    )


def test_route_reference_is_small_exact_and_controls_and_export_are_sql_free(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    _, base_reference = _load_reference(monkeypatch)
    base_reference["route_source_version"] = {
        "kpler_watermark": "kpler-v1",
        "distance_watermark": "distance-v1",
    }
    processed_df = pd.DataFrame(
        {
            "year": [2025, 2025, 2026, 2026],
            "month": [1, 4, 7, 10],
            "season": ["W", "S", "S", "W"],
            "quarter": ["Q1", "Q2", "Q3", "Q4"],
            "voyage_id": ["v1", "v2", "v3", "v4"],
            "destination_country_name": [
                "Spain",
                "Japan",
                "France",
                "Brazil",
            ],
            "distanceDirect": [100.0, 100.0, 100.0, 100.0],
            "distanceViaSuez": [120.0, np.nan, 120.0, np.nan],
            "distanceViaPanama": [np.nan, 130.0, 130.0, np.nan],
            "selected_route": [
                "Direct",
                "ViaPanama",
                "ViaSuez",
                "Direct",
            ],
        }
    )
    mapping_df = pd.DataFrame(
        {
            "country": ["Spain", "Japan", "France", "Brazil"],
            "country_name": ["Spain", "Japan", "France", "Brazil"],
            "continent": ["Europe", "Asia", "Europe", "South America"],
            "basin": ["Atlantic", "Pacific", "Atlantic", "Atlantic"],
            "shipping_region": [
                "Iberia",
                "North Asia",
                "Northwest Europe",
                "South America",
            ],
            "subcontinent": [
                "Southern Europe",
                "East Asia",
                "Western Europe",
                "South America",
            ],
            "country_classification_level1": [
                "Europe",
                "Asia",
                "Europe",
                "South America",
            ],
            "country_classification": [
                "Importer",
                "Importer",
                "Importer",
                "Importer",
            ],
        }
    )
    source_data = {
        "origin_country": "United States",
        "processed_df": processed_df,
        "mapping_df": mapping_df,
    }
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_route_source_payload",
        lambda _country: copy.deepcopy(source_data),
    )
    reference = exporter_detail.refresh_exporter_route_analysis_source(
        "United States",
        base_reference,
    )
    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_ROUTE_SOURCE_NAMESPACE,
    )
    assert len(to_json(reference).encode("utf-8")) < 10_000

    monkeypatch.setattr(
        exporter_detail,
        "process_trade_and_distance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "route controls or export queried PostgreSQL"
        ),
    )
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pytest.fail(
            "route controls queried mappings"
        ),
    )
    direct = exporter_detail.update_route_analysis_charts_and_tables(
        "Year",
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "United States",
        source_data,
    )
    cached = exporter_detail.update_route_analysis_charts_and_tables(
        "Year",
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "United States",
        reference,
    )
    assert to_json(cached) == to_json(direct)

    direct_export = exporter_detail.export_route_analysis_to_excel(
        1,
        "Year",
        "destination_country_name",
        "United States",
        source_data,
    )
    cached_export = exporter_detail.export_route_analysis_to_excel(
        1,
        "Year",
        "destination_country_name",
        "United States",
        reference,
    )
    assert _workbook_cells(cached_export) == _workbook_cells(
        direct_export
    )


def test_diversion_reference_is_small_exact_and_sql_free(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    source_data = {
        "origin_country": "United States",
        "main_data": [{
            "Diversion date": "2026-06-01",
            "Vessel": "Test Vessel",
            "State": "Loaded",
            "Charterer": "Test Charterer",
            "Cubic Meters": 174500.0,
            "Origin location": "Sabine Pass",
            "Origin country": "United States",
            "Origin date": "2026-05-20",
            "Diverted from location": "Gate",
            "Diverted from country": "Netherlands",
            "Diverted from date": "2026-06-10",
            "New destination location": "Futtsu",
            "New destination country": "Japan",
            "New destination date": "2026-06-20",
            "Added shipping days": 10,
        }],
        "charts_data": [{
            "Diversion_month": "2026-06-01 00:00:00",
            "basin_combo": "Atlantic -> Pacific",
            "region_combo": "Northwest Europe -> North Asia",
            "country_combo": "Netherlands -> Japan",
            "Added shipping days": 10,
            "Cubic Meters": 174500.0,
        }],
    }
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_diversion_source_version",
        lambda: "diversion-v1",
    )
    captured_versions = []

    def build_diversion(*_args, **kwargs):
        captured_versions.append(kwargs.get("source_version"))
        return copy.deepcopy(source_data)

    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_diversion_payload",
        build_diversion,
    )
    reference = exporter_detail.refresh_exporter_diversion_source(
        0,
        "United States",
    )
    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_DIVERSION_SOURCE_NAMESPACE,
    )
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert captured_versions == ["diversion-v1"]
    assert to_json(exporter_detail.update_diversion_ui(None, "basin_combo")) == to_json(
        exporter_detail.update_diversion_ui(
            {"main_data": [], "charts_data": []},
            "basin_combo",
        )
    )

    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pytest.fail(
            "diversion controls and export must not query SQL"
        ),
    )
    direct_ui = exporter_detail.update_diversion_ui(
        source_data,
        "region_combo",
    )
    cached_ui = exporter_detail.update_diversion_ui(
        reference,
        "region_combo",
    )
    assert to_json(cached_ui) == to_json(direct_ui)
    direct_export = exporter_detail.export_diversion_summary_to_excel(
        1,
        source_data,
        direct_ui[1],
    )
    cached_export = exporter_detail.export_diversion_summary_to_excel(
        1,
        reference,
        cached_ui[1],
    )
    assert _workbook_cells(cached_export) == _workbook_cells(
        direct_export
    )

    corrupt_reference = dict(reference, revision="missing")
    corrupt_ui = exporter_detail.update_diversion_ui(
        corrupt_reference,
        "basin_combo",
    )
    assert exporter_detail.EXPORTER_DIVERSION_RECOVERY_MESSAGE in to_json(
        corrupt_ui
    )
    with pytest.raises(exporter_detail.PreventUpdate):
        exporter_detail.export_diversion_summary_to_excel(
            1,
            corrupt_reference,
            cached_ui[1],
        )


@pytest.mark.parametrize("workers", [1, 4, 8])
def test_diversion_source_is_single_flight_under_concurrency(
    workers,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    counter = {"count": 0}
    counter_lock = threading.Lock()

    def build_payload(*_args, **_kwargs):
        with counter_lock:
            counter["count"] += 1
        time.sleep(0.04)
        return {
            "origin_country": "United States",
            "main_data": [],
            "charts_data": [],
        }

    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_diversion_source_version",
        lambda: f"diversion-single-flight-{workers}",
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_diversion_payload",
        build_payload,
    )
    with ThreadPoolExecutor(max_workers=workers) as pool:
        references = list(
            pool.map(
                lambda _index: (
                    exporter_detail.refresh_exporter_diversion_source(
                        0,
                        "United States",
                    )
                ),
                range(workers),
            )
        )

    assert counter["count"] == 1
    assert all(
        snapshots.snapshot_is_resolvable(reference)
        for reference in references
    )
    assert len({
        (reference["source_key"], reference["revision"])
        for reference in references
    }) == 1


@pytest.mark.parametrize("mode", ["missing", "corrupt"])
def test_allocation_reference_failure_is_visible_without_sql_fallback(
    mode,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    source_data = _allocation_source()
    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(source_data["run_metadata"]),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        lambda *_args, **_kwargs: copy.deepcopy(source_data),
    )
    reference = exporter_detail.refresh_destination_forecast_source(
        "United States",
        0,
    )
    _delete_or_corrupt_reference(reference, mode)
    monkeypatch.setattr(
        exporter_detail,
        "fetch_destination_forecast_summary_data",
        lambda *_args, **_kwargs: pytest.fail(
            "corrupt allocation reference queried PostgreSQL"
        ),
    )

    output = exporter_detail.update_destination_forecast_summary_table(
        "United States",
        [],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        reference,
    )

    assert (
        exporter_detail.EXPORTER_ALLOCATION_RECOVERY_MESSAGE
        in to_json(output)
    )


def test_nonempty_forecast_round_trips_exactly_and_chart_does_not_query_sql(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    payload = _base_payload()
    payload["woodmac_export_forecast"] = _forecast_frame()
    _, reference = _load_reference(monkeypatch, payload=payload)

    restarted = exporter_detail._load_exporter_detail_forecast_data(reference)
    pd.testing.assert_frame_equal(
        restarted,
        _forecast_frame(),
        check_dtype=True,
        check_exact=True,
    )

    monkeypatch.setattr(
        exporter_detail,
        "fetch_woodmac_country_export_forecast_data",
        lambda _country: pytest.fail("chart callback queried WoodMac"),
    )
    exporter_detail.update_exporter_detail_supply_charts(
        reference,
        30,
        "mcm_d",
        ["2025", "2026"],
    )


def test_forecast_month_is_part_of_source_key(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    payload = _base_payload()
    payload["woodmac_export_forecast"] = _forecast_frame()
    build_calls = []
    month = {"value": "2026-07-01"}
    monkeypatch.setattr(
        exporter_detail,
        "_exporter_detail_forecast_month_token",
        lambda: month["value"],
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda country: build_calls.append(country) or copy.deepcopy(payload),
    )

    first = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        0,
    )
    month["value"] = "2026-08-01"
    second = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        0,
    )

    assert first["source_key"] != second["source_key"]
    assert build_calls == ["United States", "United States"]


def test_reference_matches_legacy_for_selectors_charts_summaries_tables_and_supply_export(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    payload, reference = _load_reference(monkeypatch)
    monkeypatch.setattr(
        exporter_detail,
        "fetch_woodmac_country_export_forecast_data",
        lambda _country: pd.DataFrame(),
    )

    legacy_selectors = (
        exporter_detail.update_exporter_detail_supply_year_selector(
            payload,
            ["2025"],
        ),
        exporter_detail.update_exporter_detail_continent_year_selector(
            payload,
            ["2025"],
        ),
    )
    reference_selectors = (
        exporter_detail.update_exporter_detail_supply_year_selector(
            reference,
            ["2025"],
        ),
        exporter_detail.update_exporter_detail_continent_year_selector(
            reference,
            ["2025"],
        ),
    )
    assert to_json(reference_selectors) == to_json(legacy_selectors)

    legacy_charts = exporter_detail.update_exporter_detail_supply_charts(
        payload,
        30,
        "mcm_d",
        ["2025", "2026"],
    )
    reference_charts = exporter_detail.update_exporter_detail_supply_charts(
        reference,
        30,
        "mcm_d",
        ["2025", "2026"],
    )
    assert to_json(reference_charts) == to_json(legacy_charts)

    legacy_summaries = (
        exporter_detail.refresh_exporter_detail_summary_data_stores(
            "United States",
            30,
            exporter_detail.DEFAULT_DESTINATION_LEVEL,
            payload,
        )
    )
    reference_summaries = (
        exporter_detail.refresh_exporter_detail_summary_data_stores(
            "United States",
            30,
            exporter_detail.DEFAULT_DESTINATION_LEVEL,
            reference,
        )
    )
    assert _without_loaded_at(reference_summaries[0]) == _without_loaded_at(
        legacy_summaries[0]
    )
    assert _without_loaded_at(reference_summaries[1]) == _without_loaded_at(
        legacy_summaries[1]
    )

    destination_legacy = exporter_detail.update_destination_summary_table(
        "United States",
        30,
        ["Europe"],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        "previous",
        5,
        3,
        3,
        legacy_summaries[0],
    )
    destination_reference = exporter_detail.update_destination_summary_table(
        "United States",
        30,
        ["Europe"],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        "previous",
        5,
        3,
        3,
        reference_summaries[0],
    )
    assert to_json(destination_reference) == to_json(destination_legacy)

    origin_legacy = exporter_detail.update_origin_plant_summary_table(
        "United States",
        30,
        ["Gulf"],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        "previous",
        5,
        3,
        3,
        legacy_summaries[1],
    )
    origin_reference = exporter_detail.update_origin_plant_summary_table(
        "United States",
        30,
        ["Gulf"],
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        "previous",
        5,
        3,
        3,
        reference_summaries[1],
    )
    assert to_json(origin_reference) == to_json(origin_legacy)

    legacy_export = exporter_detail.export_supply_analysis_to_excel(
        1,
        "United States",
        30,
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        payload,
    )
    reference_export = exporter_detail.export_supply_analysis_to_excel(
        1,
        "United States",
        30,
        exporter_detail.DEFAULT_DESTINATION_LEVEL,
        "mcm_d",
        reference,
    )
    assert _workbook_cells(reference_export) == _workbook_cells(legacy_export)


@pytest.mark.parametrize("mode", ["missing", "corrupt"])
def test_missing_or_corrupt_reference_has_explicit_recovery_and_no_raw_fallback(
    mode,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    _, reference = _load_reference(monkeypatch)
    _delete_or_corrupt_reference(reference, mode)

    expected_selector = (
        [{
            "label": exporter_detail.EXPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE,
            "value": "__snapshot_unavailable__",
            "disabled": True,
        }],
        [],
    )
    assert (
        exporter_detail.update_exporter_detail_supply_year_selector(
            reference,
            [],
        )
        == expected_selector
    )
    assert (
        exporter_detail.update_exporter_detail_continent_year_selector(
            reference,
            [],
        )
        == expected_selector
    )

    chart_result = exporter_detail.update_exporter_detail_supply_charts(
        reference,
        30,
        "mcm_d",
        [],
    )
    assert (
        chart_result[-1].children
        == exporter_detail.EXPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE
    )
    assert chart_result[-1].role == "alert"

    summary_result = (
        exporter_detail.refresh_exporter_detail_summary_data_stores(
            "United States",
            30,
            exporter_detail.DEFAULT_DESTINATION_LEVEL,
            reference,
        )
    )
    assert all(
        item["error"]
        == exporter_detail.EXPORTER_DETAIL_SNAPSHOT_RECOVERY_MESSAGE
        for item in summary_result
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Cached exporter-detail data is unavailable",
    ):
        exporter_detail.export_supply_analysis_to_excel(
            1,
            "United States",
            30,
            exporter_detail.DEFAULT_DESTINATION_LEVEL,
            "mcm_d",
            reference,
        )


def test_wrong_namespace_and_nonresolvable_reference_are_rejected(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    wrong_reference = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": "another-page-v1",
        "source_key": "source",
        "revision": "00000000-0000-4000-8000-000000000000",
        "shared": True,
    }
    with pytest.raises(snapshots.SnapshotUnavailable):
        exporter_detail._resolve_exporter_detail_base_data(wrong_reference)

    payload = _base_payload()
    local_reference = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": exporter_detail.EXPORTER_DETAIL_BASE_NAMESPACE,
        "source_key": "source",
        "revision": 1,
        "shared": False,
    }
    monkeypatch.setattr(
        exporter_detail,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (local_reference, payload),
    )
    with pytest.raises(snapshots.SnapshotUnavailable):
        exporter_detail.refresh_exporter_detail_base_data(
            "United States",
            0,
        )

    malformed_reference, _ = snapshots.get_or_build_snapshot(
        exporter_detail.engine,
        namespace=exporter_detail.EXPORTER_DETAIL_BASE_NAMESPACE,
        source_key="malformed-payload",
        builder=lambda: {},
        force=True,
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Cached exporter-detail data is unavailable",
    ):
        exporter_detail._resolve_exporter_detail_base_data(
            malformed_reference
        )


@pytest.mark.parametrize("workers", [1, 2, 4])
def test_base_loader_is_single_flight_for_one_two_and_four_callers(
    workers,
    monkeypatch,
    persistent_exporter_detail_cache,
):
    counter = {"count": 0}
    counter_lock = threading.Lock()

    def build_payload(_country):
        with counter_lock:
            counter["count"] += 1
        time.sleep(0.04)
        return _base_payload()

    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        build_payload,
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: f"single-flight-{workers}",
    )

    with ThreadPoolExecutor(max_workers=workers) as pool:
        references = list(
            pool.map(
                lambda _index: (
                    exporter_detail.refresh_exporter_detail_base_data(
                        "United States",
                        0,
                    )
                ),
                range(workers),
            )
        )

    assert counter["count"] == 1
    assert all(
        snapshots.snapshot_is_resolvable(reference)
        for reference in references
    )
    assert len(
        {
            (
                reference["source_key"],
                reference["revision"],
            )
            for reference in references
        }
    ) == 1


def test_global_refresh_is_checked_once_and_forces_one_rebuild(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    refresh_calls = []
    build_calls = []
    monkeypatch.setattr(
        exporter_detail,
        "_was_global_refresh_triggered",
        lambda: refresh_calls.append(True) or True,
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda country: build_calls.append(country) or _base_payload(country),
    )

    reference = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        1,
    )

    assert refresh_calls == [True]
    assert build_calls == ["United States"]
    assert snapshots.snapshot_is_resolvable(reference)


def test_route_diversion_and_maintenance_remain_outside_base_snapshot_and_exports_are_stable(
    monkeypatch,
):
    independent_callbacks = (
        exporter_detail.update_route_analysis_charts_and_tables,
        exporter_detail.export_route_analysis_to_excel,
        exporter_detail.process_diversion_data,
        exporter_detail.update_diversion_ui,
        exporter_detail.export_diversion_summary_to_excel,
        exporter_detail.update_maintenance_table,
        exporter_detail.update_maintenance_seasonal_chart,
    )
    assert all(
        "_resolve_exporter_detail_base_data"
        not in callback_function.__code__.co_names
        for callback_function in independent_callbacks
    )

    route_frame = pd.DataFrame(
        {
            "year": [2025, 2025, 2026, 2026],
            "voyage_id": ["v1", "v2", "v3", "v4"],
            "destination_country_name": [
                "Spain",
                "Japan",
                "France",
                "Brazil",
            ],
            "destination_shipping_region": [
                "Iberia",
                "North Asia",
                "Northwest Europe",
                "South America",
            ],
            "distanceDirect": [100.0, 100.0, 100.0, 100.0],
            "distanceViaSuez": [120.0, np.nan, 120.0, np.nan],
            "distanceViaPanama": [np.nan, 130.0, 130.0, np.nan],
        }
    )
    monkeypatch.setattr(
        exporter_detail,
        "process_trade_and_distance_data",
        lambda *_args, **_kwargs: route_frame.copy(),
    )
    route_first = exporter_detail.export_route_analysis_to_excel(
        1,
        "Year",
        "destination_country_name",
        "United States",
    )
    route_second = exporter_detail.export_route_analysis_to_excel(
        1,
        "Year",
        "destination_country_name",
        "United States",
    )
    assert _workbook_cells(route_first) == _workbook_cells(route_second)

    diversion_store = {
        "main_data": [
            {
                "Diversion date": "2026-07-20",
                "Vessel": "Example LNG",
                "State": "Loaded",
                "Cubic Meters": 174500.25,
                "Added shipping days": 2.5,
            }
        ]
    }
    diversion_columns = [
        {"field": "Diversion date"},
        {"field": "Vessel"},
        {"field": "State"},
        {"field": "Cubic Meters"},
        {"field": "Added shipping days"},
    ]
    diversion_first = exporter_detail.export_diversion_summary_to_excel(
        1,
        diversion_store,
        diversion_columns,
    )
    diversion_second = exporter_detail.export_diversion_summary_to_excel(
        1,
        diversion_store,
        diversion_columns,
    )
    assert _workbook_cells(diversion_first) == _workbook_cells(
        diversion_second
    )


def test_maintenance_controls_reuse_current_raw_source_with_exact_output(
    monkeypatch,
):
    raw_data = pd.DataFrame(
        {
            "id_plant": [1],
            "plant_name": ["Sabine Pass"],
            "country_name": ["United States"],
            "id_lng_train": [11],
            "lng_train_name_short": ["Train 1"],
            "year": [2026],
            "month": [7],
            "year_actual_forecast": ["Forecast"],
            "total_mtpa": [0.75],
            "metric_comment": ["Planned"],
            "train_capacity_mmtpa": [4.5],
            "train_capacity_mcmd": [16.77],
            "plant_capacity_mmtpa": [27.0],
            "plant_capacity_mcmd": [100.62],
            "date": pd.to_datetime(["2026-07-01"]),
        }
    )
    fetches = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or raw_data.copy()
        ),
    )
    baseline = exporter_detail.update_maintenance_table(
        "United States",
        "mt",
        3,
        3,
        [],
    )
    assert fetches == [True]

    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "maintenance controls reread the source"
        ),
    )
    cached = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=baseline[2],
    )

    assert to_json(cached[:2]) == to_json(baseline[:2])
    assert cached[2] == baseline[2]


@pytest.mark.parametrize("error_message", ["source down", ""])
def test_maintenance_source_error_retries_on_next_control_change(
    monkeypatch,
    error_message,
):
    attempts = []

    def fail_source(*_args, **_kwargs):
        attempts.append("failed")
        raise RuntimeError(error_message)

    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        fail_source,
    )
    failed = exporter_detail.update_maintenance_table(
        "United States",
        "mcm_d",
        3,
        3,
        [],
    )
    assert failed[2]["error"] == error_message

    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            attempts.append("retried") or pd.DataFrame()
        ),
    )
    retried = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=failed[2],
    )

    assert attempts == ["failed", "retried"]
    assert "No maintenance data available" in to_json(retried[0])
    assert "error" not in retried[2]


def test_maintenance_database_failure_reaches_retryable_error_store(
    monkeypatch,
):
    monkeypatch.setattr(
        exporter_detail.engine,
        "connect",
        lambda: (_ for _ in ()).throw(RuntimeError("alias database down")),
    )
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("maintenance database down")
        ),
    )

    failed = exporter_detail.update_maintenance_table(
        "United States",
        "mcm_d",
        3,
        3,
        [],
    )

    assert failed[2]["error"] == "maintenance database down"
    assert "Error loading maintenance data" in to_json(failed[0])


def test_source_context_refresh_generation_is_not_sticky_for_later_selections(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    build_calls = []
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda country: (
            build_calls.append(country) or _base_payload(country)
        ),
    )
    watermark = {
        "kpler_watermark": "2026-07-24T00:00:00Z",
        "woodmac_watermark": "2026-07-23T00:00:00Z",
        "distance_watermark": "2026-07-22T00:00:00Z",
    }
    initial_context = exporter_detail._build_exporter_detail_source_context(
        watermark
    )
    refresh_context = exporter_detail._build_exporter_detail_source_context(
        watermark,
        force_refresh=True,
    )

    exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=initial_context,
    )
    exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=refresh_context,
    )
    exporter_detail.refresh_exporter_detail_base_data(
        "Qatar",
        source_context=refresh_context,
    )
    exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=refresh_context,
    )

    assert build_calls == ["United States", "United States", "Qatar"]


def test_mapping_fingerprint_invalidates_base_and_route_snapshots(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    base_builds = []
    route_builds = []
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        lambda country: (
            base_builds.append(country) or _base_payload(country)
        ),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_route_source_payload",
        lambda country: (
            route_builds.append(country)
            or {
                "origin_country": country,
                "processed_df": pd.DataFrame(),
                "mapping_df": pd.DataFrame(),
            }
        ),
    )
    source_watermark = {
        "kpler_watermark": "kpler-v1",
        "woodmac_watermark": "woodmac-v1",
        "distance_watermark": "distance-v1",
        "mapping_fingerprint": "mapping-v1",
    }
    initial_context = (
        exporter_detail._build_exporter_detail_source_context(
            source_watermark
        )
    )
    changed_context = (
        exporter_detail._build_exporter_detail_source_context(
            {
                **source_watermark,
                "mapping_fingerprint": "mapping-v2",
            }
        )
    )

    first_base = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=initial_context,
    )
    changed_base = exporter_detail.refresh_exporter_detail_base_data(
        "United States",
        source_context=changed_context,
    )
    first_route = exporter_detail.refresh_exporter_route_analysis_source(
        "United States",
        initial_context,
    )
    changed_route = exporter_detail.refresh_exporter_route_analysis_source(
        "United States",
        changed_context,
    )

    assert base_builds == ["United States", "United States"]
    assert route_builds == ["United States", "United States"]
    assert first_base["source_key"] != changed_base["source_key"]
    assert first_route["source_key"] != changed_route["source_key"]


def test_exporter_route_falls_back_when_context_watermark_is_unavailable(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    metadata_reads = []
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_route_source_version",
        lambda: (
            metadata_reads.append(True)
            or {
                "kpler_watermark": "kpler-v1",
                "distance_watermark": "distance-v1",
                "mapping_fingerprint": "mapping-v1",
            }
        ),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_route_source_payload",
        lambda country: {
            "origin_country": country,
            "processed_df": pd.DataFrame(),
            "mapping_df": pd.DataFrame(),
        },
    )
    degraded_context = (
        exporter_detail._build_exporter_detail_source_context(
            "metadata-unavailable"
        )
    )

    reference = exporter_detail.refresh_exporter_route_analysis_source(
        "United States",
        degraded_context,
    )

    assert metadata_reads == [True]
    assert snapshots.is_snapshot_reference(
        reference,
        exporter_detail.EXPORTER_ROUTE_SOURCE_NAMESPACE,
    )


def test_country_catalog_exposes_unavailable_source_watermark_failure(monkeypatch):
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pd.DataFrame(
            {"origin_country_name": ["Qatar", "United States"]}
        ),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: (_ for _ in ()).throw(RuntimeError("watermark down")),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_maintenance_source_version",
        lambda: (_ for _ in ()).throw(RuntimeError("maintenance down")),
    )

    options, selected, source_context = (
        exporter_detail.initialize_country_dropdown(0)
    )

    assert options == [
        {"label": "Qatar", "value": "Qatar"},
        {"label": "United States", "value": "United States"},
    ]
    assert selected == "United States"
    assert source_context["source_watermark"] is None
    assert source_context["source_revision"]["status"] == "unavailable"
    assert "unavailable" in source_context["source_revision"]["message"]


def test_country_catalog_reuses_last_good_revision_as_stale(monkeypatch):
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda *_args, **_kwargs: pd.DataFrame(
            {"origin_country_name": ["Qatar", "United States"]}
        ),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: (_ for _ in ()).throw(RuntimeError("watermark down")),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_maintenance_source_version",
        lambda: {"revision": "maintenance-v1"},
    )
    previous_context = (
        exporter_detail._build_exporter_detail_source_context(
            {"kpler_watermark": "kpler-v1"}
        )
    )
    existing_options = [
        {"label": "Qatar", "value": "Qatar"},
        {"label": "United States", "value": "United States"},
    ]

    options, selected, source_context = (
        exporter_detail.initialize_country_dropdown(
            0,
            existing_options,
            "Qatar",
            previous_context,
        )
    )

    assert options == existing_options
    assert selected == "Qatar"
    assert source_context["source_watermark"] == {
        "kpler_watermark": "kpler-v1"
    }
    assert source_context["source_revision"]["status"] == "stale"


def test_exporter_maintenance_reference_is_small_exact_and_refreshes_once(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    raw_data = pd.DataFrame(
        {
            "id_plant": [1],
            "plant_name": ["Sabine Pass"],
            "country_name": ["United States"],
            "id_lng_train": [11],
            "lng_train_name_short": ["Train 1"],
            "year": [2026],
            "month": [7],
            "year_actual_forecast": ["Forecast"],
            "total_mtpa": [0.75],
            "metric_comment": ["Planned"],
            "train_capacity_mmtpa": [4.5],
            "train_capacity_mcmd": [16.77],
            "plant_capacity_mmtpa": [27.0],
            "plant_capacity_mcmd": [100.62],
            "date": pd.to_datetime(["2026-07-01"]),
        }
    )
    fetches = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or raw_data.copy()
        ),
    )
    legacy = exporter_detail.update_maintenance_table(
        "United States",
        "mcm_d",
        3,
        3,
        [],
    )
    version = {
        "unplanned_watermark": "u1",
        "unplanned_row_count": 1,
        "planned_watermark": "p1",
        "planned_row_count": 1,
        "train_capacity_watermark": "t1",
        "train_capacity_row_count": 1,
        "plant_capacity_watermark": "s1",
        "plant_capacity_row_count": 1,
        "mapping_fingerprint": "m1",
    }
    context = exporter_detail._build_exporter_detail_source_context(
        "base-v1",
        maintenance_source_version=version,
    )
    snapshotted = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        source_context=context,
    )

    assert to_json(snapshotted[:2]) == to_json(legacy[:2])
    assert snapshots.is_snapshot_reference(
        snapshotted[2],
        exporter_detail.EXPORTER_MAINTENANCE_SOURCE_NAMESPACE,
    )
    assert len(to_json(snapshotted[2]).encode("utf-8")) < 1_000
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: pytest.fail(
            "maintenance reference reread SQL"
        ),
    )
    cached = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        maintenance_raw_data=snapshotted[2],
        source_context=context,
    )
    assert to_json(cached[:2]) == to_json(legacy[:2])

    refreshed_fetches = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            refreshed_fetches.append(True) or raw_data.copy()
        ),
    )
    refresh_context = exporter_detail._build_exporter_detail_source_context(
        "base-v1",
        force_refresh=True,
        maintenance_source_version=version,
    )
    refreshed = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        maintenance_raw_data=snapshotted[2],
        source_context=refresh_context,
    )
    exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=refreshed[2],
        source_context=refresh_context,
    )
    assert refreshed_fetches == [True]


def test_exporter_maintenance_metadata_failure_does_not_run_unversioned_query(
    monkeypatch,
):
    fetches = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_train_maintenance_data",
        lambda *_args, **_kwargs: (
            fetches.append(True) or pd.DataFrame()
        ),
    )
    initial_context = (
        exporter_detail._build_exporter_detail_source_context(
            "metadata-unavailable",
            maintenance_source_version=None,
        )
    )

    first = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        source_context=initial_context,
    )
    exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=first[2],
        source_context=initial_context,
    )

    refresh_context = (
        exporter_detail._build_exporter_detail_source_context(
            "metadata-unavailable",
            force_refresh=True,
            maintenance_source_version=None,
        )
    )
    refreshed = exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mcm_d",
        3,
        3,
        [],
        maintenance_raw_data=first[2],
        source_context=refresh_context,
    )
    exporter_detail._update_maintenance_table_from_source(
        "United States",
        "mt",
        3,
        3,
        [],
        maintenance_raw_data=refreshed[2],
        source_context=refresh_context,
    )

    assert fetches == []
    assert first[2]["error"]
    assert refreshed[2]["error"]


def test_exporter_persistent_sources_track_mapping_and_refresh_generations(
    monkeypatch,
    persistent_exporter_detail_cache,
):
    allocation_source = _allocation_source()
    allocation_metadata = []
    for fingerprint in (
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ):
        metadata = copy.deepcopy(allocation_source["run_metadata"])
        metadata["mapping_fingerprint"] = fingerprint
        allocation_metadata.append(metadata)
    allocation_builds = []
    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: allocation_metadata.pop(0),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        lambda *_args, **kwargs: (
            allocation_builds.append(
                kwargs["run_metadata"]["mapping_fingerprint"]
            )
            or copy.deepcopy(allocation_source)
        ),
    )

    initial_context = {"refresh_generation": None}
    first_refresh_context = {
        "refresh_generation": "refresh-generation-1"
    }
    second_refresh_context = {
        "refresh_generation": "refresh-generation-2"
    }
    source_contexts = (
        initial_context,
        initial_context,
        first_refresh_context,
        first_refresh_context,
        second_refresh_context,
    )
    allocation_refs = [
        exporter_detail.refresh_destination_forecast_source(
            "United States",
            source_context,
        )
        for source_context in source_contexts
    ]

    assert allocation_builds == [
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ]
    assert allocation_refs[0]["source_key"] != allocation_refs[1][
        "source_key"
    ]
    assert allocation_refs[1]["source_key"] != allocation_refs[2][
        "source_key"
    ]
    assert allocation_refs[2]["source_key"] == allocation_refs[3][
        "source_key"
    ]
    assert allocation_refs[3]["source_key"] != allocation_refs[4][
        "source_key"
    ]

    diversion_versions = [
        {
            "diversion_watermark": "diversion-v1",
            "country_mapping_fingerprint": fingerprint,
            "location_mapping_fingerprint": "location-v1",
        }
        for fingerprint in (
            "mapping-v1",
            "mapping-v2",
            "mapping-v2",
            "mapping-v2",
            "mapping-v2",
        )
    ]
    diversion_versions_for_fetch = copy.deepcopy(diversion_versions)
    diversion_builds = []
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_diversion_source_version",
        lambda: diversion_versions.pop(0),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_diversion_payload",
        lambda country, **kwargs: (
            diversion_builds.append(
                kwargs["source_version"][
                    "country_mapping_fingerprint"
                ]
            )
            or {
                "origin_country": country,
                "main_data": [],
                "charts_data": [],
            }
        ),
    )

    diversion_refs = [
        exporter_detail.refresh_exporter_diversion_source(
            source_context,
            "United States",
        )
        for source_context in source_contexts
    ]

    assert diversion_builds == [
        "mapping-v1",
        "mapping-v2",
        "mapping-v2",
        "mapping-v2",
    ]
    assert diversion_refs[0]["source_key"] != diversion_refs[1][
        "source_key"
    ]
    assert diversion_refs[1]["source_key"] != diversion_refs[2][
        "source_key"
    ]
    assert diversion_refs[2]["source_key"] == diversion_refs[3][
        "source_key"
    ]
    assert diversion_refs[3]["source_key"] != diversion_refs[4][
        "source_key"
    ]

    captured_params = {}
    monkeypatch.setattr(
        exporter_detail.pd,
        "read_sql",
        lambda _query, _engine, params: (
            captured_params.update(params) or pd.DataFrame()
        ),
    )
    exporter_detail._fetch_exporter_diversion_rows(
        "United States",
        diversion_versions_for_fetch[0],
    )
    assert captured_params["source_version"] == "diversion-v1"

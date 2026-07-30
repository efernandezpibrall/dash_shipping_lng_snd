import base64
from datetime import date
from io import BytesIO

from openpyxl import load_workbook
import pandas as pd
import pytest
from dash import html

from pages import importer_detail, importers


def _snapshot_pair(current_date, baseline_date=None):
    current_date = pd.Timestamp(current_date)
    pair = {
        'current_snapshot_id': 200,
        'current_snapshot_date_utc': current_date.date(),
        'current_snapshot_timestamp_utc': current_date.replace(
            hour=5,
            minute=34,
            second=12,
            microsecond=778724,
        ),
        'current_facts_retained': True,
        'baseline_snapshot_id': None,
        'baseline_snapshot_date_utc': None,
        'baseline_snapshot_timestamp_utc': None,
        'baseline_facts_retained': None,
    }
    if baseline_date is not None:
        baseline_date = pd.Timestamp(baseline_date)
        pair.update({
            'baseline_snapshot_id': 199,
            'baseline_snapshot_date_utc': baseline_date.date(),
            'baseline_snapshot_timestamp_utc': baseline_date.replace(
                hour=5,
                minute=36,
                second=57,
                microsecond=998945,
            ),
            'baseline_facts_retained': True,
        })
    return pair


@pytest.mark.parametrize(
    (
        'current_date',
        'baseline_date',
        'expected_status',
        'expected_gap',
    ),
    [
        ('2026-07-30', '2026-07-29', 'exact', 1),
        ('2026-07-27', '2026-07-24', 'exact', 1),
        ('2026-07-26', '2026-07-24', 'exact', 1),
        ('2026-07-30', '2026-07-28', 'fallback', 2),
        ('2026-07-30', None, 'unavailable', None),
    ],
)
def test_importer_source_state_selects_previous_weekday_or_fallback(
    current_date,
    baseline_date,
    expected_status,
    expected_gap,
):
    state = importers._build_importers_source_state(
        _snapshot_pair(current_date, baseline_date),
        refresh_token='refresh-1',
    )

    assert state['format'] == importers.IMPORTERS_SOURCE_STATE_FORMAT
    assert state['current_snapshot']['snapshot_id'] == 200
    assert state['current_snapshot']['snapshot_date_utc'] == current_date
    assert state['watermark'].endswith('05:34:12.778724')
    assert state['baseline_status'] == expected_status
    assert state['business_day_gap'] == expected_gap
    if baseline_date is None:
        assert state['baseline_snapshot'] is None
    else:
        assert state['baseline_snapshot']['snapshot_id'] == 199
        assert (
            state['baseline_snapshot']['snapshot_date_utc']
            == baseline_date
        )


def test_source_cache_key_is_versioned_and_includes_exact_pair():
    source_state = importers._build_importers_source_state(
        _snapshot_pair('2026-07-30', '2026-07-29')
    )
    changed_pair = _snapshot_pair('2026-07-30', '2026-07-29')
    changed_pair['baseline_snapshot_timestamp_utc'] = (
        pd.Timestamp('2026-07-29T06:00:00')
    )
    changed_state = importers._build_importers_source_state(
        changed_pair
    )

    assert importers.IMPORTERS_SOURCE_NAMESPACE.endswith('-v3')
    assert importers.IMPORTERS_PERIOD_NAMESPACE.endswith('-v3')
    assert (
        importers._importers_source_snapshot_key(source_state)
        != importers._importers_source_snapshot_key(changed_state)
    )
    assert (
        importers._importers_source_snapshot_key(source_state)
        == importers._importers_source_snapshot_key({
            **source_state,
            'refresh_token': 'ignored',
        })
    )


def test_scoped_trade_query_binds_exact_snapshot_and_bounded_window(
    monkeypatch,
):
    captured = {}

    def read_sql(query, _engine, params=None):
        captured['statement'] = str(query)
        captured['params'] = params
        return pd.DataFrame(columns=[
            'end_date',
            'cargo_mcm',
            'origin_country',
            'origin_continent_chart',
            'origin_continent',
            'origin_shipping_region',
            'origin_basin',
            'origin_subcontinent',
            'origin_classification_level1',
            'origin_classification',
            'destination_country_name',
        ])

    monkeypatch.setattr(importer_detail.pd, 'read_sql', read_sql)
    importer_detail._fetch_importer_scoped_trades(
        importer_detail.engine,
        ['France'],
        min_end_date=date(2026, 6, 30),
        max_end_date=date(2026, 7, 29),
        snapshot_timestamp_utc='2026-07-29T05:36:57.998945',
        delivered_only=True,
        include_destination_context=True,
    )

    assert (
        'kt.upload_timestamp_utc = :snapshot_timestamp_utc'
        in captured['statement']
    )
    assert 'kt."end"::date <= :max_end_date' in captured['statement']
    assert captured['params']['min_end_date'] == date(2026, 6, 30)
    assert captured['params']['max_end_date'] == date(2026, 7, 29)
    assert (
        captured['params']['snapshot_timestamp_utc']
        == '2026-07-29T05:36:57.998945'
    )


def test_catalog_ranking_query_uses_atomic_current_snapshot(
    monkeypatch,
):
    captured = {}

    def read_sql(query, _engine, params=None):
        captured['statement'] = str(query)
        captured['params'] = params
        return pd.DataFrame(columns=[
            'destination_country_name',
            'avg_30d_mcmd',
        ])

    monkeypatch.setattr(importers.pd, 'read_sql', read_sql)
    importers._fetch_importers_catalog_ranking_source_df(
        '2026-07-30T05:34:12.778724',
        '2026-07-30',
    )

    assert (
        'CAST(:snapshot_timestamp_utc AS timestamptz)'
        in captured['statement']
    )
    assert captured['params'] == {
        'snapshot_timestamp_utc': (
            '2026-07-30T05:34:12.778724'
        ),
        'ranking_start_date': date(2026, 7, 1),
        'ranking_end_date': date(2026, 7, 31),
    }


def _trade_frame(as_of_date, daily_by_origin):
    rows = []
    for flow_date in pd.date_range(
        pd.Timestamp(as_of_date) - pd.Timedelta(days=29),
        pd.Timestamp(as_of_date),
    ):
        for origin_country, daily_mcmd in daily_by_origin.items():
            rows.append({
                'end_date': flow_date,
                'cargo_mcm': float(daily_mcmd),
                'origin_country': origin_country,
                'destination_country_name': 'France',
            })
    return pd.DataFrame(rows)


def test_period_payload_uses_exact_pair_and_outer_joins_changes(
    monkeypatch,
):
    current_df = _trade_frame(
        '2026-07-30',
        {'A': 10, 'B': 5},
    )
    baseline_df = _trade_frame(
        '2026-07-29',
        {'A': 8, 'C': 4},
    )
    query_calls = []

    def fetch_scoped_trades(*_args, **kwargs):
        query_calls.append(kwargs)
        if kwargs.get('snapshot_timestamp_utc', '').startswith(
            '2026-07-29'
        ):
            return baseline_df.copy()
        return current_df.copy()

    monkeypatch.setattr(
        importers,
        '_fetch_importer_scoped_trades',
        fetch_scoped_trades,
    )
    source_state = importers._build_importers_source_state(
        _snapshot_pair('2026-07-30', '2026-07-29')
    )
    payload = importers._build_period_payload(
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        'Country',
        'origin_country_name',
        'show_all',
        45,
        source_state,
    )

    assert payload['format'] == importers.IMPORTERS_PERIOD_PAYLOAD_FORMAT
    assert payload['snapshot_comparison']['status'] == 'exact'
    assert len(query_calls) == 2
    assert query_calls[0]['snapshot_timestamp_utc'].startswith(
        '2026-07-30'
    )
    assert query_calls[0]['max_end_date'] == '2026-07-30'
    assert query_calls[1]['min_end_date'] == date(2026, 6, 30)
    assert query_calls[1]['max_end_date'] == '2026-07-29'

    records_by_origin = {
        record['country']: record
        for record in payload['show_all'][0]['records']
    }
    assert records_by_origin['A']['Δ 30D vs PBD'] == pytest.approx(2)
    assert records_by_origin['A']['Δ 7D vs PBD'] == pytest.approx(2)
    assert records_by_origin['B']['Δ 30D vs PBD'] == pytest.approx(5)
    assert records_by_origin['B']['Δ 7D vs PBD'] == pytest.approx(5)
    assert records_by_origin['C']['Δ 30D vs PBD'] == pytest.approx(-4)
    assert records_by_origin['C']['Δ 7D vs PBD'] == pytest.approx(-4)
    assert '45D' in records_by_origin['A']


def test_small_origin_grouping_reuses_current_vintage_taxonomy():
    current_df = pd.concat([
        _trade_frame('2026-07-30', {'Small': 1}),
        _trade_frame('2026-07-30', {'Large': 20}),
    ], ignore_index=True)
    current_df['origin_shipping_region'] = 'Atlantic'
    baseline_df = _trade_frame(
        '2026-07-29',
        {'Small': 50, 'Large': 20},
    )
    baseline_df['origin_shipping_region'] = 'Atlantic'

    grouped_current, grouping_config = (
        importers.group_small_importer_origin_countries(
            current_df,
            'origin_shipping_region',
            as_of_date='2026-07-30',
            return_grouping_config=True,
        )
    )
    grouped_baseline = (
        importers.group_small_importer_origin_countries(
            baseline_df,
            'origin_shipping_region',
            grouping_config=grouping_config,
        )
    )

    assert 'Rest of countries' in set(
        grouped_current['origin_country']
    )
    assert 'Rest of countries' in set(
        grouped_baseline['origin_country']
    )
    assert 'Small' not in set(grouped_baseline['origin_country'])
    assert 'Large' in set(grouped_baseline['origin_country'])


def test_rolling_windows_use_inclusive_snapshot_boundaries():
    rows = []
    for flow_date in pd.date_range('2026-06-30', '2026-07-30'):
        rows.append({
            'end_date': flow_date,
            'cargo_mcm': (
                310.0
                if flow_date == pd.Timestamp('2026-06-30')
                else 10.0
            ),
            'origin_country': 'A',
            'destination_country_name': 'France',
        })
    summary_df = (
        importers.build_importer_origin_summary_from_scoped_trades(
            pd.DataFrame(rows),
            rolling_window_days=30,
            origin_level='origin_country_name',
            current_date='2026-07-30',
        )
    )

    assert summary_df.loc[0, '30D'] == 10
    assert summary_df.loc[0, '7D'] == 10


def _period_payload(status='exact'):
    baseline_snapshot = (
        {
            'snapshot_id': 199,
            'snapshot_date_utc': '2026-07-29',
            'snapshot_timestamp_utc': (
                '2026-07-29T05:36:57.998945'
            ),
        }
        if status in {'exact', 'fallback'}
        else None
    )
    records = [
        {
            'continent': 'A',
            'country': 'A',
            "Q2'26": 70,
            "Q2'25": 50,
            "Jun'26": 70,
            "Jun'25": 50,
            '30D': 75,
            '30D_PP': 60,
            '30D_Y1': 50,
            "W29'26": 75,
            "W28'26": 65,
            "W29'25": 50,
            '7D': 70,
            '7D_PP': 60,
            '7D_Y1': 50,
            '30D_PBD_CURRENT': 75,
            '7D_PBD_CURRENT': 70,
            '30D_PBD': 50,
            '7D_PBD': 90,
            'Δ 30D vs PBD': 25,
            'Δ 7D vs PBD': -20,
        },
        {
            'continent': 'B',
            'country': 'B',
            "Q2'26": 30,
            "Q2'25": 50,
            "Jun'26": 30,
            "Jun'25": 50,
            '30D': 25,
            '30D_PP': 40,
            '30D_Y1': 50,
            "W29'26": 25,
            "W28'26": 35,
            "W29'25": 50,
            '7D': 30,
            '7D_PP': 40,
            '7D_Y1': 50,
            '30D_PBD_CURRENT': 25,
            '7D_PBD_CURRENT': 30,
            '30D_PBD': 50,
            '7D_PBD': 10,
            'Δ 30D vs PBD': -25,
            'Δ 7D vs PBD': 20,
        },
    ]
    if status == 'unavailable':
        for record in records:
            for column_name in (
                *importers.IMPORTER_PERIOD_PBD_CURRENT_COLUMNS,
                *importers.IMPORTER_PERIOD_PBD_REFERENCE_COLUMNS,
                *importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS,
            ):
                record[column_name] = float('nan')
    return {
        'format': importers.IMPORTERS_PERIOD_PAYLOAD_FORMAT,
        'active_grouping_mode': 'show_all',
        'show_all': [{
            'label': 'France',
            'key': 'France',
            'records': records,
        }],
        'group_small_countries': [],
        'snapshot_comparison': {
            'status': status,
            'current_snapshot': {
                'snapshot_id': 200,
                'snapshot_date_utc': '2026-07-30',
                'snapshot_timestamp_utc': (
                    '2026-07-30T05:34:12.778724'
                ),
            },
            'baseline_snapshot': baseline_snapshot,
            'business_day_gap': 2 if status == 'fallback' else 1,
        },
    }


@pytest.mark.parametrize(
    'comparison_basis',
    ['levels', 'previous_period', 'same_period_last_year'],
)
def test_pbd_columns_remain_last_and_market_share_is_pp(
    comparison_basis,
):
    component, records = importers.update_period_analysis_table(
        _period_payload(),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        'mcm_d',
        30,
        'show_all',
        'percentage',
        comparison_basis,
        5,
        3,
        3,
    )

    assert isinstance(component, html.Div)
    assert list(records[0])[-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )
    assert '30D_PBD' not in records[0]
    child_records = {
        record['Aggregation'].strip(): record
        for record in records
        if record['Aggregation'].strip() in {'A', 'B'}
    }
    assert child_records['A']['Δ 30D vs PBD'] == 25
    assert child_records['A']['Δ 7D vs PBD'] == -20
    assert child_records['B']['Δ 30D vs PBD'] == -25
    assert child_records['B']['Δ 7D vs PBD'] == 20

    grid = component.children[0]
    assert [column['field'] for column in grid.columnDefs][-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )
    grid_children = {
        record['Aggregation'].strip(): record
        for record in grid.rowData
        if record['Aggregation'].strip() in {'A', 'B'}
    }
    assert grid_children['A']['Δ 30D vs PBD'] == '+25 pp'
    assert grid_children['A']['Δ 7D vs PBD'] == '-20 pp'


def _component_text(component):
    if component is None:
        return ''
    if isinstance(component, str):
        return component
    if isinstance(component, (list, tuple)):
        return ' '.join(_component_text(item) for item in component)
    return _component_text(getattr(component, 'children', None))


def test_unavailable_baseline_shows_dash_and_explicit_warning():
    component, records = importers.update_period_analysis_table(
        _period_payload(status='unavailable'),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        'mcm_d',
        30,
        'show_all',
        'absolute',
        'levels',
        5,
        3,
        3,
    )

    assert list(records[0])[-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )
    grid = component.children[0]
    assert all(
        row['Δ 30D vs PBD'] == '—'
        and row['Δ 7D vs PBD'] == '—'
        for row in grid.rowData
    )
    assert 'PBD baseline unavailable' in _component_text(component)


@pytest.mark.parametrize('volume_metric', ['mcm_d', 'mt', 'mtpa'])
def test_pbd_volume_delta_uses_selected_unit(volume_metric):
    _component, records = importers.update_period_analysis_table(
        _period_payload(),
        ['France'],
        [{
            'key': 'France',
            'label': 'France',
            'destination_countries': ['France'],
        }],
        volume_metric,
        30,
        'show_all',
        'absolute',
        'levels',
        5,
        3,
        3,
    )
    child_a = next(
        record
        for record in records
        if record['Aggregation'].strip() == 'A'
    )
    expected = (
        25.0
        if volume_metric == 'mcm_d'
        else round(
            25.0
            * importer_detail.DAYS_PER_YEAR
            / importer_detail.MCM_PER_MT,
            1,
        )
    )
    assert child_a['Δ 30D vs PBD'] == pytest.approx(expected)


def test_pbd_market_share_handles_zero_denominators_as_add_remove():
    display_df = pd.DataFrame([
        {
            'Importer': '▶ France',
            'Aggregation': 'Total',
            '30D_PBD_CURRENT': 100,
            '7D_PBD_CURRENT': 0,
            '30D_PBD': 0,
            '7D_PBD': 100,
        },
        {
            'Importer': '',
            'Aggregation': '    A',
            '30D_PBD_CURRENT': 100,
            '7D_PBD_CURRENT': 0,
            '30D_PBD': 0,
            '7D_PBD': 100,
        },
    ])
    percentage_df = (
        importers._apply_importer_period_pbd_percentage_view(
            display_df
        )
    )
    percentage_df = (
        importers._recalculate_importer_period_pbd_deltas(
            percentage_df
        )
    )

    assert percentage_df.loc[1, 'Δ 30D vs PBD'] == 100
    assert percentage_df.loc[1, 'Δ 7D vs PBD'] == -100


def test_period_export_keeps_pbd_columns_at_end(monkeypatch):
    captured = {}

    def send_export(export_df, filename_prefix, sheet_name):
        captured['columns'] = list(export_df.columns)
        captured['sheet_name'] = sheet_name
        return {'ok': True}

    monkeypatch.setattr(
        importers,
        '_send_export_dataframe',
        send_export,
    )
    display_records = [{
        'Importer': 'Global',
        'Aggregation': '',
        '30D': 10,
        '7D': 12,
        'Δ 30D vs PBD': 2,
        'Δ 7D vs PBD': -1,
    }]
    result = importers.export_period_analysis_to_excel(
        1,
        display_records,
        'origin_shipping_region',
        30,
        'absolute',
        'levels',
    )

    assert result == {'ok': True}
    assert captured['sheet_name'] == 'Period Analysis'
    assert captured['columns'][-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )


def test_period_excel_sheet_reconciles_rendered_pbd_values():
    display_records = [{
        'Importer': 'Global',
        'Aggregation': '',
        '30D': 10,
        '7D': 12,
        'Δ 30D vs PBD': 2,
        'Δ 7D vs PBD': -1,
    }]
    download = importers.export_period_analysis_to_excel(
        1,
        display_records,
        'origin_shipping_region',
        30,
        'absolute',
        'levels',
    )
    workbook = load_workbook(
        BytesIO(base64.b64decode(download['content']))
    )
    worksheet = workbook['Period Analysis']
    rows = list(worksheet.iter_rows(values_only=True))

    assert list(rows[0])[-2:] == list(
        importers.IMPORTER_PERIOD_PBD_DELTA_COLUMNS
    )
    assert rows[1][-2:] == (2, -1)

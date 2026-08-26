# Consolidated from test_arrow_payload.py.

import copy

import numpy as np
import pandas as pd
import pytest

from utils import arrow_payload
from utils import dashboard_snapshot_cache as snapshots


def test_dataframe_arrow_round_trip_preserves_exact_pandas_contract():
    index = pd.Index([30, 10, 20], name="row_id")
    frame = pd.DataFrame(
        {
            "category": pd.Categorical(
                ["loaded", "waiting", None],
                categories=["waiting", "loaded", "idle"],
                ordered=True,
            ),
            "nullable_int": pd.array([3, pd.NA, 2], dtype="Int64"),
            "nullable_bool": pd.array([True, pd.NA, False], dtype="boolean"),
            "quantity": np.array([3.5, np.nan, 2.25], dtype="float64"),
            "date": pd.to_datetime(
                ["2026-03-01", None, "2026-02-01"],
                utc=True,
            ),
            "label": pd.Series(["C", None, "B"], index=index, dtype="string"),
        },
        index=index,
    )

    encoded = arrow_payload.encode_arrow_dataframe(frame)
    restored = arrow_payload.decode_arrow_dataframe(encoded)

    assert encoded["format"] == arrow_payload.ARROW_DATAFRAME_FORMAT
    pd.testing.assert_frame_equal(restored, frame, check_categorical=True)


def test_dataframe_mapping_decodes_only_selected_frames():
    selected = pd.DataFrame({"value": [1, 2]})
    untouched = pd.DataFrame({"value": [3, 4]})
    packed = arrow_payload.pack_dataframe_mapping(
        {"selected": selected, "untouched": untouched, "status": "ok"},
        dataframe_keys=["selected"],
    )

    restored = arrow_payload.unpack_dataframe_mapping(
        packed,
        dataframe_keys=["selected"],
    )

    pd.testing.assert_frame_equal(restored["selected"], selected)
    assert restored["untouched"] is untouched
    assert restored["status"] == "ok"


def test_record_cube_arrow_round_trip_preserves_sparse_rows_and_scalars():
    records = {
        "Global": [
            {
                "value": np.float32(1.5),
                "nullable": pd.NA,
                "missing": None,
            },
            {"value": float("nan"), "date": pd.NaT},
        ],
        "Empty": [],
    }
    legacy_cube = snapshots.pack_record_mapping(records)

    restored_cube = arrow_payload.decode_arrow_record_cube(
        arrow_payload.encode_arrow_record_cube(legacy_cube)
    )
    restored = snapshots.unpack_record_mapping(restored_cube)

    assert list(restored) == ["Global", "Empty"]
    assert list(restored["Global"][0]) == list(records["Global"][0])
    assert isinstance(restored["Global"][0]["value"], np.float32)
    assert restored["Global"][0]["value"].tobytes() == (
        records["Global"][0]["value"].tobytes()
    )
    assert restored["Global"][0]["nullable"] is pd.NA
    assert restored["Global"][0]["missing"] is None
    assert np.isnan(restored["Global"][1]["value"])
    assert restored["Global"][1]["date"] is pd.NaT
    assert restored["Empty"] == []


def test_corrupt_arrow_payload_fails_closed():
    encoded = arrow_payload.encode_arrow_dataframe(
        pd.DataFrame({"value": [1, 2, 3]})
    )
    corrupt = copy.deepcopy(encoded)
    corrupt["payload"] = corrupt["payload"][:-8] + b"corrupt!"

    with pytest.raises(
        arrow_payload.ArrowPayloadError,
        match="corrupt|schema",
    ):
        arrow_payload.decode_arrow_dataframe(corrupt)


# Consolidated from test_dashboard_page_snapshot_fallbacks.py.

import pandas as pd
import pytest

from pages import supply
from utils import dashboard_snapshot_cache as snapshots
from utils import market_balance_data


def test_supply_shared_provider_failure_is_fail_closed(monkeypatch):
    monkeypatch.setattr(
        supply,
        "_get_provider_flow_snapshot",
        lambda **_kwargs: (_ for _ in ()).throw(RuntimeError("shared unavailable")),
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        supply.load_balance_source_data(None)


def test_market_provider_failure_uses_original_source_functions(monkeypatch):
    frames = {
        "woodmac_export": pd.DataFrame({"value": [1]}),
        "woodmac_import": pd.DataFrame({"value": [2]}),
        "ea_export": pd.DataFrame({"value": [3]}),
        "ea_import": pd.DataFrame({"value": [4]}),
        "mapping": pd.DataFrame({"country_name": ["Qatar"]}),
    }
    monkeypatch.setattr(
        market_balance_data,
        "get_provider_flow_snapshot",
        lambda **_kwargs: (_ for _ in ()).throw(RuntimeError("shared unavailable")),
    )
    monkeypatch.setattr(
        market_balance_data,
        "build_provider_flow_payload",
        lambda: (
            {"current_ea": {"run_id": 42}},
            {
                **frames,
                "woodmac_export_options": {"short_term": [], "long_term": []},
                "woodmac_import_options": {"short_term": [], "long_term": []},
                "ea_export_options": [],
                "ea_import_options": [],
                "ea_comparison_runs": [],
                "current_ea": {"run_id": 42, "snapshot_at": "2026-07-16T00:00:00Z"},
                "errors": {},
            },
        ),
    )

    payload = market_balance_data._resolve_latest_provider_flow_payload()

    for key, expected in frames.items():
        pd.testing.assert_frame_equal(payload[key], expected)
    assert payload["errors"] == {}


# Consolidated from test_detail_metric_policy.py.

import json

import pandas as pd
import pytest
from openpyxl import Workbook

from pages import exporter_detail, importer_detail
from utils import detail_volume_metrics


DETAIL_MODULES = (exporter_detail, importer_detail)


def _walk_components(component):
    if component is None:
        return
    if isinstance(component, (list, tuple)):
        for child in component:
            yield from _walk_components(child)
        return
    yield component
    children = getattr(component, "children", None)
    if children is not None:
        yield from _walk_components(children)


def _detail_component_text(component):
    return json.dumps(component.to_plotly_json(), default=str)


def test_detail_pages_share_the_canonical_metric_policy_implementation():
    for module in DETAIL_MODULES:
        assert module.VOLUME_CONVERSIONS is detail_volume_metrics.VOLUME_CONVERSIONS
        assert module.VOLUME_METRIC_OPTIONS is detail_volume_metrics.VOLUME_METRIC_OPTIONS
        assert (
            module._get_detail_volume_metric_info
            is detail_volume_metrics.get_volume_metric_info
        )
        assert (
            module._convert_detail_volume_dataframe
            is detail_volume_metrics.convert_volume_metric_dataframe
        )
        assert (
            module._format_detail_table_metric_value
            is detail_volume_metrics.format_table_metric_value
        )

    assert (
        importer_detail.get_volume_metric_info
        is detail_volume_metrics.get_volume_metric_info
    )
    assert (
        importer_detail.convert_volume_metric_dataframe
        is detail_volume_metrics.convert_volume_metric_dataframe
    )


def test_detail_layouts_do_not_include_obsolete_maintenance_refresh_stores():
    exporter_ids = [
        getattr(component, "id", None)
        for component in _walk_components(exporter_detail.layout)
    ]
    importer_ids = [
        getattr(component, "id", None)
        for component in _walk_components(importer_detail.layout)
    ]
    assert "maintenance-style-refresh-store" not in exporter_ids
    assert "imp-maintenance-style-refresh-store" not in importer_ids


@pytest.mark.parametrize("module", DETAIL_MODULES)
def test_detail_metric_policy_and_bcm_conversion(module):
    assert module.VOLUME_METRIC_OPTIONS == [
        {"label": "mcm/d", "value": "mcm_d"},
        {"label": "bcm", "value": "bcm"},
        {"label": "MT", "value": "mt"},
        {"label": "MMTPA", "value": "mtpa"},
    ]
    assert {
        metric: (
            module._get_detail_volume_metric_info(metric)["quantity_kind"],
            module._get_detail_volume_metric_display_precision(metric),
            module._get_detail_volume_metric_plotly_format(metric),
        )
        for metric in ("mcm_d", "bcm", "mt", "mtpa")
    } == {
        "mcm_d": ("rate", 0, ",.0f"),
        "bcm": ("period_volume", 1, ",.1f"),
        "mt": ("period_volume", 1, ",.1f"),
        "mtpa": ("rate", 1, ",.1f"),
    }
    assert module._get_detail_volume_metric_factor("bcm", period_days=30) == pytest.approx(0.03)
    assert module._get_detail_volume_metric_factor("mt", period_days=30) == pytest.approx(30 / 1360)
    assert module._convert_detail_volume_series(
        pd.Series([100.0]), "bcm", period_days=30
    ).iloc[0] == pytest.approx(3.0)
    assert module._normalize_detail_volume_metric("unknown") == "mcm_d"


def test_detail_selector_labels_are_consistent():
    exporter_filter = exporter_detail._build_exporter_detail_filter_bar()
    exporter_components = list(_walk_components(exporter_filter))
    exporter_metric = next(
        component
        for component in exporter_components
        if getattr(component, "id", None) == "volume-metric-dropdown"
    )
    assert exporter_metric.options == exporter_detail.VOLUME_METRIC_OPTIONS
    assert "Metric" in _detail_component_text(exporter_filter)
    assert '"Volume"' not in _detail_component_text(exporter_filter)

    importer_metric = next(
        component
        for component in _walk_components(importer_detail.layout)
        if getattr(component, "id", None) == "imp-volume-metric-dropdown"
    )
    assert importer_metric.options == importer_detail.VOLUME_METRIC_OPTIONS


@pytest.mark.parametrize("module,prefix", [
    (exporter_detail, "LNG Export Analysis"),
    (importer_detail, "LNG Import Analysis"),
])
def test_detail_titles_and_scalar_precision(module, prefix):
    assert module.update_supply_analysis_title(30, "mcm_d") == (
        f"{prefix} - 30-Day Rolling Average + WoodMac Forecast"
    )
    assert module.update_supply_analysis_title(30, "mtpa") == (
        f"{prefix} - 30-Day Rolling Average + WoodMac Forecast"
    )
    assert module.update_supply_analysis_title(30, "bcm") == (
        f"{prefix} - 30-Day Rolling Volume + WoodMac Forecast"
    )
    assert module.update_supply_analysis_title(30, "mt") == (
        f"{prefix} - 30-Day Rolling Volume + WoodMac Forecast"
    )
    assert module.update_supply_analysis_title(999, "bcm") == (
        f"{prefix} - 180-Day Rolling Volume + WoodMac Forecast"
    )
    assert module._format_detail_metric_value(10.4, "mcm_d") == "10 mcm/d"
    assert module._format_detail_metric_value(2.25, "bcm") == "2.2 bcm"
    assert module._format_detail_metric_value(-0.04, "bcm") == "0.0 bcm"
    assert "+0" not in _detail_component_text(
        module._build_detail_delta_pill(
            "MoM", {"delta": -0.04, "pct": -0.4}, "bcm"
        )
    )
    assert "-0" not in _detail_component_text(
        module._build_detail_delta_pill(
            "MoM", {"delta": -0.04, "pct": -0.4}, "bcm"
        )
    )


def _chart_history(value=10.0):
    dates = pd.to_datetime(["2025-01-01", "2025-02-01"])
    return pd.DataFrame({
        "date": dates,
        "year": dates.year,
        "day_of_year": dates.dayofyear,
        "month_day": dates.strftime("%b %d"),
        "rolling_avg": [value, value],
        "is_forecast": [False, False],
    })


def _empty_forecast():
    return pd.DataFrame(columns=[
        "date", "year", "day_of_year", "month_day", "mcmd",
        "is_forecast", "source",
    ])


@pytest.mark.parametrize("module,chart_builder", [
    (exporter_detail, exporter_detail._create_total_export_chart_with_woodmac_forecast),
    (importer_detail, importer_detail._create_total_import_chart_with_woodmac_forecast),
])
@pytest.mark.parametrize("metric,expected_y,number_format", [
    ("mcm_d", 10.0, ",.0f"),
    ("bcm", 0.3, ",.1f"),
    ("mt", 10 * 30 / 1360, ",.1f"),
    ("mtpa", 10 * 365.25 / 1360, ",.1f"),
])
def test_detail_supply_charts_follow_metric_precision(
    module, chart_builder, metric, expected_y, number_format
):
    figure = chart_builder(
        _chart_history(),
        _empty_forecast(),
        metric,
        selected_years=["2025"],
        rolling_window_days=30,
    )
    value_traces = [trace for trace in figure.data if trace.hovertemplate]
    assert value_traces
    assert all(f"%{{y:{number_format}}}" in trace.hovertemplate for trace in value_traces)
    assert figure.layout.yaxis.tickformat == number_format
    assert float(value_traces[0].y[0]) == pytest.approx(expected_y)


def test_detail_continent_charts_and_mix_use_metric_precision():
    dates = pd.to_datetime(["2025-01-01", "2025-02-01"])
    exporter_frame = pd.DataFrame({
        "date": dates,
        "continent_destination": ["Asia", "Asia"],
        "year": dates.year,
        "day_of_year": dates.dayofyear,
        "month_day": dates.strftime("%b %d"),
        "rolling_avg": [10.0, 10.0],
        "is_forecast": [False, False],
    })
    exporter_figure = exporter_detail.create_continent_destination_chart(
        "Test",
        rolling_window_days=30,
        volume_metric="bcm",
        selected_years=["2025"],
        continent_df=exporter_frame,
    )
    assert exporter_figure.layout.yaxis.tickformat == ",.1f"
    assert all(
        "%{y:,.1f}" in trace.hovertemplate
        for trace in exporter_figure.data
        if trace.hovertemplate
    )
    assert "0.3" in _detail_component_text(
        exporter_detail._build_exporter_detail_continent_mix_table(
            exporter_frame,
            "bcm",
            rolling_window_days=30,
        )
    )

    importer_frame = exporter_frame.rename(
        columns={"continent_destination": "continent_origin"}
    )
    importer_figure = importer_detail.create_continent_origin_chart(
        "Test",
        ["Test"],
        rolling_window_days=30,
        volume_metric="bcm",
        selected_years=["2025"],
        continent_df=importer_frame,
    )
    assert importer_figure.layout.yaxis.tickformat == ",.1f"
    assert all(
        "%{y:,.1f}" in trace.hovertemplate
        for trace in importer_figure.data
        if trace.hovertemplate
    )
    assert "0.3" in _detail_component_text(
        importer_detail._build_importer_origin_mix_table(
            importer_frame,
            "bcm",
            rolling_window_days=30,
        )
    )


@pytest.mark.parametrize("module", DETAIL_MODULES)
def test_detail_period_volume_grid_uses_period_days_and_fixed_decimals(module):
    raw = pd.DataFrame({
        "Continent": ["Global"],
        "Country": [""],
        "30D": [100.0],
        "7D": [100.0],
        "Δ 7D-30D": [0.0],
    })
    converted = module._convert_detail_period_display_df(
        raw,
        "bcm",
        rolling_window_days=30,
    )
    assert converted.loc[0, "30D"] == pytest.approx(3.0)
    assert converted.loc[0, "7D"] == pytest.approx(0.7)
    assert pd.isna(converted.loc[0, "Δ 7D-30D"])
    columns = [
        {"name": column, "id": column, "type": "text" if column in {"Continent", "Country"} else "numeric"}
        for column in converted.columns
    ]
    grid_df, _ = module._build_exporter_detail_period_grid_display(
        converted,
        columns,
        volume_metric="bcm",
    )
    assert grid_df.loc[0, "30D"] == "3.0"
    assert grid_df.loc[0, "7D"] == "0.7"
    assert grid_df.loc[0, "Δ 7D-30D"] == "—"

    mcm_grid, _ = module._build_exporter_detail_period_grid_display(
        raw.drop(columns=["Δ 7D-30D"]),
        columns[:-1],
        volume_metric="mcm_d",
    )
    assert mcm_grid.loc[0, "30D"] == "100"
    assert mcm_grid.loc[0, "7D"] == "100"


def test_exporter_detail_seven_day_window_has_no_duplicate_columns():
    dates = pd.date_range("2024-07-01", "2026-07-01", freq="D")
    summary = pd.DataFrame({
        "start_date": dates,
        "cargo_mcm": 10.0,
        "continent": "A",
        "country": "B",
    })
    result = exporter_detail._build_destination_rolling_windows_pivot(
        summary,
        rolling_window_days=7,
        current_date=pd.Timestamp("2026-07-01"),
    )
    assert not result.columns.duplicated().any()
    assert "Δ 7D-7D" not in result.columns
    assert result.loc[0, "7D"] == pytest.approx(10.0)

    combined = exporter_detail.combine_destination_summary_data_hierarchical(
        pd.DataFrame(columns=["continent", "country"]),
        pd.DataFrame(columns=["continent", "country"]),
        pd.DataFrame(columns=["continent", "country"]),
        result,
        rolling_window_days=7,
    )
    assert not combined.columns.duplicated().any()
    assert "7D" in combined.columns
    assert not any(column.endswith(("_x", "_y")) for column in combined.columns)


@pytest.mark.parametrize("module", DETAIL_MODULES)
def test_detail_rolling_window_is_clamped_consistently(module):
    assert module.normalize_rolling_window_days(0) == 30
    assert module.normalize_rolling_window_days(1) == 1
    assert module.normalize_rolling_window_days(180) == 180
    assert module.normalize_rolling_window_days(181) == 180
    assert module._get_detail_rolling_metric_export_column_name(999, "bcm") == (
        "rolling_volume_180d (bcm)"
    )
    assert module._get_detail_volume_metric_factor(
        "bcm",
        period_days=module.normalize_rolling_window_days(999),
    ) == pytest.approx(0.18)


@pytest.mark.parametrize("module,date_column,builder", [
    (exporter_detail, "start_date", exporter_detail._build_chart_total_supply_df),
    (importer_detail, "end_date", importer_detail._build_importer_total_import_df),
])
def test_detail_180_day_first_visible_point_uses_complete_warmup(
    module, date_column, builder
):
    dates = pd.date_range(module.DETAIL_CHART_QUERY_START_DATE, "2021-01-01", freq="D")
    frame = pd.DataFrame({date_column: dates, "cargo_mcm": 100.0})
    result = builder(frame, rolling_window_days=180, current_date=pd.Timestamp("2021-01-01"))
    first_visible = result[result["date"] == pd.Timestamp("2021-01-01")].iloc[0]
    assert first_visible["rolling_avg"] == pytest.approx(100.0)


@pytest.mark.parametrize("module", DETAIL_MODULES)
def test_detail_export_headers_and_excel_formats_match_metric_semantics(module):
    assert module._get_detail_rolling_metric_export_column_name(30, "mcm_d") == (
        "rolling_avg_30d (mcm/d)"
    )
    assert module._get_detail_rolling_metric_export_column_name(30, "bcm") == (
        "rolling_volume_30d (bcm)"
    )
    assert module._get_detail_excel_number_format("mcm_d") == "#,##0"
    assert module._get_detail_excel_number_format("bcm") == "#,##0.0"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["rolling_volume_30d (bcm)", "date"])
    worksheet.append([3.01234, "2026-01-01"])
    module._apply_detail_excel_metric_format(
        worksheet,
        "bcm",
        metric_headers={"rolling_volume_30d (bcm)"},
    )
    assert worksheet["A2"].value == pytest.approx(3.01234)
    assert worksheet["A2"].number_format == "#,##0.0"
    assert worksheet["B2"].number_format == "General"


@pytest.mark.parametrize("module,table_builder,text_columns", [
    (
        exporter_detail,
        exporter_detail.create_destination_forecast_summary_table,
        {"Continent": "Global", "Country": ""},
    ),
    (
        importer_detail,
        importer_detail.create_origin_forecast_summary_table,
        {"Continent": "GRAND TOTAL", "Country": ""},
    ),
])
def test_detail_forecast_tables_declare_fixed_metric_precision(
    module, table_builder, text_columns
):
    frame = pd.DataFrame([
        {**text_columns, "Jan'26": 3.01234},
        {**text_columns, "Jan'26": 0.039839},
    ])
    bcm_table = table_builder(frame, "bcm")
    mcm_table = table_builder(frame, "mcm_d")
    bcm_column = next(
        column
        for column in bcm_table.columnDefs
        if column.get("field") == "Jan'26"
    )
    mcm_column = next(
        column
        for column in mcm_table.columnDefs
        if column.get("field") == "Jan'26"
    )
    assert bcm_table.rowData[0]["Jan'26"] == "3.0"
    assert bcm_table.rowData[1]["Jan'26"] == "0.0"
    assert mcm_table.rowData[0]["Jan'26"] == "3"
    assert mcm_table.rowData[1]["Jan'26"] == "0"
    assert bcm_column["sortable"] is False
    assert mcm_column["sortable"] is False
    assert "mckinsey-ag-grid-number-cell" in bcm_column["cellClass"]


def test_detail_maintenance_tables_render_fixed_metric_precision():
    exporter_spec = exporter_detail._build_maintenance_period_specs()[0]
    exporter_raw_field = exporter_detail._maintenance_raw_mcmd_field(
        exporter_spec["id"]
    )
    exporter_frame = pd.DataFrame({
        "Plant": ["Plant"],
        "Train": ["Train"],
        "Capacity": [1.2],
        "Type": ["train"],
        exporter_spec["id"]: [0.2],
        exporter_raw_field: [5.1],
    })
    exporter_table = exporter_detail.create_maintenance_summary_table(
        exporter_frame,
        volume_metric="bcm",
    )
    exporter_grid = next(
        component
        for component in _walk_components(exporter_table)
        if getattr(component, "id", None)
        == {"type": "maintenance-expandable-table", "index": 0}
    )
    exporter_column = next(
        column
        for column in exporter_grid.columnDefs
        if column.get("field") == exporter_spec["id"]
    )
    assert exporter_grid.rowData[0][exporter_spec["id"]] == "0.2"
    assert exporter_grid.rowData[0][exporter_raw_field] == pytest.approx(5.1)
    assert exporter_raw_field in json.dumps(exporter_column["cellClassRules"])
    assert exporter_raw_field in json.dumps(exporter_column["cellStyle"])

    importer_spec = importer_detail._build_importer_maintenance_period_specs()[0]
    importer_raw_field = importer_detail._maintenance_raw_mcmd_field(
        importer_spec["id"]
    )
    importer_frame = pd.DataFrame({
        "Supplier Country": ["Country"],
        "Plant": ["Plant"],
        "Train": ["Train"],
        "Type": ["train"],
        "PlantKey": ["Country::Plant"],
        importer_spec["id"]: [0.2],
        importer_raw_field: [5.1],
    })
    importer_table = importer_detail.create_maintenance_summary_table(
        importer_frame,
        volume_metric="mcm_d",
    )
    importer_grid = next(
        component
        for component in _walk_components(importer_table)
        if getattr(component, "id", None)
        == {"type": "imp-maintenance-expandable-table", "index": 0}
    )
    importer_column = next(
        column
        for column in importer_grid.columnDefs
        if column.get("field") == importer_spec["id"]
    )
    assert importer_grid.rowData[0][importer_spec["id"]] == "0"
    assert importer_grid.rowData[0][importer_raw_field] == pytest.approx(5.1)
    assert importer_raw_field in json.dumps(importer_column["cellClassRules"])
    assert importer_raw_field in json.dumps(importer_column["cellStyle"])


def test_importer_forecast_mismatch_is_calculated_before_bcm_display_rounding():
    source = {
        "run_metadata": {"run_id": "test"},
        "allocation_df": pd.DataFrame({
            "date": pd.to_datetime(["2027-01-01"]),
            "continent": ["Asia"],
            "country": ["Qatar"],
            "allocated_volume_bcm": [1.051],
        }),
        "internal_allocation_df": pd.DataFrame(),
        "demand_totals_df": pd.DataFrame({
            "date": pd.to_datetime(["2027-01-01"]),
            "forecast_demand_bcm": [1.0],
        }),
        "mapping_df": pd.DataFrame(),
    }
    _, footer_rows, _ = importer_detail._build_origin_forecast_summary_from_source(
        source,
        current_date=pd.Timestamp("2026-01-01"),
        origin_level="continent_origin_name",
    )
    mismatch = pd.DataFrame([footer_rows[-1]])
    converted = importer_detail._convert_detail_period_display_df(
        mismatch,
        "bcm",
        exclude_columns=["Continent", "Country"],
    )
    assert converted.loc[0, "2027 Avg"] == pytest.approx(0.051)
    assert importer_detail._format_detail_table_metric_value(
        converted.loc[0, "2027 Avg"],
        "bcm",
    ) == "0.1"


# Consolidated from test_detail_snapshot_precompute.py.

import copy
import os
from pathlib import Path
import subprocess
import sys

import pandas as pd
import pytest
from sqlalchemy import create_engine, text

from pages import exporter_detail, importer_detail
from utils import dashboard_snapshot_cache as snapshots
from utils import detail_snapshot_precompute as precompute


@pytest.fixture
def persistent_precompute_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "detail-precompute-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _source_watermark():
    return {
        "kpler_watermark": "2026-07-27T00:00:00Z",
        "woodmac_watermark": "2026-07-26T00:00:00Z",
        "distance_watermark": "2026-07-25T00:00:00Z",
        "mapping_fingerprint": "mapping-v1",
    }


def _maintenance_version():
    return {
        "planned_watermark": "2026-07-26T00:00:00Z",
        "planned_row_count": 4,
        "unplanned_watermark": "2026-07-26T00:00:00Z",
        "unplanned_row_count": 3,
        "train_capacity_watermark": "2026-07-26T00:00:00Z",
        "train_capacity_row_count": 2,
        "plant_capacity_watermark": "2026-07-26T00:00:00Z",
        "plant_capacity_row_count": 1,
        "mapping_fingerprint": "mapping-v1",
    }


def _allocation_run():
    return {
        "run_id": "allocation-run-v1",
        "analysis_date": pd.Timestamp("2026-07-27T00:00:00"),
        "forecast_start": pd.Timestamp("2026-07-01"),
        "forecast_end": pd.Timestamp("2028-12-01"),
    }


def _importer_catalog():
    return [
        {
            "destination_country_name": "China",
            "country": "China",
            "country_display": "China",
            "continent": "Asia",
            "subcontinent": "East Asia",
            "basin": "Pacific",
            "country_classification_level1": "Asia",
            "country_classification": "Importer",
            "shipping_region": "North Asia",
        }
    ]


def _payload(section, selection):
    return {
        "section": section,
        "selection": selection,
        "frame": pd.DataFrame(
            {
                "date": pd.to_datetime(
                    ["2026-01-01", "2026-02-01", None]
                ),
                "value": [1.25, None, 3.5],
            }
        ),
    }


def _patch_exporter_sources(monkeypatch, *, fail_builders=False):
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_detail_source_watermark",
        lambda: copy.deepcopy(_source_watermark()),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_maintenance_source_version",
        lambda: copy.deepcopy(_maintenance_version()),
    )
    monkeypatch.setattr(
        exporter_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(_allocation_run()),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_exporter_diversion_source_version",
        lambda: {"diversion_watermark": "diversion-v1"},
    )

    def builder(section):
        if fail_builders:
            return lambda *_args, **_kwargs: pytest.fail(
                f"{section} exporter builder should not run"
            )
        return lambda *_args, **_kwargs: _payload(
            section,
            "United States",
        )

    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_detail_base_payload",
        builder("base"),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_fetch_destination_forecast_source_data",
        builder("allocation"),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_maintenance_source_payload",
        builder("maintenance"),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_route_source_payload",
        builder("route"),
    )
    monkeypatch.setattr(
        exporter_detail,
        "_build_exporter_diversion_payload",
        builder("diversion"),
    )


def _patch_importer_sources(monkeypatch, *, fail_builders=False):
    monkeypatch.setattr(
        importer_detail,
        "build_destination_catalog",
        lambda _engine: copy.deepcopy(_importer_catalog()),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_detail_source_watermark",
        lambda: copy.deepcopy(_source_watermark()),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_maintenance_source_version",
        lambda: copy.deepcopy(_maintenance_version()),
    )
    monkeypatch.setattr(
        importer_detail,
        "fetch_latest_supply_allocation_run_metadata",
        lambda _engine: copy.deepcopy(_allocation_run()),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_importer_diversion_source_version",
        lambda: {"diversion_watermark": "diversion-v1"},
    )

    def builder(section):
        if fail_builders:
            return lambda *_args, **_kwargs: pytest.fail(
                f"{section} importer builder should not run"
            )
        return lambda *_args, **_kwargs: _payload(section, "China")

    monkeypatch.setattr(
        importer_detail,
        "_build_import_analysis_base_payload",
        builder("base"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_fetch_origin_forecast_source_data",
        builder("allocation"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_maintenance_source_payload",
        builder("maintenance"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_route_source_payload",
        builder("route"),
    )
    monkeypatch.setattr(
        importer_detail,
        "_build_importer_diversion_payload",
        builder("diversion"),
    )


def _summary_references(summary):
    return {
        (item["page"], item["section"]): (
            item["namespace"],
            item["source_key"],
            item["revision"],
        )
        for item in summary["snapshots"]
    }


def _reference_identity(reference):
    return (
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )


def test_precompute_survives_cache_reopen_and_normal_loaders_do_not_rebuild(
    monkeypatch,
    persistent_precompute_cache,
):
    _patch_exporter_sources(monkeypatch)
    _patch_importer_sources(monkeypatch)

    summary = precompute.precompute_detail_snapshots(
        exporter_detail,
        importer_detail,
        exporter_countries=["United States"],
        importer_targets=[("Country", "China")],
    )

    assert summary["status"] == "ready"
    assert len(summary["snapshots"]) == 10
    assert summary["sql_audit"]["rejected_statement_count"] == 0
    assert summary["cache"]["volume_bytes"] > 0
    expected = _summary_references(summary)

    # Simulate a fresh Dash worker: only persistent records may satisfy these
    # normal page-loader calls.
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    _patch_exporter_sources(monkeypatch, fail_builders=True)
    _patch_importer_sources(monkeypatch, fail_builders=True)

    exporter_context = precompute.build_exporter_source_context(
        exporter_detail
    )
    importer_context = precompute.build_importer_source_context(
        importer_detail
    )
    catalog = _importer_catalog()

    exporter_refs = {
        "base": exporter_detail.refresh_exporter_detail_base_data(
            "United States",
            source_context=exporter_context,
        ),
        "allocation": exporter_detail.refresh_destination_forecast_source(
            "United States",
            exporter_context,
        ),
        "maintenance": (
            exporter_detail._get_exporter_maintenance_source_reference(
                "United States",
                exporter_context,
            )[0]
        ),
        "route": exporter_detail.refresh_exporter_route_analysis_source(
            "United States",
            source_context=exporter_context,
        ),
        "diversion": exporter_detail.refresh_exporter_diversion_source(
            exporter_context,
            "United States",
        ),
    }
    importer_refs = {
        "base": importer_detail.refresh_import_analysis_base_data(
            0,
            "country",
            "China",
            catalog,
            source_context=importer_context,
        ),
        "allocation": importer_detail.refresh_origin_forecast_source(
            importer_context,
            "country",
            "China",
            catalog,
        ),
        "maintenance": (
            importer_detail._get_importer_maintenance_source_reference(
                ["China"],
                importer_context,
            )[0]
        ),
        "route": importer_detail.refresh_importer_route_analysis_source(
            "country",
            "China",
            catalog,
            importer_context,
        ),
        "diversion": importer_detail.refresh_importer_diversion_source(
            importer_context,
            "country",
            "China",
            catalog,
        ),
    }

    for section, reference in exporter_refs.items():
        assert _reference_identity(reference) == expected[
            ("exporter", section)
        ]
        payload = snapshots.resolve_snapshot(
            reference,
            exporter_detail.engine,
        )
        manifest = snapshots.resolve_snapshot_manifest(
            reference,
            exporter_detail.engine,
        )
        assert isinstance(payload["frame"], pd.DataFrame)
        assert manifest

    for section, reference in importer_refs.items():
        assert _reference_identity(reference) == expected[
            ("importer", section)
        ]
        payload = snapshots.resolve_snapshot(
            reference,
            importer_detail.engine,
        )
        manifest = snapshots.resolve_snapshot_manifest(
            reference,
            importer_detail.engine,
        )
        assert isinstance(payload["frame"], pd.DataFrame)
        assert manifest


def test_precompute_fails_closed_when_local_persistence_is_disabled(
    monkeypatch,
):
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "0")

    with pytest.raises(
        precompute.DetailSnapshotPrecomputeError,
        match="legacy Postgres backend",
    ):
        precompute.precompute_detail_snapshots(
            exporter_detail,
            importer_detail,
            exporter_countries=[],
            importer_targets=[],
        )


def test_sql_audit_allows_reads_and_rejects_ddl():
    engine = create_engine("sqlite:///:memory:")
    try:
        with precompute.audit_read_only_sql([engine]) as audit:
            with engine.connect() as connection:
                assert connection.execute(text("SELECT 1")).scalar() == 1
                with pytest.raises(
                    precompute.DetailSnapshotPrecomputeError,
                    match="CREATE",
                ):
                    connection.execute(
                        text("CREATE TABLE forbidden (id INTEGER)")
                    )
        assert audit.statement_count == 2
        assert audit.rejected_statement_count == 1
    finally:
        engine.dispose()


def test_cli_help_does_not_import_dashboard_pages():
    repository_root = Path(__file__).resolve().parents[1]
    result = subprocess.run(
        [
            sys.executable,
            str(
                repository_root
                / "scripts"
                / "precompute_detail_snapshots.py"
            ),
            "--help",
        ],
        cwd=repository_root,
        env={
            **os.environ,
            "PYTHONDONTWRITEBYTECODE": "1",
        },
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0
    assert "country=China" in result.stdout


# Consolidated from test_ea_run_interface.py.

import pytest

from utils import ea_run_interface
from utils import export_flow_data
from utils import import_flow_data
from utils import market_balance_data
from utils import provider_flow_snapshot
from utils import snapshot_controls


@pytest.mark.parametrize(
    "value",
    [None, True, False, 0, -1, 1.0, "1", "2026-07-16T00:00:00"],
)
def test_ea_run_id_rejects_implicit_coercions(value):
    with pytest.raises(ValueError, match="positive integer"):
        ea_run_interface.normalize_ea_run_id(value)


def test_ea_run_dropdown_uses_integer_values_and_disambiguates_timestamps():
    runs = [
        {
            "run_id": 12,
            "snapshot_at": "2026-07-16T10:00:00Z",
            "change_count": 1,
            "delete_count": 0,
        },
        {
            "run_id": 11,
            "snapshot_at": "2026-07-16T10:00:00Z",
            "change_count": 2,
            "delete_count": 1,
        },
        {
            "run_id": 10,
            "snapshot_at": "2026-07-15T10:00:00Z",
            "change_count": 3,
            "delete_count": 0,
        },
    ]

    options = snapshot_controls.build_ea_upload_dropdown_options(runs)

    assert [option["value"] for option in options] == [12, 11, 10]
    assert options[0]["label"].endswith("run 12")
    assert options[1]["label"].endswith("run 11")
    assert options[2]["label"] == "2026-07-15 10:00"


def test_hot_browser_timestamp_state_resets_to_second_changed_run():
    comparison_options = {
        "woodmac": {"short_term": [], "long_term": []},
        "ea_comparison_runs": [
            {"run_id": 12, "snapshot_at": "2026-07-16T10:00:00Z"},
            {"run_id": 11, "snapshot_at": "2026-07-15T10:00:00Z"},
        ],
    }

    resolved = snapshot_controls.resolve_snapshot_control_values(
        "ea",
        comparison_options,
        None,
        None,
        "2026-07-15T10:00:00Z",
    )

    assert resolved[5] == 11
    assert isinstance(resolved[4][0]["value"], int)


def test_flow_queries_use_only_public_ea_interfaces_and_bound_function_inputs():
    current_queries = (
        export_flow_data._build_ea_export_flow_query(),
        import_flow_data._build_ea_import_flow_query(),
    )
    historical_queries = (
        export_flow_data._build_ea_parameterized_export_flow_query(),
        import_flow_data._build_ea_parameterized_import_flow_query(),
    )

    for query in current_queries:
        assert ".ea_values_current" in query
        assert ".ea_values " not in query
    for query in historical_queries:
        assert "ea_values_at_run" in query
        assert ":ea_as_of_run_id" in query
        assert ":ea_start_date" in query
        assert ":ea_end_date" in query
        assert "array_agg(" in query
        assert "HAVING count(*) > 0" in query
        assert "CROSS JOIN LATERAL" in query
        assert ".ea_values " not in query


def test_current_metadata_is_not_derived_from_changed_comparison_runs():
    metadata = snapshot_controls.ea_metadata_from_upload_options(
        {"run_id": 20, "snapshot_at": "2026-07-16T12:00:00Z"}
    )

    assert metadata == {
        "run_id": 20,
        "snapshot_at": "2026-07-16T12:00:00Z",
        "upload_timestamp_utc": "2026-07-16T12:00:00Z",
    }


def test_new_provider_snapshot_fails_closed_if_sources_change(monkeypatch):
    captured = {
        "current_ea": {"run_id": 20, "snapshot_at": "2026-07-16T12:00:00Z"},
        "mapping_hash": "before",
    }
    monkeypatch.setattr(
        provider_flow_snapshot,
        "get_or_build_snapshot",
        lambda *args, **kwargs: (
            "reference",
            kwargs["builder"](),
        ),
    )
    monkeypatch.setattr(
        provider_flow_snapshot,
        "fetch_provider_flow_source_state",
        lambda: {**captured, "mapping_hash": "after"},
    )

    with pytest.raises(RuntimeError, match="captured page state"):
        provider_flow_snapshot.get_provider_flow_snapshot_for_state(captured)


def test_provider_snapshot_hit_reuses_captured_state_without_requery(
    monkeypatch,
):
    captured = {
        "current_ea": {
            "run_id": 20,
            "snapshot_at": "2026-07-16T12:00:00Z",
        },
        "mapping_hash": "before",
    }
    expected = (
        "reference",
        {"current_ea": captured["current_ea"]},
    )
    monkeypatch.setattr(
        provider_flow_snapshot,
        "get_snapshot_if_available",
        lambda *args, **kwargs: expected,
    )
    source_state_calls = 0

    def fail_if_called():
        nonlocal source_state_calls
        source_state_calls += 1
        raise AssertionError("cache hit must not requery source state")

    monkeypatch.setattr(
        provider_flow_snapshot,
        "fetch_provider_flow_source_state",
        fail_if_called,
    )

    assert (
        provider_flow_snapshot.get_provider_flow_snapshot_for_state(captured)
        == expected
    )
    assert source_state_calls == 0


def test_provider_source_state_hashes_effective_ea_catalog_and_selection(monkeypatch):
    executed_sql = []

    class Result:
        def scalar(self):
            return "mapping-revision"

    class Connection:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def execute(self, statement):
            executed_sql.append(str(statement))
            return Result()

    class Engine:
        def connect(self):
            return Connection()

    monkeypatch.setattr(provider_flow_snapshot, "engine", Engine())
    monkeypatch.setattr(
        provider_flow_snapshot,
        "build_resolved_ea_lng_balance_ctes",
        lambda *_args: (
            "resolved AS (SELECT '1'::text AS dataset_id, "
            "''::text AS country, ''::text AS country_iso, ''::text AS region, "
            "''::text AS sub_region, ''::text AS description, ''::text AS aspect, "
            "''::text AS aspect_subtype, ''::text AS category, "
            "''::text AS category_subtype, ''::text AS frequency, "
            "''::text AS lifecycle_stage, ''::text AS source, ''::text AS unit)",
            "resolved",
        ),
    )

    assert provider_flow_snapshot._fetch_ea_balance_mapping_hash() == "mapping-revision"
    rendered = executed_sql[-1]
    for field in ("dataset_id", "country", "aspect", "category_subtype", "frequency", "unit"):
        assert field in rendered


def test_provider_source_key_carries_the_effective_ea_mapping_hash(monkeypatch):
    class Result:
        def scalar(self):
            return "country-revision"

    class Connection:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def execute(self, _statement):
            return Result()

    class Engine:
        def connect(self):
            return Connection()

    monkeypatch.setattr(provider_flow_snapshot, "engine", Engine())
    monkeypatch.setattr(
        provider_flow_snapshot, "_SOURCE_STATE_QUERIES", {"mapping_hash": object()}
    )
    monkeypatch.setattr(
        provider_flow_snapshot, "_fetch_ea_balance_mapping_hash", lambda: "ea-revision"
    )
    monkeypatch.setattr(
        provider_flow_snapshot,
        "fetch_current_ea_run",
        lambda *_args, **_kwargs: {"run_id": 42, "snapshot_at": "2026-07-16T00:00:00Z"},
    )

    state = provider_flow_snapshot.fetch_provider_flow_source_state()

    assert state["ea_balance_mapping_hash"] == "ea-revision"
    assert state["current_ea"]["run_id"] == 42


def test_pinned_market_payload_does_not_fallback_to_moving_latest(monkeypatch):
    captured = {"current_ea": {"run_id": 20}}
    monkeypatch.setattr(
        market_balance_data,
        "get_provider_flow_snapshot_for_state",
        lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("state changed")),
    )
    fallback_called = False

    def fallback():
        nonlocal fallback_called
        fallback_called = True
        return {}, {}

    monkeypatch.setattr(market_balance_data, "build_provider_flow_payload", fallback)
    with pytest.raises(RuntimeError, match="state changed"):
        market_balance_data._resolve_latest_provider_flow_payload(
            source_state=captured
        )
    assert fallback_called is False


# Consolidated from test_global_supply_comparison.py.

import pandas as pd

from utils import global_supply_comparison as comparison


def test_energy_aspects_query_uses_bounded_current_state_view():
    current_query = str(comparison.EA_GLOBAL_SUPPLY_CURRENT_QUERY)

    assert "at_lng.ea_values_current source" in current_query
    assert "at_lng.ea_values source" not in current_query
    assert "MAX(upload_timestamp_utc)" not in current_query
    assert "source.dataset_id = :dataset_id" in current_query
    assert "source.date >= CAST(:start_month AS timestamp)" in current_query
    assert "source.date < CAST(:end_month AS timestamp) + INTERVAL '1 month'" in current_query


def test_comparison_bounds_use_five_complete_trailing_years():
    start, forecast_start, end = comparison.comparison_month_bounds("2026-07-16")

    assert start == pd.Timestamp("2021-07-01")
    assert forecast_start == pd.Timestamp("2026-07-01")
    assert end == pd.Timestamp("2031-12-01")


def test_comparison_keeps_real_source_horizons_and_builds_platts_2031(monkeypatch):
    monkeypatch.setattr(
        comparison,
        "_fetch_ramp",
        lambda *args: pd.DataFrame(
            {"month": [pd.Timestamp("2026-07-01")], "monthly_mt": [40.0]}
        ),
    )
    monkeypatch.setattr(
        comparison,
        "_fetch_ea",
        lambda *args: pd.DataFrame(
            {
                "month": [pd.Timestamp("2029-12-01")],
                "monthly_mt": [50.0],
                "upload_timestamp_utc": [pd.Timestamp("2026-07-16")],
            }
        ),
    )
    monkeypatch.setattr(
        comparison,
        "_fetch_platts",
        lambda *args: (
            pd.DataFrame(
                {
                    "month": [pd.Timestamp("2030-12-01")],
                    "monthly_mt": [52.0],
                    "upload_timestamp_utc": [pd.Timestamp("2026-05-05")],
                    "vintage_date": [pd.Timestamp("2026-04-01")],
                }
            ),
            pd.DataFrame(
                {
                    "annual_mt": [660.0],
                    "vintage_date": [pd.Timestamp("2026-01-01")],
                }
            ),
            pd.DataFrame(
                {
                    "month": pd.date_range("2031-01-01", "2031-12-01", freq="MS"),
                    "monthly_mt": [55.0] * 12,
                }
            ),
        ),
    )
    monkeypatch.setattr(
        comparison,
        "_fetch_woodmac",
        lambda *args: (
            pd.DataFrame(
                {"month": [pd.Timestamp("2031-12-01")], "monthly_mt": [56.0]}
            ),
            {
                "short_term_market_outlook": "Short",
                "long_term_market_outlook": "Long",
            },
        ),
    )

    result, metadata = comparison.fetch_global_supply_comparison(
        object(), 41, as_of_month="2026-07-01"
    )

    ea = result[result["source"].eq("Energy Aspects")]
    platts = result[result["source"].eq("Platts")]
    ramp = result[result["source"].eq("Our ramp forecast")]
    assert ea["month"].max() == pd.Timestamp("2029-12-01")
    assert ramp["month"].max() == pd.Timestamp("2026-07-01")
    assert len(platts[platts["month"].dt.year.eq(2031)]) == 12
    assert set(platts.loc[platts["month"].dt.year.eq(2031), "monthly_mt"]) == {55.0}
    assert metadata["sources"]["Platts"]["last_month"] == "Dec 2031"
    assert metadata["warnings"] == []


def test_time_views_sum_only_complete_periods():
    monthly = pd.DataFrame(
        {
            "month": pd.to_datetime(
                [
                    "2027-01-01",
                    "2027-02-01",
                    "2027-03-01",
                    "2027-04-01",
                    "2027-05-01",
                ]
            ),
            "source": ["WoodMac"] * 5,
            "monthly_mt": [1.0, 2.0, 3.0, 4.0, 5.0],
        }
    )

    quarterly = comparison.aggregate_global_supply_comparison(monthly, "quarter")
    yearly = comparison.aggregate_global_supply_comparison(monthly, "year")

    assert quarterly[["period_label", "supply_mt"]].to_dict("records") == [
        {"period_label": "2027 Q1", "supply_mt": 6.0}
    ]
    assert yearly.empty


def test_lng_season_view_uses_april_to_september_and_october_to_march():
    monthly = pd.DataFrame(
        {
            "month": pd.date_range("2026-10-01", "2027-03-01", freq="MS"),
            "source": ["Our ramp forecast"] * 6,
            "monthly_mt": [2.0] * 6,
        }
    )

    seasonal = comparison.aggregate_global_supply_comparison(monthly, "season")

    assert seasonal[["period_label", "supply_mt"]].to_dict("records") == [
        {"period_label": "Winter 2026/27", "supply_mt": 12.0}
    ]


# Consolidated from test_historical_comparison_snapshot.py.

from concurrent.futures import ThreadPoolExecutor
import threading
import time

import pandas as pd
import pytest

from utils import dashboard_snapshot_cache as snapshots
from utils import historical_comparison_snapshot as comparisons


@pytest.fixture
def historical_cache(monkeypatch, tmp_path):
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(tmp_path / "historical-cache"),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    comparisons.clear_historical_comparison_source_state()
    yield
    comparisons.clear_historical_comparison_source_state()
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _reference(revision=None):
    return {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": "provider-flow-source-v2",
        "source_key": "provider-source",
        "revision": revision or snapshots._new_local_revision_token(),
        "shared": True,
        "slot": "woodmac_export",
    }


def test_warm_comparison_reuses_exact_frame_and_mapping_state(
    monkeypatch,
    historical_cache,
):
    manifest_calls = 0
    live_mapping_calls = 0
    build_calls = 0

    def manifest_state(*_args, **_kwargs):
        nonlocal manifest_calls
        manifest_calls += 1
        return {
            "source_state": {
                "mapping_hash": "country-v1",
                "ea_balance_mapping_hash": "ea-v1",
            }
        }

    def live_mapping_state():
        nonlocal live_mapping_calls
        live_mapping_calls += 1
        return {
            "mapping_hash": "country-v1",
            "ea_balance_mapping_hash": "ea-v1",
        }

    def builder():
        nonlocal build_calls
        build_calls += 1
        return pd.DataFrame(
            {
                "month": pd.to_datetime(["2026-01-01"]),
                "value": [12.5],
            }
        )

    monkeypatch.setattr(
        comparisons,
        "resolve_snapshot_manifest",
        manifest_state,
    )
    monkeypatch.setattr(
        comparisons,
        "fetch_provider_flow_mapping_state",
        live_mapping_state,
    )
    base_reference = _reference()
    first_reference, first_frame = comparisons.get_historical_comparison_frame(
        direction="supply",
        base_reference=base_reference,
        selection={"source": "ea", "run_id": 42},
        query_dependencies={
            "start_date": "2026-01-01",
            "end_date": "2026-12-31",
        },
        builder=builder,
    )
    second_reference, second_frame = comparisons.get_historical_comparison_frame(
        direction="supply",
        base_reference=base_reference,
        selection={"source": "ea", "run_id": 42},
        query_dependencies={
            "start_date": "2026-01-01",
            "end_date": "2026-12-31",
        },
        builder=builder,
    )

    assert first_reference == second_reference
    pd.testing.assert_frame_equal(first_frame, second_frame, check_exact=True)
    assert build_calls == 1
    assert manifest_calls == 2
    assert live_mapping_calls == 1


def test_base_revision_and_exact_selection_partition_snapshots(
    monkeypatch,
    historical_cache,
):
    monkeypatch.setattr(
        comparisons,
        "resolve_snapshot_manifest",
        lambda *_args, **_kwargs: {
            "source_state": {
                "mapping_hash": "country-v1",
                "ea_balance_mapping_hash": "ea-v1",
            }
        },
    )
    monkeypatch.setattr(
        comparisons,
        "fetch_provider_flow_mapping_state",
        lambda: {
            "mapping_hash": "country-v1",
            "ea_balance_mapping_hash": "ea-v1",
        },
    )
    build_calls = 0

    def builder():
        nonlocal build_calls
        build_calls += 1
        return pd.DataFrame({"value": [build_calls]})

    first_base = _reference()
    second_base = _reference()
    references = []
    for base_reference, run_id in (
        (first_base, 41),
        (first_base, 42),
        (second_base, 42),
    ):
        reference, _ = comparisons.get_historical_comparison_frame(
            direction="demand",
            base_reference=base_reference,
            selection={"source": "ea", "run_id": run_id},
            query_dependencies={"start_date": "2026-01-01"},
            builder=builder,
        )
        references.append(reference)

    assert build_calls == 3
    assert len({reference["source_key"] for reference in references}) == 3


@pytest.mark.parametrize("worker_count", (1, 4, 8))
def test_comparison_build_is_single_flight(
    monkeypatch,
    historical_cache,
    worker_count,
):
    monkeypatch.setattr(
        comparisons,
        "resolve_snapshot_manifest",
        lambda *_args, **_kwargs: {
            "source_state": {
                "mapping_hash": "country-v1",
                "ea_balance_mapping_hash": "ea-v1",
            }
        },
    )
    monkeypatch.setattr(
        comparisons,
        "fetch_provider_flow_mapping_state",
        lambda: {
            "mapping_hash": "country-v1",
            "ea_balance_mapping_hash": "ea-v1",
        },
    )
    build_calls = 0
    build_lock = threading.Lock()

    def builder():
        nonlocal build_calls
        with build_lock:
            build_calls += 1
        time.sleep(0.03)
        return pd.DataFrame({"value": [7.0]})

    base_reference = _reference()

    def load():
        return comparisons.get_historical_comparison_frame(
            direction="net-balance",
            base_reference=base_reference,
            selection={
                "source": "woodmac",
                "short_term_publication_timestamp": "2026-07-20",
                "long_term_publication_timestamp": "2026-07-01",
            },
            query_dependencies={
                "country_group": "country",
                "time_group": "yearly",
                "unit": "bcm",
            },
            builder=builder,
        )

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        results = list(executor.map(lambda _index: load(), range(worker_count)))

    assert build_calls == 1
    assert all(result[0] == results[0][0] for result in results)
    assert all(result[1].equals(results[0][1]) for result in results)


def test_changed_mapping_during_build_refuses_publication(
    monkeypatch,
    historical_cache,
):
    monkeypatch.setattr(
        comparisons,
        "resolve_snapshot_manifest",
        lambda *_args, **_kwargs: {
            "source_state": {
                "mapping_hash": "country-v1",
                "ea_balance_mapping_hash": "ea-v1",
            }
        },
    )
    monkeypatch.setattr(
        comparisons,
        "fetch_provider_flow_mapping_state",
        lambda: {
            "mapping_hash": "country-v2",
            "ea_balance_mapping_hash": "ea-v1",
        },
    )

    with pytest.raises(
        RuntimeError,
        match="mappings changed during snapshot construction",
    ):
        comparisons.get_historical_comparison_frame(
            direction="supply",
            base_reference=_reference(),
            selection={"source": "ea", "run_id": 42},
            query_dependencies={"start_date": "2026-01-01"},
            builder=lambda: pd.DataFrame({"value": [1.0]}),
        )


# Consolidated from test_lng_phys_snapshot.py.

from io import BytesIO
from pathlib import Path
from threading import Barrier

import pandas as pd
from openpyxl import load_workbook

from pages import lng_phys_snapshot as page
from utils import lng_phys_snapshot_data as data
from utils.ag_grid_tables import iter_leaf_column_defs


REPO = Path(__file__).resolve().parents[1]


def _monthly_frame(
    provider,
    country_name="China",
    year=2026,
    months=12,
    monthly_mmt=1.0,
    release_type="Forecast",
):
    return pd.DataFrame(
        {
            "country_name": [country_name] * months,
            "provider": [provider] * months,
            "month": pd.date_range(
                f"{year}-01-01", periods=months, freq="MS"
            ),
            "monthly_mmt": [monthly_mmt] * months,
            "source_vintage": ["test vintage"] * months,
            "source_upload": [pd.Timestamp("2026-07-23")] * months,
            "release_type": [release_type] * months,
        }
    )


def _walk(component):
    yield component
    children = getattr(component, "children", None)
    if children is None:
        return
    if not isinstance(children, (list, tuple)):
        children = [children]
    for child in children:
        if hasattr(child, "to_plotly_json"):
            yield from _walk(child)


def test_annualization_requires_exactly_twelve_numeric_months():
    complete = _monthly_frame("Energy Aspects", months=12)
    incomplete = _monthly_frame(
        "Platts",
        country_name="Japan",
        months=11,
    )

    result = data.annualize_monthly_demand(
        pd.concat([complete, incomplete], ignore_index=True)
    )

    ea = result[result["provider"].eq("Energy Aspects")].iloc[0]
    platts = result[result["provider"].eq("Platts")].iloc[0]
    assert bool(ea["complete_year"])
    assert ea["annual_mmt"] == 12.0
    assert not bool(platts["complete_year"])
    assert pd.isna(platts["annual_mmt"])


def test_provider_failures_are_isolated(monkeypatch):
    monkeypatch.setattr(
        data,
        "fetch_ea_monthly",
        lambda _engine: _monthly_frame("Energy Aspects"),
    )
    monkeypatch.setattr(
        data,
        "fetch_woodmac_monthly",
        lambda _engine: (_ for _ in ()).throw(RuntimeError("unavailable")),
    )
    monkeypatch.setattr(
        data,
        "fetch_platts_monthly",
        lambda _engine: _monthly_frame("Platts"),
    )

    result, warnings = data.fetch_demand_snapshot(object())

    assert set(result["provider"]) == {"Energy Aspects", "Platts"}
    assert warnings == ["WoodMac demand is unavailable."]


def test_provider_loaders_overlap_with_exact_deterministic_output(monkeypatch):
    loader_barrier = Barrier(3, timeout=2)
    provider_frames = {
        "Energy Aspects": _monthly_frame(
            "Energy Aspects",
            country_name="China",
        ),
        "WoodMac": _monthly_frame(
            "WoodMac",
            country_name="Japan",
        ),
        "Platts": _monthly_frame(
            "Platts",
            country_name="South Korea",
        ),
    }

    def load(provider):
        loader_barrier.wait()
        return provider_frames[provider]

    monkeypatch.setattr(
        data,
        "fetch_ea_monthly",
        lambda _engine: load("Energy Aspects"),
    )
    monkeypatch.setattr(
        data,
        "fetch_woodmac_monthly",
        lambda _engine: load("WoodMac"),
    )
    monkeypatch.setattr(
        data,
        "fetch_platts_monthly",
        lambda _engine: load("Platts"),
    )

    actual, warnings = data.fetch_demand_snapshot(object())
    expected = data.annualize_monthly_demand(
        pd.concat(
            [
                provider_frames[provider]
                for provider in data.PROVIDER_ORDER
            ],
            ignore_index=True,
        )
    )

    pd.testing.assert_frame_equal(actual, expected, check_exact=True)
    assert warnings == []


def test_provider_warning_order_follows_provider_order(monkeypatch):
    monkeypatch.setattr(
        data,
        "fetch_ea_monthly",
        lambda _engine: (_ for _ in ()).throw(RuntimeError("ea failed")),
    )
    monkeypatch.setattr(
        data,
        "fetch_woodmac_monthly",
        lambda _engine: data.empty_monthly_frame(),
    )
    monkeypatch.setattr(
        data,
        "fetch_platts_monthly",
        lambda _engine: None,
    )

    result, warnings = data.fetch_demand_snapshot(object())

    assert result.empty
    assert result.columns.tolist() == list(data.DEMAND_COLUMNS)
    assert warnings == [
        "Energy Aspects demand is unavailable.",
        "WoodMac returned no mapped demand data.",
        "Platts returned no mapped demand data.",
    ]


def test_ea_query_is_pinned_to_latest_accepted_full_run(monkeypatch):
    captured = {}
    monkeypatch.setattr(
        data,
        "fetch_current_ea_run",
        lambda _engine, schema: {
            "run_id": 321,
            "snapshot_at": "2026-07-23T05:32:00Z",
        },
    )

    def fake_read_sql(query, _engine, params):
        captured["query"] = str(query)
        captured["params"] = params
        return data.empty_monthly_frame()

    monkeypatch.setattr(data, "_read_sql", fake_read_sql)

    data.fetch_ea_monthly(object())

    assert captured["params"]["ea_as_of_run_id"] == 321
    assert captured["params"]["ea_source_vintage"] == "2026-07-23T05:32:00Z"
    assert "ea_values_at_run" in captured["query"]
    assert "country_mapping.country = datasets.country" in captured["query"]


def test_provider_queries_use_exact_mapping_and_source_precedence():
    source = (REPO / "utils" / "lng_phys_snapshot_data.py").read_text()
    woodmac_query = str(data.WOODMAC_DEMAND_QUERY)
    platts_query = str(data.PLATTS_DEMAND_QUERY)

    assert source.count("JOIN {DB_SCHEMA}.mappings_country") == 3
    assert "country_mapping.country = datasets.country" in str(
        data.EA_DEMAND_QUERY
    )
    assert "country_mapping.country = source.country_name" in woodmac_query
    assert "country_mapping.country = source.country_or_market" in platts_query
    assert "ROW_NUMBER() OVER" in woodmac_query
    assert "month > short_term_horizon.final_short_term_month" in woodmac_query
    assert "SUM(source.metric_value)::double precision / 12.0" in woodmac_query
    assert "latest_header" in platts_query
    assert "lnga_short_term_demand" in platts_query
    assert "Turkiye" not in source
    assert "China (Mainland)" not in source


def test_platts_query_uses_one_dataset_only_latest_header():
    platts_query = str(data.PLATTS_DEMAND_QUERY)
    header_sql, data_sql = platts_query.split(
        "SELECT\n        country_mapping.country_name",
        maxsplit=1,
    )
    normalized_header = " ".join(header_sql.split())

    assert "WITH latest_header AS MATERIALIZED" in header_sql
    assert (
        "WHERE dataset_key = 'lnga_short_term_demand' "
        "ORDER BY upload_timestamp_utc DESC NULLS LAST, "
        "vintage_date DESC NULLS LAST LIMIT 1"
    ) in normalized_header
    assert "metric =" not in header_sql
    assert "flow_type =" not in header_sql
    assert "unit =" not in header_sql
    assert "latest_upload" not in platts_query
    assert "latest_vintage" not in platts_query
    assert "CROSS JOIN latest_header" in platts_query
    assert "source.metric = 'lng_demand_forecast'" in data_sql
    assert "source.flow_type = 'demand'" in data_sql
    assert "source.unit = 'MMt'" in data_sql
    assert "GROUP BY country_mapping.country_name, source.period_start::date" in data_sql
    assert "ORDER BY country_mapping.country_name, month" in data_sql


def test_platts_loader_preserves_exact_monthly_output(monkeypatch):
    expected = _monthly_frame(
        "Platts",
        country_name="Japan",
        year=2027,
        months=2,
        monthly_mmt=1.25,
        release_type="Short Term Outlook",
    )
    captured = {}

    def fake_read_sql(query, db_engine, params):
        captured["query"] = query
        captured["engine"] = db_engine
        captured["params"] = params
        return expected

    db_engine = object()
    monkeypatch.setattr(data, "_read_sql", fake_read_sql)

    actual = data.fetch_platts_monthly(db_engine)

    pd.testing.assert_frame_equal(actual, expected, check_exact=True)
    assert captured == {
        "query": data.PLATTS_DEMAND_QUERY,
        "engine": db_engine,
        "params": {"countries": list(data.DISPLAY_COUNTRIES)},
    }


def test_demand_matrix_has_21_rows_blanks_lto_and_tooltips():
    annual = data.annualize_monthly_demand(
        pd.concat(
            [
                _monthly_frame("Energy Aspects"),
                _monthly_frame(
                    "WoodMac",
                    year=2029,
                    release_type="Long Term Outlook",
                ),
                _monthly_frame("Platts", year=2030),
            ],
            ignore_index=True,
        )
    )

    rows = data.build_demand_matrix(annual)

    assert len(rows) == 21
    assert rows[0]["Country"] == "China"
    assert rows[0]["Provider"] == "Energy Aspects"
    assert rows[0]["2026E"] == 12.0
    assert rows[0]["2030E"] is None
    assert rows[1]["__2029E_is_lto"] is True
    assert "Long Term Outlook" in rows[1]["__2029E_tooltip"]
    assert sum(bool(row["__country_group_start"]) for row in rows) == 7


def test_demand_cache_is_reused_and_global_refresh_invalidates(monkeypatch):
    calls = []
    data.clear_snapshot_caches()
    monkeypatch.setattr(data, "_cache_bucket", lambda: 10)
    monkeypatch.setattr(
        data,
        "fetch_demand_snapshot",
        lambda _engine: (
            calls.append("load")
            or pd.DataFrame(columns=data.DEMAND_COLUMNS),
            [],
        ),
    )

    data.get_demand_snapshot(0)
    data.get_demand_snapshot(0)
    data.get_demand_snapshot(1)

    assert calls == ["load", "load"]
    data.clear_snapshot_caches()


def test_next_storage_endpoints_and_negative_stockout():
    endpoints = data.next_storage_endpoints("2026-07-23")
    assert [item["date"].isoformat() for item in endpoints] == [
        "2026-10-31",
        "2027-03-31",
        "2027-10-31",
        "2028-03-31",
    ]

    frame = pd.DataFrame(
        {
            "date": [endpoints[-1]["date"]],
            "storage_pct": [-1.3],
            "storage_bcm": [-1.1],
            "upload_timestamp_utc": [pd.Timestamp("2026-07-23")],
        }
    )
    records = data.format_storage_records(frame, endpoints, "base_case")

    assert records[-1]["storage_pct"] == -1.3
    assert records[-1]["storage_bcm"] == -1.1
    assert records[-1]["stockout"]
    card = page.create_storage_card(records[-1])
    assert "phys-snapshot-storage-card-stockout" in card.className
    assert "Modelled stockout" in str(card)
    standard_card = page.create_storage_card(records[0])
    assert "Modelled endpoint" not in str(standard_card)


def test_storage_cache_separates_scenarios_and_refresh(monkeypatch):
    calls = []
    endpoints = data.next_storage_endpoints("2026-07-23")
    data.clear_snapshot_caches()
    monkeypatch.setattr(data, "_cache_bucket", lambda: 10)

    def fake_fetch(scenario, target_dates, _engine):
        calls.append((scenario, target_dates))
        return pd.DataFrame()

    monkeypatch.setattr(data, "fetch_storage_snapshot", fake_fetch)

    data.get_storage_snapshot("base_case", endpoints, 0)
    data.get_storage_snapshot("base_case", endpoints, 0)
    data.get_storage_snapshot("best_view", endpoints, 0)
    data.get_storage_snapshot("base_case", endpoints, 1)

    assert [call[0] for call in calls] == [
        "base_case",
        "best_view",
        "base_case",
    ]
    data.clear_snapshot_caches()


def test_ag_grid_contract_and_export_workbook():
    rows = data.build_demand_matrix(pd.DataFrame(columns=data.DEMAND_COLUMNS))
    rows[0]["2026E"] = 12.34
    grid = page.build_demand_grid(rows)
    leaf_columns = iter_leaf_column_defs(grid.columnDefs)

    assert len(grid.rowData) == 21
    assert grid.dashGridOptions["pagination"] is False
    assert grid.style["height"] == "744px"
    assert [column["field"] for column in leaf_columns] == (
        page.VISIBLE_DEMAND_COLUMNS
    )
    assert leaf_columns[0]["pinned"] == "left"
    assert leaf_columns[1]["pinned"] == "left"
    assert leaf_columns[2].get("pinned") is None
    assert leaf_columns[0]["cellRenderer"]["function"] == (
        "physSnapshotCountryGroupLabel(params)"
    )
    assert set(leaf_columns[0]["cellClassRules"]) == {
        "phys-snapshot-country-label-cell",
        "phys-snapshot-country-continuation-cell",
    }
    assert leaf_columns[2]["valueFormatter"]["function"] == (
        "physSnapshotOneDecimal(params)"
    )
    assert set(grid.rowClassRules) >= {
        "phys-snapshot-provider-ea-row",
        "phys-snapshot-provider-woodmac-row",
        "phys-snapshot-provider-platts-row",
    }
    assert "sortChanged" in grid.eventListeners
    assert "refreshCells" in grid.eventListeners["sortChanged"][0]
    assert grid.columnDefs[0]["headerName"] == "Market"
    assert grid.columnDefs[1]["headerName"] == "Annual LNG Imports (MMT)"

    workbook_bytes = page.build_demand_export_bytes(rows)
    workbook = load_workbook(BytesIO(workbook_bytes))
    worksheet = workbook["LNG Physical Snapshot"]

    assert worksheet.freeze_panes == "C2"
    assert [cell.value for cell in worksheet[1]] == page.VISIBLE_DEMAND_COLUMNS
    assert worksheet["C2"].value == 12.34
    assert worksheet["C2"].number_format == "#,##0.0"
    assert worksheet.max_column == 7


def test_page_uses_global_refresh_and_shipping_route_contract():
    components = {
        getattr(component, "id", None): component
        for component in _walk(page.layout)
        if getattr(component, "id", None)
    }
    assert "phys-snapshot-export-button" in components
    assert "phys-snapshot-storage-scenario" in components
    assert "phys-snapshot-demand-grid" in components
    assert components["phys-snapshot-storage-scenario"].value == "best_view"
    section_classes = [
        str(getattr(component, "className", ""))
        for component in page.layout.children
    ]
    assert next(
        index
        for index, class_name in enumerate(section_classes)
        if "phys-snapshot-storage-section" in class_name
    ) < next(
        index
        for index, class_name in enumerate(section_classes)
        if "phys-snapshot-demand-section" in class_name
    )
    demand_headers = [
        component
        for component in _walk(page.layout)
        if "phys-snapshot-section-header"
        in str(getattr(component, "className", ""))
        and "phys-snapshot-storage-section-header"
        not in str(getattr(component, "className", ""))
    ]
    assert len(demand_headers) == 1
    assert any(
        getattr(component, "id", None)
        == "phys-snapshot-demand-source-metadata"
        for component in _walk(demand_headers[0])
    )

    page_source = (REPO / "pages" / "lng_phys_snapshot.py").read_text()
    route_source = (REPO / "index_shipping_snd.py").read_text()
    styles = (REPO / "assets" / "styles.css").read_text()
    grid_functions = (
        REPO / "assets" / "lng_phys_snapshot_grid.js"
    ).read_text()

    assert page_source.count('Input("global-refresh-button"') == 2
    assert "Next four seasonal endpoints" not in page_source
    assert "phys-snapshot-refresh-button" not in page_source
    assert "import pages.lng_phys_snapshot" in route_source
    assert 'path="/lng-phys-snapshot"' in route_source
    assert "LNG Shipping - LNG Physical Snapshot" in route_source
    assert "nav-lng-phys-snapshot" in route_source
    assert "grid-template-columns: repeat(4" in styles
    assert "grid-template-columns: repeat(2" in styles
    assert "dagfuncs.physSnapshotOneDecimal" in grid_functions
    assert "dagfuncs.physSnapshotCountryGroupLabel" in grid_functions


def test_storage_failure_does_not_affect_demand_callback(monkeypatch):
    monkeypatch.setattr(
        page,
        "get_storage_snapshot",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("storage unavailable")
        ),
    )

    cards, metadata, warning = page.update_storage_snapshot(None, 0)

    assert len(cards) == 4
    assert "Best View" in metadata
    assert warning == "EU storage is unavailable."


# Consolidated from test_market_balance_comparison_concurrency.py.

from threading import Barrier, Event, Lock, enumerate as enumerate_threads

import pandas as pd
import pytest

from utils import market_balance_data


def _comparison_mapping_frame():
    return pd.DataFrame(
        {
            "country_name": ["Qatar", "United Kingdom"],
            "continent": ["Asia", "Europe"],
            "subcontinent": ["Western Asia", "Northern Europe"],
            "basin": ["Pacific", "Atlantic"],
            "shipping_region": ["Middle East", "NWE"],
            "country_classification_level1": ["Exporter", "Importer"],
            "country_classification": ["Exporter", "Importer"],
        }
    )


def _comparison_flow_frame(country, values):
    return pd.DataFrame(
        {
            "month": pd.to_datetime(["2026-01-01", "2026-02-01"]),
            "country_name": [country, country],
            "total_mmtpa": values,
        }
    )


@pytest.mark.parametrize("provider", ("woodmac", "ea"))
def test_comparison_fetches_overlap_and_match_sequential_output(
    monkeypatch,
    provider,
):
    barrier = Barrier(3, timeout=2)
    active = 0
    peak_active = 0
    active_lock = Lock()
    mapping_df = _comparison_mapping_frame()
    export_df = _comparison_flow_frame("Qatar", [12.0, 14.0])
    import_df = _comparison_flow_frame("United Kingdom", [8.0, 9.0])

    def concurrent_result(value):
        nonlocal active, peak_active
        with active_lock:
            active += 1
            peak_active = max(peak_active, active)
        try:
            barrier.wait()
            return value.copy(deep=True)
        finally:
            with active_lock:
                active -= 1

    monkeypatch.setattr(
        market_balance_data,
        "fetch_country_mapping_df",
        lambda: concurrent_result(mapping_df),
    )
    if provider == "woodmac":
        monkeypatch.setattr(
            market_balance_data,
            "fetch_woodmac_export_flow_raw_data_for_publications",
            lambda *_args: concurrent_result(export_df),
        )
        monkeypatch.setattr(
            market_balance_data,
            "fetch_woodmac_import_flow_raw_data_for_publications",
            lambda *_args: concurrent_result(import_df),
        )
        result = market_balance_data.fetch_net_balance_for_woodmac_publications(
            short_term_market_outlook="ST",
            short_term_publication_timestamp="2026-07-20T00:00:00",
            long_term_market_outlook="LT",
            long_term_publication_timestamp="2026-07-01T00:00:00",
            country_group="country",
            time_group="monthly",
            unit="mt",
        )
    else:
        prewarmed = Event()

        def prewarm(*_args):
            prewarmed.set()
            return "", "resolved_lng_balance_datasets"

        def checked_result(value):
            assert prewarmed.is_set()
            return concurrent_result(value)

        monkeypatch.setattr(
            market_balance_data,
            "build_resolved_ea_lng_balance_ctes",
            prewarm,
        )
        monkeypatch.setattr(
            market_balance_data,
            "fetch_country_mapping_df",
            lambda: checked_result(mapping_df),
        )
        monkeypatch.setattr(
            market_balance_data,
            "fetch_ea_export_flow_raw_data_for_upload",
            lambda *_args, **_kwargs: checked_result(export_df),
        )
        monkeypatch.setattr(
            market_balance_data,
            "fetch_ea_import_flow_raw_data_for_upload",
            lambda *_args, **_kwargs: checked_result(import_df),
        )
        result = market_balance_data.fetch_net_balance_for_ea_upload(
            ea_as_of_run_id=42,
            country_group="country",
            time_group="monthly",
            unit="mt",
        )

    expected = market_balance_data._build_provider_net_balance_table(
        export_df,
        import_df,
        mapping_df=mapping_df,
        country_group="country",
        time_group="monthly",
        unit="mt",
    )
    pd.testing.assert_frame_equal(result, expected, check_exact=True)
    assert peak_active == 3
    assert active == 0
    assert not any(
        thread.name.startswith("market-balance-comparison")
        for thread in enumerate_threads()
    )


def test_comparison_preserves_deterministic_failure_order(monkeypatch):
    barrier = Barrier(3, timeout=2)

    def fail(message):
        barrier.wait()
        raise RuntimeError(message)

    monkeypatch.setattr(
        market_balance_data,
        "fetch_country_mapping_df",
        lambda: fail("mapping failed"),
    )
    monkeypatch.setattr(
        market_balance_data,
        "fetch_woodmac_export_flow_raw_data_for_publications",
        lambda *_args: fail("exports failed"),
    )
    monkeypatch.setattr(
        market_balance_data,
        "fetch_woodmac_import_flow_raw_data_for_publications",
        lambda *_args: fail("imports failed"),
    )

    with pytest.raises(RuntimeError, match="mapping failed"):
        market_balance_data.fetch_net_balance_for_woodmac_publications(
            short_term_market_outlook="ST",
            short_term_publication_timestamp="2026-07-20T00:00:00",
            long_term_market_outlook="LT",
            long_term_publication_timestamp="2026-07-01T00:00:00",
            country_group="country",
            time_group="monthly",
            unit="mt",
        )

    assert not any(
        thread.name.startswith("market-balance-comparison")
        for thread in enumerate_threads()
    )


# Consolidated from test_market_balance_snapshot_refs.py.

import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import threading
import time

from dash import html
from dash._utils import to_json
from openpyxl import load_workbook
import pandas as pd
import pytest

from pages import market_balance
from utils import dashboard_snapshot_cache as snapshots


OVERVIEW_NAMESPACE = "market-balance-overview-v2"
TRADE_NAMESPACE = "market-balance-trade-v1"
COUNTRY_META_NAMESPACE = "market-balance-country-meta-v2"
COUNTRY_NAMESPACE = "market-balance-country-v2"
NAMESPACES = (
    OVERVIEW_NAMESPACE,
    TRADE_NAMESPACE,
    COUNTRY_META_NAMESPACE,
    COUNTRY_NAMESPACE,
)
REFERENCE_NAMESPACES = (
    OVERVIEW_NAMESPACE,
    TRADE_NAMESPACE,
    COUNTRY_META_NAMESPACE,
)


def _frame(**columns):
    return market_balance.serialize_frame(pd.DataFrame(columns))


def _empty_frame(*columns):
    return market_balance.serialize_frame(
        pd.DataFrame(columns=list(columns))
    )


def _overview_payload():
    balance = _frame(
        Period=["2026", "2027"],
        Supply=[12.5, 13.0],
        Demand=[10.0, 11.0],
    )
    net = _frame(
        Period=["2026", "2027"],
        Europe=[2.5, 2.0],
        Total=[2.5, 2.0],
    )
    return {
        "data": {
            "woodmac_balance": copy.deepcopy(balance),
            "ea_balance": copy.deepcopy(balance),
            "woodmac_net_balance": copy.deepcopy(net),
            "ea_net_balance": copy.deepcopy(net),
            "maintenance": _empty_frame("Period", "Total"),
            "maintenance_grouped": _empty_frame(
                "Period",
                "Metric",
                "Total",
            ),
            "maintenance_ea": _empty_frame("Period", "Total"),
            "maintenance_provider_comparison": _empty_frame(
                "Period",
                "WoodMac Unplanned",
                "Energy Aspects Unplanned",
                "Delta",
            ),
            "pacific_detail": _empty_frame(
                "Period",
                "Supply",
                "Country",
                "Provider",
            ),
            "pacific_totals": _empty_frame(
                "Period",
                "Equivalent MCM/D",
                "Provider",
            ),
        },
        "metadata": {
            "woodmac_export": {
                "short_term_publication_timestamp": "2026-07-20"
            },
            "woodmac_import": {
                "short_term_publication_timestamp": "2026-07-20"
            },
            "ea_export": {
                "upload_timestamp_utc": "2026-07-21T00:00:00"
            },
            "ea_import": {
                "upload_timestamp_utc": "2026-07-21T00:00:00"
            },
            "overview_net": {
                "country_group_label": "Classification",
                "time_group": "yearly",
                "unit": "bcm",
            },
            "comparison_options": {
                "woodmac": {
                    "short_term": [],
                    "long_term": [],
                },
                "ea_comparison_runs": [],
            },
        },
        "error": None,
    }


def _trade_payload():
    levels = _frame(
        Period=["2026", "2027"],
        Qatar=[8.0, 9.0],
        Total=[8.0, 9.0],
    )
    delta = _frame(
        Period=["2026", "2027"],
        Qatar=[1.0, -1.0],
        Total=[1.0, -1.0],
    )
    return {
        "data": {
            "exports": copy.deepcopy(levels),
            "exports_diff": copy.deepcopy(delta),
            "exports_flex": copy.deepcopy(delta),
            "imports": copy.deepcopy(levels),
            "imports_diff": copy.deepcopy(delta),
            "imports_flex": copy.deepcopy(delta),
        },
        "metadata": {
            "available_years": [2026, 2027],
            "source": "Energy Aspects",
            "unit": "bcm",
            "time_group": "yearly",
            "country_group_label": "Country",
            "export_metadata": {
                "upload_timestamp_utc": "2026-07-21T00:00:00"
            },
            "import_metadata": {
                "upload_timestamp_utc": "2026-07-21T00:00:00"
            },
            "warnings": [],
        },
        "error": None,
    }


def _country_meta_payload():
    snapshots_list = [
        {
            "run_id": 20,
            "snapshot_at": "2026-07-21T00:00:00",
        },
        {
            "run_id": 19,
            "snapshot_at": "2026-07-20T00:00:00",
        },
    ]
    return {
        "data": {
            "countries": ["Belgium", "France"],
            "snapshots": copy.deepcopy(snapshots_list),
            "country_snapshots": {
                "Belgium": copy.deepcopy(snapshots_list),
                "France": copy.deepcopy(snapshots_list),
            },
        },
        "metadata": {
            "default_country": "Belgium",
            "default_snapshot": 19,
            "current_ea": {"run_id": 20},
        },
        "error": None,
    }


def _country_payload():
    return {
        "data": {
            "current_table": _frame(
                Date=["2026-01", "2026-02"],
                Demand=[5.0, 6.0],
                Supply=[6.0, 7.0],
            ),
            "delta_table": _frame(
                Date=["2026-01", "2026-02"],
                Demand=[0.5, -0.5],
                Supply=[0.25, -0.25],
            ),
            "balance_chart": _frame(
                Date=["2026-01", "2026-02"],
                Demand=[5.0, 6.0],
                Supply=[6.0, 7.0],
                **{"Fcst Margin": [1.0, 1.0]},
            ),
            "category_charts": [
                {
                    "title": "Pipelines",
                    "chart_type": "area",
                    "frame": _frame(
                        Date=["2026-01", "2026-02"],
                        value=[1.0, 1.5],
                        series_name=["Pipe A", "Pipe A"],
                    ),
                }
            ],
        },
        "metadata": {
            "country": "Belgium",
            "level": "subtype",
            "time_group": "monthly",
            "current_snapshot": "2026-07-21T00:00:00",
            "comparison_snapshot": "2026-07-20T00:00:00",
            "requested_comparison_snapshot": "2026-07-20T00:00:00",
            "warnings": [],
            "column_styles": [],
        },
        "error": None,
    }


PAYLOAD_BUILDERS = {
    OVERVIEW_NAMESPACE: _overview_payload,
    TRADE_NAMESPACE: _trade_payload,
    COUNTRY_META_NAMESPACE: _country_meta_payload,
    COUNTRY_NAMESPACE: _country_payload,
}


@pytest.fixture
def persistent_market_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "market-balance-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    monkeypatch.setattr(
        market_balance,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    monkeypatch.setattr(
        market_balance,
        "_fetch_provider_flow_source_state",
        lambda: {"watermark": "2026-07-24T00:00:00"},
    )
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_reference(namespace, payload=None, *, key_parts=None):
    payload = payload or PAYLOAD_BUILDERS[namespace]()
    reference = market_balance._load_cached_market_store(
        namespace,
        key_parts or {"case": "representative"},
        lambda: copy.deepcopy(payload),
        source_state={"watermark": "2026-07-24T00:00:00"},
    )
    return payload, reference


def _market_workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: [
            [
                (
                    cell.value,
                    cell.data_type,
                    cell.number_format,
                )
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]
        for worksheet in workbook.worksheets
    }


def _delete_market_reference(reference, corruption_mode):
    persistent_stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if corruption_mode == "missing":
        persistent_stores.cache.delete(record_key, retry=True)
    else:
        persistent_stores.cache.set(
            record_key,
            b"corrupt",
            retry=True,
        )
    snapshots.clear_local_snapshots()


@pytest.mark.parametrize("namespace", REFERENCE_NAMESPACES)
def test_market_loaders_emit_small_resolvable_refs_and_survive_restart(
    namespace,
    persistent_market_cache,
):
    payload, reference = _load_reference(namespace)

    assert snapshots.is_snapshot_reference(reference, namespace)
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert len(to_json(reference).encode("utf-8")) < 50_000

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    assert (
        market_balance._resolve_market_store(reference, namespace)
        == payload
    )


def test_market_reference_callbacks_and_five_exports_match_raw_payloads(
    monkeypatch,
    persistent_market_cache,
):
    payloads_and_refs = {
        namespace: _load_reference(namespace)
        for namespace in NAMESPACES
    }
    overview, overview_ref = payloads_and_refs[OVERVIEW_NAMESPACE]
    trade, trade_ref = payloads_and_refs[TRADE_NAMESPACE]
    country_meta, country_meta_ref = payloads_and_refs[
        COUNTRY_META_NAMESPACE
    ]
    country, country_ref = payloads_and_refs[COUNTRY_NAMESPACE]
    comparison_df = pd.DataFrame(
        {
            "Period": ["2026", "2027"],
            "Europe": [2.0, 1.5],
            "Total": [2.0, 1.5],
        }
    )
    monkeypatch.setattr(
        market_balance,
        "_fetch_net_balance_comparison_frame",
        lambda **_kwargs: (comparison_df.copy(), None),
    )

    callback_pairs = [
        (
            market_balance.render_overview(
                "overview",
                overview,
                "Unplanned",
            ),
            market_balance.render_overview(
                "overview",
                overview_ref,
                "Unplanned",
            ),
        ),
        (
            market_balance.sync_woodmac_overview_comparison_controls(
                overview,
                "woodmac",
                None,
                None,
                None,
            ),
            market_balance.sync_woodmac_overview_comparison_controls(
                overview_ref,
                "woodmac",
                None,
                None,
                None,
            ),
        ),
        (
            market_balance.sync_ea_overview_comparison_controls(
                overview,
                "ea",
                None,
                None,
                None,
            ),
            market_balance.sync_ea_overview_comparison_controls(
                overview_ref,
                "ea",
                None,
                None,
                None,
            ),
        ),
        (
            market_balance.render_woodmac_overview_delta(
                "overview",
                overview,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "ea",
                None,
                None,
                19,
            ),
            market_balance.render_woodmac_overview_delta(
                "overview",
                overview_ref,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "ea",
                None,
                None,
                19,
            ),
        ),
        (
            market_balance.render_ea_overview_delta(
                "overview",
                overview,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "woodmac",
                None,
                None,
                None,
            ),
            market_balance.render_ea_overview_delta(
                "overview",
                overview_ref,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "woodmac",
                None,
                None,
                None,
            ),
        ),
        (
            market_balance.sync_trade_years(trade),
            market_balance.sync_trade_years(trade_ref),
        ),
        (
            market_balance.render_trade_balance(trade),
            market_balance.render_trade_balance(trade_ref),
        ),
        (
            market_balance.sync_country_controls(
                country_meta,
                "France",
            ),
            market_balance.sync_country_controls(
                country_meta_ref,
                "France",
            ),
        ),
        (
            market_balance.sync_country_snapshot_control(
                country_meta,
                "Belgium",
                19,
            ),
            market_balance.sync_country_snapshot_control(
                country_meta_ref,
                "Belgium",
                19,
            ),
        ),
        (
            market_balance.render_country_balance(country),
            market_balance.render_country_balance(country_ref),
        ),
    ]
    for raw_result, reference_result in callback_pairs:
        assert to_json(raw_result) == to_json(reference_result)

    export_pairs = [
        (
            market_balance.export_woodmac_overview_net_workbook(
                1,
                overview,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "ea",
                None,
                None,
                19,
            ),
            market_balance.export_woodmac_overview_net_workbook(
                1,
                overview_ref,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "ea",
                None,
                None,
                19,
            ),
        ),
        (
            market_balance.export_ea_overview_net_workbook(
                1,
                overview,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "woodmac",
                None,
                None,
                None,
            ),
            market_balance.export_ea_overview_net_workbook(
                1,
                overview_ref,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "woodmac",
                None,
                None,
                None,
            ),
        ),
        (
            market_balance.export_overview_workbook(1, overview),
            market_balance.export_overview_workbook(
                1,
                overview_ref,
            ),
        ),
        (
            market_balance.export_trade_workbook(1, trade),
            market_balance.export_trade_workbook(1, trade_ref),
        ),
        (
            market_balance.export_country_workbook(
                1,
                country,
                "Belgium",
            ),
            market_balance.export_country_workbook(
                1,
                country_ref,
                "Belgium",
            ),
        ),
    ]
    for raw_download, reference_download in export_pairs:
        assert _market_workbook_cells(raw_download) == _market_workbook_cells(
            reference_download
        )


@pytest.mark.parametrize("namespace", NAMESPACES)
@pytest.mark.parametrize("worker_count", (1, 2, 4))
def test_market_namespaces_single_flight_for_one_two_and_four_callers(
    namespace,
    worker_count,
    persistent_market_cache,
):
    payload = PAYLOAD_BUILDERS[namespace]()
    build_count = 0
    count_lock = threading.Lock()

    def builder():
        nonlocal build_count
        with count_lock:
            build_count += 1
        time.sleep(0.03)
        return copy.deepcopy(payload)

    def load():
        return market_balance._load_cached_market_store(
            namespace,
            {
                "case": "single-flight",
                "workers": worker_count,
            },
            builder,
            source_state={
                "watermark": "2026-07-24T00:00:00"
            },
        )

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        references = list(
            executor.map(lambda _index: load(), range(worker_count))
        )

    assert build_count == 1
    assert all(reference == references[0] for reference in references)
    assert (
        market_balance._resolve_market_store(
            references[0],
            namespace,
        )
        == payload
    )


@pytest.mark.parametrize("namespace", REFERENCE_NAMESPACES)
def test_market_loader_never_falls_back_to_raw_payload(
    namespace,
    monkeypatch,
):
    payload = PAYLOAD_BUILDERS[namespace]()
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": namespace,
        "source_key": "source",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        market_balance,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (
            non_resolvable,
            payload,
        ),
    )
    monkeypatch.setattr(
        market_balance,
        "_was_global_refresh_triggered",
        lambda: False,
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        market_balance._load_cached_market_store(
            namespace,
            {"case": "no-fallback"},
            lambda: copy.deepcopy(payload),
            source_state={"watermark": "stable"},
        )


def test_country_detail_keeps_baseline_fallback_after_benchmark_revert(
    monkeypatch,
):
    payload = _country_payload()
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": COUNTRY_NAMESPACE,
        "source_key": "source",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        market_balance,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (
            non_resolvable,
            payload,
        ),
    )
    monkeypatch.setattr(
        market_balance,
        "_was_global_refresh_triggered",
        lambda: False,
    )

    result = market_balance._load_cached_market_store(
        COUNTRY_NAMESPACE,
        {"case": "benchmark-revert"},
        lambda: copy.deepcopy(payload),
        source_state={"watermark": "stable"},
    )

    assert COUNTRY_NAMESPACE not in (
        market_balance.MARKET_BALANCE_REFERENCE_NAMESPACES
    )
    assert result == payload


@pytest.mark.parametrize("namespace", NAMESPACES)
@pytest.mark.parametrize("corruption_mode", ("missing", "corrupt"))
def test_missing_or_corrupt_market_ref_has_explicit_recovery(
    namespace,
    corruption_mode,
    persistent_market_cache,
):
    _payload, reference = _load_reference(
        namespace,
        key_parts={
            "case": "recovery",
            "mode": corruption_mode,
        },
    )
    _delete_market_reference(reference, corruption_mode)

    if namespace == OVERVIEW_NAMESPACE:
        result = market_balance.render_overview(
            "overview",
            reference,
            "Unplanned",
        )
        notice = result[1]
    elif namespace == TRADE_NAMESPACE:
        result = market_balance.render_trade_balance(reference)
        notice = result[1]
    elif namespace == COUNTRY_META_NAMESPACE:
        options, selected = market_balance.sync_country_controls(
            reference,
            "Belgium",
        )
        assert selected is None
        assert (
            options[0]["label"]
            == market_balance.MARKET_BALANCE_SNAPSHOT_RECOVERY_MESSAGE
        )
        return
    else:
        result = market_balance.render_country_balance(reference)
        notice = result[1]

    assert isinstance(notice, html.Div)
    assert (
        notice.children
        == market_balance.MARKET_BALANCE_SNAPSHOT_RECOVERY_MESSAGE
    )


def test_missing_market_refs_block_all_five_exports(
    monkeypatch,
    persistent_market_cache,
):
    references = {
        namespace: _load_reference(
            namespace,
            key_parts={"case": "export-recovery"},
        )[1]
        for namespace in (
            OVERVIEW_NAMESPACE,
            TRADE_NAMESPACE,
            COUNTRY_NAMESPACE,
        )
    }
    for reference in references.values():
        _delete_market_reference(reference, "missing")

    overview_ref = references[OVERVIEW_NAMESPACE]
    overview_export_calls = [
        lambda: market_balance.export_woodmac_overview_net_workbook(
            1,
            overview_ref,
            "2026-01-01",
            "2027-12-31",
            "yearly",
            "bcm",
            "country",
            "ea",
            None,
            None,
            19,
        ),
        lambda: market_balance.export_ea_overview_net_workbook(
            1,
            overview_ref,
            "2026-01-01",
            "2027-12-31",
            "yearly",
            "bcm",
            "country",
            "woodmac",
            None,
            None,
            None,
        ),
        lambda: market_balance.export_overview_workbook(
            1,
            overview_ref,
        ),
    ]
    export_calls = [
        *overview_export_calls,
        lambda: market_balance.export_trade_workbook(
            1,
            references[TRADE_NAMESPACE],
        ),
        lambda: market_balance.export_country_workbook(
            1,
            references[COUNTRY_NAMESPACE],
            "Belgium",
        ),
    ]
    for export_call in export_calls:
        with pytest.raises(
            snapshots.SnapshotUnavailable,
            match="Click the global Refresh",
        ):
            export_call()


@pytest.mark.parametrize("namespace", NAMESPACES)
def test_global_refresh_forces_exactly_one_market_rebuild(
    namespace,
    monkeypatch,
    persistent_market_cache,
):
    payload = PAYLOAD_BUILDERS[namespace]()
    old_reference = market_balance._load_cached_market_store(
        namespace,
        {"case": "global-refresh"},
        lambda: copy.deepcopy(payload),
        source_state={"watermark": "stable"},
    )
    build_count = 0

    def builder():
        nonlocal build_count
        build_count += 1
        refreshed = copy.deepcopy(payload)
        refreshed["metadata"]["refresh_marker"] = 1
        return refreshed

    monkeypatch.setattr(
        market_balance,
        "_was_global_refresh_triggered",
        lambda: True,
    )
    monkeypatch.setattr(
        market_balance,
        "_fetch_provider_flow_source_state",
        lambda: {"watermark": "stable"},
    )
    refreshed_reference = market_balance._load_cached_market_store(
        namespace,
        {"case": "global-refresh"},
        builder,
        source_state={"watermark": "stale"},
    )

    assert build_count == 1
    assert refreshed_reference != old_reference
    assert (
        market_balance._resolve_market_store(
            refreshed_reference,
            namespace,
        )["metadata"]["refresh_marker"]
        == 1
    )


# Consolidated from test_overview_period_month_limits.py.

from pages import exporters, importers


def test_exporters_period_selector_supports_up_to_48_months():
    options = exporters._build_supply_dest_count_options(
        exporters.SUPPLY_DEST_MAX_MONTH_COUNT
    )

    assert exporters.SUPPLY_DEST_MAX_MONTH_COUNT == 48
    assert options[0] == {'label': '1', 'value': 1}
    assert options[-1] == {'label': '48', 'value': 48}
    assert len(options) == 48
    assert exporters._coerce_supply_dest_period_count(49, 3, 48) == 48
    assert exporters.SUPPLY_DEST_PRELOAD_MONTH_COUNT == 60


def test_importers_period_selector_supports_up_to_48_months():
    options = importers._build_importer_period_count_options(
        importers.IMPORTER_PERIOD_MAX_MONTH_COUNT
    )

    assert importers.IMPORTER_PERIOD_MAX_MONTH_COUNT == 48
    assert options[0] == {'label': '1', 'value': 1}
    assert options[-1] == {'label': '48', 'value': 48}
    assert len(options) == 48
    assert importers._coerce_importer_period_count(49, 3, 48) == 48


# Consolidated from test_production.py.

import pandas as pd
import pytest

from pages import production


def _production_component_text(component):
    if component is None:
        return ""
    if isinstance(component, str):
        return component
    if isinstance(component, (list, tuple)):
        return " ".join(_production_component_text(child) for child in component)
    return _production_component_text(getattr(component, "children", None))


def test_production_groupings_reconcile_to_the_exact_run_total(monkeypatch):
    source_df = pd.DataFrame(
        [
            {
                "run_id": 12,
                "train_key": "train-1",
                "year": 2027,
                "month": 1,
                "country_name": "Mexico",
                "plant_name": "Terminal A",
                "lng_train_name_short": "Train 1",
                "total_output": 1.0,
            },
            {
                "run_id": 12,
                "train_key": "train-2",
                "year": 2027,
                "month": 1,
                "country_name": "Mexico",
                "plant_name": "Terminal A",
                "lng_train_name_short": "Train 2",
                "total_output": 2.0,
            },
            {
                "run_id": 12,
                "train_key": "train-3",
                "year": 2027,
                "month": 1,
                "country_name": "Canada",
                "plant_name": "Terminal B",
                "lng_train_name_short": "Train 1",
                "total_output": 3.0,
            },
        ]
    )
    calls = []

    def fake_fetch(**kwargs):
        calls.append(kwargs)
        return source_df.copy()

    monkeypatch.setattr(production, "fetch_capacity_ramp_production_monthly", fake_fetch)
    production._fetch_volume_country_dataframe_cached.cache_clear()

    grouped = {}
    for breakdown in ("country", "project", "train"):
        grouped[breakdown] = production._prepare_volume_country_dataframe(
            run_id=12,
            selected_unit="mtpa",
            new_capacity_only=True,
            start_year=2026,
            end_year=2031,
            breakdown=breakdown,
            refresh_token=1,
        )

    assert all(frame["total_output"].sum() == pytest.approx(6.0) for frame in grouped.values())
    assert set(grouped["country"]["country_name"]) == {"Canada", "Mexico"}
    assert set(grouped["project"]["country_name"]) == {"Terminal A", "Terminal B"}
    assert set(grouped["train"]["country_name"]) == {
        "Terminal A - Train 1",
        "Terminal A - Train 2",
        "Terminal B - Train 1",
    }
    assert all(call["run_id"] == 12 for call in calls)
    assert all(call["new_capacity_only"] is True for call in calls)


def test_capacity_scenario_selector_defaults_to_base_case(monkeypatch):
    catalog = (
        {
            "scenario_id": 7,
            "scenario_name": "Alternative",
            "display_run_id": None,
        },
        {
            "scenario_id": 2,
            "scenario_name": "Base Case",
            "display_run_id": 12,
            "display_run_status": "completed",
            "display_blocking_qa_count": 1,
        },
    )
    monkeypatch.setattr(production, "_get_capacity_ramp_catalog_cached", lambda _: catalog)

    options, selected = production.populate_capacity_scenario_options(0, None)

    assert selected == 2
    assert [option["value"] for option in options] == [7, 2]
    assert options[0]["label"] == "Alternative - No ramp run"
    assert options[1]["label"] == "Base Case - QA blocked (run 12)"


def test_blocked_fallback_banner_identifies_both_runs_and_stale_state():
    banner = production._build_ramp_run_status_banner(
        {
            "scenario_name": "Base Case",
            "latest_attempt_run_id": 13,
            "latest_attempt_run_status": "failed",
            "display_run_id": 12,
            "display_run_status": "completed",
            "display_is_current_published": False,
            "display_generator_version": "terminal_ramp_sql_inputs_v1",
            "display_generated_at": "2026-07-13T10:00:00Z",
            "display_horizon_start_month": "2026-07-01",
            "display_horizon_end_month": "2031-07-01",
            "display_monthly_row_count": 7198,
            "display_train_count": 118,
            "display_blocking_qa_count": 1,
            "display_blocking_qa_summaries": [
                {"plant_name": "Tango", "message": "Approved profile is missing."}
            ],
            "display_is_stale": True,
            "fallback_reason": "Latest attempt run 13 is failed; showing materialized run 12.",
        }
    )

    text = _production_component_text(banner)
    assert "QA-blocked ramp output shown for analysis" in text
    assert "Run 12" in text
    assert "generator terminal_ramp_sql_inputs_v1" in text
    assert "118 trains" in text
    assert "Latest attempt run 13 is failed" in text
    assert "Approved profile is missing" in text
    assert "Stale" in text


def test_excel_export_uses_store_run_id(monkeypatch):
    captured = {}

    def fake_prepare(**kwargs):
        captured.update(kwargs)
        return pd.DataFrame(
            {
                "month": pd.to_datetime(["2027-01-01"]),
                "country_name": ["Terminal A"],
                "total_output": [1.25],
            }
        )

    monkeypatch.setattr(production, "_prepare_volume_country_dataframe", fake_prepare)

    download = production.export_capacity_to_excel(
        1,
        {"display_run_id": 12, "refresh_token": 4},
        "mtpa",
        [],
        "project",
        None,
        "rest_of_world",
        [2026, 2031],
    )

    assert captured["run_id"] == 12
    assert captured["refresh_token"] == 4
    assert download["filename"].startswith("LNG_Production_Run12_PlantMatrix_MTPA_")
    assert download["base64"] is True


def test_global_supply_chart_has_one_trace_per_available_source():
    comparison_df = pd.DataFrame(
        {
            "month": pd.to_datetime(
                ["2026-07-01", "2026-08-01"] * 4
            ),
            "source": [
                "Our ramp forecast",
                "Our ramp forecast",
                "Energy Aspects",
                "Energy Aspects",
                "Platts",
                "Platts",
                "WoodMac",
                "WoodMac",
            ],
            "monthly_mt": [40.0, 41.0, 39.0, 40.0, 40.5, 41.5, 38.5, 39.5],
        }
    )
    metadata = {
        "window_start": "2021-07-01",
        "forecast_start": "2026-07-01",
        "window_end": "2031-12-01",
    }

    figure = production.create_global_supply_comparison_chart(comparison_df, metadata)

    assert [trace.name for trace in figure.data] == [
        "Our ramp forecast",
        "Energy Aspects",
        "Platts",
        "WoodMac",
    ]
    assert figure.layout.yaxis.title.text == "Monthly LNG supply (Mt/month)"
    assert pd.Timestamp(figure.layout.xaxis.range[0]) == pd.Timestamp("2021-07-01")
    assert pd.Timestamp(figure.layout.xaxis.range[1]) == pd.Timestamp("2031-12-01")


def test_global_supply_quarter_table_and_chart_use_same_period_totals():
    comparison_df = pd.DataFrame(
        {
            "month": list(pd.date_range("2027-01-01", "2027-03-01", freq="MS")) * 2,
            "source": ["Our ramp forecast"] * 3 + ["WoodMac"] * 3,
            "monthly_mt": [40.0, 41.0, 42.0, 39.0, 40.0, 41.0],
        }
    )
    metadata = {
        "window_start": "2021-07-01",
        "forecast_start": "2026-07-01",
        "window_end": "2031-12-01",
    }
    period_df = production.aggregate_global_supply_comparison(comparison_df, "quarterly")

    figure = production.create_global_supply_comparison_chart(
        period_df,
        metadata,
        "quarterly",
    )
    table = production._create_global_supply_comparison_table(period_df, "quarterly")

    assert figure.layout.yaxis.title.text == "Quarterly LNG supply (Mt/quarter)"
    assert list(figure.data[0].y) == [123.0]
    assert list(figure.data[1].y) == [120.0]
    table_grid = table.children[1]
    assert table_grid.rowData[0]["Quarter"] == "2027 Q1"
    assert table_grid.rowData[0]["Our ramp forecast"] == pytest.approx(123.0)
    assert table_grid.rowData[0]["WoodMac"] == pytest.approx(120.0)


# Consolidated from test_snapshot_manifest_resolution.py.

import pytest

from utils import dashboard_snapshot_cache as snapshots


def test_exact_manifest_survives_restart_and_newer_revision(
    monkeypatch,
    tmp_path,
):
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(tmp_path / "manifest-cache"),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()

    namespace = "provider-flow-source-v2"
    source_key = "provider-source"
    old_state = {
        "mapping_hash": "country-v1",
        "ea_balance_mapping_hash": "ea-v1",
    }
    old_reference, _ = snapshots.get_or_build_snapshot(
        None,
        namespace=namespace,
        source_key=source_key,
        builder=lambda: {"value": "old"},
        manifest={"source_state": old_state},
    )

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    assert snapshots.resolve_snapshot_manifest(
        old_reference,
        None,
        expected_namespace=namespace,
    ) == {"source_state": old_state}

    new_state = {
        "mapping_hash": "country-v2",
        "ea_balance_mapping_hash": "ea-v2",
    }
    new_reference, _ = snapshots.get_or_build_snapshot(
        None,
        namespace=namespace,
        source_key=source_key,
        builder=lambda: {"value": "new"},
        manifest={"source_state": new_state},
        force=True,
    )
    assert new_reference != old_reference
    assert snapshots.resolve_snapshot_manifest(
        old_reference,
        None,
        expected_namespace=namespace,
    ) == {"source_state": old_state}
    assert snapshots.resolve_snapshot_manifest(
        new_reference,
        None,
        expected_namespace=namespace,
    ) == {"source_state": new_state}

    latest_reference, latest_payload = snapshots.get_or_build_snapshot(
        None,
        namespace=namespace,
        source_key=source_key,
        builder=lambda: {"value": "unexpected"},
    )
    assert latest_reference == new_reference
    assert latest_payload == {"value": "new"}

    with pytest.raises(snapshots.SnapshotUnavailable):
        snapshots.resolve_snapshot_manifest(
            old_reference,
            None,
            expected_namespace="wrong-namespace",
        )
    with pytest.raises(snapshots.SnapshotUnavailable):
        snapshots.resolve_snapshot_manifest(
            {**old_reference, "revision": "corrupt"},
            None,
            expected_namespace=namespace,
        )

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


# Consolidated from test_supply_demand_snapshot_refs.py.

import base64
import copy
from concurrent.futures import ThreadPoolExecutor
import hashlib
from io import BytesIO
import threading
import time

from dash import html
from dash._utils import to_json
from openpyxl import load_workbook
import pandas as pd
import pytest

from pages import demand, supply
from utils import dashboard_snapshot_cache as snapshots
from utils import provider_flow_snapshot


PAGES = (
    (
        "supply",
        supply,
        "woodmac_export",
        "ea_export",
        supply.SUPPLY_SNAPSHOT_RECOVERY_MESSAGE,
    ),
    (
        "demand",
        demand,
        "woodmac_import",
        "ea_import",
        demand.DEMAND_SNAPSHOT_RECOVERY_MESSAGE,
    ),
)


def test_shared_comparison_component_trees_are_exact():
    components = {
        "supply_woodmac": supply._create_woodmac_comparison_section(),
        "supply_ea": supply._create_ea_comparison_section(),
        "demand_woodmac": demand._create_woodmac_comparison_section(),
        "demand_ea": demand._create_ea_comparison_section(),
        "supply_layout": supply.layout,
        "demand_layout": demand.layout,
    }
    expected_hashes = {
        "supply_woodmac": "e934821853dc4269b9be20207bfa11fdf7f94a2a01b5ee802ec5d7ed8bb117fa",
        "supply_ea": "effdaefe2bddf65b789d98e324e78540f30239b4098a6abb12ab917336cf68e8",
        "demand_woodmac": "0fdd98114d7c5daab83e29b2b9d5a1f8fd86d291428839d85c23d043f1af61a5",
        "demand_ea": "f7f27f50bc247b8a071fd30d6a9b1122880fbbc1f0bead8b8c3405c74832196c",
        "supply_layout": "ca1a022452704f123f06a39323e4fdd9255730bddabbd6e7ff51edee8c5920f7",
        "demand_layout": "7a6bc3a04df17b6cb4695fe82a1e226a5ba44e9daab495acd70f55d91533fba2",
    }

    assert {
        name: hashlib.sha256(to_json(component).encode()).hexdigest()
        for name, component in components.items()
    } == expected_hashes


def _provider_flow_frame(multiplier=1.0):
    return pd.DataFrame(
        {
            "month": pd.to_datetime(
                [
                    "2025-01-01",
                    "2025-02-01",
                    "2025-01-01",
                    "2025-02-01",
                ]
            ),
            "country_name": [
                "Qatar",
                "Qatar",
                "United States",
                "United States",
            ],
            "total_mmtpa": [
                10.25 * multiplier,
                11.5 * multiplier,
                8.75 * multiplier,
                9.5 * multiplier,
            ],
        }
    )


def _provider_mapping_frame():
    return pd.DataFrame(
        [
            {
                "country": "Qatar",
                "country_name": "Qatar",
                "continent": "Asia",
                "subcontinent": "Middle East",
                "basin": "Pacific",
                "country_classification_level1": "Middle East",
                "country_classification": "Producer",
                "shipping_region": "Arabian Gulf",
            },
            {
                "country": "USA",
                "country_name": "United States",
                "continent": "North America",
                "subcontinent": "North America",
                "basin": "Atlantic",
                "country_classification_level1": "North America",
                "country_classification": "Producer",
                "shipping_region": "US Gulf",
            },
        ]
    )


def _provider_payload():
    return {
        "woodmac_export": _provider_flow_frame(1.0),
        "ea_export": _provider_flow_frame(1.1),
        "woodmac_import": _provider_flow_frame(0.9),
        "ea_import": _provider_flow_frame(0.8),
        "mapping": _provider_mapping_frame(),
        "woodmac_export_options": {
            "short_term": [],
            "long_term": [],
        },
        "woodmac_import_options": {
            "short_term": [],
            "long_term": [],
        },
        "ea_comparison_runs": [],
        "ea_export_options": [],
        "ea_import_options": [],
        "current_ea": {
            "run_id": 154,
            "snapshot_at": "2026-07-24T14:36:47.113486Z",
            "change_count": 0,
            "delete_count": 0,
        },
        "errors": {},
    }


@pytest.fixture
def persistent_provider_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "provider-flow-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _install_snapshot_getter(
    monkeypatch,
    page_module,
    payload,
    *,
    source_key,
    build_counter=None,
    build_delay=0.0,
    force_calls=None,
):
    def builder():
        if build_counter is not None:
            with build_counter["lock"]:
                build_counter["count"] += 1
        if build_delay:
            time.sleep(build_delay)
        return copy.deepcopy(payload)

    def getter(*, force=False):
        if force_calls is not None:
            force_calls.append(force)
        return snapshots.get_or_build_snapshot(
            page_module.engine,
            namespace=provider_flow_snapshot.NAMESPACE,
            source_key=source_key,
            builder=builder,
            force=force,
            manifest={"source_state": "stable"},
        )

    monkeypatch.setattr(
        page_module,
        "_get_provider_flow_snapshot",
        getter,
    )
    return getter


def _load_page_references(
    monkeypatch,
    page_name,
    page_module,
    payload,
    *,
    source_key_suffix="representative",
):
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    _install_snapshot_getter(
        monkeypatch,
        page_module,
        payload,
        source_key=f"{page_name}-{source_key_suffix}",
    )
    return page_module.load_balance_source_data(0)


def _provider_workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: [
            [
                (
                    cell.value,
                    cell.data_type,
                    cell.number_format,
                )
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]
        for worksheet in workbook.worksheets
    }


def _delete_provider_reference(reference, corruption_mode):
    stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if corruption_mode == "missing":
        stores.cache.delete(record_key, retry=True)
    else:
        stores.cache.set(record_key, b"corrupt", retry=True)
    snapshots.clear_local_snapshots()


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_loader_emits_ordered_small_refs_and_survives_restart(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    result = _load_page_references(
        monkeypatch,
        page_name,
        page_module,
        payload,
    )
    woodmac_ref, ea_ref = result[:2]

    assert woodmac_ref["slot"] == woodmac_slot
    assert ea_ref["slot"] == ea_slot
    for reference in (woodmac_ref, ea_ref):
        assert snapshots.is_snapshot_reference(
            reference,
            provider_flow_snapshot.NAMESPACE,
        )
        assert snapshots.snapshot_is_resolvable(reference)
        assert len(to_json(reference).encode("utf-8")) < 10_000
    assert len(to_json(result).encode("utf-8")) < 50_000
    assert result[2] == page_module.get_available_countries(
        [payload[woodmac_slot], payload[ea_slot]]
    )
    assert result[4] is None

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    pd.testing.assert_frame_equal(
        page_module._deserialize_dataframe(woodmac_ref),
        payload[woodmac_slot],
        check_dtype=True,
        check_exact=True,
    )
    pd.testing.assert_frame_equal(
        page_module._deserialize_dataframe(ea_ref),
        payload[ea_slot],
        check_dtype=True,
        check_exact=True,
    )


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_reference_controls_tables_deltas_and_two_exports_match_legacy(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    result = _load_page_references(
        monkeypatch,
        page_name,
        page_module,
        payload,
        source_key_suffix="parity",
    )
    woodmac_ref, ea_ref = result[:2]
    raw_woodmac = page_module._serialize_dataframe(
        payload[woodmac_slot]
    )
    raw_ea = page_module._serialize_dataframe(payload[ea_slot])
    countries = result[2]
    lookup = result[3]
    woodmac_metadata = result[5]
    ea_metadata = result[6]
    selected = ["Qatar"]
    start_date = "2025-01-01"
    end_date = "2025-02-28"
    monkeypatch.setattr(
        page_module,
        "_fetch_comparison_raw_df",
        lambda *_args, **_kwargs: (
            payload[woodmac_slot].copy(),
            None,
        ),
    )

    callback_pairs = [
        (
            page_module.update_balance_country_options(
                countries,
                "continent",
                lookup,
                raw_woodmac,
                raw_ea,
                start_date,
                end_date,
                selected,
                selected,
            ),
            page_module.update_balance_country_options(
                countries,
                "continent",
                lookup,
                woodmac_ref,
                ea_ref,
                start_date,
                end_date,
                selected,
                selected,
            ),
        ),
        (
            page_module.update_balance_date_range(
                raw_woodmac,
                raw_ea,
                start_date,
                end_date,
            ),
            page_module.update_balance_date_range(
                woodmac_ref,
                ea_ref,
                start_date,
                end_date,
            ),
        ),
        (
            page_module.render_balance_tables(
                raw_woodmac,
                raw_ea,
                woodmac_metadata,
                ea_metadata,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
            page_module.render_balance_tables(
                woodmac_ref,
                ea_ref,
                woodmac_metadata,
                ea_metadata,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
        ),
        (
            page_module.render_comparison_delta_table(
                raw_woodmac,
                raw_ea,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
                "ea",
                None,
                None,
                153,
            ),
            page_module.render_comparison_delta_table(
                woodmac_ref,
                ea_ref,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
                "ea",
                None,
                None,
                153,
            ),
        ),
        (
            page_module.render_ea_comparison_delta_table(
                raw_woodmac,
                raw_ea,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
                "woodmac",
                None,
                None,
                None,
            ),
            page_module.render_ea_comparison_delta_table(
                woodmac_ref,
                ea_ref,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
                "woodmac",
                None,
                None,
                None,
            ),
        ),
    ]
    for raw_result, reference_result in callback_pairs:
        assert to_json(raw_result) == to_json(reference_result)

    export_pairs = [
        (
            page_module.export_woodmac_balance_excel(
                1,
                raw_woodmac,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
            page_module.export_woodmac_balance_excel(
                1,
                woodmac_ref,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
        ),
        (
            page_module.export_ea_balance_excel(
                1,
                raw_ea,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
            page_module.export_ea_balance_excel(
                1,
                ea_ref,
                selected,
                "rest_of_world",
                start_date,
                end_date,
                "monthly",
                "country",
                lookup,
            ),
        ),
    ]
    for raw_download, reference_download in export_pairs:
        assert _provider_workbook_cells(raw_download) == _provider_workbook_cells(
            reference_download
        )


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
@pytest.mark.parametrize("worker_count", (1, 2, 4))
def test_page_provider_loader_single_flight_at_one_two_and_four_callers(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    worker_count,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    build_counter = {
        "count": 0,
        "lock": threading.Lock(),
    }
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    _install_snapshot_getter(
        monkeypatch,
        page_module,
        payload,
        source_key=f"{page_name}-single-flight-{worker_count}",
        build_counter=build_counter,
        build_delay=0.04,
    )

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        results = list(
            executor.map(
                lambda _index: page_module.load_balance_source_data(0),
                range(worker_count),
            )
        )

    assert build_counter["count"] == 1
    assert all(result == results[0] for result in results)
    assert snapshots.snapshot_is_resolvable(results[0][0])
    assert snapshots.snapshot_is_resolvable(results[0][1])
    assert len(to_json(results[0]).encode("utf-8")) < 50_000


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_loader_never_uses_non_resolvable_or_legacy_fallback(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
):
    payload = _provider_payload()
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": provider_flow_snapshot.NAMESPACE,
        "source_key": f"{page_name}-non-resolvable",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        page_module,
        "_get_provider_flow_snapshot",
        lambda **_kwargs: (non_resolvable, payload),
    )
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        page_module.load_balance_source_data(0)


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
@pytest.mark.parametrize("corruption_mode", ("missing", "corrupt"))
def test_missing_or_corrupt_page_ref_has_explicit_recovery_and_blocks_exports(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    corruption_mode,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    result = _load_page_references(
        monkeypatch,
        page_name,
        page_module,
        payload,
        source_key_suffix=f"recovery-{corruption_mode}",
    )
    woodmac_ref, ea_ref = result[:2]
    countries = result[2]
    lookup = result[3]
    _delete_provider_reference(woodmac_ref, corruption_mode)

    rendered = page_module.render_balance_tables(
        woodmac_ref,
        ea_ref,
        result[5],
        result[6],
        ["Qatar"],
        "rest_of_world",
        "2025-01-01",
        "2025-02-28",
        "monthly",
        "country",
        lookup,
    )
    assert isinstance(rendered[0], html.Div)
    assert rendered[0].children == recovery_message
    assert isinstance(rendered[2], html.Div)
    assert rendered[2].children == recovery_message

    selector = page_module.update_balance_country_options(
        countries,
        "continent",
        lookup,
        woodmac_ref,
        ea_ref,
        "2025-01-01",
        "2025-02-28",
        ["Qatar"],
        ["Qatar"],
    )
    assert selector[0][0]["label"] == recovery_message
    assert selector[1] == []
    assert selector[2] is True

    export_calls = [
        lambda: page_module.export_woodmac_balance_excel(
            1,
            woodmac_ref,
            ["Qatar"],
            "rest_of_world",
            "2025-01-01",
            "2025-02-28",
            "monthly",
            "country",
            lookup,
        ),
        lambda: page_module.export_ea_balance_excel(
            1,
            ea_ref,
            ["Qatar"],
            "rest_of_world",
            "2025-01-01",
            "2025-02-28",
            "monthly",
            "country",
            lookup,
        ),
    ]
    for export_call in export_calls:
        with pytest.raises(
            snapshots.SnapshotUnavailable,
            match="Click the global Refresh",
        ):
            export_call()


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_global_refresh_reuses_unchanged_snapshot(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
    persistent_provider_cache,
):
    payload = _provider_payload()
    build_counter = {
        "count": 0,
        "lock": threading.Lock(),
    }
    force_calls = []
    _install_snapshot_getter(
        monkeypatch,
        page_module,
        payload,
        source_key=f"{page_name}-global-refresh",
        build_counter=build_counter,
        force_calls=force_calls,
    )
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    old_result = page_module.load_balance_source_data(0)
    refresh_checks = 0

    def refresh_triggered():
        nonlocal refresh_checks
        refresh_checks += 1
        return True

    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        refresh_triggered,
    )
    refreshed_result = page_module.load_balance_source_data(1)

    assert refresh_checks == 0
    assert force_calls == [False, False]
    assert build_counter["count"] == 1
    assert refreshed_result == old_result


@pytest.mark.parametrize(
    "page_name,page_module,woodmac_slot,ea_slot,recovery_message",
    PAGES,
)
def test_page_provider_failure_is_fail_closed(
    page_name,
    page_module,
    woodmac_slot,
    ea_slot,
    recovery_message,
    monkeypatch,
):
    monkeypatch.setattr(
        page_module,
        "_get_provider_flow_snapshot",
        lambda **_kwargs: (_ for _ in ()).throw(
            RuntimeError("source state changed")
        ),
    )
    monkeypatch.setattr(
        page_module,
        "_was_global_refresh_triggered",
        lambda: False,
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        page_module.load_balance_source_data(0)

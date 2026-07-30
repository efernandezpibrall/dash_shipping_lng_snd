import base64
import copy
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import threading
import time

from dash import html
from dash._utils import to_json
from openpyxl import load_workbook
import pandas as pd
import pytest

from pages import market_balance
from utils import dashboard_snapshot_cache as snapshots


OVERVIEW_NAMESPACE = "market-balance-overview-v2"
TRADE_NAMESPACE = "market-balance-trade-v1"
COUNTRY_META_NAMESPACE = "market-balance-country-meta-v2"
COUNTRY_NAMESPACE = "market-balance-country-v2"
NAMESPACES = (
    OVERVIEW_NAMESPACE,
    TRADE_NAMESPACE,
    COUNTRY_META_NAMESPACE,
    COUNTRY_NAMESPACE,
)
REFERENCE_NAMESPACES = (
    OVERVIEW_NAMESPACE,
    TRADE_NAMESPACE,
    COUNTRY_META_NAMESPACE,
)


def _frame(**columns):
    return market_balance.serialize_frame(pd.DataFrame(columns))


def _empty_frame(*columns):
    return market_balance.serialize_frame(
        pd.DataFrame(columns=list(columns))
    )


def _overview_payload():
    balance = _frame(
        Period=["2026", "2027"],
        Supply=[12.5, 13.0],
        Demand=[10.0, 11.0],
    )
    net = _frame(
        Period=["2026", "2027"],
        Europe=[2.5, 2.0],
        Total=[2.5, 2.0],
    )
    return {
        "data": {
            "woodmac_balance": copy.deepcopy(balance),
            "ea_balance": copy.deepcopy(balance),
            "woodmac_net_balance": copy.deepcopy(net),
            "ea_net_balance": copy.deepcopy(net),
            "maintenance": _empty_frame("Period", "Total"),
            "maintenance_grouped": _empty_frame(
                "Period",
                "Metric",
                "Total",
            ),
            "maintenance_ea": _empty_frame("Period", "Total"),
            "maintenance_provider_comparison": _empty_frame(
                "Period",
                "WoodMac Unplanned",
                "Energy Aspects Unplanned",
                "Delta",
            ),
            "pacific_detail": _empty_frame(
                "Period",
                "Supply",
                "Country",
                "Provider",
            ),
            "pacific_totals": _empty_frame(
                "Period",
                "Equivalent MCM/D",
                "Provider",
            ),
        },
        "metadata": {
            "woodmac_export": {
                "short_term_publication_timestamp": "2026-07-20"
            },
            "woodmac_import": {
                "short_term_publication_timestamp": "2026-07-20"
            },
            "ea_export": {
                "upload_timestamp_utc": "2026-07-21T00:00:00"
            },
            "ea_import": {
                "upload_timestamp_utc": "2026-07-21T00:00:00"
            },
            "overview_net": {
                "country_group_label": "Classification",
                "time_group": "yearly",
                "unit": "bcm",
            },
            "comparison_options": {
                "woodmac": {
                    "short_term": [],
                    "long_term": [],
                },
                "ea_comparison_runs": [],
            },
        },
        "error": None,
    }


def _trade_payload():
    levels = _frame(
        Period=["2026", "2027"],
        Qatar=[8.0, 9.0],
        Total=[8.0, 9.0],
    )
    delta = _frame(
        Period=["2026", "2027"],
        Qatar=[1.0, -1.0],
        Total=[1.0, -1.0],
    )
    return {
        "data": {
            "exports": copy.deepcopy(levels),
            "exports_diff": copy.deepcopy(delta),
            "exports_flex": copy.deepcopy(delta),
            "imports": copy.deepcopy(levels),
            "imports_diff": copy.deepcopy(delta),
            "imports_flex": copy.deepcopy(delta),
        },
        "metadata": {
            "available_years": [2026, 2027],
            "source": "Energy Aspects",
            "unit": "bcm",
            "time_group": "yearly",
            "country_group_label": "Country",
            "export_metadata": {
                "upload_timestamp_utc": "2026-07-21T00:00:00"
            },
            "import_metadata": {
                "upload_timestamp_utc": "2026-07-21T00:00:00"
            },
            "warnings": [],
        },
        "error": None,
    }


def _country_meta_payload():
    snapshots_list = [
        {
            "run_id": 20,
            "snapshot_at": "2026-07-21T00:00:00",
        },
        {
            "run_id": 19,
            "snapshot_at": "2026-07-20T00:00:00",
        },
    ]
    return {
        "data": {
            "countries": ["Belgium", "France"],
            "snapshots": copy.deepcopy(snapshots_list),
            "country_snapshots": {
                "Belgium": copy.deepcopy(snapshots_list),
                "France": copy.deepcopy(snapshots_list),
            },
        },
        "metadata": {
            "default_country": "Belgium",
            "default_snapshot": 19,
            "current_ea": {"run_id": 20},
        },
        "error": None,
    }


def _country_payload():
    return {
        "data": {
            "current_table": _frame(
                Date=["2026-01", "2026-02"],
                Demand=[5.0, 6.0],
                Supply=[6.0, 7.0],
            ),
            "delta_table": _frame(
                Date=["2026-01", "2026-02"],
                Demand=[0.5, -0.5],
                Supply=[0.25, -0.25],
            ),
            "balance_chart": _frame(
                Date=["2026-01", "2026-02"],
                Demand=[5.0, 6.0],
                Supply=[6.0, 7.0],
                **{"Fcst Margin": [1.0, 1.0]},
            ),
            "category_charts": [
                {
                    "title": "Pipelines",
                    "chart_type": "area",
                    "frame": _frame(
                        Date=["2026-01", "2026-02"],
                        value=[1.0, 1.5],
                        series_name=["Pipe A", "Pipe A"],
                    ),
                }
            ],
        },
        "metadata": {
            "country": "Belgium",
            "level": "subtype",
            "time_group": "monthly",
            "current_snapshot": "2026-07-21T00:00:00",
            "comparison_snapshot": "2026-07-20T00:00:00",
            "requested_comparison_snapshot": "2026-07-20T00:00:00",
            "warnings": [],
            "column_styles": [],
        },
        "error": None,
    }


PAYLOAD_BUILDERS = {
    OVERVIEW_NAMESPACE: _overview_payload,
    TRADE_NAMESPACE: _trade_payload,
    COUNTRY_META_NAMESPACE: _country_meta_payload,
    COUNTRY_NAMESPACE: _country_payload,
}


@pytest.fixture
def persistent_market_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "market-balance-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    monkeypatch.setattr(
        market_balance,
        "_was_global_refresh_triggered",
        lambda: False,
    )
    monkeypatch.setattr(
        market_balance,
        "_fetch_provider_flow_source_state",
        lambda: {"watermark": "2026-07-24T00:00:00"},
    )
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_reference(namespace, payload=None, *, key_parts=None):
    payload = payload or PAYLOAD_BUILDERS[namespace]()
    reference = market_balance._load_cached_market_store(
        namespace,
        key_parts or {"case": "representative"},
        lambda: copy.deepcopy(payload),
        source_state={"watermark": "2026-07-24T00:00:00"},
    )
    return payload, reference


def _workbook_cells(download):
    workbook = load_workbook(
        BytesIO(base64.b64decode(download["content"]))
    )
    return {
        worksheet.title: [
            [
                (
                    cell.value,
                    cell.data_type,
                    cell.number_format,
                )
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]
        for worksheet in workbook.worksheets
    }


def _delete_reference(reference, corruption_mode):
    persistent_stores = snapshots._get_persistent_stores()
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if corruption_mode == "missing":
        persistent_stores.cache.delete(record_key, retry=True)
    else:
        persistent_stores.cache.set(
            record_key,
            b"corrupt",
            retry=True,
        )
    snapshots.clear_local_snapshots()


@pytest.mark.parametrize("namespace", REFERENCE_NAMESPACES)
def test_market_loaders_emit_small_resolvable_refs_and_survive_restart(
    namespace,
    persistent_market_cache,
):
    payload, reference = _load_reference(namespace)

    assert snapshots.is_snapshot_reference(reference, namespace)
    assert snapshots.snapshot_is_resolvable(reference)
    assert len(to_json(reference).encode("utf-8")) < 10_000
    assert len(to_json(reference).encode("utf-8")) < 50_000

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()
    assert (
        market_balance._resolve_market_store(reference, namespace)
        == payload
    )


def test_market_reference_callbacks_and_five_exports_match_raw_payloads(
    monkeypatch,
    persistent_market_cache,
):
    payloads_and_refs = {
        namespace: _load_reference(namespace)
        for namespace in NAMESPACES
    }
    overview, overview_ref = payloads_and_refs[OVERVIEW_NAMESPACE]
    trade, trade_ref = payloads_and_refs[TRADE_NAMESPACE]
    country_meta, country_meta_ref = payloads_and_refs[
        COUNTRY_META_NAMESPACE
    ]
    country, country_ref = payloads_and_refs[COUNTRY_NAMESPACE]
    comparison_df = pd.DataFrame(
        {
            "Period": ["2026", "2027"],
            "Europe": [2.0, 1.5],
            "Total": [2.0, 1.5],
        }
    )
    monkeypatch.setattr(
        market_balance,
        "_fetch_net_balance_comparison_frame",
        lambda **_kwargs: (comparison_df.copy(), None),
    )

    callback_pairs = [
        (
            market_balance.render_overview(
                "overview",
                overview,
                "Unplanned",
            ),
            market_balance.render_overview(
                "overview",
                overview_ref,
                "Unplanned",
            ),
        ),
        (
            market_balance.sync_woodmac_overview_comparison_controls(
                overview,
                "woodmac",
                None,
                None,
                None,
            ),
            market_balance.sync_woodmac_overview_comparison_controls(
                overview_ref,
                "woodmac",
                None,
                None,
                None,
            ),
        ),
        (
            market_balance.sync_ea_overview_comparison_controls(
                overview,
                "ea",
                None,
                None,
                None,
            ),
            market_balance.sync_ea_overview_comparison_controls(
                overview_ref,
                "ea",
                None,
                None,
                None,
            ),
        ),
        (
            market_balance.render_woodmac_overview_delta(
                "overview",
                overview,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "ea",
                None,
                None,
                19,
            ),
            market_balance.render_woodmac_overview_delta(
                "overview",
                overview_ref,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "ea",
                None,
                None,
                19,
            ),
        ),
        (
            market_balance.render_ea_overview_delta(
                "overview",
                overview,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "woodmac",
                None,
                None,
                None,
            ),
            market_balance.render_ea_overview_delta(
                "overview",
                overview_ref,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "woodmac",
                None,
                None,
                None,
            ),
        ),
        (
            market_balance.sync_trade_years(trade),
            market_balance.sync_trade_years(trade_ref),
        ),
        (
            market_balance.render_trade_balance(trade),
            market_balance.render_trade_balance(trade_ref),
        ),
        (
            market_balance.sync_country_controls(
                country_meta,
                "France",
            ),
            market_balance.sync_country_controls(
                country_meta_ref,
                "France",
            ),
        ),
        (
            market_balance.sync_country_snapshot_control(
                country_meta,
                "Belgium",
                19,
            ),
            market_balance.sync_country_snapshot_control(
                country_meta_ref,
                "Belgium",
                19,
            ),
        ),
        (
            market_balance.render_country_balance(country),
            market_balance.render_country_balance(country_ref),
        ),
    ]
    for raw_result, reference_result in callback_pairs:
        assert to_json(raw_result) == to_json(reference_result)

    export_pairs = [
        (
            market_balance.export_woodmac_overview_net_workbook(
                1,
                overview,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "ea",
                None,
                None,
                19,
            ),
            market_balance.export_woodmac_overview_net_workbook(
                1,
                overview_ref,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "ea",
                None,
                None,
                19,
            ),
        ),
        (
            market_balance.export_ea_overview_net_workbook(
                1,
                overview,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "woodmac",
                None,
                None,
                None,
            ),
            market_balance.export_ea_overview_net_workbook(
                1,
                overview_ref,
                "2026-01-01",
                "2027-12-31",
                "yearly",
                "bcm",
                "country",
                "woodmac",
                None,
                None,
                None,
            ),
        ),
        (
            market_balance.export_overview_workbook(1, overview),
            market_balance.export_overview_workbook(
                1,
                overview_ref,
            ),
        ),
        (
            market_balance.export_trade_workbook(1, trade),
            market_balance.export_trade_workbook(1, trade_ref),
        ),
        (
            market_balance.export_country_workbook(
                1,
                country,
                "Belgium",
            ),
            market_balance.export_country_workbook(
                1,
                country_ref,
                "Belgium",
            ),
        ),
    ]
    for raw_download, reference_download in export_pairs:
        assert _workbook_cells(raw_download) == _workbook_cells(
            reference_download
        )


@pytest.mark.parametrize("namespace", NAMESPACES)
@pytest.mark.parametrize("worker_count", (1, 2, 4))
def test_market_namespaces_single_flight_for_one_two_and_four_callers(
    namespace,
    worker_count,
    persistent_market_cache,
):
    payload = PAYLOAD_BUILDERS[namespace]()
    build_count = 0
    count_lock = threading.Lock()

    def builder():
        nonlocal build_count
        with count_lock:
            build_count += 1
        time.sleep(0.03)
        return copy.deepcopy(payload)

    def load():
        return market_balance._load_cached_market_store(
            namespace,
            {
                "case": "single-flight",
                "workers": worker_count,
            },
            builder,
            source_state={
                "watermark": "2026-07-24T00:00:00"
            },
        )

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        references = list(
            executor.map(lambda _index: load(), range(worker_count))
        )

    assert build_count == 1
    assert all(reference == references[0] for reference in references)
    assert (
        market_balance._resolve_market_store(
            references[0],
            namespace,
        )
        == payload
    )


@pytest.mark.parametrize("namespace", REFERENCE_NAMESPACES)
def test_market_loader_never_falls_back_to_raw_payload(
    namespace,
    monkeypatch,
):
    payload = PAYLOAD_BUILDERS[namespace]()
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": namespace,
        "source_key": "source",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        market_balance,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (
            non_resolvable,
            payload,
        ),
    )
    monkeypatch.setattr(
        market_balance,
        "_was_global_refresh_triggered",
        lambda: False,
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        market_balance._load_cached_market_store(
            namespace,
            {"case": "no-fallback"},
            lambda: copy.deepcopy(payload),
            source_state={"watermark": "stable"},
        )


def test_country_detail_keeps_baseline_fallback_after_benchmark_revert(
    monkeypatch,
):
    payload = _country_payload()
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": COUNTRY_NAMESPACE,
        "source_key": "source",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        market_balance,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (
            non_resolvable,
            payload,
        ),
    )
    monkeypatch.setattr(
        market_balance,
        "_was_global_refresh_triggered",
        lambda: False,
    )

    result = market_balance._load_cached_market_store(
        COUNTRY_NAMESPACE,
        {"case": "benchmark-revert"},
        lambda: copy.deepcopy(payload),
        source_state={"watermark": "stable"},
    )

    assert COUNTRY_NAMESPACE not in (
        market_balance.MARKET_BALANCE_REFERENCE_NAMESPACES
    )
    assert result == payload


@pytest.mark.parametrize("namespace", NAMESPACES)
@pytest.mark.parametrize("corruption_mode", ("missing", "corrupt"))
def test_missing_or_corrupt_market_ref_has_explicit_recovery(
    namespace,
    corruption_mode,
    persistent_market_cache,
):
    _payload, reference = _load_reference(
        namespace,
        key_parts={
            "case": "recovery",
            "mode": corruption_mode,
        },
    )
    _delete_reference(reference, corruption_mode)

    if namespace == OVERVIEW_NAMESPACE:
        result = market_balance.render_overview(
            "overview",
            reference,
            "Unplanned",
        )
        notice = result[1]
    elif namespace == TRADE_NAMESPACE:
        result = market_balance.render_trade_balance(reference)
        notice = result[1]
    elif namespace == COUNTRY_META_NAMESPACE:
        options, selected = market_balance.sync_country_controls(
            reference,
            "Belgium",
        )
        assert selected is None
        assert (
            options[0]["label"]
            == market_balance.MARKET_BALANCE_SNAPSHOT_RECOVERY_MESSAGE
        )
        return
    else:
        result = market_balance.render_country_balance(reference)
        notice = result[1]

    assert isinstance(notice, html.Div)
    assert (
        notice.children
        == market_balance.MARKET_BALANCE_SNAPSHOT_RECOVERY_MESSAGE
    )


def test_missing_market_refs_block_all_five_exports(
    monkeypatch,
    persistent_market_cache,
):
    references = {
        namespace: _load_reference(
            namespace,
            key_parts={"case": "export-recovery"},
        )[1]
        for namespace in (
            OVERVIEW_NAMESPACE,
            TRADE_NAMESPACE,
            COUNTRY_NAMESPACE,
        )
    }
    for reference in references.values():
        _delete_reference(reference, "missing")

    overview_ref = references[OVERVIEW_NAMESPACE]
    overview_export_calls = [
        lambda: market_balance.export_woodmac_overview_net_workbook(
            1,
            overview_ref,
            "2026-01-01",
            "2027-12-31",
            "yearly",
            "bcm",
            "country",
            "ea",
            None,
            None,
            19,
        ),
        lambda: market_balance.export_ea_overview_net_workbook(
            1,
            overview_ref,
            "2026-01-01",
            "2027-12-31",
            "yearly",
            "bcm",
            "country",
            "woodmac",
            None,
            None,
            None,
        ),
        lambda: market_balance.export_overview_workbook(
            1,
            overview_ref,
        ),
    ]
    export_calls = [
        *overview_export_calls,
        lambda: market_balance.export_trade_workbook(
            1,
            references[TRADE_NAMESPACE],
        ),
        lambda: market_balance.export_country_workbook(
            1,
            references[COUNTRY_NAMESPACE],
            "Belgium",
        ),
    ]
    for export_call in export_calls:
        with pytest.raises(
            snapshots.SnapshotUnavailable,
            match="Click the global Refresh",
        ):
            export_call()


@pytest.mark.parametrize("namespace", NAMESPACES)
def test_global_refresh_forces_exactly_one_market_rebuild(
    namespace,
    monkeypatch,
    persistent_market_cache,
):
    payload = PAYLOAD_BUILDERS[namespace]()
    old_reference = market_balance._load_cached_market_store(
        namespace,
        {"case": "global-refresh"},
        lambda: copy.deepcopy(payload),
        source_state={"watermark": "stable"},
    )
    build_count = 0

    def builder():
        nonlocal build_count
        build_count += 1
        refreshed = copy.deepcopy(payload)
        refreshed["metadata"]["refresh_marker"] = 1
        return refreshed

    monkeypatch.setattr(
        market_balance,
        "_was_global_refresh_triggered",
        lambda: True,
    )
    monkeypatch.setattr(
        market_balance,
        "_fetch_provider_flow_source_state",
        lambda: {"watermark": "stable"},
    )
    refreshed_reference = market_balance._load_cached_market_store(
        namespace,
        {"case": "global-refresh"},
        builder,
        source_state={"watermark": "stale"},
    )

    assert build_count == 1
    assert refreshed_reference != old_reference
    assert (
        market_balance._resolve_market_store(
            refreshed_reference,
            namespace,
        )["metadata"]["refresh_marker"]
        == 1
    )

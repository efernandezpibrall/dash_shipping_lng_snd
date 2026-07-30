import base64
import copy
from concurrent.futures import ThreadPoolExecutor
import inspect
from io import BytesIO
import statistics
import threading
import time

from dash import html
from dash._callback import GLOBAL_CALLBACK_MAP
from dash._utils import to_json
from flask import Flask, Response
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import pytest

from pages import importers
from utils import dashboard_snapshot_cache as snapshots


def _make_overview_payload(
    *,
    entity_count=2,
    years=(2025, 2026),
    points_per_year=6,
    continents=("Africa", "Asia"),
):
    chart_entities = [
        {
            "key": f"Importer {index}",
            "label": f"Importer {index}",
            "destination_countries": [f"Country {index}"],
            "avg_30d_mcmd": float(20 - index),
            "is_global": index == 0,
        }
        for index in range(entity_count)
    ]
    table_entities = [
        dict(entity)
        for entity in chart_entities
        if not entity.get("is_global")
    ] or [dict(chart_entities[0])]
    demand_records = {}
    origin_records = {}
    for entity_index, entity in enumerate(chart_entities):
        demand_rows = []
        origin_rows = []
        for year in years:
            dates = pd.date_range(
                f"{year}-01-01",
                periods=points_per_year,
                freq="D",
            )
            for day_index, date in enumerate(dates):
                base_value = float(
                    40 + entity_index * 3 + day_index
                )
                demand_rows.append({
                    "date": date.isoformat(),
                    "year": str(year),
                    "month_day": date.strftime("%b %d"),
                    "rolling_avg": base_value,
                    "is_forecast": year == max(years),
                })
                for continent_index, continent in enumerate(continents):
                    origin_rows.append({
                        "date": date.isoformat(),
                        "year": str(year),
                        "month_day": date.strftime("%b %d"),
                        "continent_origin": continent,
                        "rolling_avg": base_value / len(continents),
                        "percentage": float(
                            100 / len(continents)
                            + continent_index
                        ),
                        "is_forecast": year == max(years),
                    })
        demand_records[entity["label"]] = demand_rows
        origin_records[entity["label"]] = origin_rows
    return {
        "chart_entities": chart_entities,
        "table_entities": table_entities,
        "demand_cube": snapshots.pack_record_mapping(demand_records),
        "origin_cube": snapshots.pack_record_mapping(origin_records),
    }


def _make_source_payload():
    return {
        "catalog_df": pd.DataFrame({
            "destination_country_name": ["Country 0", "Country 1"],
            "country_display": ["Country 0", "Country 1"],
        }),
        "ranking_df": pd.DataFrame({
            "destination_country_name": ["Country 0", "Country 1"],
            "country_display": ["Country 0", "Country 1"],
            "avg_30d_mcmd": [20.0, 19.0],
        }),
        "scoped_trades_df": pd.DataFrame({"scope": [1]}),
    }


def _install_overview_builders(
    monkeypatch,
    overview_payload,
    source_payload=None,
):
    source_payload = source_payload or _make_source_payload()
    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        lambda: copy.deepcopy(source_payload),
    )
    monkeypatch.setattr(
        importers,
        "_build_importers_overview_payload_from_source",
        lambda *_args: copy.deepcopy(overview_payload),
    )


def _build_realistic_typed_overview_payload(
    monkeypatch,
    *,
    include_codec_fields,
    timezone_aware,
):
    entity = {
        "key": "Importer 1",
        "label": "Importer 1",
        "destination_countries": ["Country 1"],
        "avg_30d_mcmd": 20.0,
        "is_global": False,
    }
    dates = [
        pd.Timestamp(
            "2026-01-01 12:30:15.123456789",
            tz="Europe/London" if timezone_aware else None,
        ),
        pd.Timestamp(
            "2026-01-02 12:30:15.987654321",
            tz="Europe/London" if timezone_aware else None,
        ),
    ]
    demand_columns = {
        "date": pd.Series(dates, dtype=object),
        "year": pd.Series(
            [np.int16(2026), np.int16(2026)],
            dtype=object,
        ),
        "month_day": ["Jan 01", "Jan 02"],
        "rolling_avg": pd.Series(
            [np.float32(12.5), np.float32(13.25)],
            dtype=object,
        ),
        "is_forecast": pd.Series(
            [np.bool_(False), np.bool_(True)],
            dtype=object,
        ),
    }
    if include_codec_fields:
        demand_columns.update({
            "missing_timestamp": pd.Series(
                [pd.NaT, pd.NaT],
                dtype=object,
            ),
            "numpy_timestamp": pd.Series(
                [
                    np.datetime64(
                        "2026-01-01T12:30:15.123456789",
                        "ns",
                    ),
                    np.datetime64("NaT", "ns"),
                ],
                dtype=object,
            ),
            "pandas_missing": pd.Series(
                [pd.NA, pd.NA],
                dtype=object,
            ),
            "nullable_text": pd.Series(
                [None, "available"],
                dtype=object,
            ),
            "numpy_int": pd.Series(
                [np.int32(7), np.int32(8)],
                dtype=object,
            ),
            "numpy_float": pd.Series(
                [np.float64(1.25), np.float64(np.nan)],
                dtype=object,
            ),
            "numpy_bool": pd.Series(
                [np.bool_(True), np.bool_(False)],
                dtype=object,
            ),
        })
    demand_df = pd.DataFrame(demand_columns)
    origin_df = demand_df.assign(
        continent_origin=["Asia", "Americas"],
        percentage=[62.5, 37.5],
    )

    monkeypatch.setattr(
        importers,
        "_build_destination_entities",
        lambda *_args, **_kwargs: [copy.deepcopy(entity)],
    )
    monkeypatch.setattr(
        importers,
        "_fetch_importer_scoped_trades",
        lambda *_args, **_kwargs: pd.DataFrame({"scope": [1]}),
    )
    monkeypatch.setattr(
        importers,
        "_filter_scoped_trades_for_entity",
        lambda scoped_df, *_args: scoped_df,
    )
    monkeypatch.setattr(
        importers,
        "_build_importer_total_import_df",
        lambda *_args, **_kwargs: demand_df.copy(),
    )
    monkeypatch.setattr(
        importers,
        "_build_importer_continent_chart_df",
        lambda *_args, **_kwargs: origin_df.copy(),
    )
    payload = importers._build_importers_overview_payload_from_source(
        _make_source_payload(),
        "Country",
        30,
    )
    if include_codec_fields:
        cube = payload["demand_cube"]
        column_indexes = {
            column: cube["columns"].index(column)
            for column in (
                "year",
                "rolling_avg",
                "is_forecast",
                "numpy_timestamp",
                "pandas_missing",
                "numpy_int",
                "numpy_float",
                "numpy_bool",
            )
        }
        typed_values = [
            {
                "year": np.int16(2026),
                "rolling_avg": np.float32(12.5),
                "is_forecast": np.bool_(False),
                "numpy_timestamp": np.datetime64(
                    "2026-01-01T12:30:15.123456789",
                    "ns",
                ),
                "pandas_missing": pd.NA,
                "numpy_int": np.int32(7),
                "numpy_float": np.float64(1.25),
                "numpy_bool": np.bool_(True),
            },
            {
                "year": np.int16(2026),
                "rolling_avg": np.float32(13.25),
                "is_forecast": np.bool_(True),
                "numpy_timestamp": np.datetime64("NaT", "ns"),
                "pandas_missing": pd.NA,
                "numpy_int": np.int32(8),
                "numpy_float": np.float64(np.nan),
                "numpy_bool": np.bool_(False),
            },
        ]
        for row, values in zip(cube["rows"], typed_values):
            for column, value in values.items():
                row[column_indexes[column]] = value
    return payload


@pytest.fixture
def persistent_importer_cache(monkeypatch, tmp_path):
    cache_directory = tmp_path / "importers-cache"
    monkeypatch.setenv(snapshots.LOCAL_PERSISTENCE_ENV, "1")
    monkeypatch.setenv(
        snapshots.LOCAL_CACHE_DIR_ENV,
        str(cache_directory),
    )
    snapshots.close_persistent_snapshot_cache()
    snapshots.clear_local_snapshots()
    yield cache_directory
    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()


def _load_representative_references(monkeypatch):
    overview_payload = _make_overview_payload()
    _install_overview_builders(monkeypatch, overview_payload)
    overview_stores = importers.refresh_overview_data(
        {"watermark": "2026-07-24T00:00:00"},
        "Country",
        30,
    )
    return overview_payload, overview_stores


def _load_legacy_v1_overview_references(overview_payload):
    reference, _payload = snapshots.get_or_build_snapshot(
        importers.engine,
        namespace="importers-overview-v1",
        source_key="legacy-v1-migration-guard-fixture",
        builder=lambda: (
            importers._prepare_importers_overview_snapshot_payload(
                copy.deepcopy(overview_payload)
            )
        ),
    )
    return (
        snapshots.with_snapshot_slot(reference, "chart_entities"),
        snapshots.with_snapshot_slot(reference, "table_entities"),
        snapshots.with_snapshot_slot(reference, "demand_cube"),
        snapshots.with_snapshot_slot(reference, "origin_cube"),
    )


def test_importer_overview_emits_small_resolvable_references_and_survives_restart(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload, overview_stores = (
        _load_representative_references(monkeypatch)
    )

    for store in overview_stores:
        assert snapshots.is_snapshot_reference(store)
        assert snapshots.snapshot_is_resolvable(store)
        assert len(to_json(store).encode("utf-8")) < 10_000
    assert len(to_json(overview_stores).encode("utf-8")) < 50_000

    snapshots.clear_local_snapshots()
    snapshots.close_persistent_snapshot_cache()

    assert importers._resolve_importers_entities_store(
        overview_stores[0],
        "chart_entities",
    ) == overview_payload["chart_entities"]
    assert importers._resolve_importers_entities_store(
        overview_stores[1],
        "table_entities",
    ) == overview_payload["table_entities"]
    assert importers._resolve_importers_chart_store(
        overview_stores[2]
    ) == snapshots.unpack_record_mapping(
        overview_payload["demand_cube"]
    )
    assert importers._resolve_importers_chart_store(
        overview_stores[3]
    ) == snapshots.unpack_record_mapping(
        overview_payload["origin_cube"]
    )


def test_legacy_v1_references_show_recovery_while_v2_references_work(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload, v2_stores = _load_representative_references(
        monkeypatch
    )
    v1_stores = _load_legacy_v1_overview_references(
        overview_payload
    )
    assert all(
        snapshots.snapshot_is_resolvable(store)
        for store in v1_stores
    )
    assert all(
        store["namespace"] == "importers-overview-v1"
        for store in v1_stores
    )

    options, selected = (
        importers.update_demand_year_selector_options(
            v1_stores[2],
            ["2026"],
        )
    )
    assert selected == []
    assert (
        options[0]["label"]
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )

    chart_notice = importers.update_demand_charts(
        v1_stores[2],
        v2_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(chart_notice, html.Div)
    assert (
        chart_notice.children
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )

    entity_notice = importers.update_demand_charts(
        v2_stores[2],
        v1_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(entity_notice, html.Div)
    assert (
        entity_notice.children
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )

    v2_options, v2_selected = (
        importers.update_demand_year_selector_options(
            v2_stores[2],
            ["2026"],
        )
    )
    assert v2_selected == ["2026"]
    assert any(
        option["value"] == "2026"
        for option in v2_options
    )
    v2_chart = importers.update_demand_charts(
        v2_stores[2],
        v2_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(v2_chart, html.Div)
    assert (
        v2_chart.children
        != importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )


def test_tagged_codec_preserves_real_builder_scalar_types_and_dataframe(
    monkeypatch,
):
    overview_payload = _build_realistic_typed_overview_payload(
        monkeypatch,
        include_codec_fields=True,
        timezone_aware=True,
    )
    prepared_payload = (
        importers._prepare_importers_overview_snapshot_payload(
            overview_payload
        )
    )
    decoded_cube = importers._decode_importers_json_payload(
        prepared_payload["demand_cube"],
        importers.IMPORTERS_RECORD_CUBE_FORMAT,
    )
    raw_records = snapshots.unpack_record_mapping(
        overview_payload["demand_cube"]
    )["Importer 1"]
    decoded_records = snapshots.unpack_record_mapping(
        decoded_cube
    )["Importer 1"]

    assert list(decoded_records[0]) == list(raw_records[0])
    assert isinstance(decoded_records[0]["date"], pd.Timestamp)
    assert (
        str(decoded_records[0]["date"].tz)
        == str(raw_records[0]["date"].tz)
        == "Europe/London"
    )
    assert (
        decoded_records[0]["date"].value
        == raw_records[0]["date"].value
    )
    assert decoded_records[0]["missing_timestamp"] is pd.NaT
    assert decoded_records[0]["pandas_missing"] is pd.NA
    assert decoded_records[0]["nullable_text"] is None
    assert type(decoded_records[0]["year"]) is np.int16
    assert type(decoded_records[0]["rolling_avg"]) is np.float32
    assert type(decoded_records[0]["is_forecast"]) is np.bool_
    assert type(decoded_records[0]["numpy_int"]) is np.int32
    assert type(decoded_records[0]["numpy_float"]) is np.float64
    assert type(decoded_records[0]["numpy_bool"]) is np.bool_
    assert (
        decoded_records[0]["numpy_timestamp"].dtype
        == raw_records[0]["numpy_timestamp"].dtype
    )
    assert (
        decoded_records[0]["numpy_timestamp"]
        == raw_records[0]["numpy_timestamp"]
    )
    assert np.isnat(decoded_records[1]["numpy_timestamp"])

    raw_frame = pd.DataFrame(raw_records)
    decoded_frame = pd.DataFrame(decoded_records)
    pd.testing.assert_frame_equal(
        decoded_frame,
        raw_frame,
        check_dtype=True,
        check_exact=True,
    )
    assert list(decoded_frame.columns) == list(raw_frame.columns)
    assert decoded_frame.isna().equals(raw_frame.isna())


def test_typed_reference_preserves_chart_csv_and_xlsx_exports(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload = _build_realistic_typed_overview_payload(
        monkeypatch,
        include_codec_fields=False,
        timezone_aware=False,
    )
    _install_overview_builders(monkeypatch, overview_payload)
    overview_stores = importers.refresh_overview_data(
        {"watermark": "typed-export-source"},
        "Country",
        30,
    )
    raw_demand = snapshots.unpack_record_mapping(
        overview_payload["demand_cube"]
    )
    reference_demand = importers._resolve_importers_chart_store(
        overview_stores[2]
    )

    raw_chart = importers.update_demand_charts(
        raw_demand,
        overview_payload["chart_entities"],
        "mcm_d",
        ["2026"],
        30,
    )
    reference_chart = importers.update_demand_charts(
        overview_stores[2],
        overview_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert to_json(reference_chart) == to_json(raw_chart)

    raw_export_frame = importers._build_chart_export_df(
        raw_demand,
        "mcm_d",
        ["2026"],
    )
    reference_export_frame = importers._build_chart_export_df(
        reference_demand,
        "mcm_d",
        ["2026"],
    )
    pd.testing.assert_frame_equal(
        reference_export_frame,
        raw_export_frame,
        check_dtype=True,
        check_exact=True,
    )
    assert (
        reference_export_frame.to_csv(index=False)
        == raw_export_frame.to_csv(index=False)
    )

    raw_download = importers.export_demand_to_excel(
        1,
        raw_demand,
        "mcm_d",
        ["2026"],
        30,
    )
    reference_download = importers.export_demand_to_excel(
        1,
        overview_stores[2],
        "mcm_d",
        ["2026"],
        30,
    )
    raw_workbook = load_workbook(
        BytesIO(base64.b64decode(raw_download["content"]))
    )
    reference_workbook = load_workbook(
        BytesIO(base64.b64decode(reference_download["content"]))
    )
    raw_sheet = raw_workbook["Demand"]
    reference_sheet = reference_workbook["Demand"]

    def worksheet_cells(worksheet):
        return [
            [
                (
                    cell.value,
                    cell.data_type,
                    cell.number_format,
                )
                for cell in row
            ]
            for row in worksheet.iter_rows()
        ]

    assert worksheet_cells(reference_sheet) == worksheet_cells(raw_sheet)
    headers = [cell.value for cell in raw_sheet[1]]
    date_column = headers.index("date") + 1
    raw_date_cell = raw_sheet.cell(row=2, column=date_column)
    reference_date_cell = reference_sheet.cell(
        row=2,
        column=date_column,
    )
    assert raw_date_cell.data_type == reference_date_cell.data_type == "d"
    assert (
        raw_date_cell.number_format
        == reference_date_cell.number_format
        != "General"
    )


def test_importer_overview_never_falls_back_to_raw_payloads(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload = _make_overview_payload()
    non_resolvable = {
        "format": snapshots.REFERENCE_FORMAT,
        "namespace": importers.IMPORTERS_OVERVIEW_NAMESPACE,
        "source_key": "source",
        "revision": snapshots._new_local_revision_token(),
        "shared": False,
    }
    monkeypatch.setattr(
        importers,
        "_get_or_build_snapshot",
        lambda *_args, **_kwargs: (
            non_resolvable,
            overview_payload,
        ),
    )

    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        importers.refresh_overview_data(
            {"watermark": "stable"},
            "Country",
            30,
        )


def test_legacy_and_reference_downstream_outputs_are_exactly_equal(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload, overview_stores = (
        _load_representative_references(monkeypatch)
    )
    raw_chart_entities = overview_payload["chart_entities"]
    raw_demand = snapshots.unpack_record_mapping(
        overview_payload["demand_cube"]
    )
    raw_origin = snapshots.unpack_record_mapping(
        overview_payload["origin_cube"]
    )
    selected_years = ["2025", "2026"]

    assert to_json(
        importers.update_demand_year_selector_options(
            raw_demand,
            selected_years,
        )
    ) == to_json(
        importers.update_demand_year_selector_options(
            overview_stores[2],
            selected_years,
        )
    )
    assert to_json(
        importers.update_origin_year_selector_options(
            raw_origin,
            selected_years,
        )
    ) == to_json(
        importers.update_origin_year_selector_options(
            overview_stores[3],
            selected_years,
        )
    )
    assert to_json(
        importers.update_demand_charts(
            raw_demand,
            raw_chart_entities,
            "mcm_d",
            selected_years,
            30,
        )
    ) == to_json(
        importers.update_demand_charts(
            overview_stores[2],
            overview_stores[0],
            "mcm_d",
            selected_years,
            30,
        )
    )
    assert to_json(
        importers.update_origin_continent_charts(
            raw_origin,
            raw_chart_entities,
            "mcm_d",
            selected_years,
            "absolute",
            30,
        )
    ) == to_json(
        importers.update_origin_continent_charts(
            overview_stores[3],
            overview_stores[0],
            "mcm_d",
            selected_years,
            "absolute",
            30,
        )
    )

    def capture_export(export_df, filename_prefix, sheet_name):
        return {
            "columns": list(export_df.columns),
            "records": export_df.to_dict("records"),
            "filename_prefix": filename_prefix,
            "sheet_name": sheet_name,
        }

    monkeypatch.setattr(
        importers,
        "_send_export_dataframe",
        capture_export,
    )
    assert importers.export_demand_to_excel(
        1,
        raw_demand,
        "mcm_d",
        selected_years,
        30,
    ) == importers.export_demand_to_excel(
        1,
        overview_stores[2],
        "mcm_d",
        selected_years,
        30,
    )
    assert importers.export_origin_continent_to_excel(
        1,
        raw_origin,
        "mcm_d",
        selected_years,
        "absolute",
        30,
    ) == importers.export_origin_continent_to_excel(
        1,
        overview_stores[3],
        "mcm_d",
        selected_years,
        "absolute",
        30,
    )


@pytest.mark.parametrize("corruption_mode", ["missing", "corrupt"])
def test_missing_or_corrupt_reference_shows_explicit_refresh_recovery(
    monkeypatch,
    persistent_importer_cache,
    corruption_mode,
):
    _overview_payload, overview_stores = (
        _load_representative_references(monkeypatch)
    )
    stores = snapshots._get_persistent_stores()
    reference = overview_stores[2]
    record_key = snapshots._disk_record_key(
        reference["namespace"],
        reference["source_key"],
        reference["revision"],
    )
    if corruption_mode == "missing":
        stores.cache.delete(record_key, retry=True)
    else:
        stores.cache.set(record_key, b"corrupt", retry=True)
    snapshots.clear_local_snapshots()

    options, selected = (
        importers.update_demand_year_selector_options(
            overview_stores[2],
            ["2026"],
        )
    )
    assert selected == []
    assert (
        options[0]["label"]
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    chart_notice = importers.update_demand_charts(
        overview_stores[2],
        overview_stores[0],
        "mcm_d",
        ["2026"],
        30,
    )
    assert isinstance(chart_notice, html.Div)
    assert (
        chart_notice.children
        == importers.IMPORTERS_SNAPSHOT_RECOVERY_MESSAGE
    )
    with pytest.raises(
        snapshots.SnapshotUnavailable,
        match="Click the global Refresh",
    ):
        importers.export_demand_to_excel(
            1,
            overview_stores[2],
            "mcm_d",
            ["2026"],
            30,
        )


def test_source_cache_reuses_reads_until_watermark_or_date_changes(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload = _make_overview_payload()
    source_payload = _make_source_payload()
    source_calls = 0
    derived_calls = []

    def source_builder():
        nonlocal source_calls
        source_calls += 1
        return copy.deepcopy(source_payload)

    def derived_builder(_source, classification_mode, rolling_avg_days):
        derived_calls.append((classification_mode, rolling_avg_days))
        return copy.deepcopy(overview_payload)

    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        source_builder,
    )
    monkeypatch.setattr(
        importers,
        "_build_importers_overview_payload_from_source",
        derived_builder,
    )

    source_state = {
        "watermark": "2026-07-24T00:00:00",
        "as_of_date": "2026-07-25",
    }
    first = importers.refresh_overview_data(
        source_state,
        "Country",
        30,
    )
    warm = importers.refresh_overview_data(
        source_state,
        "Country",
        30,
    )
    changed_rolling = importers.refresh_overview_data(
        source_state,
        "Country",
        45,
    )
    changed_classification = importers.refresh_overview_data(
        source_state,
        "Classification Level 1",
        45,
    )
    next_source = importers.refresh_overview_data(
        {
            "watermark": "2026-07-25T00:00:00",
            "as_of_date": "2026-07-25",
        },
        "Country",
        30,
    )

    assert source_calls == 2
    assert derived_calls == [
        ("Country", 30),
        ("Country", 45),
        ("Classification Level 1", 45),
        ("Country", 30),
    ]
    assert warm == first
    assert changed_rolling[2]["source_key"] != first[2]["source_key"]
    assert (
        changed_classification[2]["source_key"]
        != changed_rolling[2]["source_key"]
    )
    assert next_source[2]["source_key"] != first[2]["source_key"]
    assert all(
        snapshots.snapshot_is_resolvable(store)
        for store in next_source
    )


def test_explicit_refresh_forces_source_and_derived_once_per_new_state(
    monkeypatch,
    persistent_importer_cache,
):
    overview_payload = _make_overview_payload()
    source_payload = _make_source_payload()
    source_calls = 0
    derived_calls = []

    def source_builder():
        nonlocal source_calls
        source_calls += 1
        return copy.deepcopy(source_payload)

    def derived_builder(_source, classification_mode, rolling_avg_days):
        derived_calls.append((classification_mode, rolling_avg_days))
        return copy.deepcopy(overview_payload)

    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        source_builder,
    )
    monkeypatch.setattr(
        importers,
        "_build_importers_overview_payload_from_source",
        derived_builder,
    )
    triggered_id = "imp-overview-source-state-store"
    monkeypatch.setattr(
        importers,
        "_importers_overview_triggered_id",
        lambda: triggered_id,
    )

    base_state = {
        "watermark": "same-watermark",
        "as_of_date": "2026-07-25",
    }
    initial = importers.refresh_overview_data(
        base_state,
        "Country",
        30,
    )
    warm = importers.refresh_overview_data(
        base_state,
        "Country",
        30,
    )
    assert warm == initial
    assert source_calls == 1
    assert derived_calls == [("Country", 30)]

    first_refresh = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-one"},
        "Country",
        30,
    )
    second_refresh = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-two"},
        "Country",
        30,
    )
    assert source_calls == 3
    assert derived_calls == [
        ("Country", 30),
        ("Country", 30),
        ("Country", 30),
    ]
    assert first_refresh[2]["source_key"] != initial[2]["source_key"]
    assert (
        second_refresh[2]["source_key"]
        != first_refresh[2]["source_key"]
    )
    assert all(
        store["namespace"] == importers.IMPORTERS_OVERVIEW_NAMESPACE
        for store in first_refresh + second_refresh
    )

    triggered_id = "imp-overview-classification-mode"
    classification = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-two"},
        "Classification Level 1",
        30,
    )
    assert source_calls == 3
    assert derived_calls[-1] == ("Classification Level 1", 30)

    triggered_id = "imp-overview-rolling-window-days-input"
    rolling = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-two"},
        "Classification Level 1",
        45,
    )
    assert source_calls == 3
    assert derived_calls[-1] == ("Classification Level 1", 45)

    replay = importers.refresh_overview_data(
        {**base_state, "refresh_token": "refresh-two"},
        "Classification Level 1",
        45,
    )
    assert replay == rolling
    assert source_calls == 3
    assert len(derived_calls) == 5
    assert all(
        snapshots.snapshot_is_resolvable(store)
        for store in classification
    )


def test_source_state_refresh_token_is_unique_and_excluded_from_key(
    monkeypatch,
):
    monkeypatch.setattr(
        importers,
        "_fetch_importers_source_watermark",
        lambda: pd.Timestamp("2026-07-25T00:00:00"),
    )
    is_refresh = False
    monkeypatch.setattr(
        importers,
        "_was_global_refresh_triggered",
        lambda: is_refresh,
    )

    initial = importers.load_importers_overview_source_state(0)
    assert initial["refresh_token"] is None

    is_refresh = True
    first = importers.load_importers_overview_source_state(1)
    second = importers.load_importers_overview_source_state(2)
    assert first["refresh_token"]
    assert second["refresh_token"]
    assert first["refresh_token"] != second["refresh_token"]
    assert (
        importers._importers_source_snapshot_key(initial)
        == importers._importers_source_snapshot_key(first)
        == importers._importers_source_snapshot_key(second)
    )


def test_period_callback_builds_renders_and_exports_populated_payload(
    monkeypatch,
    persistent_importer_cache,
):
    period_callback = GLOBAL_CALLBACK_MAP[
        "imp-overview-period-data-store.data"
    ]
    period_inputs = {
        (item["id"], item["property"])
        for item in period_callback["inputs"]
    }
    assert (
        "imp-overview-table-entities-store",
        "data",
    ) in period_inputs
    assert (
        "global-refresh-button",
        "n_clicks",
    ) in period_inputs
    assert period_callback["state"] == [{
        "id": "imp-overview-source-state-store",
        "property": "data",
    }]

    _overview_payload, overview_stores = (
        _load_representative_references(monkeypatch)
    )
    scoped_trade_calls = []

    def fetch_scoped_trades(*_args, **kwargs):
        scoped_trade_calls.append(kwargs)
        return pd.DataFrame({"scope": [1]})

    summary_df = pd.DataFrame([{
        "continent": "Asia",
        "country": "Qatar",
        "Q2'26": 90.0,
        "Jun'26": 30.0,
        "30D": 30.0,
        "W29'26": 7.0,
        "7D": 7.0,
    }])
    monkeypatch.setattr(
        importers,
        "_fetch_importer_scoped_trades",
        fetch_scoped_trades,
    )
    monkeypatch.setattr(
        importers,
        "_filter_scoped_trades_for_entity",
        lambda scoped_df, *_args: scoped_df,
    )
    monkeypatch.setattr(
        importers,
        "build_importer_origin_summary_from_scoped_trades",
        lambda *_args, **_kwargs: summary_df.copy(),
    )

    current_payload = importers.refresh_period_data(
        overview_stores[1],
        "Country",
        "origin_shipping_region",
        30,
        "show_all",
        0,
        {"watermark": "2026-07-24T00:00:00"},
    )
    assert snapshots.is_snapshot_reference(current_payload)
    resolved_payload = snapshots.resolve_snapshot(
        current_payload,
        importers.engine,
        expected_namespace=importers.IMPORTERS_PERIOD_NAMESPACE,
    )
    assert resolved_payload["active_grouping_mode"] == "show_all"
    assert len(resolved_payload["show_all"]) == 1
    assert resolved_payload["show_all"][0]["records"] == (
        summary_df.to_dict("records")
    )
    assert scoped_trade_calls == [{
        "delivered_only": True,
        "include_destination_context": True,
        "selected_destination_aggregation": "country",
    }]

    current_component, current_display = (
        importers.update_period_analysis_table(
            current_payload,
            [],
            overview_stores[1],
            "mcm_d",
            30,
            "show_all",
            "absolute",
            "levels",
            5,
            3,
            3,
        )
    )
    assert current_display
    assert isinstance(current_component, html.Div)
    assert current_component.className == (
        "importer-period-table-shell"
    )
    assert current_display[0]["Importer"] == importers.IMPORTER_GLOBAL_LABEL

    monkeypatch.setattr(
        importers,
        "_send_export_dataframe",
        lambda export_df, filename_prefix, sheet_name: {
            "records": export_df.to_dict("records"),
            "filename_prefix": filename_prefix,
            "sheet_name": sheet_name,
        },
    )
    export_result = importers.export_period_analysis_to_excel(
        1,
        current_display,
        "origin_shipping_region",
        30,
        "absolute",
        "levels",
    )
    assert export_result["records"] == current_display
    assert export_result["sheet_name"] == "Period Analysis"


def test_source_build_uses_two_catalog_reads_and_defers_scoped_sql(
    monkeypatch,
):
    barrier = threading.Barrier(2)
    loader_threads = set()
    read_calls = []
    mapping_source = pd.DataFrame([
        {
            "country_name": "Alpha Display",
            "country": "Alpha",
            "continent": "Asia",
            "subcontinent": "Eastern Asia",
            "basin": "Pacific",
            "country_classification_level1": "Core",
            "country_classification": "Tier 1",
            "shipping_region": "North Asia",
        },
        {
            "country_name": "Beta Display",
            "country": "Beta",
            "continent": "Europe",
            "subcontinent": "Northern Europe",
            "basin": "Atlantic",
            "country_classification_level1": "Growth",
            "country_classification": "Tier 2",
            "shipping_region": "NWE",
        },
    ])
    catalog_ranking_source = pd.DataFrame({
        "destination_country_name": ["Alpha", "Beta"],
        "avg_30d_mcmd": [12.5, np.nan],
    })

    def load_catalog_ranking():
        read_calls.append("catalog_ranking")
        loader_threads.add(threading.get_ident())
        barrier.wait(timeout=2)
        return catalog_ranking_source.copy()

    def load_mappings():
        read_calls.append("mappings")
        loader_threads.add(threading.get_ident())
        barrier.wait(timeout=2)
        return mapping_source.copy()

    monkeypatch.setattr(
        importers,
        "_fetch_importers_catalog_ranking_source_df",
        load_catalog_ranking,
    )
    monkeypatch.setattr(
        importers,
        "_fetch_importers_mapping_source_df",
        load_mappings,
    )
    payload = importers._build_importers_source_payload()

    assert sorted(read_calls) == [
        "catalog_ranking",
        "mappings",
    ]
    assert len(loader_threads) == 2
    assert payload["scoped_trades_df"].empty
    assert payload["catalog_df"].to_dict("records") == [
        {
            "destination_country_name": "Alpha",
            "country": "Alpha",
            "country_display": "Alpha Display",
            "continent": "Asia",
            "subcontinent": "Eastern Asia",
            "basin": "Pacific",
            "country_classification_level1": "Core",
            "country_classification": "Tier 1",
            "shipping_region": "North Asia",
        },
        {
            "destination_country_name": "Beta",
            "country": "Beta",
            "country_display": "Beta Display",
            "continent": "Europe",
            "subcontinent": "Northern Europe",
            "basin": "Atlantic",
            "country_classification_level1": "Growth",
            "country_classification": "Tier 2",
            "shipping_region": "NWE",
        },
    ]
    pd.testing.assert_frame_equal(
        payload["ranking_df"],
        pd.DataFrame({
            "destination_country_name": ["Alpha"],
            "avg_30d_mcmd": [12.5],
            "country_display": ["Alpha Display"],
        }),
        check_dtype=True,
        check_exact=True,
    )


def test_mapping_lookup_preserves_all_russia_and_singapore_aliases():
    mapping_source = pd.DataFrame([
        {
            "country_name": "Russia",
            "country": "Russia",
            "continent": "Europe",
            "shipping_region": "Russia",
            "basin": "Atlantic Basin",
            "subcontinent": "Eastern Europe",
            "country_classification_level1": "Russia",
            "country_classification": "Russia",
        },
        {
            "country_name": "Russia",
            "country": "Russian Federation",
            "continent": "Europe",
            "shipping_region": "Russia",
            "basin": "Atlantic Basin",
            "subcontinent": "Eastern Europe",
            "country_classification_level1": "Russia",
            "country_classification": "Russia",
        },
        {
            "country_name": "Singapore",
            "country": "Singapore",
            "continent": "Asia",
            "shipping_region": "SE Asia",
            "basin": "Pacific Basin",
            "subcontinent": "Southeast Asia",
            "country_classification_level1": "South East Asia",
            "country_classification": "SE Asia",
        },
        {
            "country_name": "Singapore",
            "country": "Singapore Republic",
            "continent": "Asia",
            "shipping_region": "SE Asia",
            "basin": "Pacific Basin",
            "subcontinent": "Southeast Asia",
            "country_classification_level1": "South East Asia",
            "country_classification": "SE Asia",
        },
    ])

    lookup = importers._build_importer_mapping_lookup_from_source(
        mapping_source
    ).set_index("mapping_key")

    assert {
        "Russia",
        "Russian Federation",
        "Singapore",
        "Singapore Republic",
    }.issubset(lookup.index)
    assert (
        lookup.loc[
            "Russian Federation",
            "origin_classification_level1",
        ]
        == "Russia"
    )
    assert (
        lookup.loc[
            "Singapore Republic",
            "origin_classification_level1",
        ]
        == "South East Asia"
    )


def test_source_and_derived_keys_have_only_approved_dependencies():
    source_state = {
        "watermark": "revision-a",
        "as_of_date": "2026-07-25",
    }
    base_key = importers._importers_source_snapshot_key(source_state)
    assert base_key == importers._importers_source_snapshot_key({
        **source_state,
        "classification_mode": "Country",
        "rolling_avg_days": 30,
        "refresh_token": "excluded-from-key",
        "ignored": "value",
    })
    assert base_key != importers._importers_source_snapshot_key({
        **source_state,
        "watermark": "revision-b",
    })
    assert base_key != importers._importers_source_snapshot_key({
        **source_state,
        "as_of_date": "2026-07-26",
    })

    source_reference = {
        "namespace": importers.IMPORTERS_SOURCE_NAMESPACE,
        "source_key": base_key,
        "revision": "source-revision-a",
    }
    derived_key = importers._importers_overview_snapshot_key(
        source_reference,
        "Country",
        30,
    )
    assert derived_key != importers._importers_overview_snapshot_key(
        source_reference,
        "Country",
        45,
    )
    assert derived_key != importers._importers_overview_snapshot_key(
        source_reference,
        "Classification Level 1",
        30,
    )
    assert derived_key != importers._importers_overview_snapshot_key(
        {**source_reference, "revision": "source-revision-b"},
        "Country",
        30,
    )


def test_year_selectors_resolve_compact_metadata_without_decoding_cubes(
    monkeypatch,
    persistent_importer_cache,
):
    payload, stores = _load_representative_references(monkeypatch)
    expected_demand = importers.update_demand_year_selector_options(
        snapshots.unpack_record_mapping(payload["demand_cube"]),
        ["2026"],
    )
    expected_origin = importers.update_origin_year_selector_options(
        snapshots.unpack_record_mapping(payload["origin_cube"]),
        ["2026"],
    )

    def fail_if_cube_is_resolved(_charts_data):
        raise AssertionError("selector decompressed the chart cube")

    monkeypatch.setattr(
        importers,
        "_resolve_importers_chart_store",
        fail_if_cube_is_resolved,
    )
    assert to_json(importers.update_demand_year_selector_options(
        stores[2],
        ["2026"],
    )) == to_json(expected_demand)
    assert to_json(importers.update_origin_year_selector_options(
        stores[3],
        ["2026"],
    )) == to_json(expected_origin)


def test_chart_callbacks_prepare_each_entity_once_and_keep_public_signatures(
    monkeypatch,
):
    payload = _make_overview_payload(entity_count=4)
    entities = payload["chart_entities"]
    demand = snapshots.unpack_record_mapping(payload["demand_cube"])
    origin = snapshots.unpack_record_mapping(payload["origin_cube"])

    assert list(inspect.signature(
        importers.create_importer_demand_chart
    ).parameters) == [
        "data",
        "volume_metric",
        "selected_years",
        "rolling_avg_days",
    ]
    assert list(inspect.signature(
        importers.get_importer_demand_chart_header_metrics
    ).parameters) == [
        "data",
        "volume_metric",
        "selected_years",
        "rolling_avg_days",
    ]
    assert list(inspect.signature(
        importers.create_importer_origin_continent_chart
    ).parameters) == [
        "data",
        "chart_type",
        "volume_metric",
        "selected_years",
    ]
    assert list(inspect.signature(
        importers._calculate_origin_continent_kpis
    ).parameters) == [
        "data",
        "chart_type",
        "volume_metric",
        "selected_years",
    ]

    demand_calls = 0
    original_demand_prepare = (
        importers._prepare_importer_demand_chart_dataframe
    )

    def count_demand_prepare(*args, **kwargs):
        nonlocal demand_calls
        demand_calls += 1
        return original_demand_prepare(*args, **kwargs)

    monkeypatch.setattr(
        importers,
        "_prepare_importer_demand_chart_dataframe",
        count_demand_prepare,
    )
    importers.update_demand_charts(
        demand,
        entities,
        "mcm_d",
        ["2025", "2026"],
        30,
    )
    assert demand_calls == len(entities)

    origin_calls = 0
    original_origin_prepare = (
        importers._prepare_importer_origin_chart_dataframe
    )

    def count_origin_prepare(*args, **kwargs):
        nonlocal origin_calls
        origin_calls += 1
        return original_origin_prepare(*args, **kwargs)

    monkeypatch.setattr(
        importers,
        "_prepare_importer_origin_chart_dataframe",
        count_origin_prepare,
    )
    importers.update_origin_continent_charts(
        origin,
        entities,
        "mcm_d",
        ["2025", "2026"],
        "absolute",
        30,
    )
    assert origin_calls == len(entities)


def test_overview_callback_is_triggered_once_through_source_state():
    assert importers.IMPORTERS_OVERVIEW_NAMESPACE == (
        "importers-overview-v3"
    )
    callback_key = next(
        key
        for key in GLOBAL_CALLBACK_MAP
        if "imp-overview-chart-entities-store.data" in key
    )
    callback_inputs = GLOBAL_CALLBACK_MAP[callback_key]["inputs"]
    assert callback_inputs == [
        {
            "id": "imp-overview-source-state-store",
            "property": "data",
        },
        {
            "id": "imp-overview-classification-mode",
            "property": "value",
        },
        {
            "id": "imp-overview-rolling-window-days-input",
            "property": "value",
        },
    ]


@pytest.mark.parametrize("pool_size", [1, 2, 4])
def test_importer_overview_loader_single_flight_at_pool_sizes(
    monkeypatch,
    persistent_importer_cache,
    pool_size,
):
    payload = _make_overview_payload()
    source_payload = _make_source_payload()
    source_build_calls = 0
    derived_build_calls = 0
    build_lock = threading.Lock()

    def source_builder():
        nonlocal source_build_calls
        with build_lock:
            source_build_calls += 1
        time.sleep(0.05)
        return copy.deepcopy(source_payload)

    def derived_builder(*_args):
        nonlocal derived_build_calls
        with build_lock:
            derived_build_calls += 1
        time.sleep(0.05)
        return copy.deepcopy(payload)

    monkeypatch.setattr(
        importers,
        "_build_importers_source_payload",
        source_builder,
    )
    monkeypatch.setattr(
        importers,
        "_build_importers_overview_payload_from_source",
        derived_builder,
    )

    def load():
        return importers.refresh_overview_data(
            {"watermark": "concurrency-stable-source"},
            "Country",
            30,
        )

    with ThreadPoolExecutor(max_workers=pool_size) as executor:
        results = list(executor.map(lambda _index: load(), range(pool_size)))

    assert source_build_calls == 1
    assert derived_build_calls == 1
    assert all(result == results[0] for result in results)
    assert all(
        snapshots.snapshot_is_resolvable(store)
        for store in results[0]
    )
    assert len(to_json(results[0]).encode("utf-8")) < 50_000


def test_importer_warm_loader_response_benchmark_gate(
    monkeypatch,
    persistent_importer_cache,
):
    template = _make_overview_payload(
        entity_count=12,
        years=(2022, 2023, 2024, 2025, 2026),
        points_per_year=120,
        continents=("Africa", "Americas", "Asia", "Europe"),
    )

    def build_payload(*_args):
        return copy.deepcopy(template)

    _install_overview_builders(monkeypatch, template)
    baseline_payload = build_payload()

    app = Flask(__name__)

    @app.get("/baseline")
    def baseline_response():
        response_payload = (
            baseline_payload["chart_entities"],
            baseline_payload["table_entities"],
            snapshots.unpack_record_mapping(
                baseline_payload["demand_cube"]
            ),
            snapshots.unpack_record_mapping(
                baseline_payload["origin_cube"]
            ),
        )
        return Response(
            to_json(response_payload),
            mimetype="application/json",
        )

    @app.get("/reference")
    def reference_response():
        return Response(
            to_json(
                importers.refresh_overview_data(
                    {"watermark": "stable-source"},
                    "Country",
                    30,
                )
            ),
            mimetype="application/json",
        )

    client = app.test_client()
    assert client.get("/baseline").status_code == 200
    assert client.get("/reference").status_code == 200

    timings = {"baseline": [], "reference": []}
    sizes = {"baseline": [], "reference": []}
    for _iteration in range(6):
        for mode in ("baseline", "reference"):
            started = time.perf_counter()
            response = client.get(f"/{mode}")
            elapsed = time.perf_counter() - started
            assert response.status_code == 200
            timings[mode].append(elapsed)
            sizes[mode].append(len(response.data))

    baseline_seconds = statistics.median(timings["baseline"])
    reference_seconds = statistics.median(timings["reference"])
    improvement = (
        baseline_seconds - reference_seconds
    ) / baseline_seconds
    reference_bytes = max(sizes["reference"])
    baseline_bytes = max(sizes["baseline"])
    print(
        "IMPORTERS_WARM_BENCHMARK "
        f"baseline_ms={baseline_seconds * 1000:.2f} "
        f"reference_ms={reference_seconds * 1000:.2f} "
        f"improvement_pct={improvement * 100:.2f} "
        f"baseline_bytes={baseline_bytes} "
        f"reference_bytes={reference_bytes}"
    )

    assert reference_bytes < 50_000
    assert improvement > 0.10
